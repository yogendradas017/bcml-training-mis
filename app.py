import os
import re
import secrets
import sqlite3
import io
from datetime import date, datetime
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, flash, send_file, g)
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'bcml-tms-2627-xK9pQ')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_default_db = os.path.join(BASE_DIR, 'data', 'training.db')
_env_db = os.environ.get('DATABASE_PATH')
if _env_db:
    try:
        os.makedirs(os.path.dirname(_env_db), exist_ok=True)
        DB_PATH = _env_db
    except Exception:
        DB_PATH = _default_db  # fallback if path is invalid
else:
    DB_PATH = _default_db
TEMP_UPLOAD_DIR = os.path.join(BASE_DIR, 'data', 'tmp_uploads')

PLANTS = [
    {'id': 1,  'name': 'Balrampur',  'unit_code': 'BCM'},
    {'id': 2,  'name': 'Babhnan',    'unit_code': 'BBN'},
    {'id': 3,  'name': 'Rauzagaon',  'unit_code': 'RCM'},
    {'id': 4,  'name': 'Maizapur',   'unit_code': 'MZP'},
    {'id': 5,  'name': 'Mankapur',   'unit_code': 'MCM'},
    {'id': 6,  'name': 'Gularia',    'unit_code': 'GCM'},
    {'id': 7,  'name': 'Tulsipur',   'unit_code': 'TCM'},
    {'id': 8,  'name': 'Kumbhi',     'unit_code': 'KCM'},
    {'id': 9,  'name': 'Haidergarh', 'unit_code': 'HCM'},
    {'id': 10, 'name': 'Akbarpur',   'unit_code': 'ACM'},
]
PLANT_MAP = {p['id']: p for p in PLANTS}

PROG_TYPES   = ['Behavioural/Leadership', 'Cane', 'Commercial', 'EHS/HR', 'IT', 'Technical']
MODES        = ['Classroom', 'OJT', 'SOP', 'Online']
LEVELS       = ['General', 'Specialized']
AUDIENCES    = ['Blue Collared', 'White Collared', 'Common']
STATUSES     = ['To Be Planned', 'Conducted', 'Re-Scheduled', 'Cancelled']
INT_EXT      = ['Internal', 'External', 'Online']
MONTHS_FY    = ['April','May','June','July','August','September',
                'October','November','December','January','February','March']
CAL_NEW      = ['Calendar Program', 'New Program']
GENDERS      = ['Male', 'Female', 'Others']
TYPE_ABBREV  = {
    'Behavioural/Leadership': 'BEH', 'Cane': 'CAN', 'Commercial': 'COM',
    'EHS/HR': 'EHS', 'IT': 'ITC', 'Technical': 'TEC'
}

# ─── DB ──────────────────────────────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        db.close()

def _migrate_tni_unique(db):
    """Add UNIQUE(plant_id, emp_code, programme_name) to tni if not present."""
    has_unique = db.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='index' AND tbl_name='tni' "
        "AND sql LIKE '%emp_code%' AND sql LIKE '%programme_name%'"
    ).fetchone()[0]
    if has_unique:
        return
    # Recreate tni with unique constraint, keeping one row per unique key (latest id)
    db.executescript('''
        CREATE TABLE tni_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plant_id INTEGER NOT NULL,
            emp_code TEXT NOT NULL,
            programme_name TEXT NOT NULL,
            prog_type TEXT, mode TEXT, target_month TEXT,
            planned_hours REAL DEFAULT 0,
            created_at TEXT DEFAULT (date('now')),
            UNIQUE(plant_id, emp_code, programme_name)
        );
        INSERT OR IGNORE INTO tni_new
            (plant_id, emp_code, programme_name, prog_type, mode, target_month, planned_hours, created_at)
        SELECT plant_id, emp_code, programme_name, prog_type, mode, target_month, planned_hours, created_at
        FROM tni ORDER BY id;
        DROP TABLE tni;
        ALTER TABLE tni_new RENAME TO tni;
        CREATE INDEX IF NOT EXISTS idx_tni_plant ON tni(plant_id);
        CREATE INDEX IF NOT EXISTS idx_tni_dedup ON tni(plant_id, emp_code, programme_name);
    ''')

def _cleanse_master_spelling(db):
    """Fix known misspelled words inside programme_master entries using _WORD_FIXES."""
    rows = db.execute('SELECT id, name FROM programme_master').fetchall()
    for row in rows:
        cleaned = _smart_title(_apply_word_fixes(row['name']))
        if cleaned != row['name']:
            # Check no duplicate already exists before renaming
            clash = db.execute(
                'SELECT id FROM programme_master WHERE plant_id=(SELECT plant_id FROM programme_master WHERE id=?) AND LOWER(name)=LOWER(?) AND id!=?',
                (row['id'], cleaned, row['id'])
            ).fetchone()
            if not clash:
                db.execute('UPDATE programme_master SET name=? WHERE id=?', (cleaned, row['id']))
    db.commit()

def _migrate_tni_source(db):
    """Add source column to tni table if not present (migration for existing DBs)."""
    cols = [row[1] for row in db.execute("PRAGMA table_info(tni)").fetchall()]
    if 'source' not in cols:
        db.execute("ALTER TABLE tni ADD COLUMN source TEXT DEFAULT 'TNI Driven'")
        db.commit()

def _cleanse_programme_names(db, plant_id=None):
    """Fix programme name casing/typos in tni, emp_training, calendar against master lists.
    Exact case-insensitive fixes applied automatically; fuzzy fixes only at 0.88+ confidence.
    Returns dict of {plant_id: {fixed, merged}} counts.
    """
    from difflib import get_close_matches as gcm
    report = {}
    plants = [plant_id] if plant_id else [
        r[0] for r in db.execute('SELECT DISTINCT plant_id FROM programme_master').fetchall()]

    for pid in plants:
        master = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (pid,)).fetchall()]
        if not master:
            continue
        master_lower_map = {m.lower(): m for m in master}
        master_lower = list(master_lower_map.keys())
        fixed = 0; merged = 0

        for table in ('tni', 'emp_training', 'calendar'):
            rows = db.execute(f'SELECT id, programme_name FROM {table} WHERE plant_id=?', (pid,)).fetchall()
            for row in rows:
                raw = row['programme_name'] or ''
                if not raw:
                    continue
                raw_lower = raw.lower()

                # Exact case-insensitive match
                if raw_lower in master_lower_map:
                    canonical = master_lower_map[raw_lower]
                    if canonical != raw:
                        db.execute(f'UPDATE {table} SET programme_name=? WHERE id=?', (canonical, row['id']))
                        fixed += 1
                else:
                    # High-confidence fuzzy only (0.88+) to avoid wrong bulk changes
                    m = gcm(raw_lower, master_lower, n=1, cutoff=0.88)
                    if m:
                        canonical = master_lower_map[m[0]]
                        db.execute(f'UPDATE {table} SET programme_name=? WHERE id=?', (canonical, row['id']))
                        fixed += 1

        # Merge duplicate TNI entries created by the name fix
        dupes = db.execute('''
            SELECT emp_code, programme_name, MIN(id) as keep_id, COUNT(*) as cnt
            FROM tni WHERE plant_id=?
            GROUP BY emp_code, programme_name HAVING cnt > 1
        ''', (pid,)).fetchall()
        for d in dupes:
            db.execute('DELETE FROM tni WHERE plant_id=? AND emp_code=? AND programme_name=? AND id != ?',
                       (pid, d['emp_code'], d['programme_name'], d['keep_id']))
            merged += 1

        db.commit()
        report[pid] = {'fixed': fixed, 'merged': merged}

    return report

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    with open(os.path.join(BASE_DIR, 'schema.sql')) as f:
        db.executescript(f.read())
    _migrate_tni_unique(db)
    _migrate_tni_source(db)
    # Seed plants
    for p in PLANTS:
        db.execute('INSERT OR IGNORE INTO plants(id,name,unit_code) VALUES(?,?,?)',
                   (p['id'], p['name'], p['unit_code']))
    # Default users
    users = [('central', 'bcml@1234', 'central', None),
             ('admin',   'admin@bcml', 'admin',   None)]
    for p in PLANTS:
        users.append((p['name'].lower(), 'bcml@1234', 'spoc', p['id']))
    for u in users:
        db.execute('INSERT OR IGNORE INTO users(username,password,role,plant_id) VALUES(?,?,?,?)',
                   (u[0], generate_password_hash(u[1]), u[2], u[3]))
    # Seed Balrampur master programmes (plant_id=1) from hardcoded list if empty
    if db.execute('SELECT COUNT(*) FROM programme_master WHERE plant_id=1').fetchone()[0] == 0:
        for name in MASTER_PROGRAMMES:
            db.execute('INSERT OR IGNORE INTO programme_master(plant_id,name) VALUES(1,?)', (name,))
    db.commit()
    # Fix spelling in master list first, then cleanse TNI/calendar against it
    _cleanse_master_spelling(db)
    _cleanse_programme_names(db)
    db.commit()
    db.close()

# ─── AUTH ─────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def spoc_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') not in ('spoc', 'admin'):
            flash('Access denied.', 'danger')
            return redirect(url_for('central_dashboard'))
        if session.get('role') == 'admin' and not session.get('plant_id'):
            flash('Please select a plant first to access SPOC functions.', 'warning')
            return redirect(url_for('central_dashboard'))
        return f(*args, **kwargs)
    return decorated

def central_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') not in ('central', 'admin'):
            flash('Access denied.', 'danger')
            return redirect(url_for('spoc_dashboard'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('This action requires admin access.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') in ('central', 'admin'):
        return redirect(url_for('central_dashboard'))
    return redirect(url_for('spoc_dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        if user and check_password_hash(user['password'], password):
            session.clear()
            session['user_id']  = user['id']
            session['username'] = user['username']
            session['role']     = user['role']
            session['plant_id'] = user['plant_id']
            if user['plant_id']:
                session['plant_name'] = PLANT_MAP[user['plant_id']]['name']
                session['unit_code']  = PLANT_MAP[user['plant_id']]['unit_code']
            if user['role'] in ('central', 'admin'):
                return redirect(url_for('central_dashboard'))
            return redirect(url_for('spoc_dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── ADMIN PLANT SELECTOR ─────────────────────────────────────────────────────

@app.route('/admin/plant/<int:plant_id>')
def admin_select_plant(plant_id):
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    plant = PLANT_MAP.get(plant_id)
    if not plant:
        flash('Plant not found.', 'danger')
        return redirect(url_for('central_dashboard'))
    session['plant_id']   = plant['id']
    session['plant_name'] = plant['name']
    session['unit_code']  = plant['unit_code']
    flash(f"Now viewing as SPOC for {plant['name']}. Use 'Switch Plant' in sidebar to go back.", 'info')
    return redirect(url_for('spoc_dashboard'))

@app.route('/admin/clear-plant')
def admin_clear_plant():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    session.pop('plant_id',   None)
    session.pop('plant_name', None)
    session.pop('unit_code',  None)
    return redirect(url_for('central_dashboard'))

# ─── SPOC DASHBOARD ───────────────────────────────────────────────────────────

@app.route('/dashboard')
@spoc_required
def spoc_dashboard():
    plant_id = session['plant_id']
    db = get_db()
    stats = {
        'total_emp':    db.execute('SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)).fetchone()[0],
        'blue_collar':  db.execute("SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='Blue Collared'", (plant_id,)).fetchone()[0],
        'white_collar': db.execute("SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='White Collared'", (plant_id,)).fetchone()[0],
        'tni_count':    db.execute('SELECT COUNT(DISTINCT emp_code || "|" || programme_name) FROM tni WHERE plant_id=?', (plant_id,)).fetchone()[0],
        'sessions':     db.execute('SELECT COUNT(*) FROM calendar WHERE plant_id=?', (plant_id,)).fetchone()[0],
        'conducted':    db.execute("SELECT COUNT(*) FROM calendar WHERE plant_id=? AND status='Conducted'", (plant_id,)).fetchone()[0],
        'trainings':    db.execute('SELECT COUNT(*) FROM emp_training WHERE plant_id=?', (plant_id,)).fetchone()[0],
        'manhours':     db.execute('SELECT COALESCE(SUM(hrs),0) FROM emp_training WHERE plant_id=?', (plant_id,)).fetchone()[0],
    }
    return render_template('dashboard.html', stats=stats)

# ─── EMPLOYEE MASTER ──────────────────────────────────────────────────────────

@app.route('/employees')
@spoc_required
def employees():
    plant_id = session['plant_id']
    db = get_db()
    show_exited = request.args.get('show_exited', '0') == '1'
    if show_exited:
        emps = db.execute('SELECT * FROM employees WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()
    else:
        emps = db.execute('SELECT * FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name', (plant_id,)).fetchall()
    # Recently exited (within 7 days) — shown as undo strip in active view
    recent_exited = db.execute(
        "SELECT * FROM employees WHERE plant_id=? AND is_active=0 AND exit_date >= date('now','-7 days') ORDER BY exit_date DESC",
        (plant_id,)).fetchall()
    return render_template('employees.html', employees=emps, show_exited=show_exited,
                           recent_exited=recent_exited,
                           genders=GENDERS, today=str(date.today()))

@app.route('/employees/add', methods=['POST'])
@spoc_required
def add_employee():
    plant_id = session['plant_id']
    f = request.form
    db = get_db()
    collar = normalise_collar(f.get('collar',''))
    try:
        db.execute('''INSERT INTO employees
            (plant_id,emp_code,name,designation,grade,collar,department,section,
             category,gender,physically_handicapped,remarks)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
            (plant_id, f['emp_code'].strip(), f['name'].strip(),
             f.get('designation',''), f.get('grade',''), collar,
             f.get('department',''), f.get('section',''), f.get('category',''),
             f.get('gender',''), f.get('physically_handicapped','No'),
             f.get('remarks','')))
        db.commit()
        flash(f"Employee {f['name'].strip()} added successfully.", 'success')
    except sqlite3.IntegrityError:
        flash(f"Employee code {f['emp_code'].strip()} already exists.", 'danger')
    return redirect(url_for('employees'))

@app.route('/employees/<int:emp_id>/exit', methods=['POST'])
@spoc_required
def exit_employee(emp_id):
    db = get_db()
    exit_date   = request.form.get('exit_date', str(date.today()))
    exit_reason = request.form.get('exit_reason', '')
    if exit_date > str(date.today()):
        flash('Exit date cannot be a future date.', 'danger')
        return redirect(url_for('employees'))
    if not exit_reason.strip():
        flash('Exit reason is mandatory for attrition analysis.', 'danger')
        return redirect(url_for('employees'))
    db.execute('UPDATE employees SET is_active=0, exit_date=?, exit_reason=? WHERE id=? AND plant_id=?',
               (exit_date, exit_reason, emp_id, session['plant_id']))
    db.commit()
    flash('Employee marked as exited.', 'warning')
    return redirect(url_for('employees'))

@app.route('/employees/<int:emp_id>/reactivate', methods=['POST'])
@spoc_required
def reactivate_employee(emp_id):
    db = get_db()
    db.execute('UPDATE employees SET is_active=1, exit_date=NULL, exit_reason=NULL WHERE id=? AND plant_id=?',
               (emp_id, session['plant_id']))
    db.commit()
    flash('Employee reactivated.', 'success')
    return redirect(url_for('employees') + '?show_exited=1')

# ─── TNI TRACKING ─────────────────────────────────────────────────────────────

@app.route('/tni')
@spoc_required
def tni():
    plant_id = session['plant_id']
    db = get_db()
    total = db.execute(
        'SELECT COUNT(DISTINCT emp_code || "|" || programme_name) FROM tni WHERE plant_id=?',
        (plant_id,)).fetchone()[0]
    emps = db.execute('SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name', (plant_id,)).fetchall()
    programmes = _get_programme_names(plant_id, db)
    depts = [r[0] for r in db.execute(
        'SELECT DISTINCT department FROM employees WHERE plant_id=? AND department IS NOT NULL AND department != "" ORDER BY department',
        (plant_id,)).fetchall()]

    # Count programme names in TNI that don't exactly match the master list
    master_lower = set(r[0].lower() for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=?', (plant_id,)).fetchall())
    dirty_names = []
    if master_lower:
        tni_names = [r[0] for r in db.execute(
            'SELECT DISTINCT programme_name FROM tni WHERE plant_id=?', (plant_id,)).fetchall()]
        dirty_names = [n for n in tni_names if n.lower() not in master_lower]

    # Count duplicate TNI rows (same emp_code + programme_name)
    dup_count = db.execute('''
        SELECT COALESCE(SUM(cnt - 1), 0)
        FROM (SELECT COUNT(*) as cnt FROM tni WHERE plant_id=?
              GROUP BY emp_code, programme_name HAVING cnt > 1)
    ''', (plant_id,)).fetchone()[0]

    return render_template('tni.html', total=total,
                           employees=emps, programmes=programmes,
                           prog_types=PROG_TYPES, modes=MODES, months=MONTHS_FY,
                           departments=depts, dirty_names=dirty_names,
                           dup_count=dup_count)

def _tni_filters(plant_id):
    """Build WHERE clause + params from current request args for TNI queries."""
    q         = request.args.get('q', '').strip()
    collar    = request.args.get('collar', '')
    dept      = request.args.get('dept', '')
    ptype     = request.args.get('type', '')
    mode      = request.args.get('mode', '')
    month     = request.args.get('month', '')
    completed = request.args.get('completed', '')

    where  = ['t.plant_id=?']
    params = [plant_id]
    if collar: where.append('e.collar=?');          params.append(collar)
    if dept:   where.append('e.department=?');       params.append(dept)
    if ptype:  where.append('t.prog_type=?');        params.append(ptype)
    if mode:   where.append('t.mode=?');             params.append(mode)
    # target_month filter removed
    if q:
        where.append('(COALESCE(e.name,"") LIKE ? OR t.emp_code LIKE ? OR t.programme_name LIKE ?)')
        like = f'%{q}%'; params += [like, like, like]
    if completed == 'Yes':
        where.append('et.emp_code IS NOT NULL')
    elif completed == 'Pending':
        where.append('et.emp_code IS NULL')
    return ' AND '.join(where), params

@app.route('/tni/data')
@spoc_required
def tni_data():
    plant_id = session['plant_id']
    db       = get_db()
    page     = max(1, int(request.args.get('page', 1)))
    per_page = 30

    where_clause, params = _tni_filters(plant_id)

    join_sql = f'''
        FROM tni t
        JOIN (SELECT MAX(id) as max_id FROM tni
              WHERE plant_id=?
              GROUP BY emp_code, programme_name) dedup ON t.id = dedup.max_id
        LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        LEFT JOIN (SELECT DISTINCT emp_code, programme_name
                   FROM emp_training WHERE plant_id=?) et
               ON et.emp_code=t.emp_code AND et.programme_name=t.programme_name
        WHERE {where_clause}
    '''
    # join_sql has 2 leading plant_id params (dedup subquery + emp_training subquery)
    join_params = [plant_id, plant_id] + params

    # total count — fast SQL COUNT
    total = db.execute(f'SELECT COUNT(*) {join_sql}', join_params).fetchone()[0]

    offset = (page - 1) * per_page
    rows_raw = db.execute(
        f'''SELECT t.id, t.emp_code, t.programme_name, t.prog_type, t.mode,
                   t.planned_hours, t.source,
                   e.name, e.collar, e.department,
                   CASE WHEN et.emp_code IS NOT NULL THEN 'Yes' ELSE 'Pending' END AS completed
            {join_sql}
            ORDER BY t.id DESC LIMIT ? OFFSET ?''',
        join_params + [per_page, offset]
    ).fetchall()

    rows = [{
        'id':             r['id'],
        'emp_code':       r['emp_code'],
        'name':           r['name'] or r['emp_code'],
        'collar':         r['collar'] or '',
        'department':     r['department'] or '',
        'programme_name': r['programme_name'],
        'prog_type':      r['prog_type'] or '',
        'mode':           r['mode'] or '',
        'planned_hours':  r['planned_hours'],
        'source':         r['source'] or 'TNI Driven',
        'completed':      r['completed'],
        'delete_url':     url_for('delete_tni', tni_id=r['id']),
    } for r in rows_raw]

    return jsonify({'total': total, 'page': page, 'per_page': per_page, 'rows': rows})

@app.route('/tni/add', methods=['POST'])
@spoc_required
def add_tni():
    plant_id = session['plant_id']
    f = request.form
    db = get_db()
    source    = f.get('source', 'TNI Driven').strip() or 'TNI Driven'
    prog_name = _canonical_prog(f['programme_name'].strip(), plant_id, db)
    db.execute(
        '''INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours,source)
           VALUES(?,?,?,?,?,?,?)''',
        (plant_id, f['emp_code'], prog_name,
         f.get('prog_type',''), f.get('mode',''),
         float(f.get('planned_hours') or 0), source)
    )
    # Non-TNI sources: ensure programme exists in master list
    if source != 'TNI Driven':
        db.execute('INSERT OR IGNORE INTO programme_master(plant_id,name) VALUES(?,?)',
                   (plant_id, prog_name))
    db.commit()
    flash('TNI entry added.', 'success')
    return redirect(url_for('tni'))

_NON_TNI_SOURCES = ('Corp Driven', 'Unit Driven', 'Compliance Driven')

@app.route('/tni/<int:tni_id>/set-source', methods=['POST'])
@spoc_required
def tni_set_source(tni_id):
    plant_id = session['plant_id']
    data     = request.get_json(silent=True) or {}
    source   = (data.get('source') or '').strip()
    if source not in _NON_TNI_SOURCES:
        return jsonify({'ok': False, 'error': 'Invalid source'}), 400
    db  = get_db()
    row = db.execute('SELECT programme_name FROM tni WHERE id=? AND plant_id=?',
                     (tni_id, plant_id)).fetchone()
    if not row:
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    db.execute('UPDATE tni SET source=? WHERE id=? AND plant_id=?',
               (source, tni_id, plant_id))
    # Ensure programme is in master list when given a non-TNI source
    db.execute('INSERT OR IGNORE INTO programme_master(plant_id,name) VALUES(?,?)',
               (plant_id, row['programme_name']))
    db.commit()
    return jsonify({'ok': True})

@app.route('/tni/<int:tni_id>/delete', methods=['POST'])
@spoc_required
def delete_tni(tni_id):
    db = get_db()
    db.execute('DELETE FROM tni WHERE id=? AND plant_id=?', (tni_id, session['plant_id']))
    db.commit()
    if _is_ajax():
        return '', 204
    flash('TNI entry deleted.', 'warning')
    return redirect(url_for('tni'))

@app.route('/tni/bulk-delete', methods=['POST'])
@spoc_required
def tni_bulk_delete():
    plant_id = session['plant_id']
    db = get_db()
    if request.form.get('select_all') == '1':
        where_clause, params = _tni_filters(plant_id)
        join_sql = f'''
            FROM tni t
            LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            LEFT JOIN (SELECT DISTINCT emp_code, programme_name
                       FROM emp_training WHERE plant_id=?) et
                   ON et.emp_code=t.emp_code AND et.programme_name=t.programme_name
            WHERE {where_clause}
        '''
        count = db.execute(f'SELECT COUNT(*) {join_sql}', [plant_id] + params).fetchone()[0]
        if count:
            db.execute(f'DELETE FROM tni WHERE id IN (SELECT t.id {join_sql})', [plant_id] + params)
            db.commit()
            flash(f'{count} TNI entries deleted.', 'warning')
    else:
        ids = request.form.getlist('ids[]')
        if ids:
            # Delete in chunks of 900 to stay under SQLite variable limit
            deleted = 0
            for i in range(0, len(ids), 900):
                chunk = ids[i:i+900]
                ph = ','.join('?' * len(chunk))
                db.execute(f'DELETE FROM tni WHERE id IN ({ph}) AND plant_id=?', chunk + [plant_id])
                deleted += len(chunk)
            db.commit()
            flash(f'{deleted} TNI entries deleted.', 'warning')
    return redirect(url_for('tni'))

@app.route('/tni/cleanse', methods=['GET', 'POST'])
@spoc_required
def tni_cleanse():
    plant_id = session['plant_id']
    db = get_db()

    # quick=1 applies immediately without preview (used from warning banner)
    if request.args.get('quick') == '1':
        result = _cleanse_programme_names(db, plant_id=plant_id)
        r = result.get(plant_id, {'fixed': 0, 'merged': 0})
        flash(f'Data cleanse complete: {r["fixed"]} programme name(s) corrected, '
              f'{r["merged"]} duplicate(s) merged.', 'success')
        return redirect(url_for('tni'))

    if request.method == 'POST':
        result = _cleanse_programme_names(db, plant_id=plant_id)
        r = result.get(plant_id, {'fixed': 0, 'merged': 0})
        flash(f'Data cleanse complete: {r["fixed"]} programme name(s) corrected, '
              f'{r["merged"]} duplicate(s) merged.', 'success')
        return redirect(url_for('tni'))

    # Preview: show what will be fixed before confirming
    from difflib import get_close_matches as gcm
    master = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()]
    preview = []
    if master:
        master_lower_map = {m.lower(): m for m in master}
        master_lower = list(master_lower_map.keys())
        seen = set()
        rows = db.execute(
            'SELECT DISTINCT programme_name FROM tni WHERE plant_id=?', (plant_id,)).fetchall()
        for row in rows:
            raw = row['programme_name'] or ''
            if not raw or raw in seen:
                continue
            seen.add(raw)
            raw_lower = raw.lower()
            if raw_lower in master_lower_map:
                canonical = master_lower_map[raw_lower]
                if canonical != raw:
                    preview.append({'original': raw, 'fixed': canonical, 'how': 'Case correction'})
            else:
                m = gcm(raw_lower, master_lower, n=1, cutoff=0.88)
                if m:
                    canonical = master_lower_map[m[0]]
                    preview.append({'original': raw, 'fixed': canonical, 'how': 'Spelling correction'})

    return render_template('tni_cleanse.html', preview=preview)

@app.route('/tni/duplicates')
@spoc_required
def tni_duplicates():
    plant_id = session['plant_id']
    db = get_db()
    rows = db.execute('''
        SELECT t.emp_code,
               MAX(e.name) as emp_name,
               t.programme_name,
               COUNT(*) as cnt,
               GROUP_CONCAT(t.mode, ' / ') as modes,
               GROUP_CONCAT(t.id) as ids
        FROM tni t
        LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=?
        GROUP BY t.emp_code, t.programme_name
        HAVING cnt > 1
        ORDER BY cnt DESC, emp_name
    ''', (plant_id,)).fetchall()
    total_extra = sum(r['cnt'] - 1 for r in rows)
    return render_template('tni_duplicates.html', rows=rows, total_extra=total_extra)

@app.route('/tni/duplicates/delete', methods=['POST'])
@spoc_required
def tni_duplicates_delete():
    plant_id = session['plant_id']
    db = get_db()
    deleted = 0
    rows = db.execute('''
        SELECT GROUP_CONCAT(id ORDER BY id) as ids
        FROM tni
        WHERE plant_id=?
        GROUP BY emp_code, programme_name
        HAVING COUNT(*) > 1
    ''', (plant_id,)).fetchall()
    for r in rows:
        id_list = [int(x) for x in r['ids'].split(',')]
        keep = id_list[0]
        remove = id_list[1:]
        ph = ','.join('?' * len(remove))
        db.execute(f'DELETE FROM tni WHERE id IN ({ph}) AND plant_id=?', remove + [plant_id])
        deleted += len(remove)
    db.commit()
    flash(f'{deleted} duplicate TNI entries removed.', 'success')
    return redirect(url_for('tni'))

# ─── PROGRAMME MASTER ─────────────────────────────────────────────────────────

@app.route('/programme-master')
@spoc_required
def programme_master():
    plant_id = session['plant_id']
    db = get_db()
    db.execute('''CREATE TABLE IF NOT EXISTS programme_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plant_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        prog_type TEXT, mode TEXT,
        created_at TEXT DEFAULT (date('now')),
        UNIQUE(plant_id, name))''')
    progs = db.execute(
        'SELECT * FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()
    return render_template('programme_master.html', progs=progs,
                           prog_types=PROG_TYPES, modes=MODES)

@app.route('/programme-master/add', methods=['POST'])
@spoc_required
def programme_master_add():
    plant_id = session['plant_id']
    name = request.form.get('name', '').strip()
    prog_type = request.form.get('prog_type', '').strip()
    mode = request.form.get('mode', '').strip()
    if not name:
        flash('Programme name is required.', 'danger')
        return redirect(url_for('programme_master'))
    db = get_db()
    # Ensure table exists (handles live DBs that predate this migration)
    db.execute('''CREATE TABLE IF NOT EXISTS programme_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plant_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        prog_type TEXT, mode TEXT,
        created_at TEXT DEFAULT (date('now')),
        UNIQUE(plant_id, name))''')
    try:
        existing = db.execute('SELECT id FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
                              (plant_id, name)).fetchone()
        if existing:
            db.execute('UPDATE programme_master SET prog_type=?, mode=? WHERE id=?',
                       (prog_type or None, mode or None, existing['id']))
            db.commit()
            flash(f'"{name}" updated.', 'success')
        else:
            db.execute('INSERT INTO programme_master(plant_id,name,prog_type,mode) VALUES(?,?,?,?)',
                       (plant_id, name, prog_type or None, mode or None))
            db.commit()
            flash(f'"{name}" added to master list.', 'success')
    except Exception as e:
        import logging; logging.error(f'programme_master_add error: {e}')
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('programme_master'))

@app.route('/programme-master/<int:prog_id>/rename', methods=['POST'])
@spoc_required
def programme_master_rename(prog_id):
    plant_id = session['plant_id']
    new_name = request.json.get('name', '').strip() if request.is_json else request.form.get('name', '').strip()
    if not new_name:
        return jsonify(ok=False, error='Name cannot be empty'), 400
    db = get_db()
    clash = db.execute('SELECT id FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?) AND id!=?',
                       (plant_id, new_name, prog_id)).fetchone()
    if clash:
        return jsonify(ok=False, error=f'"{new_name}" already exists in master list'), 409
    db.execute('UPDATE programme_master SET name=? WHERE id=? AND plant_id=?', (new_name, prog_id, plant_id))
    db.commit()
    return jsonify(ok=True, name=new_name)

@app.route('/programme-master/<int:prog_id>/delete', methods=['POST'])
@spoc_required
def programme_master_delete(prog_id):
    plant_id = session['plant_id']
    db = get_db()
    db.execute('DELETE FROM programme_master WHERE id=? AND plant_id=?', (prog_id, plant_id))
    db.commit()
    flash('Programme removed from master list.', 'warning')
    return redirect(url_for('programme_master'))

@app.route('/programme-master/bulk-delete', methods=['POST'])
@spoc_required
def programme_master_bulk_delete():
    plant_id = session['plant_id']
    ids = request.form.getlist('ids[]')
    if not ids:
        flash('No programmes selected.', 'warning')
        return redirect(url_for('programme_master'))
    try:
        ids_int = [int(i) for i in ids]
    except ValueError:
        flash('Invalid selection.', 'danger')
        return redirect(url_for('programme_master'))
    db = get_db()
    db.executemany('DELETE FROM programme_master WHERE id=? AND plant_id=?',
                   [(i, plant_id) for i in ids_int])
    db.commit()
    flash(f'{len(ids_int)} programme(s) deleted from master list.', 'warning')
    return redirect(url_for('programme_master'))

@app.route('/programme-master/bulk', methods=['POST'])
@spoc_required
def programme_master_bulk():
    plant_id = session['plant_id']
    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('programme_master'))
    try:
        import pandas as _pd
        fname = f.filename.lower()
        if fname.endswith('.csv'):
            df = _pd.read_csv(f, dtype=str).fillna('')
        else:
            df = _pd.read_excel(f, dtype=str).fillna('')
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('programme_master'))

    # Auto-detect name column
    cols_lower = {c.strip().lower(): c for c in df.columns}
    name_col = next((cols_lower[k] for k in ['programme name','program name','name','training name','course name'] if k in cols_lower), None)
    type_col = next((cols_lower[k] for k in ['type of programme','type','prog type','programme type'] if k in cols_lower), None)
    mode_col = next((cols_lower[k] for k in ['mode','training mode','delivery mode'] if k in cols_lower), None)

    if not name_col:
        flash(f'Could not find a "Programme Name" column. Columns found: {", ".join(df.columns.tolist()[:10])}', 'danger')
        return redirect(url_for('programme_master'))

    db = get_db()
    inserted = skipped = 0
    for _, row in df.iterrows():
        name = str(row.get(name_col, '')).strip()
        if not name or name.lower() in ('nan', 'none', ''):
            continue
        prog_type = str(row.get(type_col, '')).strip() if type_col else ''
        mode      = str(row.get(mode_col, '')).strip() if mode_col else ''
        try:
            db.execute('INSERT INTO programme_master(plant_id,name,prog_type,mode) VALUES(?,?,?,?)',
                       (plant_id, name, prog_type or None, mode or None))
            inserted += 1
        except Exception:
            skipped += 1
    db.commit()
    flash(f'{inserted} programmes added. {skipped} already existed (skipped).', 'success' if inserted else 'warning')
    return redirect(url_for('programme_master'))

@app.route('/programme-master/template')
@spoc_required
def programme_master_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Programme Master'
    hdr_fill = PatternFill('solid', fgColor='1A1F35')
    headers = ['Programme Name', 'Type of Programme', 'Mode']
    widths  = [45, 25, 15]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    # Sample rows
    samples = [('Fire Safety', 'EHS/HR', 'Classroom'),
               ('5-S Management', 'EHS/HR', 'OJT'),
               ('Advanced Excel', 'IT', 'Classroom')]
    for row in samples:
        ws.append(row)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_type = DataValidation(type='list', formula1=f'"{",".join(PROG_TYPES)}"', allow_blank=True)
    dv_mode = DataValidation(type='list', formula1=f'"{",".join(MODES)}"', allow_blank=True)
    dv_type.sqref = 'B2:B500'; dv_mode.sqref = 'C2:C500'
    ws.add_data_validation(dv_type); ws.add_data_validation(dv_mode)
    ws.freeze_panes = 'A2'
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Programme_Master_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/programme-master/export')
@spoc_required
def programme_master_export():
    plant_id = session['plant_id']
    plant_name = session.get('plant_name', 'Plant')
    db = get_db()
    progs = db.execute('SELECT name, prog_type, mode, created_at FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Programme Master'
    hdr_fill = PatternFill('solid', fgColor='1A1F35')
    headers = ['#', 'Programme Name', 'Type of Programme', 'Mode', 'Added On']
    widths  = [5, 45, 25, 15, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    for i, r in enumerate(progs, 1):
        ws.append([i, r['name'], r['prog_type'] or '', r['mode'] or '', r['created_at'] or ''])
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'Programme_Master_{plant_name}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/tni/template')
@spoc_required
def tni_template():
    from openpyxl.worksheet.datavalidation import DataValidation

    plant_id = session['plant_id']
    db       = get_db()

    # fetch active employees for this plant
    emps = db.execute(
        'SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name',
        (plant_id,)).fetchall()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data Entry ──────────────────────────────────────
    ws = wb.active
    ws.title = 'TNI Data'

    # header
    headers = ['Employee Code', 'Employee Name (auto)', 'Programme Name',
               'Type of Programme', 'Mode', 'Target Month', 'Planned Hours']
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # column widths
    widths = [18, 28, 36, 24, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # lock header row
    ws.freeze_panes = 'A2'

    # ── Sheet 2: Employee List (hidden) ──────────────────────────
    ws_emp = wb.create_sheet('_EmpList')
    ws_emp.sheet_state = 'hidden'
    for r, emp in enumerate(emps, 1):
        c = ws_emp.cell(row=r, column=1, value=str(emp['emp_code']))
        c.number_format = '@'  # force text so VLOOKUP TEXT() match works
        ws_emp.cell(row=r, column=2, value=emp['name'])

    emp_count = len(emps)

    # ── Sheet 3: Valid Values (hidden) ───────────────────────────
    ws_vals = wb.create_sheet('_ValidValues')
    ws_vals.sheet_state = 'hidden'
    prog_types = PROG_TYPES
    modes      = MODES
    months     = MONTHS_FY
    for r, v in enumerate(prog_types, 1): ws_vals.cell(row=r, column=1, value=v)
    for r, v in enumerate(modes, 1):      ws_vals.cell(row=r, column=2, value=v)
    for r, v in enumerate(months, 1):     ws_vals.cell(row=r, column=3, value=v)

    # ── Sheet 4: Programme Master (hidden) ───────────────────────
    master_progs = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
    ).fetchall()] or MASTER_PROGRAMMES
    ws_prog = wb.create_sheet('_ProgList')
    ws_prog.sheet_state = 'hidden'
    for r, v in enumerate(master_progs, 1):
        ws_prog.cell(row=r, column=1, value=v)

    # named ranges for validation
    wb.defined_names['EmpCodes']   = openpyxl.workbook.defined_name.DefinedName(
        'EmpCodes',   attr_text=f'_EmpList!$A$1:$A${emp_count}')
    wb.defined_names['ProgTypes']  = openpyxl.workbook.defined_name.DefinedName(
        'ProgTypes',  attr_text=f'_ValidValues!$A$1:$A${len(prog_types)}')
    wb.defined_names['ModeList']   = openpyxl.workbook.defined_name.DefinedName(
        'ModeList',   attr_text=f'_ValidValues!$B$1:$B${len(modes)}')
    wb.defined_names['MonthList']  = openpyxl.workbook.defined_name.DefinedName(
        'MonthList',  attr_text=f'_ValidValues!$C$1:$C${len(months)}')
    wb.defined_names['ProgList']   = openpyxl.workbook.defined_name.DefinedName(
        'ProgList',   attr_text=f'_ProgList!$A$1:$A${len(master_progs)}')

    # ── Data Validations ─────────────────────────────────────────
    max_rows = 2000

    # Col A — Employee Code dropdown
    dv_emp = DataValidation(type='list', formula1='EmpCodes', allow_blank=False,
                            showErrorMessage=True, errorTitle='Invalid Employee',
                            error='Select a valid Employee Code from the dropdown.',
                            showDropDown=False)
    dv_emp.sqref = f'A2:A{max_rows}'
    ws.add_data_validation(dv_emp)

    # Col C — Programme Name dropdown (from master list)
    if master_progs:
        dv_prog = DataValidation(type='list', formula1='ProgList', allow_blank=False,
                                 showErrorMessage=True, errorTitle='Programme Not in Master',
                                 error='Select a programme from the dropdown. If missing, add it to Programme Master first.',
                                 showDropDown=False)
        dv_prog.sqref = f'C2:C{max_rows}'
        ws.add_data_validation(dv_prog)

    # Col D — Type of Programme
    dv_type = DataValidation(type='list', formula1='ProgTypes', allow_blank=False,
                             showErrorMessage=True, errorTitle='Invalid Type',
                             error='Select from: ' + ', '.join(prog_types),
                             showDropDown=False)
    dv_type.sqref = f'D2:D{max_rows}'
    ws.add_data_validation(dv_type)

    # Col E — Mode
    dv_mode = DataValidation(type='list', formula1='ModeList', allow_blank=True,
                             showErrorMessage=True, errorTitle='Invalid Mode',
                             error='Select from: ' + ', '.join(modes),
                             showDropDown=False)
    dv_mode.sqref = f'E2:E{max_rows}'
    ws.add_data_validation(dv_mode)

    # Col F — Month
    dv_month = DataValidation(type='list', formula1='MonthList', allow_blank=True,
                              showErrorMessage=True, errorTitle='Invalid Month',
                              error='Select a valid month from the dropdown.',
                              showDropDown=False)
    dv_month.sqref = f'F2:F{max_rows}'
    ws.add_data_validation(dv_month)

    # Col G — Hours: decimal > 0
    dv_hrs = DataValidation(type='decimal', operator='greaterThan', formula1='0',
                            allow_blank=True,
                            showErrorMessage=True, errorTitle='Invalid Hours',
                            error='Enter a number greater than 0, e.g. 4 or 2.5')
    dv_hrs.sqref = f'G2:G{max_rows}'
    ws.add_data_validation(dv_hrs)

    # Col B — auto-fill name via VLOOKUP
    # TEXT(A{r},"0") normalises both text and number entries so type mismatch doesn't cause "Not Found"
    for r in range(2, max_rows + 1):
        ws.cell(row=r, column=2).value = (
            f'=IF(A{r}="","",IFERROR(VLOOKUP(TEXT(A{r},"0"),_EmpList!$A:$B,2,0),"Not Found"))'
        )
        ws.cell(row=r, column=2).font      = Font(color='1F4E79', italic=True)
        ws.cell(row=r, column=2).fill      = PatternFill('solid', fgColor='EFF6FF')
        ws.cell(row=r, column=7).value     = 0  # default hours

    # instruction row
    ws.cell(row=1, column=1).comment = None  # clear any existing
    note_cell = ws.cell(row=max_rows + 2, column=1,
                        value='⚠ Do not add columns. Do not delete hidden sheets. Column B auto-fills from Employee Code.')
    note_cell.font = Font(italic=True, color='FF0000', size=9)
    ws.merge_cells(f'A{max_rows+2}:G{max_rows+2}')

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out, download_name='TNI_Bulk_Upload_Template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/tni/bulk', methods=['POST'])
@spoc_required
def tni_bulk_upload():
    plant_id = session['plant_id']
    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('tni'))
    try:
        df = _read_upload_file(f)
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('tni'))

    db       = get_db()
    inserted = 0
    errors   = []
    for i, row in df.iterrows():
        emp_code  = _clean(row, ['employee code', 'emp code', 'empcode', 'employee_code'])
        prog_name = _clean(row, ['programme name', 'program name', 'programme_name', 'training name'])
        prog_type = _clean(row, ['type of programme', 'type', 'prog type', 'programme type'])
        mode      = _clean(row, ['mode'])
        hours     = _safe_float(_clean(row, ['planned hours', 'hours', 'hrs'])) or 0

        if not emp_code or not prog_name:
            errors.append(f'Row {i+2}: Employee Code and Programme Name are required.')
            continue
        emp = db.execute('SELECT 1 FROM employees WHERE emp_code=? AND plant_id=? AND is_active=1',
                         (emp_code, plant_id)).fetchone()
        if not emp:
            errors.append(f'Row {i+2}: Employee {emp_code} not found in your plant.')
            continue
        prog_name = _canonical_prog(prog_name, plant_id, db)
        db.execute('INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours) VALUES(?,?,?,?,?,?)',
                   (plant_id, emp_code, prog_name, prog_type, mode, hours))
        inserted += 1
    db.commit()
    if errors:
        if inserted:
            flash(f'Bulk upload complete: {inserted} TNI entries added. {len(errors)} rows had errors — downloading error report.', 'warning')
        return _error_excel_response(errors, inserted, 'TNI_Upload_Errors.xlsx')
    flash(f'Bulk upload complete: {inserted} TNI entries added successfully.', 'success')
    return redirect(url_for('tni'))

@app.route('/tni/fresh-upload', methods=['GET', 'POST'])
@admin_required
def tni_fresh_upload():
    plant_id = session['plant_id']
    db = get_db()

    if request.method == 'GET':
        return render_template('tni_fresh_upload.html')

    # ── Step 2: User clicked Confirm ──────────────────────────────────────────
    confirm_token = request.form.get('confirm')
    if confirm_token:
        ext      = session.get('fresh_upload_ext', '.xlsx')
        tmp_path = os.path.join(TEMP_UPLOAD_DIR, f'tni_fresh_{confirm_token}{ext}')
        if not os.path.exists(tmp_path):
            flash('Session expired — please re-upload the file.', 'danger')
            return redirect(url_for('tni_fresh_upload'))
        try:
            df = _read_upload_file_path(tmp_path)
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('tni_fresh_upload'))

        result = _process_fresh_tni(df, plant_id, db)
        rows   = result['valid_rows']

        db.execute('DELETE FROM tni WHERE plant_id=?', (plant_id,))
        db.execute('DELETE FROM programme_master WHERE plant_id=?', (plant_id,))

        for r in rows:
            db.execute(
                'INSERT INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours) VALUES(?,?,?,?,?,?)',
                (plant_id, r['emp_code'], r['programme_name'], r['prog_type'], r['mode'], r['hours'])
            )
        for prog in result['unique_progs']:
            db.execute('INSERT OR IGNORE INTO programme_master(plant_id,name) VALUES(?,?)', (plant_id, prog))
        db.commit()

        try:
            os.remove(tmp_path)
        except Exception:
            pass

        flash(
            f'Fresh upload complete: {len(rows)} TNI entries saved. '
            f'{len(result["unique_progs"])} unique programmes are now your master list.',
            'success'
        )
        return redirect(url_for('tni'))

    # ── Step 1: File uploaded → parse, show preview ────────────────────────────
    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('tni_fresh_upload'))

    os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)
    token    = secrets.token_hex(16)
    ext      = os.path.splitext(f.filename)[1].lower() or '.xlsx'
    tmp_path = os.path.join(TEMP_UPLOAD_DIR, f'tni_fresh_{token}{ext}')
    f.save(tmp_path)
    session['fresh_upload_token'] = token
    session['fresh_upload_ext']   = ext

    try:
        df = _read_upload_file_path(tmp_path)
    except Exception as e:
        try: os.remove(tmp_path)
        except Exception: pass
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('tni_fresh_upload'))

    result = _process_fresh_tni(df, plant_id, db)
    return render_template(
        'tni_fresh_upload.html',
        preview=True,
        token=token,
        total_rows=result['total_rows'],
        valid_rows=result['valid_rows'],
        error_rows=result['error_rows'],
        name_corrections=result['name_corrections'],
        unique_progs=result['unique_progs'],
        duplicate_count=result['duplicate_count'],
    )

# ─── TRAINING CALENDAR ────────────────────────────────────────────────────────

@app.route('/calendar')
@spoc_required
def training_calendar():
    plant_id = session['plant_id']
    db = get_db()

    # Auto-update statuses from 2C
    _sync_calendar_from_2c(plant_id, db)

    sessions = db.execute('SELECT * FROM calendar WHERE plant_id=? ORDER BY id DESC', (plant_id,)).fetchall()
    # TNI programme demand — distinct employees per programme (deduped)
    demand_map = {}
    for row in db.execute('SELECT programme_name, COUNT(DISTINCT emp_code) as cnt FROM tni WHERE plant_id=? GROUP BY programme_name', (plant_id,)):
        demand_map[row['programme_name']] = row['cnt']

    # Dropdown uses master list only (canonical names); TNI demand shown via badge
    master_programmes = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()] or MASTER_PROGRAMMES
    all_cal_programmes = master_programmes
    # TNI programmes still needed for demand_map and coverage panel
    tni_programmes = [r[0] for r in db.execute(
        'SELECT DISTINCT programme_name FROM tni WHERE plant_id=? ORDER BY programme_name', (plant_id,))]
    tni_set = set(p.lower() for p in tni_programmes)

    # Coverage summary: per programme — demand vs total planned PAX vs conducted PAX
    cov_rows = []
    pax_map = {}   # programme → {planned_pax, conducted_pax, sessions}
    for s in sessions:
        p = s['programme_name']
        if p not in pax_map:
            pax_map[p] = {'sessions': 0, 'planned_pax': 0, 'conducted_pax': 0}
        pax_map[p]['sessions']     += 1
        pax_map[p]['planned_pax']  += (s['planned_pax'] or 0)
        if s['status'] == 'Conducted':
            pax_map[p]['conducted_pax'] += (s['planned_pax'] or 0)
    for prog, d in demand_map.items():
        pm = pax_map.get(prog, {'sessions': 0, 'planned_pax': 0, 'conducted_pax': 0})
        planned_pax = pm['planned_pax']
        gap         = max(0, d - planned_pax)
        pct         = min(100, round(planned_pax / d * 100)) if d > 0 else 0
        cov_rows.append({'name': prog, 'demand': d,
                         'sessions': pm['sessions'], 'planned_pax': planned_pax,
                         'conducted_pax': pm['conducted_pax'],
                         'gap': gap, 'pct': pct,
                         'over': max(0, planned_pax - d)})
    cov_rows.sort(key=lambda x: x['gap'], reverse=True)  # biggest gap first

    return render_template('calendar.html', sessions=sessions, demand_map=demand_map,
                           tni_programmes=tni_programmes,
                           all_cal_programmes=all_cal_programmes, cov_rows=cov_rows,
                           prog_types=PROG_TYPES, modes=MODES, levels=LEVELS,
                           audiences=AUDIENCES, months=MONTHS_FY, statuses=STATUSES)

@app.route('/calendar/add', methods=['POST'])
@spoc_required
def add_calendar():
    plant_id = session['plant_id']
    f = request.form
    db = get_db()
    prog_name = _canonical_prog(f['programme_name'].strip(), plant_id, db)
    prog_type = f.get('prog_type', '')

    prog_code    = _get_or_create_prog_code(plant_id, prog_name, prog_type, db)
    session_code = _new_session_code(plant_id, prog_code, db)

    db.execute('''INSERT INTO calendar
        (plant_id,prog_code,session_code,source,programme_name,prog_type,
         planned_month,plan_start,plan_end,time_from,time_to,duration_hrs,
         level,mode,target_audience,planned_pax,trainer_vendor,status)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (plant_id, prog_code, session_code,
         f.get('source','TNI'), prog_name, prog_type,
         f.get('planned_month',''), f.get('plan_start',''), f.get('plan_end',''),
         f.get('time_from',''), f.get('time_to',''),
         float(f.get('duration_hrs') or 0),
         f.get('level',''), f.get('mode',''), f.get('target_audience',''),
         int(f.get('planned_pax') or 0), f.get('trainer_vendor',''),
         'To Be Planned'))
    db.commit()
    flash(f'Session {session_code} added to calendar.', 'success')
    return redirect(url_for('training_calendar'))

@app.route('/calendar/<int:cal_id>/delete', methods=['POST'])
@spoc_required
def delete_calendar(cal_id):
    db = get_db()
    db.execute('DELETE FROM calendar WHERE id=? AND plant_id=?', (cal_id, session['plant_id']))
    db.commit()
    if _is_ajax():
        return '', 204
    flash('Calendar entry deleted.', 'warning')
    return redirect(url_for('training_calendar'))

@app.route('/calendar/<int:cal_id>/edit', methods=['POST'])
@spoc_required
def edit_calendar(cal_id):
    plant_id = session['plant_id']
    f = request.form
    db = get_db()
    db.execute('''UPDATE calendar SET
        programme_name=?, prog_type=?, source=?, planned_month=?,
        plan_start=?, plan_end=?, time_from=?, time_to=?,
        duration_hrs=?, level=?, mode=?, target_audience=?,
        planned_pax=?, trainer_vendor=?, status=?
        WHERE id=? AND plant_id=?''',
        (_canonical_prog(f.get('programme_name','').strip(), plant_id, db), f.get('prog_type',''),
         f.get('source','TNI'), f.get('planned_month',''),
         f.get('plan_start',''), f.get('plan_end',''),
         f.get('time_from',''), f.get('time_to',''),
         float(f.get('duration_hrs') or 0), f.get('level',''),
         f.get('mode',''), f.get('target_audience',''),
         int(f.get('planned_pax') or 0), f.get('trainer_vendor',''),
         f.get('status','To Be Planned'),
         cal_id, plant_id))
    db.commit()
    flash('Session updated.', 'success')
    return redirect(url_for('training_calendar'))

@app.route('/calendar/bulk-delete', methods=['POST'])
@spoc_required
def calendar_bulk_delete():
    plant_id = session['plant_id']
    ids = request.form.getlist('ids[]')
    if ids:
        db = get_db()
        deleted = 0
        for i in range(0, len(ids), 900):
            chunk = ids[i:i+900]
            ph = ','.join('?' * len(chunk))
            db.execute(f'DELETE FROM calendar WHERE id IN ({ph}) AND plant_id=?', chunk + [plant_id])
            deleted += len(chunk)
        db.commit()
        flash(f'{deleted} calendar sessions deleted.', 'warning')
    return redirect(url_for('training_calendar'))

# ─── EMPLOYEE TRAINING (2A) ───────────────────────────────────────────────────

@app.route('/training')
@spoc_required
def emp_training():
    plant_id = session['plant_id']
    db = get_db()
    records = db.execute('''
        SELECT t.*, e.name as emp_name, e.designation, e.grade, e.collar,
               e.department, e.section
        FROM emp_training t
        LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=?
        ORDER BY t.id DESC
    ''', (plant_id,)).fetchall()

    emps = db.execute('SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name', (plant_id,)).fetchall()
    sessions_list = db.execute("SELECT session_code, programme_name FROM calendar WHERE plant_id=? ORDER BY session_code", (plant_id,)).fetchall()
    return render_template('training_2a.html', records=records, employees=emps,
                           sessions=sessions_list, months=MONTHS_FY)

@app.route('/training/add', methods=['POST'])
@spoc_required
def add_emp_training():
    plant_id = session['plant_id']
    f = request.form
    db = get_db()
    emp_code     = f['emp_code']
    session_code = f.get('session_code', '').strip()
    start_date   = f.get('start_date', '')
    end_date     = f.get('end_date', '')

    # Auto-fill from calendar
    prog_name = _canonical_prog(f.get('programme_name', '').strip(), plant_id, db)
    prog_type = level = mode = cal_new = ''
    if session_code:
        cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                         (session_code, plant_id)).fetchone()
        if cal:
            prog_name = cal['programme_name']
            prog_type = cal['prog_type']
            level     = cal['level']
            mode      = cal['mode']
            cal_new   = 'Calendar Program'
            if not start_date:
                start_date = cal['plan_start'] or ''
            if not end_date:
                end_date = cal['plan_end'] or ''

    month = _date_to_month(start_date)

    db.execute('''INSERT INTO emp_training
        (plant_id,emp_code,session_code,programme_name,start_date,end_date,
         hrs,prog_type,level,mode,cal_new,pre_rating,post_rating,venue,month)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (plant_id, emp_code, session_code, prog_name,
         start_date, end_date, float(f.get('hrs') or 0),
         prog_type, level, mode, cal_new,
         _safe_float(f.get('pre_rating')), _safe_float(f.get('post_rating')),
         f.get('venue',''), month))
    db.commit()
    flash('Training record added.', 'success')
    return redirect(url_for('emp_training'))

@app.route('/training/<int:rec_id>/delete', methods=['POST'])
@spoc_required
def delete_emp_training(rec_id):
    db = get_db()
    db.execute('DELETE FROM emp_training WHERE id=? AND plant_id=?', (rec_id, session['plant_id']))
    db.commit()
    if _is_ajax():
        return '', 204
    flash('Training record deleted.', 'warning')
    return redirect(url_for('emp_training'))

@app.route('/training/bulk-delete', methods=['POST'])
@spoc_required
def training_bulk_delete():
    plant_id = session['plant_id']
    ids = request.form.getlist('ids[]')
    if ids:
        db = get_db()
        deleted = 0
        for i in range(0, len(ids), 900):
            chunk = ids[i:i+900]
            ph = ','.join('?' * len(chunk))
            db.execute(f'DELETE FROM emp_training WHERE id IN ({ph}) AND plant_id=?', chunk + [plant_id])
            deleted += len(chunk)
        db.commit()
        flash(f'{deleted} training records deleted.', 'warning')
    return redirect(url_for('emp_training'))

@app.route('/training/template')
@spoc_required
def training_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2A_Bulk_Upload'
    headers = ['Employee Code', 'Session Code (optional)', 'Programme Name',
               'Start Date (YYYY-MM-DD)', 'End Date (YYYY-MM-DD)',
               'Hours', 'Venue', 'Pre-Session Rating (1-5)', 'Post-Session Rating (1-5)']
    hdr_fill = PatternFill('solid', start_color='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF')
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        ws.column_dimensions[get_column_letter(i)].width = 26
    samples = [
        ['21700011', 'BCM/EHS/001/B01', 'Fire Safety Training', '2026-06-10', '2026-06-10', 4, 'Training Hall', 3.5, 4.2],
        ['21101568', '', 'MS Office Basics', '2026-07-05', '2026-07-06', 8, 'Computer Lab', '', 4.0],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws['A5'] = 'NOTE: Session Code is optional. If provided, Programme Name/Type/Mode auto-fill from Calendar.'
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, download_name='2A_Training_Bulk_Upload_Template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/training/bulk', methods=['POST'])
@spoc_required
def training_bulk_upload():
    plant_id = session['plant_id']
    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('emp_training'))
    try:
        df = _read_upload_file(f)
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('emp_training'))

    db       = get_db()
    inserted = 0
    errors   = []
    for i, row in df.iterrows():
        emp_code     = _clean(row, ['employee code', 'emp code', 'empcode'])
        session_code = _clean(row, ['session code', 'session code (optional)', 'sessioncode'])
        prog_name    = _clean(row, ['programme name', 'program name', 'training name'])
        start_date   = _clean(row, ['start date', 'start date (yyyy-mm-dd)', 'startdate', 'date'])
        end_date     = _clean(row, ['end date', 'end date (yyyy-mm-dd)', 'enddate'])
        hrs          = _safe_float(_clean(row, ['hours', 'hrs', 'duration'])) or 0
        venue        = _clean(row, ['venue'])
        pre_r        = _safe_float(_clean(row, ['pre-session rating', 'pre rating', 'pre_rating']))
        post_r       = _safe_float(_clean(row, ['post-session rating', 'post rating', 'post_rating']))

        if not emp_code:
            errors.append(f'Row {i+2}: Employee Code is required.')
            continue
        emp = db.execute('SELECT 1 FROM employees WHERE emp_code=? AND plant_id=? AND is_active=1',
                         (emp_code, plant_id)).fetchone()
        if not emp:
            errors.append(f'Row {i+2}: Employee {emp_code} not found.')
            continue

        prog_type = level = mode = cal_new = ''
        if session_code:
            cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                             (session_code, plant_id)).fetchone()
            if cal:
                prog_name = prog_name or cal['programme_name']
                prog_type = cal['prog_type']
                level     = cal['level']
                mode      = cal['mode']
                cal_new   = 'Calendar Program'
                start_date = start_date or cal['plan_start'] or ''
                end_date   = end_date or cal['plan_end'] or ''

        if not prog_name:
            errors.append(f'Row {i+2}: Programme Name required (no session code matched).')
            continue

        if not cal_new:
            prog_name = _canonical_prog(prog_name, plant_id, db)
        month = _date_to_month(start_date)
        db.execute('''INSERT INTO emp_training
            (plant_id,emp_code,session_code,programme_name,start_date,end_date,
             hrs,prog_type,level,mode,cal_new,pre_rating,post_rating,venue,month)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (plant_id, emp_code, session_code, prog_name,
             start_date, end_date, hrs, prog_type, level, mode, cal_new,
             pre_r, post_r, venue, month))
        inserted += 1
    db.commit()
    if errors:
        if inserted:
            flash(f'Bulk upload complete: {inserted} records added. {len(errors)} rows had errors — downloading error report.', 'warning')
        return _error_excel_response(errors, inserted, 'Training2A_Upload_Errors.xlsx')
    flash(f'Bulk upload complete: {inserted} training records added successfully.', 'success')
    return redirect(url_for('emp_training'))

# ─── PROGRAMME DETAILS (2C) ───────────────────────────────────────────────────

@app.route('/programme')
@spoc_required
def programme_details():
    plant_id = session['plant_id']
    db = get_db()
    records = db.execute('''
        SELECT p.*,
               (SELECT COUNT(*) FROM emp_training t WHERE t.session_code=p.session_code AND t.plant_id=p.plant_id) as participants,
               (SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t WHERE t.session_code=p.session_code AND t.plant_id=p.plant_id) as man_hours
        FROM programme_details p
        WHERE p.plant_id=?
        ORDER BY p.id DESC
    ''', (plant_id,)).fetchall()

    cal_sessions = db.execute(
        "SELECT session_code, programme_name FROM calendar WHERE plant_id=? ORDER BY session_code",
        (plant_id,)).fetchall()
    return render_template('programme_2c.html', records=records,
                           cal_sessions=cal_sessions,
                           int_ext=INT_EXT, audiences=AUDIENCES, months=MONTHS_FY)

@app.route('/programme/add', methods=['POST'])
@spoc_required
def add_programme_details():
    plant_id = session['plant_id']
    f = request.form
    db = get_db()
    session_code = f['session_code'].strip()

    if db.execute('SELECT 1 FROM programme_details WHERE session_code=? AND plant_id=?',
                  (session_code, plant_id)).fetchone():
        flash(f'Session {session_code} already recorded. Edit the existing entry.', 'warning')
        return redirect(url_for('programme_details'))

    cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                     (session_code, plant_id)).fetchone()
    prog_name = cal['programme_name'] if cal else f.get('programme_name','')
    prog_type = cal['prog_type'] if cal else ''
    level     = cal['level'] if cal else ''
    cal_new   = 'Calendar Program' if cal else 'New Program'
    mode      = cal['mode'] if cal else ''
    audience  = cal['target_audience'] if cal else ''

    db.execute('''INSERT INTO programme_details
        (plant_id,session_code,programme_name,prog_type,level,cal_new,mode,
         start_date,end_date,audience,hours_actual,faculty_name,int_ext,cost,
         venue,course_feedback,faculty_feedback,trainer_fb_participants,trainer_fb_facilities)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (plant_id, session_code, prog_name, prog_type, level, cal_new, mode,
         f.get('start_date',''), f.get('end_date',''), audience,
         float(f.get('hours_actual') or 0), f.get('faculty_name',''),
         f.get('int_ext',''), float(f.get('cost') or 0),
         f.get('venue',''),
         _safe_float(f.get('course_feedback')),
         _safe_float(f.get('faculty_feedback')),
         _safe_float(f.get('trainer_fb_participants')),
         _safe_float(f.get('trainer_fb_facilities'))))
    db.commit()

    # Update calendar status
    act_date = f.get('start_date','')
    db.execute("UPDATE calendar SET status='Conducted' WHERE session_code=? AND plant_id=?",
               (session_code, plant_id))
    db.commit()
    flash(f'Programme {session_code} details saved.', 'success')
    return redirect(url_for('programme_details'))

@app.route('/programme/<int:rec_id>/delete', methods=['POST'])
@spoc_required
def delete_programme(rec_id):
    db = get_db()
    rec = db.execute('SELECT session_code FROM programme_details WHERE id=? AND plant_id=?',
                     (rec_id, session['plant_id'])).fetchone()
    if rec:
        db.execute('DELETE FROM programme_details WHERE id=? AND plant_id=?', (rec_id, session['plant_id']))
        db.execute("UPDATE calendar SET status='To Be Planned' WHERE session_code=? AND plant_id=?",
                   (rec['session_code'], session['plant_id']))
        db.commit()
    if _is_ajax():
        return '', 204
    flash('Programme record deleted.', 'warning')
    return redirect(url_for('programme_details'))

@app.route('/programme/bulk-delete', methods=['POST'])
@spoc_required
def programme_bulk_delete():
    plant_id = session['plant_id']
    ids = request.form.getlist('ids[]')
    if ids:
        ph = ','.join('?' * len(ids))
        db = get_db()
        recs = db.execute(f'SELECT session_code FROM programme_details WHERE id IN ({ph}) AND plant_id=?',
                          ids + [plant_id]).fetchall()
        for r in recs:
            db.execute("UPDATE calendar SET status='To Be Planned' WHERE session_code=? AND plant_id=?",
                       (r['session_code'], plant_id))
        db.execute(f'DELETE FROM programme_details WHERE id IN ({ph}) AND plant_id=?', ids + [plant_id])
        db.commit()
        flash(f'{len(ids)} programme records deleted.', 'warning')
    return redirect(url_for('programme_details'))

# ─── MONTHLY SUMMARY ──────────────────────────────────────────────────────────

@app.route('/summary')
@spoc_required
def monthly_summary():
    plant_id     = session['plant_id']
    sel_month    = request.args.get('month', '')
    db           = get_db()
    summary_rows = _calc_summary(plant_id, sel_month, db)
    totals       = _calc_totals(summary_rows)
    compliance   = _calc_compliance(plant_id, db)
    return render_template('summary.html', summary_rows=summary_rows,
                           totals=totals, compliance=compliance,
                           months=MONTHS_FY, selected_month=sel_month,
                           prog_types=PROG_TYPES)

# ─── CENTRAL DASHBOARD ────────────────────────────────────────────────────────

@app.route('/central')
@central_required
def central_dashboard():
    db = get_db()
    plant_summaries = []
    for p in PLANTS:
        pid = p['id']
        bc  = db.execute("SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='Blue Collared'", (pid,)).fetchone()[0]
        wc  = db.execute("SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='White Collared'", (pid,)).fetchone()[0]
        sessions_cnt  = db.execute("SELECT COUNT(*) FROM calendar WHERE plant_id=?", (pid,)).fetchone()[0]
        conducted_cnt = db.execute("SELECT COUNT(*) FROM calendar WHERE plant_id=? AND status='Conducted'", (pid,)).fetchone()[0]
        manhours = db.execute("SELECT COALESCE(SUM(hrs),0) FROM emp_training WHERE plant_id=?", (pid,)).fetchone()[0]
        bc_hrs   = db.execute("SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id WHERE t.plant_id=? AND e.collar='Blue Collared'", (pid,)).fetchone()[0]
        wc_hrs   = db.execute("SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id WHERE t.plant_id=? AND e.collar='White Collared'", (pid,)).fetchone()[0]
        bc_mandate = bc * 12
        wc_mandate = wc * 24
        bc_pct = round((bc_hrs / bc_mandate * 100), 1) if bc_mandate else 0
        wc_pct = round((wc_hrs / wc_mandate * 100), 1) if wc_mandate else 0
        plant_summaries.append({**p,
            'blue_collar': bc, 'white_collar': wc, 'total_emp': bc + wc,
            'sessions': sessions_cnt, 'conducted': conducted_cnt,
            'manhours': round(manhours, 1),
            'bc_pct': bc_pct, 'wc_pct': wc_pct
        })
    grand = {
        'total_emp': sum(p['total_emp'] for p in plant_summaries),
        'manhours':  round(sum(p['manhours'] for p in plant_summaries), 1),
        'sessions':  sum(p['sessions'] for p in plant_summaries),
        'conducted': sum(p['conducted'] for p in plant_summaries),
    }

    # Quarterly review (FY quarters)
    Q_MONTHS = [
        ('Q1 (Apr–Jun)', ['April','May','June']),
        ('Q2 (Jul–Sep)', ['July','August','September']),
        ('Q3 (Oct–Dec)', ['October','November','December']),
        ('Q4 (Jan–Mar)', ['January','February','March']),
    ]
    quarterly = []
    for qname, months in Q_MONTHS:
        ph = ','.join('?'*len(months))
        sc = db.execute(f"SELECT COUNT(*) FROM calendar WHERE status='Conducted' AND planned_month IN ({ph})", months).fetchone()[0]
        mh = db.execute(f"SELECT COALESCE(SUM(hrs),0) FROM emp_training WHERE month IN ({ph})", months).fetchone()[0]
        # Per-plant breakdown for this quarter
        plant_q = []
        for p in plant_summaries:
            pid = p['id']
            sc_p = db.execute(f"SELECT COUNT(*) FROM calendar WHERE plant_id=? AND status='Conducted' AND planned_month IN ({ph})", [pid]+months).fetchone()[0]
            mh_p = db.execute(f"SELECT COALESCE(SUM(hrs),0) FROM emp_training WHERE plant_id=? AND month IN ({ph})", [pid]+months).fetchone()[0]
            plant_q.append({'name': p['name'], 'unit_code': p['unit_code'], 'id': p['id'],
                            'sessions': sc_p, 'manhours': round(mh_p, 1)})
        quarterly.append({'quarter': qname, 'sessions': sc, 'manhours': round(mh, 1), 'plants': plant_q})

    return render_template('central.html', plants=plant_summaries, grand=grand, quarterly=quarterly)

@app.route('/central/plant/<int:plant_id>')
@central_required
def central_plant_view(plant_id):
    if plant_id not in PLANT_MAP:
        flash('Plant not found.', 'danger')
        return redirect(url_for('central_dashboard'))
    plant     = PLANT_MAP[plant_id]
    db        = get_db()
    sel_month = request.args.get('month', '')
    summary_rows = _calc_summary(plant_id, sel_month, db)
    totals       = _calc_totals(summary_rows)
    compliance   = _calc_compliance(plant_id, db)
    return render_template('central_plant.html', plant=plant,
                           summary_rows=summary_rows, totals=totals,
                           compliance=compliance, months=MONTHS_FY,
                           selected_month=sel_month)

# ─── EXPORT TO EXCEL ──────────────────────────────────────────────────────────

@app.route('/export/<int:plant_id>')
@login_required
def export_excel(plant_id):
    if session.get('role') == 'spoc' and session.get('plant_id') != plant_id:
        flash('Access denied.', 'danger')
        return redirect(url_for('spoc_dashboard'))
    plant = PLANT_MAP.get(plant_id)
    if not plant:
        flash('Plant not found.', 'danger')
        return redirect(url_for('index'))

    from openpyxl.cell import WriteOnlyCell
    db  = get_db()
    fy  = '2026-27'
    # write_only=True streams row-by-row → much less RAM
    wb  = openpyxl.Workbook(write_only=True)

    H_FONT = Font(bold=True, color='FFFFFF', size=10)
    H_FILL = PatternFill('solid', fgColor='1F4E79')
    H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    T_FONT  = Font(bold=True, size=11)

    def hc(ws, val):
        """Styled header cell."""
        c = WriteOnlyCell(ws, value=val)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = H_ALIGN
        return c

    def tc(ws, val):
        """Title cell."""
        c = WriteOnlyCell(ws, value=val)
        c.font = T_FONT
        return c

    pname = plant['name'].upper()

    # ── Sheet 1: Employee Master ──────────────────────────────────────────
    ws1 = wb.create_sheet('EMP_MASTER')
    ws1.append([tc(ws1, 'BALRAMPUR CHINI MILLS LIMITED')])
    ws1.append([tc(ws1, f'{pname} — EMPLOYEE MASTER | FY {fy}')])
    ws1.append([])
    ws1.append([hc(ws1, h) for h in
        ['Sr.','Emp Code','Name','Designation','Grade','Collar',
         'Department','Section','Category','Gender','PH',
         'Exit Date','Exit Reason','Remarks']])
    for r, e in enumerate(db.execute(
            'SELECT * FROM employees WHERE plant_id=? ORDER BY name', (plant_id,)), 1):
        ws1.append([r, e['emp_code'], e['name'], e['designation'] or '',
                    e['grade'] or '', e['collar'] or '', e['department'] or '',
                    e['section'] or '', e['category'] or '', e['gender'] or '',
                    e['physically_handicapped'] or '',
                    e['exit_date'] or '', e['exit_reason'] or '', e['remarks'] or ''])

    # ── Sheet 2: TNI ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet('TNI_Tracking')
    ws2.append([tc(ws2, 'BALRAMPUR CHINI MILLS LIMITED')])
    ws2.append([tc(ws2, f'{pname} — TNI TRACKING | FY {fy}')])
    ws2.append([])
    ws2.append([hc(ws2, h) for h in
        ['Sr.','Emp Code','Name','Designation','Grade','Collar','Dept',
         'Section','Programme Name','Type','Mode','Target Month','Planned Hrs','Completed?']])
    # Pre-build completion set — one query instead of N queries
    done_set = set(
        (row[0], row[1]) for row in db.execute(
            'SELECT emp_code, programme_name FROM emp_training WHERE plant_id=?', (plant_id,))
    )
    for r, t in enumerate(db.execute('''
            SELECT t.*,e.name,e.designation,e.grade,e.collar,e.department,e.section
            FROM tni t LEFT JOIN employees e
              ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=?''', (plant_id,)), 1):
        ws2.append([r, t['emp_code'], t['name'] or '', t['designation'] or '',
                    t['grade'] or '', t['collar'] or '', t['department'] or '',
                    t['section'] or '', t['programme_name'], t['prog_type'] or '',
                    t['mode'] or '', t['target_month'] or '', t['planned_hours'],
                    'Yes' if (t['emp_code'], t['programme_name']) in done_set else 'No'])

    # ── Sheet 3: Calendar ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('Cal_Plan_vs_Actual')
    ws3.append([tc(ws3, 'BALRAMPUR CHINI MILLS LIMITED')])
    ws3.append([tc(ws3, f'{pname} — TRAINING CALENDAR | FY {fy}')])
    ws3.append([])
    ws3.append([hc(ws3, h) for h in
        ['S/N','PROG CODE','SESSION CODE','Source','Programme Name','Type',
         'Planned Month','Plan Start','Plan End','Duration(Hrs)','Level','Mode',
         'Target Audience','Planned Pax','Trainer/Vendor','STATUS','Actual Date','Actual Pax']])
    act_pax_map = {row[0]: row[1] for row in db.execute(
        'SELECT session_code, COUNT(*) FROM emp_training WHERE plant_id=? GROUP BY session_code',
        (plant_id,))}
    pd_date_map = {row[0]: row[1] for row in db.execute(
        'SELECT session_code, start_date FROM programme_details WHERE plant_id=?', (plant_id,))}
    for r, c in enumerate(db.execute(
            'SELECT * FROM calendar WHERE plant_id=? ORDER BY id', (plant_id,)), 1):
        ws3.append([r, c['prog_code'], c['session_code'], c['source'] or '',
                    c['programme_name'], c['prog_type'] or '', c['planned_month'] or '',
                    c['plan_start'] or '', c['plan_end'] or '',
                    c['duration_hrs'], c['level'] or '', c['mode'] or '',
                    c['target_audience'] or '', c['planned_pax'],
                    c['trainer_vendor'] or '', c['status'] or '',
                    pd_date_map.get(c['session_code'], ''),
                    act_pax_map.get(c['session_code'], 0)])

    # ── Sheet 4: 2A Employee Training ─────────────────────────────────────
    ws4 = wb.create_sheet('2A_Emp_Training')
    ws4.append([tc(ws4, 'BALRAMPUR CHINI MILLS LIMITED')])
    ws4.append([tc(ws4, f'{pname} — 2A: EMPLOYEE TRAINING DETAILS | FY {fy}')])
    ws4.append([])
    ws4.append([hc(ws4, h) for h in
        ['Sr.','Emp Code','Name','Designation','Grade','Collar','Dept','Section',
         'Start Date','End Date','Hrs','Programme Name','Type','Level','Mode',
         'Cal/New','Pre Rating','Post Rating','Venue','Month']])
    for r, t in enumerate(db.execute('''
            SELECT t.*,e.name as emp_name,e.designation,e.grade,e.collar,
                   e.department,e.section
            FROM emp_training t LEFT JOIN employees e
              ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? ORDER BY t.id''', (plant_id,)), 1):
        ws4.append([r, t['emp_code'], t['emp_name'] or '', t['designation'] or '',
                    t['grade'] or '', t['collar'] or '', t['department'] or '',
                    t['section'] or '', t['start_date'] or '', t['end_date'] or '',
                    t['hrs'], t['programme_name'], t['prog_type'] or '',
                    t['level'] or '', t['mode'] or '', t['cal_new'] or '',
                    t['pre_rating'], t['post_rating'], t['venue'] or '', t['month'] or ''])

    # ── Sheet 5: 2C Programme Details ─────────────────────────────────────
    ws5 = wb.create_sheet('2C_Programme')
    ws5.append([tc(ws5, 'BALRAMPUR CHINI MILLS LIMITED')])
    ws5.append([tc(ws5, f'{pname} — 2C: PROGRAMME-WISE DETAILS | FY {fy}')])
    ws5.append([])
    ws5.append([hc(ws5, h) for h in
        ['Sr.','Session Code','Programme Name','Type','Level','Cal/New','Mode',
         'Start Date','End Date','Audience','Hours Actual','Faculty Name','Int/Ext',
         'Cost (Rs.)','Venue','Course FB','Faculty FB','Trainer FB-Participants',
         'Trainer FB-Facilities','Participants','Man-Hours']])
    pax_map = {row[0]: row[1] for row in db.execute(
        'SELECT session_code, COUNT(*) FROM emp_training WHERE plant_id=? GROUP BY session_code',
        (plant_id,))}
    hrs_map = {row[0]: row[1] for row in db.execute(
        'SELECT session_code, COALESCE(SUM(hrs),0) FROM emp_training WHERE plant_id=? GROUP BY session_code',
        (plant_id,))}
    for r, p in enumerate(db.execute(
            'SELECT * FROM programme_details WHERE plant_id=? ORDER BY id', (plant_id,)), 1):
        ws5.append([r, p['session_code'], p['programme_name'], p['prog_type'] or '',
                    p['level'] or '', p['cal_new'] or '', p['mode'] or '',
                    p['start_date'] or '', p['end_date'] or '', p['audience'] or '',
                    p['hours_actual'], p['faculty_name'] or '', p['int_ext'] or '',
                    p['cost'], p['venue'] or '', p['course_feedback'],
                    p['faculty_feedback'], p['trainer_fb_participants'],
                    p['trainer_fb_facilities'],
                    pax_map.get(p['session_code'], 0),
                    round(hrs_map.get(p['session_code'], 0), 1)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"BCML_{plant['unit_code']}_Training_MIS_{fy.replace('-','')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/calendar/template')
@spoc_required
def calendar_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Calendar_Bulk_Upload'
    headers = ['Programme Name', 'Type of Programme', 'Source', 'Planned Month',
               'Plan Start (YYYY-MM-DD)', 'Plan End (YYYY-MM-DD)', 'Duration (Hrs)',
               'Level', 'Mode', 'Target Audience', 'Planned Pax', 'Trainer/Vendor']
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF')
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        ws.column_dimensions[get_column_letter(i)].width = 24
    samples = [
        ['Fire Safety Training', 'EHS/HR', 'TNI', 'June', '2026-06-10', '2026-06-10', 4, 'General', 'Classroom', 'Blue Collared', 30, 'Internal Faculty'],
        ['Leadership Skills', 'Behavioural/Leadership', 'Management', 'July', '2026-07-05', '2026-07-06', 8, 'Specialized', 'Classroom', 'White Collared', 20, 'External Vendor'],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws['A5'] = 'VALID Types: Behavioural/Leadership | Cane | Commercial | EHS/HR | IT | Technical'
    ws['A6'] = 'VALID Modes: Classroom | OJT | SOP | Online'
    ws['A7'] = 'VALID Audience: Blue Collared | White Collared | Common'
    ws['A8'] = 'VALID Months: April | May | June | July | August | September | October | November | December | January | February | March'
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, download_name='Calendar_Bulk_Upload_Template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/calendar/bulk', methods=['POST'])
@spoc_required
def calendar_bulk_upload():
    plant_id = session['plant_id']
    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('training_calendar'))
    try:
        df = _read_upload_file(f)
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('training_calendar'))
    db = get_db(); inserted = 0; errors = []
    for i, row in df.iterrows():
        prog_name = _clean(row, ['programme name', 'programme_name', 'program name'])
        prog_type = _clean(row, ['type of programme', 'type', 'prog type'])
        source    = _clean(row, ['source']) or 'TNI'
        month     = _clean(row, ['planned month', 'month'])
        plan_start= _clean(row, ['plan start (yyyy-mm-dd)', 'plan start', 'start date'])
        plan_end  = _clean(row, ['plan end (yyyy-mm-dd)', 'plan end', 'end date'])
        duration  = _safe_float(_clean(row, ['duration (hrs)', 'duration', 'hrs'])) or 0
        level     = _clean(row, ['level'])
        mode      = _clean(row, ['mode'])
        audience  = _clean(row, ['target audience', 'audience'])
        pax       = int(_safe_float(_clean(row, ['planned pax', 'pax'])) or 0)
        trainer   = _clean(row, ['trainer/vendor', 'trainer', 'vendor'])
        if not prog_name:
            errors.append(f'Row {i+2}: Programme Name is required.')
            continue
        if not prog_type:
            errors.append(f'Row {i+2}: Type of Programme is required.')
            continue
        prog_name    = _canonical_prog(prog_name, plant_id, db)
        prog_code    = _get_or_create_prog_code(plant_id, prog_name, prog_type, db)
        session_code = _new_session_code(plant_id, prog_code, db)
        db.execute('''INSERT INTO calendar
            (plant_id,prog_code,session_code,source,programme_name,prog_type,
             planned_month,plan_start,plan_end,duration_hrs,level,mode,
             target_audience,planned_pax,trainer_vendor,status)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'To Be Planned')''',
            (plant_id, prog_code, session_code, source, prog_name, prog_type,
             month, plan_start, plan_end, duration, level, mode, audience, pax, trainer))
        inserted += 1
    db.commit()
    if errors:
        if inserted:
            flash(f'Bulk upload complete: {inserted} sessions added. {len(errors)} rows had errors — downloading error report.', 'warning')
        return _error_excel_response(errors, inserted, 'Calendar_Upload_Errors.xlsx')
    flash(f'Bulk upload complete: {inserted} sessions added to calendar.', 'success')
    return redirect(url_for('training_calendar'))

@app.route('/programme/template')
@spoc_required
def programme_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2C_Bulk_Upload'
    headers = ['Session Code', 'Actual Start Date (YYYY-MM-DD)', 'Actual End Date (YYYY-MM-DD)',
               'Actual Hours', 'Faculty Name', 'Internal/External', 'Cost (Rs.)', 'Venue',
               'Course Feedback (1-5)', 'Faculty Feedback (1-5)',
               'Trainer FB Participants (1-5)', 'Trainer FB Facilities (1-5)']
    hdr_fill = PatternFill('solid', fgColor='6B3FA0')
    hdr_font = Font(bold=True, color='FFFFFF')
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        ws.column_dimensions[get_column_letter(i)].width = 26
    samples = [
        ['BCM/EHS/001/B01', '2026-06-10', '2026-06-10', 4, 'Mr. Ramesh Kumar', 'Internal', 0, 'Training Hall', 4.2, 4.0, 3.8, 4.1],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws['A4'] = 'NOTE: Session Code must exist in Training Calendar. Internal/External options: Internal | External | Online'
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, download_name='2C_Programme_Bulk_Upload_Template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/programme/bulk', methods=['POST'])
@spoc_required
def programme_bulk_upload():
    plant_id = session['plant_id']
    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('programme_details'))
    try:
        df = _read_upload_file(f)
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('programme_details'))
    db = get_db(); inserted = 0; errors = []
    for i, row in df.iterrows():
        sc         = _clean(row, ['session code', 'session_code'])
        start_date = _clean(row, ['actual start date (yyyy-mm-dd)', 'start date', 'actual start date'])
        end_date   = _clean(row, ['actual end date (yyyy-mm-dd)', 'end date', 'actual end date'])
        hrs        = _safe_float(_clean(row, ['actual hours', 'hours', 'hrs'])) or 0
        faculty    = _clean(row, ['faculty name', 'faculty'])
        int_ext    = _clean(row, ['internal/external', 'int/ext', 'internal external'])
        cost       = _safe_float(_clean(row, ['cost (rs.)', 'cost'])) or 0
        venue      = _clean(row, ['venue'])
        cfb        = _safe_float(_clean(row, ['course feedback (1-5)', 'course feedback', 'course fb']))
        ffb        = _safe_float(_clean(row, ['faculty feedback (1-5)', 'faculty feedback', 'faculty fb']))
        tfbp       = _safe_float(_clean(row, ['trainer fb participants (1-5)', 'trainer fb participants']))
        tfbf       = _safe_float(_clean(row, ['trainer fb facilities (1-5)', 'trainer fb facilities']))
        if not sc:
            errors.append(f'Row {i+2}: Session Code is required.')
            continue
        cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?', (sc, plant_id)).fetchone()
        if not cal:
            errors.append(f'Row {i+2}: Session Code {sc} not found in Calendar.')
            continue
        if db.execute('SELECT 1 FROM programme_details WHERE session_code=? AND plant_id=?', (sc, plant_id)).fetchone():
            errors.append(f'Row {i+2}: Session {sc} already has programme details recorded.')
            continue
        if not start_date:
            errors.append(f'Row {i+2}: Actual Start Date is required.')
            continue
        db.execute('''INSERT INTO programme_details
            (plant_id,session_code,programme_name,prog_type,level,cal_new,mode,
             start_date,end_date,audience,hours_actual,faculty_name,int_ext,cost,
             venue,course_feedback,faculty_feedback,trainer_fb_participants,trainer_fb_facilities)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (plant_id, sc, cal['programme_name'], cal['prog_type'], cal['level'],
             'Calendar Program', cal['mode'], start_date, end_date,
             cal['target_audience'], hrs, faculty, int_ext, cost, venue, cfb, ffb, tfbp, tfbf))
        db.execute("UPDATE calendar SET status='Conducted' WHERE session_code=? AND plant_id=?",
                   (sc, plant_id))
        inserted += 1
    db.commit()
    if errors:
        if inserted:
            flash(f'Bulk upload complete: {inserted} programme records saved. {len(errors)} rows had errors — downloading error report.', 'warning')
        return _error_excel_response(errors, inserted, 'Programme2C_Upload_Errors.xlsx')
    flash(f'Bulk upload complete: {inserted} programme records saved.', 'success')
    return redirect(url_for('programme_details'))

# ─── API (AJAX auto-fill) ─────────────────────────────────────────────────────

@app.route('/api/employee/<emp_code>')
@login_required
def api_employee(emp_code):
    plant_id = session.get('plant_id')
    if not plant_id:
        return jsonify({})
    db = get_db()
    emp = db.execute('SELECT * FROM employees WHERE emp_code=? AND plant_id=? AND is_active=1',
                     (emp_code.strip(), plant_id)).fetchone()
    if not emp:
        return jsonify({})
    return jsonify({
        'name': emp['name'], 'designation': emp['designation'],
        'grade': emp['grade'], 'collar': emp['collar'],
        'department': emp['department'], 'section': emp['section'],
        'category': emp['category'], 'gender': emp['gender']
    })

@app.route('/api/session/<path:session_code>')
@login_required
def api_session(session_code):
    plant_id = session.get('plant_id')
    if not plant_id:
        return jsonify({})
    db = get_db()
    cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                     (session_code.strip(), plant_id)).fetchone()
    if not cal:
        return jsonify({})
    return jsonify({
        'programme_name': cal['programme_name'], 'prog_type': cal['prog_type'],
        'level': cal['level'], 'mode': cal['mode'],
        'plan_start': cal['plan_start'], 'plan_end': cal['plan_end'],
        'duration_hrs': cal['duration_hrs'], 'target_audience': cal['target_audience']
    })

@app.route('/api/employees_list')
@login_required
def api_employees_list():
    plant_id = session.get('plant_id')
    if not plant_id:
        return jsonify([])
    db = get_db()
    emps = db.execute('SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name',
                      (plant_id,)).fetchall()
    return jsonify([{'code': e['emp_code'], 'name': e['name']} for e in emps])

@app.route('/api/emp-lookup')
@spoc_required
def api_emp_lookup():
    plant_id = session['plant_id']
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({'name': None})
    db  = get_db()
    emp = db.execute('SELECT name FROM employees WHERE emp_code=? AND plant_id=? AND is_active=1',
                     (code, plant_id)).fetchone()
    return jsonify({'name': emp['name'] if emp else None})

@app.route('/api/programme-list')
@spoc_required
def api_programme_list():
    plant_id = session['plant_id']
    db = get_db()
    master = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
    ).fetchall()] or MASTER_PROGRAMMES
    return jsonify(master)

@app.route('/api/tni-coverage')
@spoc_required
def api_tni_coverage():
    from difflib import get_close_matches as gcm
    plant_id  = session['plant_id']
    prog_name = request.args.get('q', '').strip()
    if not prog_name:
        return jsonify({})
    db  = get_db()
    fy  = _current_fy()

    # Fuzzy-resolve name against TNI if exact not found
    canonical = prog_name
    exact = db.execute('SELECT 1 FROM tni WHERE plant_id=? AND programme_name=? LIMIT 1',
                       (plant_id, prog_name)).fetchone()
    if not exact:
        all_names = [r[0] for r in db.execute(
            'SELECT DISTINCT programme_name FROM tni WHERE plant_id=?', (plant_id,))]
        m = gcm(prog_name.lower(), [n.lower() for n in all_names], n=1, cutoff=0.65)
        if m:
            canonical = all_names[[n.lower() for n in all_names].index(m[0])]

    demand           = db.execute('SELECT COUNT(DISTINCT emp_code) FROM tni WHERE plant_id=? AND programme_name=?',
                                  (plant_id, canonical)).fetchone()[0]
    sessions_planned = db.execute('SELECT COUNT(*) FROM calendar WHERE plant_id=? AND programme_name=? AND session_code LIKE ?',
                                  (plant_id, canonical, f'%/{fy}/%')).fetchone()[0]
    covered          = db.execute('SELECT COUNT(DISTINCT emp_code) FROM emp_training WHERE plant_id=? AND programme_name=?',
                                  (plant_id, canonical)).fetchone()[0]
    uncovered = max(0, demand - covered)
    pct       = round(covered / demand * 100) if demand > 0 else 0

    # Most common prog_type, mode, collar combination
    meta = db.execute('''
        SELECT t.prog_type, t.mode, e.collar, COUNT(*) as cnt
        FROM tni t
        LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND t.programme_name=?
        GROUP BY t.prog_type, t.mode, e.collar
        ORDER BY cnt DESC LIMIT 1
    ''', (plant_id, canonical)).fetchone()

    # Most common target_month
    month_row = db.execute('''
        SELECT target_month, COUNT(*) as cnt FROM tni
        WHERE plant_id=? AND programme_name=? AND target_month IS NOT NULL AND target_month != ''
        GROUP BY target_month ORDER BY cnt DESC LIMIT 1
    ''', (plant_id, canonical)).fetchone()

    # Average planned hours
    hrs_row = db.execute('''
        SELECT AVG(planned_hours) as avg_hrs FROM tni
        WHERE plant_id=? AND programme_name=? AND planned_hours > 0
    ''', (plant_id, canonical)).fetchone()

    # Fallback to programme_master for prog_type/mode if TNI has no data
    if not meta or not meta['prog_type']:
        pm = db.execute('SELECT prog_type, mode FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
                        (plant_id, canonical)).fetchone()
    else:
        pm = None

    prog_type = (meta['prog_type'] if meta else '') or (pm['prog_type'] if pm else '')
    mode      = (meta['mode']      if meta else '') or (pm['mode']      if pm else '')

    collar_map = {'Blue Collared': 'Blue Collared', 'White Collared': 'White Collared'}
    audience  = collar_map.get((meta['collar'] or '') if meta else '', 'Common') if meta else ''

    # Derive source from prog_type
    pt_lower = prog_type.lower() if prog_type else ''
    if 'statutory' in pt_lower or 'compliance' in pt_lower or 'legal' in pt_lower:
        source = 'Compliance Driven'
    elif demand > 0:
        source = 'TNI Driven'
    else:
        source = ''

    avg_hrs = round(hrs_row['avg_hrs'], 1) if hrs_row and hrs_row['avg_hrs'] else 0

    return jsonify({
        'demand': demand, 'sessions_planned': sessions_planned,
        'covered': covered, 'uncovered': uncovered, 'pct': pct,
        'prog_type':    prog_type,
        'mode':         mode,
        'audience':     audience,
        'source':       source,
        'target_month': month_row['target_month'] if month_row else '',
        'avg_hrs':      avg_hrs,
    })

@app.route('/intelligence')
@spoc_required
def programme_intelligence():
    plant_id = session['plant_id']
    db       = get_db()
    fy       = _current_fy()

    # ── Unique programmes in TNI & those still in demand ──────────────────────
    unique_progs = db.execute(
        'SELECT COUNT(DISTINCT programme_name) FROM tni WHERE plant_id=?',
        (plant_id,)).fetchone()[0]

    # all TNI programmes with demand
    tni_rows = db.execute(
        'SELECT programme_name, COUNT(DISTINCT emp_code) as demand FROM tni WHERE plant_id=? GROUP BY programme_name ORDER BY demand DESC',
        (plant_id,)).fetchall()

    # covered per programme (from 2A)
    covered_map = {}
    for r in db.execute('SELECT programme_name, COUNT(DISTINCT emp_code) as cnt FROM emp_training WHERE plant_id=? GROUP BY programme_name', (plant_id,)):
        covered_map[r['programme_name']] = r['cnt']

    # sessions planned per programme this FY
    sessions_map = {}
    for r in db.execute('SELECT programme_name, COUNT(*) as cnt FROM calendar WHERE plant_id=? AND session_code LIKE ? GROUP BY programme_name',
                        (plant_id, f'%/{fy}/%')):
        sessions_map[r['programme_name']] = r['cnt']

    programmes = []
    total_needs   = 0
    total_covered = 0
    for r in tni_rows:
        name      = r['programme_name']
        demand    = r['demand']
        covered   = covered_map.get(name, 0)
        planned   = sessions_map.get(name, 0)
        uncovered = max(0, demand - covered)
        pct       = round(covered / demand * 100) if demand > 0 else 0
        if demand < 30:  status = 'Small Group'
        elif pct >= 80:  status = 'On Track'
        elif pct >= 30:  status = 'In Progress'
        else:            status = 'Big Ticket'
        total_needs   += demand
        total_covered += min(covered, demand)
        programmes.append({'name': name, 'demand': demand, 'covered': covered,
                           'planned': planned, 'uncovered': uncovered, 'pct': pct, 'status': status})

    total_uncovered  = max(0, total_needs - total_covered)
    progs_in_demand  = sum(1 for p in programmes if p['uncovered'] > 0)
    big_tickets      = sum(1 for p in programmes if p['status'] == 'Big Ticket')

    _status_order = {'Big Ticket': 0, 'In Progress': 1, 'On Track': 2, 'Small Group': 3}
    programmes.sort(key=lambda p: (_status_order.get(p['status'], 9), -p['demand']))

    # ── Sessions by mode this FY ───────────────────────────────────────────────
    _mode_keys = ['Classroom', 'OJT', 'Online', 'SOP']
    mode_map   = {m: {'planned': 0, 'conducted': 0} for m in _mode_keys}
    mode_map['Other'] = {'planned': 0, 'conducted': 0}
    for r in db.execute(
            'SELECT mode, status, COUNT(*) as cnt FROM calendar WHERE plant_id=? AND session_code LIKE ? GROUP BY mode, status',
            (plant_id, f'%/{fy}/%')):
        key = r['mode'] if r['mode'] in _mode_keys else 'Other'
        mode_map[key]['planned'] += r['cnt']
        if r['status'] == 'Conducted':
            mode_map[key]['conducted'] += r['cnt']
    session_modes = [{'mode': k, **v, 'pct': round(v['conducted']/v['planned']*100) if v['planned'] else 0}
                     for k, v in mode_map.items() if v['planned'] > 0]
    total_sessions = sum(v['planned'] for v in mode_map.values())

    return render_template('intelligence.html', fy=fy,
                           unique_progs=unique_progs, progs_in_demand=progs_in_demand,
                           total_needs=total_needs, total_covered=total_covered,
                           total_uncovered=total_uncovered,
                           total_sessions=total_sessions, session_modes=session_modes,
                           big_tickets=big_tickets, programmes=programmes)

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _is_ajax():
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'

def _read_upload_file(file_storage):
    import pandas as pd
    fname = file_storage.filename.lower()
    if fname.endswith('.csv'):
        return pd.read_csv(file_storage, dtype=str).fillna('')
    else:
        return pd.read_excel(file_storage, dtype=str).fillna('')

def _read_upload_file_path(path):
    import pandas as pd
    if path.lower().endswith('.csv'):
        return pd.read_csv(path, dtype=str).fillna('')
    else:
        return pd.read_excel(path, dtype=str).fillna('')

def _poka_yoke_clean_prog(name):
    """Normalize a programme name: strip, collapse spaces, word-spell-fix, smart title case."""
    if not name:
        return ''
    s = re.sub(r'[\x00-\x1f\x7f]', '', str(name).strip())
    s = re.sub(r'\s+', ' ', s).strip()
    s = _apply_word_fixes(s)   # fix known misspellings word-by-word
    return _smart_title(s)

def _process_fresh_tni(df, plant_id, db):
    """
    Parse a TNI DataFrame with 100% poka-yoke cleaning.
    Returns dict: total_rows, valid_rows, error_rows, name_corrections, unique_progs, duplicate_count.
    """
    emp_rows = db.execute(
        'SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)
    ).fetchall()
    emp_map   = {r['emp_code']: r['name'] for r in emp_rows}
    emp_upper = {k.upper(): k for k in emp_map}

    cols     = df.columns.tolist()
    col_emp  = _detect_col(cols, ['emp code','employee code','empcode','staff code','emp id','employee id','code'])
    col_prog = _detect_col(cols, ['programme name','program name','training name','course name','training need','training'])
    col_type = _detect_col(cols, ['type of programme','type','programme type','prog type','training type','category'])
    col_mode = _detect_col(cols, ['mode','training mode','delivery mode'])
    col_hrs  = _detect_col(cols, ['planned hours','hours','hrs','duration'])

    def gv(row, col):
        if not col: return ''
        v = str(row.get(col, '') or '').strip()
        return '' if v.lower() in ('nan', 'none', '0', '') else v

    valid_rows       = []
    error_rows       = []
    name_corrections = {}
    seen             = set()
    duplicate_count  = 0

    for i, row in df.iterrows():
        raw_emp  = gv(row, col_emp)
        raw_prog = gv(row, col_prog)
        prog_type = gv(row, col_type)
        mode      = gv(row, col_mode)
        hours     = _safe_float(gv(row, col_hrs)) or 0

        if not raw_emp and not raw_prog:
            continue

        if not raw_emp:
            error_rows.append({'row': i+2, 'emp_code': '', 'prog_name': raw_prog, 'reason': 'Employee Code missing'})
            continue
        if not raw_prog:
            error_rows.append({'row': i+2, 'emp_code': raw_emp, 'prog_name': '', 'reason': 'Programme Name missing'})
            continue

        clean_emp = raw_emp
        if raw_emp not in emp_map:
            if raw_emp.upper() in emp_upper:
                clean_emp = emp_upper[raw_emp.upper()]
            else:
                error_rows.append({'row': i+2, 'emp_code': raw_emp, 'prog_name': raw_prog,
                                   'reason': f'Employee "{raw_emp}" not found in this plant'})
                continue

        cleaned = _poka_yoke_clean_prog(raw_prog)
        if cleaned != raw_prog:
            name_corrections[raw_prog] = cleaned

        key = (clean_emp, cleaned.lower())
        if key in seen:
            duplicate_count += 1
            continue
        seen.add(key)

        valid_rows.append({
            'emp_code': clean_emp, 'programme_name': cleaned,
            'prog_type': prog_type, 'mode': mode, 'hours': hours,
        })

    return {
        'total_rows':       len(df),
        'valid_rows':       valid_rows,
        'error_rows':       error_rows,
        'name_corrections': name_corrections,
        'unique_progs':     sorted(set(r['programme_name'] for r in valid_rows)),
        'duplicate_count':  duplicate_count,
    }

def _clean(row, keys):
    for k in keys:
        for col in row.index:
            if str(col).strip().lower() == k:
                val = str(row[col]).strip()
                return '' if val.lower() in ('nan', 'none', '') else val
    return ''

def normalise_collar(val):
    v = str(val).strip().upper()
    if any(x in v for x in ['WHITE', 'WC', 'W C']):
        return 'White Collared'
    if any(x in v for x in ['BLUE', 'BC', 'B C']):
        return 'Blue Collared'
    return val.strip()

def _safe_float(val):
    try:
        return float(val) if val and str(val).strip() != '' else None
    except (ValueError, TypeError):
        return None

def _error_excel_response(errors, inserted, download_name='Upload_Errors.xlsx'):
    """Return an Excel file response listing failed rows with reasons."""
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Failed Rows'
    # summary row
    ws.append([f'{inserted} rows imported successfully. {len(errors)} rows failed — details below.'])
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:C1')
    ws.append([])
    # header
    hdr = ['Row #', 'Error Reason', 'Tip']
    ws.append(hdr)
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=3, column=c)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='C0392B')
        cell.alignment = Alignment(horizontal='center')
    # error rows
    for err in errors:
        # parse "Row N: message"
        parts = err.split(':', 1)
        row_ref = parts[0].strip() if len(parts) == 2 else ''
        reason  = parts[1].strip() if len(parts) == 2 else err
        # generate tip
        tip = ''
        rl  = reason.lower()
        if 'not found in your plant'   in rl: tip = 'Check employee code is registered under this plant in Employee Master'
        elif 'required'                in rl: tip = 'This column must not be empty'
        elif 'month'                   in rl: tip = 'Use: April / May / June / July / August / September / October / November / December / January / February / March'
        elif 'type' in rl or 'prog'    in rl: tip = 'Use: Behavioural/Leadership | Cane | Commercial | EHS/HR | IT | Technical'
        elif 'mode'                    in rl: tip = 'Use: Classroom | OJT | SOP | Online'
        elif 'date'                    in rl: tip = 'Use format YYYY-MM-DD'
        elif 'hours' in rl or 'hrs'    in rl: tip = 'Must be a number e.g. 4 or 2.5'
        elif 'session'                 in rl: tip = 'Session Code must already exist in Training Calendar'
        elif 'employee'                in rl: tip = 'Employee Code must exist in Employee Master for this plant'
        ws.append([row_ref, reason, tip])
    # column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 65
    # zebra stripes
    for row in ws.iter_rows(min_row=4):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor='FFF5F5')
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out, download_name=download_name, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _date_to_month(date_str):
    if not date_str:
        return ''
    try:
        d = datetime.strptime(str(date_str)[:10], '%Y-%m-%d')
        return d.strftime('%B')
    except Exception:
        return ''

def _get_programme_names(plant_id, db):
    rows = db.execute('SELECT DISTINCT programme_name FROM calendar WHERE plant_id=? ORDER BY programme_name', (plant_id,)).fetchall()
    return [r['programme_name'] for r in rows]

def _get_or_create_prog_code(plant_id, prog_name, prog_type, db):
    existing = db.execute('SELECT prog_code FROM calendar WHERE plant_id=? AND programme_name=? LIMIT 1',
                          (plant_id, prog_name)).fetchone()
    if existing:
        return existing['prog_code']
    unit_code = PLANT_MAP[plant_id]['unit_code']
    abbrev    = TYPE_ABBREV.get(prog_type, 'GEN')
    count     = db.execute("SELECT COUNT(DISTINCT prog_code) FROM calendar WHERE plant_id=? AND prog_code LIKE ?",
                           (plant_id, f'{unit_code}/{abbrev}/%')).fetchone()[0]
    return f'{unit_code}/{abbrev}/{count+1:03d}'

def _current_fy():
    today = date.today()
    y = today.year
    return f'{str(y-1)[2:]}-{str(y)[2:]}' if today.month < 4 else f'{str(y)[2:]}-{str(y+1)[2:]}'

def _new_session_code(plant_id, prog_code, db):
    fy    = _current_fy()
    count = db.execute('SELECT COUNT(*) FROM calendar WHERE plant_id=? AND prog_code=? AND session_code LIKE ?',
                       (plant_id, prog_code, f'{prog_code}/{fy}/%')).fetchone()[0]
    return f'{prog_code}/{fy}/B{count+1:02d}'

def _sync_calendar_from_2c(plant_id, db):
    db.execute('''UPDATE calendar SET status='Conducted'
        WHERE plant_id=? AND session_code IN
        (SELECT session_code FROM programme_details WHERE plant_id=?)
        AND status='To Be Planned' ''', (plant_id, plant_id))
    db.commit()

def _calc_summary(plant_id, month_filter, db):
    rows = []
    for pt in PROG_TYPES:
        clause = "AND t.month=?" if month_filter else ""
        params_bc = [plant_id, pt, 'Blue Collared'] + ([month_filter] if month_filter else [])
        params_wc = [plant_id, pt, 'White Collared'] + ([month_filter] if month_filter else [])
        params_all = [plant_id, pt] + ([month_filter] if month_filter else [])

        bc_progs = db.execute(f'''SELECT COUNT(DISTINCT t.session_code) FROM emp_training t
            JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? AND t.prog_type=? AND e.collar=? {clause}''', params_bc).fetchone()[0]
        wc_progs = db.execute(f'''SELECT COUNT(DISTINCT t.session_code) FROM emp_training t
            JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? AND t.prog_type=? AND e.collar=? {clause}''', params_wc).fetchone()[0]
        bc_seats = db.execute(f'''SELECT COUNT(*) FROM emp_training t
            JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? AND t.prog_type=? AND e.collar=? {clause}''', params_bc).fetchone()[0]
        wc_seats = db.execute(f'''SELECT COUNT(*) FROM emp_training t
            JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? AND t.prog_type=? AND e.collar=? {clause}''', params_wc).fetchone()[0]
        bc_hrs = db.execute(f'''SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t
            JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? AND t.prog_type=? AND e.collar=? {clause}''', params_bc).fetchone()[0]
        wc_hrs = db.execute(f'''SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t
            JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? AND t.prog_type=? AND e.collar=? {clause}''', params_wc).fetchone()[0]
        # Internal/External counts from 2C — filter by month derived from start_date
        month_clause_2c = "AND strftime('%m', p.start_date) = strftime('%m', '2000-' || ? || '-01')" if month_filter else ""
        int_prog = db.execute(f'''SELECT COUNT(DISTINCT p.session_code) FROM programme_details p
            WHERE p.plant_id=? AND p.prog_type=? AND p.int_ext='Internal' {month_clause_2c}''',
            [plant_id, pt] + ([month_filter] if month_filter else [])).fetchone()[0]
        ext_prog = db.execute(f'''SELECT COUNT(DISTINCT p.session_code) FROM programme_details p
            WHERE p.plant_id=? AND p.prog_type=? AND p.int_ext='External' {month_clause_2c}''',
            [plant_id, pt] + ([month_filter] if month_filter else [])).fetchone()[0]
        rows.append({
            'prog_type': pt,
            'bc_progs': bc_progs, 'wc_progs': wc_progs,
            'int_prog': int_prog, 'ext_prog': ext_prog,
            'total_prog': bc_progs + wc_progs,
            'bc_seats': bc_seats, 'wc_seats': wc_seats,
            'total_seats': bc_seats + wc_seats,
            'bc_hrs': round(bc_hrs, 1), 'wc_hrs': round(wc_hrs, 1),
            'total_hrs': round(bc_hrs + wc_hrs, 1)
        })
    return rows

def _calc_totals(rows):
    t = {k: 0 for k in rows[0]} if rows else {}
    t['prog_type'] = 'TOTAL'
    for r in rows:
        for k, v in r.items():
            if k != 'prog_type':
                t[k] = round(t.get(k, 0) + (v or 0), 1)
    return t

def _calc_compliance(plant_id, db):
    bc = db.execute("SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='Blue Collared'", (plant_id,)).fetchone()[0]
    wc = db.execute("SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='White Collared'", (plant_id,)).fetchone()[0]
    bc_act = db.execute('''SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t
        JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND e.collar='Blue Collared' ''', (plant_id,)).fetchone()[0]
    wc_act = db.execute('''SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t
        JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND e.collar='White Collared' ''', (plant_id,)).fetchone()[0]
    bc_mandate = bc * 12
    wc_mandate = wc * 24
    return {
        'bc_emp': bc, 'wc_emp': wc,
        'bc_mandate': bc_mandate, 'wc_mandate': wc_mandate,
        'bc_actual': round(bc_act, 1), 'wc_actual': round(wc_act, 1),
        'bc_pct': round(bc_act / bc_mandate * 100, 1) if bc_mandate else 0,
        'wc_pct': round(wc_act / wc_mandate * 100, 1) if wc_mandate else 0,
        'total_pct': round((bc_act + wc_act) / (bc_mandate + wc_mandate) * 100, 1) if (bc_mandate + wc_mandate) else 0
    }

# ─── MICROSOFT FORMS IMPORT ───────────────────────────────────────────────────

# MS Forms Excel export: columns = ID, Start time, Completion time, Email, Name, [questions...]
# We detect question columns by matching known header keywords.

_MSFORMS_SKIP_HEADERS = {'id','start time','completion time','email','name','responder'}

def _parse_msforms_excel(file_storage, plant_id, db):
    import pandas as pd, io as _io
    raw = file_storage.read()
    try:
        df = pd.read_excel(_io.BytesIO(raw), dtype=str).fillna('')
    except Exception as e:
        raise ValueError(f'Could not read file: {e}')

    emp_rows = db.execute('SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)).fetchall()
    emp_map  = {r['emp_code']: r['name'] for r in emp_rows}

    # Map column headers to TNI fields
    field_keywords = {
        'emp_code':       ['emp code','employee code','empcode','staff code','employee id'],
        'programme_name': ['programme name','program name','training name','course name','training need'],
        'prog_type':      ['type of programme','programme type','training type','type'],
        'mode':           ['mode','training mode','delivery mode'],
        'hours':          ['planned hours','hours','hrs','duration'],
    }
    col_map = {}
    for col in df.columns:
        cl = str(col).strip().lower()
        if cl in _MSFORMS_SKIP_HEADERS:
            continue
        for field, kws in field_keywords.items():
            if field not in col_map:
                for kw in kws:
                    if kw in cl or cl in kw:
                        col_map[field] = col
                        break

    inserted, errors = 0, []
    for i, row in df.iterrows():
        def gv(field):
            c = col_map.get(field)
            v = str(row.get(c, '') or '').strip() if c else ''
            return '' if v.lower() in ('nan','none') else v

        emp_code     = gv('emp_code')
        prog_name    = gv('programme_name')
        prog_type    = gv('prog_type')
        mode         = gv('mode')
        hours        = _safe_float(gv('hours')) or 0.0

        if not emp_code or not prog_name:
            errors.append(f'Row {i+2}: Employee Code and Programme Name are required.')
            continue
        if emp_code not in emp_map:
            errors.append(f'Row {i+2}: Employee code "{emp_code}" not found in your plant.')
            continue
        db.execute('INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours) VALUES(?,?,?,?,?,?)',
                   (plant_id, emp_code, prog_name, prog_type, mode, hours))
        inserted += 1

    db.commit()
    return inserted, errors


@app.route('/tni/msforms', methods=['GET'])
@spoc_required
def tni_msforms():
    plant_id   = session['plant_id']
    plant_name = session.get('plant_name', '')
    db = get_db()
    emp_count  = db.execute('SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)).fetchone()[0]
    return render_template('tni_msforms.html', plant_name=plant_name,
                           emp_count=emp_count, prog_types=PROG_TYPES,
                           modes=MODES, months=MONTHS_FY)


@app.route('/tni/msforms/import', methods=['POST'])
@spoc_required
def tni_msforms_import():
    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('tni_msforms'))
    plant_id = session['plant_id']
    db = get_db()
    try:
        inserted, errors = _parse_msforms_excel(f, plant_id, db)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('tni_msforms'))

    if errors:
        return _error_excel_response(errors, inserted, 'MSForms_Import_Errors.xlsx')
    flash(f'Microsoft Forms import complete: {inserted} TNI entries added.', 'success')
    return redirect(url_for('tni'))


# ─── SMART TNI ANALYZER ──────────────────────────────────────────────────────
import json as _json, uuid as _uuid, io as _io
from difflib import get_close_matches

def _canonical_prog(raw_name, plant_id, db):
    """Return best canonical programme name from master, falling back to smart title case."""
    if not raw_name or not raw_name.strip():
        return raw_name
    from difflib import get_close_matches as gcm
    master = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
    ).fetchall()] or MASTER_PROGRAMMES
    master_lower = [m.lower() for m in master]
    # Word-correct the raw name before matching so misspellings find their master entry
    corrected = _apply_word_fixes(raw_name.strip())
    raw_lower  = corrected.lower()
    if raw_lower in master_lower:
        return master[master_lower.index(raw_lower)]
    m = gcm(raw_lower, master_lower, n=1, cutoff=0.65)
    if m:
        return master[master_lower.index(m[0])]
    return _smart_title(corrected)

def _fuzzy_fix(val, valid_list):
    """Return (fixed_val, was_changed). None if no confident match."""
    if not val: return '', False
    vl = val.strip().lower()
    for v in valid_list:
        if v.lower() == vl: return v, False          # exact match
    for v in valid_list:
        if vl in v.lower() or v.lower() in vl: return v, True   # substring
    m = get_close_matches(vl, [v.lower() for v in valid_list], n=1, cutoff=0.55)
    if m:
        idx = [v.lower() for v in valid_list].index(m[0])
        return valid_list[idx], True
    return val, False  # can't fix → return original so caller can flag

def _detect_col(columns, keywords):
    """Return first column name whose header matches any keyword (partial, case-insensitive)."""
    for col in columns:
        cl = str(col).strip().lower()
        for kw in keywords:
            if kw in cl or cl in kw:
                return col
    return None

# ── Master Programme List ─────────────────────────────────────────────────────
_ACRONYMS = {
    'PPE','SOP','EHS','OJT','DCS','UPS','VFD','DG','SLD','AC','DC','GST','ISO',
    'HR','IT','MBC','FFT','MIST','DM','ETP','CPU','CGCB','MSDS','OFSAM','ZFD',
    'STD2SD','FCS','RTD','TC','KNO3','MOP','PDM','5S','5-S','JCB','PM','R&M',
    'AI','ML','KPI','GMP','BOD','COD','TOC','TDS','ROI','MIS','SAP','ERR','ERB',
    'CCTV','GPS','QR','LED','LCD','CRM','ERP','LMS','HRM','WMS','PLC','SCADA',
}

# Hardcoded domain-specific misspellings (fast path, no fuzzy needed)
_WORD_FIXES = {
    'techqnique':'Technique','tecnique':'Technique','techique':'Technique',
    'technqiue':'Technique','teqnique':'Technique','technque':'Technique',
    'grainding':'Grinding','graining':'Grinding','granding':'Grinding',
    'grindig':'Grinding','grindding':'Grinding',
    'hyigene':'Hygiene','hyegiene':'Hygiene','hygeine':'Hygiene','higiene':'Hygiene',
    'maitenance':'Maintenance','maintenace':'Maintenance','maintainance':'Maintenance',
    'mantenance':'Maintenance',
    'safty':'Safety','saftey':'Safety',
    'operartion':'Operation','operetion':'Operation','opertion':'Operation',
    'managment':'Management','managament':'Management','mangement':'Management',
    'trainning':'Training','traning':'Training',
    'awarness':'Awareness','awreness':'Awareness',
    'handeling':'Handling','handlng':'Handling',
    'electrcial':'Electrical','eletrical':'Electrical',
    'chemcial':'Chemical','chemicle':'Chemical',
    'equpiment':'Equipment','equipement':'Equipment','equipmnet':'Equipment',
    'proceudre':'Procedure','proceedure':'Procedure',
    'complience':'Compliance','compliace':'Compliance',
    'enviroment':'Environment','enviromental':'Environmental',
    'knowlege':'Knowledge','knwoledge':'Knowledge','knoweldge':'Knowledge',
    'monitorng':'Monitoring','monitering':'Monitoring',
    'buidling':'Building','buldling':'Building',
    'confind':'Confined','confinde':'Confined','condfined':'Confined',
    'chocking':'Choking','chocing':'Choking',
    'equipments':'Equipment',   # Equipment is uncountable
    'lubricants':'Lubricants',  # already correct, keep
}


_STRIP_CHARS = '.,;:()/'

def _smart_title(s):
    """Title case: preserve acronyms, handle slashes, keep small words lowercase mid-title."""
    _SMALL = frozenset({'a','an','the','and','or','but','nor','for','yet','so',
                        'at','by','in','of','on','to','as','is','it',
                        'with','from','into','onto','off','per','via'})

    def _tw(w, is_first):
        if not w:
            return w
        # Words like "MBC/Belt" — split on slash, process each part, rejoin
        if '/' in w:
            parts = w.split('/')
            return '/'.join(_tw(p, is_first and j == 0) for j, p in enumerate(parts))
        lstripped = w.lstrip(_STRIP_CHARS)
        prefix    = w[:len(w) - len(lstripped)]
        core      = lstripped.rstrip(_STRIP_CHARS)
        suffix    = lstripped[len(core):]
        if not core:
            return w
        core_up  = core.upper()
        core_low = core.lower()
        if core_up in _ACRONYMS:
            return prefix + core_up + suffix
        # Auto-preserve any all-uppercase word (e.g. "AI", "KPI", "GMP") even if not in list
        if core == core_up and len(core) >= 2 and core.isalpha():
            return prefix + core_up + suffix
        if core_low == 'ph':
            return prefix + 'pH' + suffix
        if not is_first and core_low in _SMALL and not prefix and not suffix:
            return core_low
        return prefix + core.capitalize() + suffix

    return ' '.join(_tw(w, i == 0) for i, w in enumerate(s.split()))

def _apply_word_fixes(s):
    """
    Three-layer word-level spell correction:
    1. Hardcoded _WORD_FIXES for instant known errors
    2. Master vocab fuzzy correction (≥0.82 similarity) for anything in MASTER_PROGRAMMES words
    3. Pass-through for everything else
    Words < 4 chars and acronyms are never changed.
    """
    if not s:
        return s
    from difflib import get_close_matches as _gcm
    words = s.split()
    out   = []
    for w in words:
        lstripped = w.lstrip(_STRIP_CHARS)
        prefix    = w[:len(w) - len(lstripped)]
        core      = lstripped.rstrip(_STRIP_CHARS)
        suffix    = lstripped[len(core):]
        core_low  = core.lower()

        if not core or len(core) < 4 or core.upper() in _ACRONYMS:
            out.append(w)
            continue

        # Layer 1: hardcoded fixes
        fix = _WORD_FIXES.get(core_low)
        if fix:
            out.append(prefix + fix + suffix)
            continue

        # Layer 2: already correct per master vocab
        if core_low in _MASTER_VOCAB:
            out.append(w)
            continue

        # Layer 3: vocabulary fuzzy match (only if master vocab is populated)
        if _MASTER_VOCAB:
            m = _gcm(core_low, _MASTER_VOCAB.keys(), n=1, cutoff=0.82)
            if m:
                out.append(prefix + _MASTER_VOCAB[m[0]] + suffix)
                continue

        out.append(w)
    return ' '.join(out)

MASTER_PROGRAMMES = [
    "5-S Management","5S Commercial","Advance Practice Pathology In Agriculture",
    "Advanced Excel","Alignment Of Pumps, Fans And Gear Boxes With Motors",
    "An Overview Of Bio-Pesticides And Its Classification",
    "Bagasse Feeding As Per Boiler Requirement","Bagasse Handling",
    "Basic Fire Safety Awareness","Basic Knowledge Of Hardware",
    "Basic Of DCS Maintenance","Basic Of Maintenance And Programming Of Electronic Governor",
    "Behaviour Based Safety","Boiler Operation & Maintenance","Breakdown Handling",
    "Budget Planning","CGCB Guideline","CPU Operation",
    "Cane Quality & Minimise The Cut To Crush Period",
    "Checking Of Tube Cleaning/Tube Choking","Chemical Safety","Communication Skill",
    "Concept Of Fitting Methodology","Condenser Maintenance And Testing",
    "Condition Monitoring & Log Book Maintenance",
    "Condition Monitoring System Of Equipment","Confined Space",
    "Control Of Insect Pest & Diseases","Control Of Maintenance Schedule",
    "DCS Maintenance & Programming","DM Plant Operation & Maintenance",
    "Dismantling And Fitting Of Pump, Gear Boxes And Fans",
    "ETP Operation","Economiser Maintenance","Electrical Safety",
    "Emergency Awareness During Running Plant","Emergency Management",
    "FFT Efficiency","Failure Analysis Of Process Material",
    "Fire Fighting Equipment Technique","Fire Safety","Flow Measurement",
    "GST","General Checking Of Equipment In Running Season",
    "General Safety Awareness","Good Knowledge Of Metals","HR SOP",
    "Handling Of All Testing Equipment","Handling Of Juice And Mud Removal System",
    "Health Monitoring/Condition Monitoring Of Running Equipment",
    "Hot Work (Gas Cutting, Building And Grinding)",
    "How To Collect Samples","How To Handle Emergency Situation During Running Plant",
    "How To Identify Cane Diseases","Hydraulic Testing And Vacuum Trial Of Pan",
    "ISO General Awareness","IT SOP","Implementation Of OFSAM",
    "Importance Of Machine Guarding","Improvement Of Fermentation Efficiency",
    "Improvement Steam Economy","Income Tax",
    "Industrial Hygiene And 5-S Management",
    "Inspections And Testing Of Lifting Tool And Tackles Lifting Operation",
    "Inventory Management","Irrigation Automation","JCB Operation","JCB Pumping",
    "Juice Analysis","Knowledge About Cleaning And Ability To Conduct Schedule Checking Of Engines",
    "Knowledge About Cleaning, Switch Gear Panels And Motors Checklist",
    "Knowledge About Lab Apparatus & Equipment",
    "Knowledge And Application Of Condensate Removal System",
    "Knowledge And Operations Of Turbine","Knowledge Of Boiler Water Treatment",
    "Knowledge Of DCS Hardware And Panel Wiring",
    "Knowledge Of Electrical Equipment (Motor, Transformers, DG), SLD And Electrical Logics",
    "Knowledge Of Field Instruments Like Pressure Transmitters, Temperature Transmitters, RTD/TC, I To P Converters, Control Valves, Loop Testing",
    "Knowledge Of Importance Of Aeration In Melt",
    "Knowledge Of Industrial Lubricants Properties",
    "Knowledge Of Lighting And AC Systems",
    "Knowledge Of MBC/Belt Conveyor Health Monitoring - Gearbox, Chain Condition, Rake Condition, Idlers And Belt",
    "Knowledge Of Maintenance Of Pumps","Knowledge Of Measuring Instruments",
    "Knowledge Of Molasses Brix & Purity",
    "Knowledge Of Operating Parameter And Quality Of Steam And Cooling Water",
    "Knowledge Of Operation Of MIST","Knowledge Of Pumps And Its Parts",
    "Knowledge Of Supersaturation Zones During Pan Boiling",
    "Knowledge Of Three Motion Hydraulic Cane Unloader Operations",
    "Knowledge Of Upgraded Technology Related IT",
    "Knowledge Of Wire Rope Sling Size By Weight And Job Wise",
    "Knowledge Of Working Tools, Tackles And Fasteners",
    "Labour Laws","Leadership Quality","Legal Compliance",
    "Maintain Brix Of Magma","Maintain Temperature From Juice Heater",
    "Maintaining Load Of Machine By Operating The Feed Valve",
    "Maintenance And Programming Of Electronic Governor",
    "Maintenance Effectiveness","Maintenance Of AC & DC Drives",
    "Maintenance Of Electrical Machine And Switch Gears",
    "Maintenance Of Power Turbine",
    "Maintenance Of Safety Valves And Checking For Its Perfection",
    "Maintenance Of UPS & Battery","Manufacturing Process",
    "Massecuite Curing Temperature And Its Impact","Material Handling",
    "Material Handling (Manual And Mechanical)","Measurement Maintenance",
    "Mill Efficiency","Molasses Conditioning Temperature And Brix",
    "Monitoring And Ensuring Proper Operations Of Bagasse Belt Conveyors",
    "New Methodology In Soil Testing","New Varietal Trial","New Wage Code",
    "Nil Safety","Operation & Maintenance Of Boiler",
    "Operation Of Boiler From Cold To Pressurization",
    "Operation Of Turbine Within Controlled Parameters",
    "Operational Behaviour Awareness",
    "Optimum Temperature And pH Adjustment During Defecation",
    "Organic Waste Management","Overhauling And Maintenance Of Centrifugal Machine",
    "Ownership","PPE Awareness Use Inspection And Handling","Personal Effectiveness",
    "Planning Of Different Massecuite Boiling To Control The Material Load",
    "PowerPoint","Premium Potash Fertilisers: SOP Vs KNO3 Vs MOP Vs PDM",
    "Preventive Maintenance",
    "Preventive Maintenance And Condition Monitoring Of Electrical Equipment",
    "Problem Solving","Process Parameter","Process Safety Management",
    "Purification Of Condensate Removal System","Qualitimetry",
    "Quality Communication & Industrial Security","Quality R&M And Operation",
    "Raw Material Analysis","Removal Of Juice And Filter Cake From The System",
    "Repair & Maintenance Of Workshop Machine","Reporting Of Non EHS Lapses",
    "Reporting Of Non Performance Of Chemicals","Road Safety & Defensive Driving",
    "SOP Boiler","SOP Distillation And Fermentation Operation","SOP ETP",
    "SOP Electrical","SOP Instrumentation","SOP Mill House","SOP PM Module",
    "SOP Sales","SOP Store","SOP Workshop","STD2SD","Safety Induction",
    "Sample Collection","Sample Collection Methods As Per SOP",
    "Sampling & Its Importance","Screen Checking And Molasses Purity Control",
    "Self Discipline","Start And Stop The Boiler","Switch Gear Maintenance & Testing",
    "Team Work","Theft Prevention","To Control Fermentation Process Parameter",
    "To Maintain Brix Of Massecuite Of Dropping Pan",
    "To Maintain Brix Of Syrup At Fix",
    "To Maintain Chemical Dosing As Per Requirement",
    "To Maintain Temperature & pH Of Juice",
    "To Maintain Temperature & pH Of Juice/Melt",
    "To Maintain Temperature And pH Of Juice","Treated Water Parameters",
    "Understanding Of DCS Logic And Graphics","Use Of MSDS",
    "Use Of Optimum Dose Of Flocculant","VFD Maintenance","VFD Operation & Maintenance",
    "Water Management","Withdrawal Of Scum From FCS Clarifier","ZFD",
]
_MASTER_LOWER = [p.lower() for p in MASTER_PROGRAMMES]

def _build_master_vocab():
    """Build word-level vocabulary from MASTER_PROGRAMMES for fuzzy spell correction."""
    _skip = frozenset({'a','an','the','and','or','but','of','in','to','at','by','as','is','it',
                       'for','on','per','via','from','into','onto','off','with','how','its'})
    vocab = {}
    for prog in MASTER_PROGRAMMES:
        for w in re.split(r'[ /,;:()\-]+', prog):
            core = re.sub(r'[.,;:()/&\-]', '', w).strip()
            if core and len(core) >= 4 and core.upper() not in _ACRONYMS and core.lower() not in _skip:
                vocab[core.lower()] = core
    return vocab

_MASTER_VOCAB = _build_master_vocab()

def _smart_analyze_rows(df, plant_id, db):
    from difflib import get_close_matches as gcm
    _prog_cache = {}  # cache: raw_lower → canonical name

    # Union of DB master + hardcoded list so every known programme is matchable
    _db_master = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,))]
    _combined_seen = dict.fromkeys(_db_master + MASTER_PROGRAMMES)
    _active_master       = list(_combined_seen.keys())
    _active_master_lower = [p.lower() for p in _active_master]

    def _match_master(raw_lower):
        if raw_lower in _prog_cache:
            return _prog_cache[raw_lower]
        m = gcm(raw_lower, _active_master_lower, n=1, cutoff=0.65)
        result = _active_master[_active_master_lower.index(m[0])] if m else None
        _prog_cache[raw_lower] = result
        return result
    emp_rows  = db.execute('SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)).fetchall()
    emp_map   = {r['emp_code']: r['name'] for r in emp_rows}
    emp_upper = {k.upper(): k for k in emp_map}

    # Known programme names from calendar + existing TNI for spell-fix
    known_progs = [r[0] for r in db.execute(
        'SELECT DISTINCT programme_name FROM tni WHERE plant_id=? UNION '
        'SELECT DISTINCT programme_name FROM calendar WHERE plant_id=?', (plant_id, plant_id))]

    # Detect columns once — works on any file format
    cols = df.columns.tolist()
    col_emp   = _detect_col(cols, ['emp code','employee code','empcode','staff code','emp id','employee id','code'])
    col_prog  = _detect_col(cols, ['programme name','program name','training name','course name','training need','training'])
    col_type  = _detect_col(cols, ['type of programme','type','programme type','prog type','training type','category'])
    col_mode  = _detect_col(cols, ['mode','training mode','delivery mode'])
    col_hrs   = _detect_col(cols, ['planned hours','hours','hrs','duration'])

    # If neither critical column found, raise helpful error showing actual columns
    if not col_emp and not col_prog:
        col_list = ', '.join(f'"{c}"' for c in cols[:15])
        raise ValueError(
            f'Could not detect Employee Code or Programme Name columns. '
            f'Columns found in file: {col_list}. '
            f'Try using "Skip top rows" if your file has a title row above the headers.'
        )

    def gv(row, col):
        if not col: return ''
        v = str(row.get(col, '') or '').strip()
        return '' if v.lower() in ('nan','none','0','') else v

    results = []
    for i, row in df.iterrows():
        raw_emp   = gv(row, col_emp)
        raw_prog  = gv(row, col_prog)
        raw_type  = gv(row, col_type)
        raw_mode  = gv(row, col_mode)
        raw_hrs   = gv(row, col_hrs)

        if not any([raw_emp, raw_prog, raw_type, raw_mode]):
            continue

        fixes   = []   # list of {field, original, fixed, how}
        issues  = []   # unfixable problems
        status  = 'ok'

        # ── Employee Code ──────────────────────────────────────────
        clean_emp = raw_emp
        if not raw_emp:
            issues.append('Employee Code is missing')
            status = 'error'
        elif raw_emp in emp_map:
            pass  # exact match
        elif raw_emp.upper() in emp_upper:
            clean_emp = emp_upper[raw_emp.upper()]
            fixes.append({'field':'Employee Code','original':raw_emp,'fixed':clean_emp,'how':'Capitalisation corrected'})
            if status == 'ok': status = 'fixed'
        else:
            issues.append(f'Employee code "{raw_emp}" not found in this plant')
            status = 'error'

        emp_name = emp_map.get(clean_emp, '')

        # ── Programme Name ─────────────────────────────────────────
        clean_prog = raw_prog
        if not raw_prog:
            issues.append('Programme Name is missing')
            status = 'error'
        else:
            # Apply word-level spell fixes BEFORE fuzzy matching so e.g.
            # "techqnique" → "Technique" before comparing against master
            word_fixed = _apply_word_fixes(raw_prog.strip())
            raw_lower  = word_fixed.lower()
            # Step 1: fuzzy match against master list (cached)
            best = _match_master(raw_lower)
            if best is not None:
                if best.lower() != raw_prog.strip().lower():
                    fixes.append({'field':'Programme Name','original':raw_prog,'fixed':best,'how':'Matched to master list'})
                    if status == 'ok': status = 'fixed'
                clean_prog = best
            else:
                # Step 2: fallback — apply word fixes + smart title case
                titled = _smart_title(word_fixed)
                if titled != raw_prog:
                    fixes.append({'field':'Programme Name','original':raw_prog,'fixed':titled,'how':'Spelling/case corrected (not in master list)'})
                    clean_prog = titled
                    if status == 'ok': status = 'fixed'
                else:
                    clean_prog = titled
                # Flag: not found in master list at all
                if status not in ('error',):
                    issues.append(f'"{clean_prog}" not found in Programme Master — verify spelling or add it to master list')
                    if status == 'ok': status = 'warning'

        # ── Programme Type ─────────────────────────────────────────
        clean_type, type_changed = _fuzzy_fix(raw_type, PROG_TYPES) if raw_type else ('', False)
        if raw_type and clean_type not in PROG_TYPES:
            issues.append(f'Unknown programme type: "{raw_type}" — could not auto-fix')
            if status == 'ok': status = 'error'
        elif raw_type and type_changed:
            fixes.append({'field':'Type of Programme','original':raw_type,'fixed':clean_type,'how':'Auto-matched to standard value'})
            if status == 'ok': status = 'fixed'

        # ── Mode ───────────────────────────────────────────────────
        clean_mode, mode_changed = _fuzzy_fix(raw_mode, MODES) if raw_mode else ('', False)
        if raw_mode and clean_mode not in MODES:
            issues.append(f'Unknown mode: "{raw_mode}" — could not auto-fix')
            if status == 'ok': status = 'error'
        elif raw_mode and mode_changed:
            fixes.append({'field':'Mode','original':raw_mode,'fixed':clean_mode,'how':'Auto-matched to standard value'})
            if status == 'ok': status = 'fixed'

        hours = _safe_float(raw_hrs) or 0.0

        results.append({
            'row_num':        i + 2,
            'status':         status,   # ok | fixed | error
            'fixes':          fixes,
            'issues':         issues,
            'emp_code':       clean_emp,
            'emp_name':       emp_name,
            'programme_name': clean_prog,
            'prog_type':      clean_type or raw_type,
            'mode':           clean_mode or raw_mode,
            'planned_hours':  hours,
        })
    return results


def _error_excel_for_tni(error_rows, dup_rows=None, plant_id=None, db=None):
    """Generate a pre-filled correction Excel for rows that couldn't be auto-fixed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rows To Fix'

    from openpyxl.worksheet.datavalidation import DataValidation
    hdr_fill = PatternFill('solid', fgColor='7F1D1D')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    headers  = ['Row #','Employee Code','Programme Name','Type of Programme',
                'Mode','Planned Hours','Issue(s) Found']
    col_w    = [7, 16, 34, 22, 14, 14, 60]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    err_fill = PatternFill('solid', fgColor='FEF2F2')
    for ri, r in enumerate(error_rows, 2):
        vals = [r['row_num'], r['emp_code'], r['programme_name'],
                r['prog_type'], r['mode'], r['planned_hours'],
                '; '.join(r['issues'])]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = err_fill
            if ci == 7:
                cell.font = Font(color='DC2626', size=10, italic=True)
                cell.alignment = Alignment(wrap_text=True)

    last_row = len(error_rows) + 1

    # Programme Name dropdown from master list
    if plant_id and db and last_row > 1:
        master_progs = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
        ).fetchall()] or MASTER_PROGRAMMES
        ws_pl = wb.create_sheet('_ProgList')
        ws_pl.sheet_state = 'hidden'
        for idx, v in enumerate(master_progs, 1):
            ws_pl.cell(row=idx, column=1, value=v)
        dv_prog = DataValidation(type='list',
                                 formula1=f'_ProgList!$A$1:$A${len(master_progs)}',
                                 allow_blank=True, showDropDown=False)
        dv_prog.sqref = f'C2:C{last_row}'
        ws.add_data_validation(dv_prog)

    # Dropdowns (only if there are actual error rows)
    if last_row > 1:
        dv_type = DataValidation(type='list', formula1=f'"{",".join(PROG_TYPES)}"', allow_blank=True)
        dv_mode = DataValidation(type='list', formula1=f'"{",".join(MODES)}"',      allow_blank=True)
        dv_type.sqref = f'D2:D{last_row}'
        dv_mode.sqref = f'E2:E{last_row}'
        ws.add_data_validation(dv_type)
        ws.add_data_validation(dv_mode)

    # Instructions sheet
    ws2 = wb.create_sheet('How To Fix')
    tips = [
        'These rows could NOT be auto-fixed. Please correct column by column:',
        '',
        'Employee Code — must exactly match the code in TMS employee master',
        'Programme Name — use the dropdown (click the cell) to pick from the master list',
        'Type of Programme — use the dropdown (click the cell)',
        'Mode — use the dropdown',
        '',
        'After fixing, save and re-upload on the TNI Analyzer page.',
        'Column G (Issues) shows exactly what was wrong — read it before fixing.',
    ]
    for ri, t in enumerate(tips, 1):
        c = ws2.cell(row=ri, column=1, value=t)
        if ri == 1: c.font = Font(bold=True, color='7F1D1D', size=12)
        ws2.column_dimensions['A'].width = 80

    # ── Duplicates sheet ──────────────────────────────────────────────────────
    if dup_rows:
        ws3 = wb.create_sheet('Duplicates')
        dup_hdr_fill = PatternFill('solid', fgColor='92400E')
        dup_headers  = ['Row #', 'Employee Code', 'Employee Name', 'Programme Name',
                        'Type', 'Mode', 'Duplicate Type']
        dup_col_w    = [7, 16, 28, 34, 22, 14, 60]
        for ci, (h, w) in enumerate(zip(dup_headers, dup_col_w), 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = dup_hdr_fill
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.row_dimensions[1].height = 24
        ws3.freeze_panes = 'A2'
        dup_fill = PatternFill('solid', fgColor='FFFBEB')
        for ri, r in enumerate(dup_rows, 2):
            vals = [r.get('row_num', ''), r['emp_code'], r.get('emp_name', ''),
                    r['programme_name'], r.get('prog_type', ''), r.get('mode', ''),
                    r['dup_type']]
            for ci, v in enumerate(vals, 1):
                cell = ws3.cell(row=ri, column=ci, value=v)
                cell.fill = dup_fill
                if ci == 7:
                    cell.font = Font(color='92400E', size=10, italic=True)
        ws3.cell(row=len(dup_rows)+3, column=1,
                 value='These rows were NOT imported. Fix or confirm before re-uploading.').font = Font(bold=True, color='92400E')

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


@app.route('/tni/analyze', methods=['GET', 'POST'])
@spoc_required
def tni_analyze():
    if request.method == 'GET':
        return render_template('tni_analyze.html', step='upload')

    f = request.files.get('file')
    if not f or f.filename == '':
        flash('No file selected.', 'danger')
        return render_template('tni_analyze.html', step='upload')
    skip = int(request.form.get('skip_rows', 0))
    try:
        raw   = f.read()
        fname = f.filename.lower()
        import pandas as _pd  # noqa: PLC0415
        if fname.endswith('.csv'):
            df = _pd.read_csv(_io.BytesIO(raw), dtype=str, skiprows=skip).fillna('')
        else:
            df = _pd.read_excel(_io.BytesIO(raw), dtype=str, skiprows=skip).fillna('')
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return render_template('tni_analyze.html', step='upload')

    plant_id = session['plant_id']
    db       = get_db()
    try:
        rows = _smart_analyze_rows(df, plant_id, db)
    except Exception as e:
        import traceback, logging
        logging.error(traceback.format_exc())
        flash(f'Analysis error: {e}', 'danger')
        return render_template('tni_analyze.html', step='upload')

    if not rows:
        col_list = ', '.join(f'"{c}"' for c in df.columns.tolist()[:15])
        flash(f'No data rows found. Columns detected: {col_list}. '
              f'Try increasing "Skip top rows" if headers are not on row 1.', 'warning')
        return render_template('tni_analyze.html', step='upload')

    aid = str(_uuid.uuid4())
    with open(os.path.join(BASE_DIR, 'data', f'tni_analyze_{aid}.json'), 'w') as fp:
        _json.dump(rows, fp)

    ok_count    = sum(1 for r in rows if r['status'] == 'ok')
    fixed_count = sum(1 for r in rows if r['status'] == 'fixed')
    warn_count  = sum(1 for r in rows if r['status'] == 'warning')
    err_count   = sum(1 for r in rows if r['status'] == 'error')

    plant_id_r  = session['plant_id']
    db_r        = get_db()
    master_progs = [r[0] for r in db_r.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id_r,)
    ).fetchall()] or MASTER_PROGRAMMES
    # Programmes confirmed clean in this upload (ok + fixed rows)
    upload_progs_lower = set(
        r['programme_name'].lower() for r in rows if r['status'] in ('ok', 'fixed')
    )
    return render_template('tni_analyze.html', step='review',
                           rows=rows, aid=aid,
                           ok_count=ok_count, fixed_count=fixed_count,
                           warn_count=warn_count, err_count=err_count,
                           master_progs=master_progs,
                           upload_progs_lower=upload_progs_lower,
                           prog_types=PROG_TYPES, modes=MODES, months=MONTHS_FY)


@app.route('/tni/analyze/confirm', methods=['POST'])
@spoc_required
def tni_analyze_confirm():
    aid  = request.form.get('aid', '')
    path = os.path.join(BASE_DIR, 'data', f'tni_analyze_{aid}.json')
    if not aid or not os.path.exists(path):
        flash('Session expired. Please re-upload.', 'danger')
        return redirect(url_for('tni_analyze'))

    with open(path) as fp:
        rows = _json.load(fp)

    plant_id  = session['plant_id']
    db        = get_db()
    inserted  = 0

    # ── Collect inline programme corrections submitted by user ─────────────────
    # fix_prog_{row_num} = corrected programme name chosen from dropdown
    corrections = {}
    for k, v in request.form.items():
        if k.startswith('fix_prog_') and v.strip():
            try:
                corrections[int(k[9:])] = v.strip()
            except ValueError:
                pass

    # ── Categorise blocked rows (errors + uncorrected warnings) ───────────────
    err_rows = []
    for r in rows:
        if r['status'] == 'error':
            err_rows.append(r)
        elif r['status'] == 'warning' and r['row_num'] not in corrections:
            err_rows.append(r)

    # ── Duplicate detection ───────────────────────────────────────────────────
    existing = set()
    for er in db.execute('SELECT emp_code, programme_name FROM tni WHERE plant_id=?', (plant_id,)):
        existing.add((er['emp_code'].strip().upper(), er['programme_name'].strip().lower()))

    dup_rows    = []
    updated     = 0
    seen_batch  = {}

    for row in rows:
        if row['status'] == 'error':
            continue
        if row['status'] == 'warning':
            fix = corrections.get(row['row_num'])
            if not fix:
                continue   # still blocked
            prog_name = fix
        else:
            prog_name = row['programme_name']

        key = (row['emp_code'].strip().upper(), prog_name.strip().lower())
        if key in seen_batch:
            dup_rows.append({**row, 'programme_name': prog_name,
                'dup_type': f'Same employee+programme already at Row {seen_batch[key]} in this file — first entry was imported, this row skipped'})
            continue
        seen_batch[key] = row['row_num']
        if key in existing:
            db.execute('''UPDATE tni SET prog_type=?, mode=?, planned_hours=?
                          WHERE plant_id=? AND UPPER(emp_code)=? AND LOWER(programme_name)=?''',
                       (row['prog_type'], row['mode'], row['planned_hours'],
                        plant_id, row['emp_code'].upper(), prog_name.lower()))
            updated += 1
        else:
            db.execute('INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours) VALUES(?,?,?,?,?,?)',
                       (plant_id, row['emp_code'], prog_name,
                        row['prog_type'], row['mode'], row['planned_hours']))
            inserted += 1

    db.commit()
    try: os.remove(path)
    except: pass

    if err_rows or dup_rows:
        buf = _error_excel_for_tni(err_rows, dup_rows=dup_rows, plant_id=plant_id, db=db)
        parts = []
        if err_rows:  parts.append(f'{len(err_rows)} errors')
        if dup_rows:  parts.append(f'{len(dup_rows)} duplicates in file')
        flash(f'{inserted} new + {updated} updated. {" & ".join(parts)} — downloading report.', 'warning')
        return send_file(buf, as_attachment=True,
                         download_name='TNI_Import_Issues.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    flash(f'{inserted} new entries added, {updated} existing entries updated — all clean!', 'success')
    return redirect(url_for('tni'))


# ─── ENTRY POINT ─────────────────────────────────────────────────────────────
try:
    init_db()
except Exception as _e:
    import logging; logging.error(f'init_db failed: {_e}')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
