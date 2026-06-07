"""Generate Excel listing all manual SPOC tasks before TMS, with time estimates."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SPOC Manual Effort"

# Header
headers = [
    "S.No", "Stage", "Task", "Description / Sub-activities",
    "Frequency", "Time per occurrence (hrs)", "Occurrences / year",
    "Total hrs / year per plant", "Tools used", "Pain points / risk"
]

# Task data: pre-TMS manual workflow per plant SPOC
tasks = [
    # Stage 1 — Employee Master
    (1, "Employee Master", "Annual master refresh",
     "Pull active headcount from HR / payroll. Map emp_code, name, designation, dept, collar (BC/WC). Reconcile new joiners + exits.",
     "Yearly", 8, 1, 8, "MS Excel, HR system exports",
     "Manual VLOOKUP errors. Stale exits not removed. Collar misclassified."),
    (2, "Employee Master", "Monthly delta update",
     "Add new joiners. Mark exits with date + reason. Update designation changes.",
     "Monthly", 2, 12, 24, "MS Excel",
     "Missed updates → trained-employee count off."),

    # Stage 2 — TNI
    (3, "TNI Collection", "Circulate TNI template",
     "Email blank TNI template to each HoD/manager. Follow-up calls. Multiple revisions.",
     "Yearly", 4, 1, 4, "Outlook, Excel",
     "HoDs delay 2-4 weeks. Versions multiply."),
    (4, "TNI Collection", "HoD-wise TNI receipt + chase",
     "Track who has submitted. Send reminders. Personal visits.",
     "Yearly", 12, 1, 12, "Outlook, phone",
     "5-10 reminders typical. SPOC time lost."),
    (5, "TNI Consolidation", "Merge HoD sheets into one",
     "Copy-paste rows from 10-20 HoD sheets into single master TNI. Reconcile column naming.",
     "Yearly", 8, 1, 8, "MS Excel",
     "Column mismatch. Rows lost. Duplicate emp+programme entries."),
    (6, "TNI Validation", "Emp code verification",
     "VLOOKUP each emp_code against employee master. Flag missing / mistyped.",
     "Yearly", 4, 1, 4, "MS Excel",
     "Wrong codes silently accepted. Affects coverage denominator."),
    (7, "TNI Validation", "Programme name normalisation",
     "Manually fix typo variants (e.g. 'EHS' vs 'E.H.S.' vs 'Environment Health Safety').",
     "Yearly", 6, 1, 6, "MS Excel, find-replace",
     "Spelling variants → orphan programmes → coverage % wrong."),
    (8, "TNI Validation", "Duplicate detection",
     "Sort by emp+programme. Remove duplicates manually.",
     "Yearly", 2, 1, 2, "MS Excel",
     "Duplicates inflate nominated count."),
    (9, "TNI Approval", "Cluster head / Plant head sign-off",
     "Print, take signature, scan back, file. Multiple iterations if rejected.",
     "Yearly", 6, 1, 6, "Print, scan, email",
     "Sign-off delays 1-2 weeks."),

    # Stage 3 — Programme Master
    (10, "Programme Master", "Extract unique programmes",
     "Pivot TNI by programme_name. List distinct programmes. Assign prog_type, mode, hrs, level.",
     "Yearly", 6, 1, 6, "MS Excel pivot",
     "Programmes not in master cause 2A/2C entry errors."),
    (11, "Programme Master", "Faculty/vendor mapping",
     "For each programme: identify trainer (internal/external), get rate, terms.",
     "Yearly", 16, 1, 16, "Email, phone, Excel",
     "Faculty availability changes. Quotes stale."),

    # Stage 4 — Calendar
    (12, "Calendar Planning", "Annual calendar draft",
     "Spread programmes across 12 months. Avoid harvest season, audits, festivals. Match faculty availability.",
     "Yearly", 12, 1, 12, "MS Excel",
     "Clash with production planning. Faculty no-show."),
    (13, "Calendar Planning", "Session-code assignment",
     "Manually generate session codes (UNIT/TYPE/NNN/FY/B01). Maintain running serial.",
     "Per session", 0.1, 250, 25, "MS Excel",
     "Duplicate codes. Wrong serial. Hard to track."),
    (14, "Calendar Planning", "Stakeholder circulation",
     "Email calendar to HoDs, plant head, central L&D. Collect feedback. Revise.",
     "Yearly", 6, 1, 6, "Outlook",
     "Multiple revisions. Version drift."),

    # Stage 5 — Pre-session
    (15, "Session Prep", "Nominee list per session",
     "Filter TNI for this programme → list of nominees → check who's available that day → finalise.",
     "Per session", 1, 250, 250, "MS Excel filter, phone",
     "Last-minute swaps. Wrong people show up."),
    (16, "Session Prep", "Venue + logistics booking",
     "Book training hall. Order refreshments. Arrange projector, printouts.",
     "Per session", 0.5, 250, 125, "Phone, email, paper",
     "Venue clash. Equipment failure."),
    (17, "Session Prep", "Faculty confirmation",
     "Confirm faculty 2 days before. Send agenda. Share attendee count.",
     "Per session", 0.3, 250, 75, "Phone, email",
     "Faculty cancellation last minute."),
    (18, "Session Prep", "Attendance register printout",
     "Print attendance sheet with emp_code + name + signature column.",
     "Per session", 0.2, 250, 50, "MS Word, printer",
     "Wrong list printed. Names misspelled."),
    (19, "Session Prep", "Pre-test / material printout",
     "Print pre-tests, handouts, certificates.",
     "Per session", 0.5, 250, 125, "Word, printer",
     "Printer down. Material outdated."),

    # Stage 6 — Day of session
    (20, "Day-of Session", "On-site coordination",
     "Receive attendees. Faculty intro. Manage logistics during session.",
     "Per session", 2, 250, 500, "In-person",
     "Late attendees. Faculty issues."),
    (21, "Day-of Session", "Attendance register signing",
     "Circulate register. Ensure all sign.",
     "Per session", 0.3, 250, 75, "Paper register",
     "Proxy signatures. Missed signers. Lost register."),
    (22, "Day-of Session", "Feedback form distribution",
     "Print, distribute, collect feedback forms at end.",
     "Per session", 0.5, 250, 125, "Paper",
     "Low response rate. Forms lost."),

    # Stage 7 — 2A entry
    (23, "2A Entry", "Transcribe paper register to Excel",
     "Type each attendee row by row from paper register into Sheet 2A. Map signature → present.",
     "Per session", 1.5, 250, 375, "MS Excel",
     "Transcription errors. Names mistyped. Wrong emp_code."),
    (24, "2A Entry", "Hours computation",
     "Calculate hours per attendee (full vs partial). Manual entry.",
     "Per session", 0.5, 250, 125, "MS Excel",
     "Hour padding. Inconsistent calc."),

    # Stage 8 — 2C entry
    (25, "2C Entry", "Programme details sheet fill",
     "Fill 2C: programme name, dates, hours, mode, faculty, vendor, cost, int/ext.",
     "Per session", 0.5, 250, 125, "MS Excel",
     "Missing fields. Wrong faculty name."),
    (26, "2C Entry", "Faculty signoff collection",
     "Print 2C draft, take faculty sign, scan, file.",
     "Per session", 0.5, 250, 125, "Print, scan",
     "Faculty delays. Lost signed copies."),
    (27, "2C Entry", "Vendor invoice attachment",
     "Collect invoice from vendor. Match cost. File copy.",
     "Per external session", 0.5, 80, 40, "Email, file",
     "Invoice delays. Mismatches with cost in 2C."),

    # Stage 9 — Monthly compilation
    (28, "Monthly Summary", "Consolidate month's 2A + 2C",
     "Combine all session 2A + 2C entries of the month. Build pivot.",
     "Monthly", 3, 12, 36, "MS Excel pivot",
     "Formula errors. Pivot scope drifts."),
    (29, "Monthly Summary", "Manhours + person-seats calc",
     "Sum manhours by prog_type × collar. Manual lookup against employee master.",
     "Monthly", 2, 12, 24, "MS Excel",
     "Lookup misses → undercounted manhours."),
    (30, "Monthly Summary", "Coverage % calc",
     "For each prog_type+collar: distinct trained / distinct nominated × 100. VLOOKUP-heavy.",
     "Monthly", 3, 12, 36, "MS Excel",
     "Formula bugs. Coverage misreported."),
    (31, "Monthly Summary", "Discrepancy reconciliation",
     "Cross-check summary numbers with raw 2A. Investigate gaps.",
     "Monthly", 2, 12, 24, "MS Excel",
     "Time-consuming. Repeat errors."),

    # Stage 10 — Central / dashboard
    (32, "Central Reporting", "Submit plant report to Central",
     "Email monthly Excel to central L&D. Respond to queries.",
     "Monthly", 1, 12, 12, "Outlook",
     "Delays. Format inconsistencies."),
    (33, "Central Reporting", "Central consolidation across plants",
     "[Central team] Merge 10 plant sheets into one. Build cross-plant dashboard manually.",
     "Monthly", 8, 12, 96, "MS Excel, PowerPoint",
     "Central effort multiplies. Errors compound."),
    (34, "Central Reporting", "Query / clarification loop",
     "Plants and central exchange clarifications over email/phone.",
     "Monthly", 2, 12, 24, "Email, phone",
     "Repeated cycles. Days lost."),

    # Stage 11 — Quarterly review
    (35, "Quarterly Review", "Quarterly compliance pack",
     "Build quarterly review deck. Prep slides for plant head.",
     "Quarterly", 6, 4, 24, "PowerPoint, Excel",
     "Inconsistent narrative. Last-minute data fixes."),
    (36, "Quarterly Review", "Review meeting + MoM",
     "Attend review, take MoM, distribute action items.",
     "Quarterly", 3, 4, 12, "Word, meeting",
     "Action items not tracked."),

    # Stage 12 — Audit
    (37, "Audit Prep", "ISO / customer audit evidence",
     "Compile attendance registers, 2C signed copies, vendor invoices, certificates, feedback forms.",
     "Yearly (2-3 audits)", 16, 3, 48, "Physical files, Excel",
     "Missing documents. Lost files. Re-creation under pressure."),
    (38, "Audit Prep", "NCR / observation closure",
     "Respond to audit observations. Provide gap evidence. Re-submit.",
     "Per audit", 8, 3, 24, "Word, Excel",
     "Multiple iterations. Re-opens."),

    # Stage 13 — Year-end
    (39, "Year-End", "FY data archival",
     "Archive prior FY files. Lock data. Prepare new FY workbook.",
     "Yearly", 4, 1, 4, "File system, Excel",
     "Data lost. Old links broken."),
    (40, "Year-End", "Carry-forward analysis",
     "Identify incomplete TNI items. Carry to next FY plan.",
     "Yearly", 4, 1, 4, "MS Excel",
     "Carryover missed."),

    # Stage 14 — Misc ongoing
    (41, "Ongoing", "Ad-hoc queries from HR / management",
     "Pull specific reports: dept training, individual employee record, vendor cost.",
     "Weekly", 1, 50, 50, "Excel filters",
     "Repeat effort. No single source of truth."),
    (42, "Ongoing", "Certificate generation + dispatch",
     "Generate certificates per attendee. Print, sign, distribute.",
     "Per session", 0.5, 250, 125, "Word mail-merge, printer",
     "Mail-merge errors. Lost in distribution."),
    (43, "Ongoing", "TNI mid-year correction requests",
     "Handle requests to add/remove TNI line items mid-year.",
     "Monthly", 2, 12, 24, "Email, Excel",
     "Untracked. Audit gap."),
    (44, "Ongoing", "Communication / coordination",
     "Email threads, calls, follow-ups across plant + central + faculty + vendor.",
     "Daily", 0.5, 250, 125, "Outlook, phone",
     "Time sink. No traceability."),
]

# Build sheet
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E78")
border_thin = Border(left=Side(style='thin', color='B0B0B0'),
                     right=Side(style='thin', color='B0B0B0'),
                     top=Side(style='thin', color='B0B0B0'),
                     bottom=Side(style='thin', color='B0B0B0'))

for col, h in enumerate(headers, 1):
    c = ws.cell(1, col, h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border_thin

for r, row in enumerate(tasks, 2):
    for col, val in enumerate(row, 1):
        c = ws.cell(r, col, val)
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = border_thin
        if col in (6, 7, 8):
            c.alignment = Alignment(horizontal='right', vertical='top')

# Totals row
total_row = len(tasks) + 2
ws.cell(total_row, 1, "TOTAL")
ws.cell(total_row, 8, f"=SUM(H2:H{len(tasks)+1})")
for col in range(1, 11):
    c = ws.cell(total_row, col)
    c.font = Font(bold=True, size=12)
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    c.border = border_thin
    c.alignment = Alignment(horizontal='right' if col == 8 else 'left', vertical='center')

# Column widths
widths = [6, 22, 30, 60, 18, 14, 14, 16, 28, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws.freeze_panes = "A2"

# Row heights
ws.row_dimensions[1].height = 36
for r in range(2, total_row + 1):
    ws.row_dimensions[r].height = 50

# Second sheet — summary by stage
ws2 = wb.create_sheet("Summary by Stage")
ws2.append(["Stage", "Tasks", "Total hrs/yr/plant"])
from collections import defaultdict
agg = defaultdict(lambda: [0, 0])
for t in tasks:
    stage = t[1]
    agg[stage][0] += 1
    agg[stage][1] += t[7]
for stage, (n, h) in agg.items():
    ws2.append([stage, n, h])
ws2.append(["TOTAL", sum(v[0] for v in agg.values()), sum(v[1] for v in agg.values())])

for col in range(1, 4):
    ws2.cell(1, col).font = header_font
    ws2.cell(1, col).fill = header_fill
    ws2.cell(1, col).alignment = Alignment(horizontal='center')
    ws2.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 18

# Bold total
last = ws2.max_row
for col in range(1, 4):
    ws2.cell(last, col).font = Font(bold=True)
    ws2.cell(last, col).fill = PatternFill("solid", fgColor="FFF2CC")

# Third sheet — assumptions
ws3 = wb.create_sheet("Assumptions")
ws3.append(["Item", "Value"])
ws3.append(["Plant SPOC count", "1 per plant (10 plants total at BCML)"])
ws3.append(["Annual sessions per plant (avg)", "250"])
ws3.append(["External (paid) sessions per plant", "~80 (32%)"])
ws3.append(["TNI nominees per plant", "1500-3000 unique"])
ws3.append(["Audits per year", "ISO + customer + internal = ~3"])
ws3.append(["Time basis", "Indian PSU/manufacturing benchmark; observed averages"])
ws3.append(["Conversion: hours/year/plant → person-days", "Divide by 8"])
ws3.append(["Conversion: per BCML (10 plants)", "Multiply per-plant by 10"])
for col in (1, 2):
    ws3.column_dimensions[get_column_letter(col)].width = 40
ws3.cell(1, 1).font = header_font; ws3.cell(1, 1).fill = header_fill
ws3.cell(1, 2).font = header_font; ws3.cell(1, 2).fill = header_fill

out = r"docs\SPOC_Manual_Effort_Pre_TMS.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total tasks: {len(tasks)}")
print(f"Total hrs/year/plant: {sum(t[7] for t in tasks)}")
print(f"Person-days/year/plant: {sum(t[7] for t in tasks)/8:.1f}")
print(f"Across 10 plants (man-days/yr): {sum(t[7] for t in tasks)*10/8:.0f}")
