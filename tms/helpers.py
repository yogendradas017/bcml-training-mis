import os
import re
import io
import glob
import time
import json as _json
import uuid as _uuid
from datetime import date, datetime
from difflib import get_close_matches

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import request, send_file

from tms.constants import (
    BASE_DIR, TEMP_UPLOAD_DIR, PLANT_MAP, TYPE_ABBREV,
    PROG_TYPES, MODES, MONTHS_FY, MONTH_NUM, NON_TNI_SOURCES
)

# ── Upload helpers ────────────────────────────────────────────────────────────

def _is_ajax():
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


# Bound bulk-upload size: MAX_CONTENT_LENGTH caps the raw bytes, but a small
# spreadsheet can still expand to a huge row count that OOMs the single worker
# or makes the per-row fuzzy match O(rows × master) catastrophic for ALL tenants.
# Cap the row count post-parse (generous: full headcount × several programmes).
_MAX_UPLOAD_ROWS = 60000


def _enforce_row_cap(df):
    if len(df) > _MAX_UPLOAD_ROWS:
        raise ValueError(
            f'File has {len(df):,} rows — the maximum per upload is '
            f'{_MAX_UPLOAD_ROWS:,}. Split it into smaller files.')
    return df


def _read_upload_file(file_storage):
    import pandas as pd
    fname = file_storage.filename.lower()
    if fname.endswith('.csv'):
        return _enforce_row_cap(pd.read_csv(file_storage, dtype=str).fillna(''))
    return _enforce_row_cap(pd.read_excel(file_storage, dtype=str).fillna(''))


def _read_upload_file_path(path):
    import pandas as pd
    if path.lower().endswith('.csv'):
        return _enforce_row_cap(pd.read_csv(path, dtype=str).fillna(''))
    return _enforce_row_cap(pd.read_excel(path, dtype=str).fillna(''))


def _clean(row, keys):
    cols = row.keys() if hasattr(row, 'keys') else row.index
    for k in keys:
        for col in cols:
            if str(col).strip().lower() == k:
                val = str(row[col]).strip()
                return '' if val.lower() in ('nan', 'none', '') else val
    return ''


def normalise_collar(val):
    v = str(val).strip().upper()
    if any(x in v for x in ['WHITE', 'WC', 'W C']):
        return 'White Collared'
    if any(x in v for x in ['BLUE', 'BC', 'B C']):
        return 'Blue Collared'
    return val.strip()


def _safe_float(val):
    try:
        return float(val) if val and str(val).strip() != '' else None
    except (ValueError, TypeError):
        return None


def _current_fy():
    """Returns (fy_start, fy_end) as 'YYYY-MM-DD' strings for the current financial year (Apr–Mar).
    Uses IST wall-clock — Render runs UTC and would roll the FY 5.5h early otherwise."""
    today = _today_ist()
    yr = today.year if today.month >= 4 else today.year - 1
    return f'{yr}-04-01', f'{yr+1}-03-31'


def _in_current_fy(date_str):
    """True if date_str falls within the current FY, or is empty/None."""
    if not date_str:
        return True
    try:
        d = date.fromisoformat(str(date_str)[:10])
        s, e = _current_fy()
        return date.fromisoformat(s) <= d <= date.fromisoformat(e)
    except (ValueError, TypeError):
        return False


def _tni_is_locked():
    """True if today is past March 31 of the current FY (TNI write window closed).
    Uses IST so the lock flips at IST midnight, not 18:30 IST (UTC midnight)."""
    _, fy_end = _current_fy()
    return _today_ist() > date.fromisoformat(fy_end)


def _recompute_session_actuals(plant_id, session_code, db):
    """Refresh calendar.actual_pax + actual_hrs from emp_training. Idempotent.

    Host-aware: a central (plant 99) session's attendees are stored under their
    HOME plant_id with host_plant_id=99, so matching only plant_id=99 would miss
    them all and leave the central calendar's actuals stuck at 0. Match
    (plant_id OR host_plant_id) = the calendar row's plant. For ordinary plant
    sessions host_plant_id is NULL, so the OR adds nothing — safe both ways."""
    if not session_code:
        return
    r = db.execute(
        'SELECT COUNT(*) AS pax, COALESCE(SUM(hrs),0) AS hrs '
        'FROM emp_training WHERE session_code=? AND (plant_id=? OR host_plant_id=?)',
        (session_code, plant_id, plant_id)).fetchone()
    db.execute(
        'UPDATE calendar SET actual_pax=?, actual_hrs=? WHERE session_code=? AND plant_id=?',
        (r['pax'], r['hrs'], session_code, plant_id))


def _date_to_month(date_str):
    if not date_str:
        return ''
    try:
        d = datetime.strptime(str(date_str)[:10], '%Y-%m-%d')
        return d.strftime('%B')
    except Exception:
        return ''


# ── IST-aware datetime helpers ──
# Render runs in UTC; user-facing dates/times must be IST. Naive datetime.now()
# / date.today() drift by 5.5 hours and have already caused production bugs
# (lockout view skew, "session has not started" gate false-rejecting at 00:00-05:30 IST,
# month-end KPI snapback). Always use _now_ist / _today_ist for user-facing values.
def _now_ist():
    """Current datetime in IST (Asia/Kolkata). Returns timezone-naive datetime
    representing IST wall-clock time — drop-in for datetime.now() but stable
    on UTC servers."""
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo('Asia/Kolkata')).replace(tzinfo=None)
    except Exception:
        from datetime import timedelta as _td
        return datetime.utcnow() + _td(hours=5, minutes=30)


def _today_ist():
    """Current date in IST. Drop-in for date.today()."""
    return _now_ist().date()


def _time_to_minutes(hhmm):
    """'HH:MM' → integer minutes since midnight. None on bad input."""
    if not hhmm or ':' not in str(hhmm):
        return None
    try:
        h, m = str(hhmm).strip().split(':')[:2]
        h, m = int(h), int(m)
        if 0 <= h <= 23 and 0 <= m <= 59:
            return h * 60 + m
    except (ValueError, TypeError):
        pass
    return None


def _validate_time_vs_duration(time_from, time_to, total_hours,
                                start_date='', end_date='',
                                tolerance_min=15):
    """Cross-check that (End Time − Start Time) × days ≈ total_hours.

    Used by Calendar / 2C validators (2A skips this — per-person hrs is
    apples-to-oranges vs session window). All inputs optional — passes when
    not enough data to check.

    For multi-day sessions, time_from/time_to is per-day window; total_hours
    is cumulative across days. So expected = per_day_hrs × days.

    Required-pair: if exactly one of time_from / time_to is set, blocks with
    a 'must provide both' error.

    Returns (ok: bool, msg: str). tolerance_min default = 15 min slack.
    """
    tf_set, tt_set = bool(time_from), bool(time_to)
    if tf_set != tt_set:  # exactly one set
        which = 'Start Time' if tt_set else 'End Time'
        return False, f'{which} is required when the other is provided.'
    if not (time_from and time_to and total_hours):
        return True, ''
    try:
        total = float(total_hours)
    except (ValueError, TypeError):
        return True, ''
    if total <= 0:
        return True, ''
    fmin = _time_to_minutes(time_from)
    tmin = _time_to_minutes(time_to)
    if fmin is None or tmin is None:
        return True, ''
    if tmin <= fmin:
        return False, f'End Time ({time_to}) must be after Start Time ({time_from}).'
    per_day_hrs = (tmin - fmin) / 60.0

    days = 1
    if start_date and end_date:
        try:
            from datetime import date as _d
            d1 = _d.fromisoformat(str(start_date)[:10])
            d2 = _d.fromisoformat(str(end_date)[:10])
            days = max(1, (d2 - d1).days + 1)
        except (ValueError, TypeError):
            pass

    expected = per_day_hrs * days
    diff_min = abs(expected - total) * 60
    if diff_min > tolerance_min:
        day_part = f' × {days} day(s)' if days > 1 else ''
        return False, (
            f'Time window does not match Duration. '
            f'{time_from}–{time_to}{day_part} = {expected:.2f} hrs, '
            f'but Duration is set to {total:g} hrs. '
            f'Fix one: Start Time, End Time, or Duration (Hrs).'
        )
    return True, ''


def _parse_date_strict(raw):
    """Accept DD-MM-YYYY only (bulk upload user input). Returns YYYY-MM-DD for storage.
    Returns '' for empty. Raises ValueError with message on bad format."""
    s = str(raw).strip()[:10] if raw else ''
    if not s or s in ('nan', 'None', '-'):
        return ''
    try:
        return datetime.strptime(s, '%d-%m-%Y').strftime('%Y-%m-%d')
    except ValueError:
        raise ValueError(f'"{s}" is not DD-MM-YYYY')


def _get_programme_names(plant_id, db):
    rows = db.execute(
        'SELECT DISTINCT programme_name FROM calendar WHERE plant_id=? ORDER BY programme_name',
        (plant_id,)).fetchall()
    return [r['programme_name'] for r in rows]


# ── Audience derivation ───────────────────────────────────────────────────────

def _derive_audience(plant_id, prog_name, db):
    collars = db.execute('''
        SELECT DISTINCT e.collar FROM tni t
        JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND LOWER(t.programme_name)=LOWER(?) AND t.fy_year=?
          AND e.collar IS NOT NULL AND e.collar != ''
    ''', (plant_id, prog_name, _fy_label())).fetchall()
    collar_set = {r['collar'] for r in collars}
    if not collar_set:
        return None
    if 'Blue Collared' in collar_set and 'White Collared' in collar_set:
        return 'Common'
    if 'Blue Collared' in collar_set:
        return 'Blue Collared'
    return 'White Collared'


# ── Session code helpers ──────────────────────────────────────────────────────

def _fy_label():
    """Returns short FY label like '26-27' for use in session codes.
    Uses IST so session-code FY rolls at IST midnight, not UTC midnight."""
    today = _today_ist()
    y = today.year
    return f'{str(y-1)[2:]}-{str(y)[2:]}' if today.month < 4 else f'{str(y)[2:]}-{str(y+1)[2:]}'


def _fy_label_long():
    """Returns long FY label like '2026–27' (en-dash) for UI display.
    Single source of truth — never hardcode in templates."""
    today = _today_ist()
    y = today.year
    start_yr = y - 1 if today.month < 4 else y
    end_yr = start_yr + 1
    return f'{start_yr}–{str(end_yr)[2:]}'  # en-dash matches historical literals


def coverage_universe(db, plant_id, fy_start, fy_end, fy):
    """CANONICAL TNI-coverage inputs — the SINGLE source of truth used identically
    by the Monthly Summary, Dashboard QC charts, and Programme Intelligence, so the
    coverage % is the same number in every module.

    Returns (nominations, trained):
      nominations — one dict {emp, prog, prog_type, collar} per current-FY
        'TNI Driven' nomination of an ACTIVE Blue/White-collar employee. `prog`
        is lower-cased for case-insensitive matching.
      trained — set of (emp, prog) with a 2A (emp_training) attendance record for
        that programme dated WITHIN the current FY (attended-within-FY rule, per
        CLAUDE.md _calc_compliance — not conducted-gated).
    A nomination is 'covered' iff its (emp, prog) is in `trained`. Coverage % for
    any slice = covered nominations / total nominations in that slice.
    """
    nominations = [
        {'emp': r['emp'], 'prog': r['prog'], 'prog_display': r['prog_display'],
         'prog_type': r['pt'], 'collar': r['collar']}
        for r in db.execute('''
            SELECT t.emp_code AS emp, LOWER(t.programme_name) AS prog,
                   t.programme_name AS prog_display,
                   t.prog_type AS pt, e.collar AS collar
            FROM tni t
            JOIN employees e ON e.emp_code = t.emp_code AND e.plant_id = t.plant_id
            WHERE t.plant_id = ? AND t.fy_year = ? AND t.source = 'TNI Driven'
              AND e.is_active = 1
              AND e.collar IN ('Blue Collared', 'White Collared')
        ''', (plant_id, fy)).fetchall()
    ]
    trained = {
        (r['emp'], r['prog']) for r in db.execute('''
            SELECT DISTINCT emp_code AS emp, LOWER(programme_name) AS prog
            FROM emp_training
            WHERE plant_id = ? AND start_date BETWEEN ? AND ?
        ''', (plant_id, fy_start, fy_end)).fetchall()
    }
    return nominations, trained


def _get_or_create_prog_code(plant_id, prog_name, prog_type, db):
    # Atomic: reuse existing prog_code if the programme already has one.
    existing = db.execute(
        'SELECT prog_code FROM calendar WHERE plant_id=? AND programme_name=? LIMIT 1',
        (plant_id, prog_name)).fetchone()
    if existing:
        return existing['prog_code']
    unit_code = PLANT_MAP[plant_id]['unit_code']
    abbrev    = TYPE_ABBREV.get(prog_type, 'GEN')
    prefix    = f'{unit_code}/{abbrev}/'
    # Use MAX of the numeric suffix instead of COUNT(DISTINCT) — COUNT is racy
    # when two workers insert concurrently (both see N, both pick N+1, collision).
    # MAX+1 is also racy on its own, so caller wraps session_code creation in a
    # retry loop guarded by the UNIQUE(session_code) constraint, which catches
    # any duplicate prog_code attempt as a downstream session_code conflict.
    row = db.execute(
        "SELECT MAX(CAST(SUBSTR(prog_code, ?+1) AS INTEGER)) AS mx "
        "FROM calendar WHERE plant_id=? AND prog_code LIKE ?",
        (len(prefix), plant_id, f'{prefix}%')).fetchone()
    nxt = (row['mx'] or 0) + 1
    return f'{prefix}{nxt:03d}'


def _new_session_code(plant_id, prog_code, db):
    """Generate next session_code for (plant_id, prog_code) in current FY.
    Uses MAX(suffix)+1; if two workers race, the UNIQUE(session_code) constraint
    on calendar will reject one — caller must retry. We probe with a few attempts
    here so most callers don't need their own loop."""
    import sqlite3
    fy     = _fy_label()
    prefix = f'{prog_code}/{fy}/B'
    for attempt in range(5):
        row = db.execute(
            "SELECT MAX(CAST(SUBSTR(session_code, ?+1) AS INTEGER)) AS mx "
            "FROM calendar WHERE plant_id=? AND prog_code=? AND session_code LIKE ?",
            (len(prefix), plant_id, prog_code, f'{prefix}%')).fetchone()
        nxt  = (row['mx'] or 0) + 1 + attempt   # bump on retry to dodge races
        code = f'{prefix}{nxt:02d}'
        # Pre-check uniqueness inside same txn — cheap guard; UNIQUE constraint
        # on calendar.session_code is the final authority.
        clash = db.execute(
            'SELECT 1 FROM calendar WHERE session_code=? LIMIT 1', (code,)).fetchone()
        if not clash:
            return code
    # Last-resort: include microsecond to guarantee uniqueness rather than crash.
    from datetime import datetime
    return f'{prefix}{(row["mx"] or 0)+1:02d}X{datetime.utcnow().strftime("%f")}'


# ── Calendar sync ─────────────────────────────────────────────────────────────

def _sync_calendar_from_2c(plant_id, db):
    # Guard: never overwrite verify-queue / terminal states. Only promote
    # in-flight statuses (e.g. 'To Be Planned', 'Planned') to 'Conducted'.
    # Without this guard, a session sitting in 'Awaiting Verification' would
    # be silently flipped to 'Conducted' on any calendar page load, bypassing
    # the SPOC verification step. Lapsed / Cancelled / Re-Scheduled are
    # terminal and must also be preserved.
    db.execute('''UPDATE calendar SET status='Conducted'
        WHERE plant_id=? AND session_code IN
        (SELECT session_code FROM programme_details WHERE plant_id=?)
        AND status NOT IN ('Awaiting Verification','Conducted','Lapsed','Cancelled','Re-Scheduled')''',
        (plant_id, plant_id))
    # Audience sync — set-based. Was a per-programme loop calling _derive_audience
    # (a tni⋈employees join with non-sargable LOWER()) once per distinct calendar
    # programme: M programmes = M join queries = the old ~4-5s /calendar load.
    # Now: one GROUP BY derives every audience, one SELECT reads current calendar
    # audiences, and only true diffs are batched via executemany.
    fy = _fy_label()
    aud_rows = db.execute('''
        SELECT LOWER(t.programme_name) AS pname,
               CASE
                 WHEN MAX(CASE WHEN e.collar='Blue Collared'  THEN 1 ELSE 0 END)=1
                  AND MAX(CASE WHEN e.collar='White Collared' THEN 1 ELSE 0 END)=1 THEN 'Common'
                 WHEN MAX(CASE WHEN e.collar='Blue Collared'  THEN 1 ELSE 0 END)=1 THEN 'Blue Collared'
                 WHEN MAX(CASE WHEN e.collar='White Collared' THEN 1 ELSE 0 END)=1 THEN 'White Collared'
                 ELSE NULL
               END AS aud
        FROM tni t
        JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND t.fy_year=?
          AND e.collar IN ('Blue Collared','White Collared')
        GROUP BY LOWER(t.programme_name)
    ''', (plant_id, fy)).fetchall()
    aud_map = {r['pname']: r['aud'] for r in aud_rows if r['aud']}
    if aud_map:
        # Current calendar audiences, keyed by lowercase programme name. Only rows
        # whose derived audience differs from the stored value get an UPDATE.
        cur = db.execute(
            'SELECT id, LOWER(programme_name) AS pname, target_audience FROM calendar WHERE plant_id=?',
            (plant_id,)).fetchall()
        updates = [(aud_map[c['pname']], c['id'])
                   for c in cur
                   if c['pname'] in aud_map and c['target_audience'] != aud_map[c['pname']]]
        if updates:
            db.executemany('UPDATE calendar SET target_audience=? WHERE id=?', updates)
    db.commit()


# ── Programme master helpers ──────────────────────────────────────────────────

def _sync_master_from_tni(plant_id, db):
    fy = _fy_label()
    rows = db.execute('''
        SELECT programme_name,
               (SELECT prog_type FROM tni t2
                WHERE t2.plant_id=t.plant_id AND t2.programme_name=t.programme_name
                  AND t2.fy_year=? AND t2.prog_type IS NOT NULL AND t2.prog_type != ""
                GROUP BY prog_type ORDER BY COUNT(*) DESC LIMIT 1) AS top_type,
               CASE WHEN EXISTS(
                   SELECT 1 FROM tni t3 WHERE t3.plant_id=t.plant_id
                   AND t3.programme_name=t.programme_name AND t3.fy_year=?
                   AND t3.source='TNI Driven'
               ) THEN 'TNI Requirement' ELSE 'New Requirement' END AS derived_source
        FROM tni t
        WHERE plant_id=? AND fy_year=? AND programme_name IS NOT NULL AND programme_name != ""
        GROUP BY programme_name
    ''', (fy, fy, plant_id, fy)).fetchall()
    if not rows:
        return
    # Batch INSERT — single executemany instead of N individual INSERTs
    db.executemany(
        'INSERT OR IGNORE INTO programme_master(plant_id, name, prog_type, source) VALUES(?,?,?,?)',
        [(plant_id, r['programme_name'], r['top_type'], r['derived_source']) for r in rows]
    )
    for r in rows:
        if r['top_type']:
            db.execute(
                'UPDATE programme_master SET prog_type=? WHERE plant_id=? AND LOWER(name)=LOWER(?) AND (prog_type IS NULL OR prog_type="")',
                (r['top_type'], plant_id, r['programme_name']))
        db.execute(
            'UPDATE programme_master SET source=? WHERE plant_id=? AND LOWER(name)=LOWER(?)',
            (r['derived_source'], plant_id, r['programme_name']))


def _prog_in_use(prog_name, plant_id, db):
    for table in ('tni', 'calendar', 'emp_training'):
        if db.execute(
                f'SELECT 1 FROM {table} WHERE plant_id=? AND LOWER(programme_name)=LOWER(?) LIMIT 1',
                (plant_id, prog_name)).fetchone():
            return True
    return False


# ── Summary calculations ──────────────────────────────────────────────────────

def _calc_summary(plant_id, month_filter, db):
    """Per-prog_type Summary table. Refactored 2026-05-30: was 10 queries × 6
    prog_types = 60 round-trips; now 4 GROUP BY queries × 1 = 4 round-trips.
    Same outputs, ~15× faster on cold disk."""
    fy   = _fy_label()
    # FY date bounds — man-hours/seats must be scoped to the current financial
    # year to match the Dashboard gauge and the /central card (which are
    # FY-bound). Without this the same plant showed an all-time Summary total
    # next to an FY dashboard total.
    fy_start, fy_end = _current_fy()
    mn = MONTH_NUM.get(month_filter, '') if month_filter else ''
    # ms_* clauses force empty result when a month is selected but missing
    # from MONTH_NUM (preserves old 'AND 1=0' semantics).
    month_pd_clause      = f"AND strftime('%m', p.start_date)='{mn}'"  if mn else ("AND 1=0" if month_filter else "")
    month_et_clause      = f"AND strftime('%m', t.start_date)='{mn}'"  if mn else ("AND 1=0" if month_filter else "")
    month_central_clause = f"AND strftime('%m', et.start_date)='{mn}'" if mn else ("AND 1=0" if month_filter else "")
    month_pdx_clause     = f"AND strftime('%m', pd.start_date)='{mn}'" if mn else ("AND 1=0" if month_filter else "")

    # Query 1: programme_details aggregates per prog_type.
    # Count only CONDUCTED programmes (policy: a 2C row sitting in 'Awaiting
    # Verification' must NOT be counted until Central verifies it — this matches
    # the Dashboard/export 'Conducted' gate so the screens agree). LEFT JOIN so
    # legacy 'New Program' 2C rows that have no calendar entry are still counted.
    pd_rows = db.execute(f'''
        SELECT p.prog_type AS prog_type,
               COUNT(DISTINCT CASE WHEN p.audience='Blue Collared'  THEN p.programme_name END) AS bc_progs,
               COUNT(DISTINCT CASE WHEN p.audience='White Collared' THEN p.programme_name END) AS wc_progs,
               COUNT(DISTINCT CASE WHEN p.audience='Common'         THEN p.programme_name END) AS common_progs,
               COUNT(DISTINCT p.programme_name) AS total_progs,
               COUNT(DISTINCT CASE WHEN p.int_ext='Internal' THEN p.programme_name END) AS int_prog,
               COUNT(DISTINCT CASE WHEN p.int_ext='External' THEN p.programme_name END) AS ext_prog
        FROM programme_details p
        LEFT JOIN calendar c
               ON c.session_code=p.session_code AND c.plant_id=p.plant_id
        WHERE p.plant_id=? {month_pd_clause}
          AND (c.session_code IS NULL OR c.status='Conducted')
        GROUP BY p.prog_type
    ''', [plant_id]).fetchall()
    pd_map = {r['prog_type']: dict(r) for r in pd_rows}

    # Query 2: central-hosted programmes (exclude already-in-2C). Group by
    # prog_type + programme_name; classify collar mix in Python.
    central_rows = db.execute(f'''
        SELECT LOWER(et.prog_type) AS pt_lc, et.programme_name,
               SUM(CASE WHEN e.collar='Blue Collared'  THEN 1 ELSE 0 END) AS bc_cnt,
               SUM(CASE WHEN e.collar='White Collared' THEN 1 ELSE 0 END) AS wc_cnt
        FROM emp_training et
        LEFT JOIN employees e
               ON e.emp_code=et.emp_code AND e.plant_id=et.plant_id
        WHERE et.plant_id=? AND et.host_plant_id=99
          AND et.start_date BETWEEN ? AND ?
          {month_central_clause}
          AND NOT EXISTS (
              SELECT 1 FROM programme_details pd
              WHERE pd.plant_id=et.plant_id
              AND LOWER(pd.programme_name)=LOWER(et.programme_name)
              AND LOWER(pd.prog_type)=LOWER(et.prog_type)
              {month_pdx_clause}
          )
        GROUP BY pt_lc, et.programme_name
    ''', [plant_id, fy_start, fy_end]).fetchall()
    # Per-prog_type tally: {pt_lc: {'bc':n, 'wc':n, 'common':n, 'total':n}}
    central_tally = {}
    for r in central_rows:
        bc_cnt, wc_cnt = r['bc_cnt'] or 0, r['wc_cnt'] or 0
        bucket = central_tally.setdefault(r['pt_lc'],
                    {'bc': 0, 'wc': 0, 'common': 0, 'total': 0})
        if bc_cnt and wc_cnt:
            bucket['common'] += 1; bucket['total'] += 1
        elif bc_cnt:
            bucket['bc'] += 1;     bucket['total'] += 1
        elif wc_cnt:
            bucket['wc'] += 1;     bucket['total'] += 1

    # Query 3: emp_training seats + hrs per prog_type + collar
    seat_rows = db.execute(f'''
        SELECT t.prog_type, e.collar,
               COUNT(*) AS seats,
               COALESCE(SUM(t.hrs),0) AS hrs
        FROM emp_training t
        JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND t.start_date BETWEEN ? AND ? {month_et_clause}
          AND e.collar IN ('Blue Collared', 'White Collared')
        GROUP BY t.prog_type, e.collar
    ''', [plant_id, fy_start, fy_end]).fetchall()
    seat_map = {}
    for r in seat_rows:
        seat_map.setdefault(r['prog_type'], {})[r['collar']] = (r['seats'], r['hrs'])

    # Query 4: TNI nominations + fulfilment per prog_type + collar.
    # Uses the CANONICAL coverage rule (coverage_universe) so this % is identical
    # to the Dashboard QC and Programme Intelligence for the same plant.
    _noms, _trained = coverage_universe(db, plant_id, fy_start, fy_end, fy)
    tni_map = {}
    for n in _noms:
        cell = tni_map.setdefault(n['prog_type'], {}).setdefault(n['collar'], [0, 0])
        cell[0] += 1
        if (n['emp'], n['prog']) in _trained:
            cell[1] += 1

    # Assemble rows in PROG_TYPES order
    rows = []
    for pt in PROG_TYPES:
        pd_row = pd_map.get(pt, {})
        bc_progs     = pd_row.get('bc_progs', 0)
        wc_progs     = pd_row.get('wc_progs', 0)
        common_progs = pd_row.get('common_progs', 0)
        total_progs  = pd_row.get('total_progs', 0)
        int_prog     = pd_row.get('int_prog', 0)
        ext_prog     = pd_row.get('ext_prog', 0)

        ct = central_tally.get(pt.lower(), {})
        bc_progs    += ct.get('bc', 0)
        wc_progs    += ct.get('wc', 0)
        common_progs+= ct.get('common', 0)
        total_progs += ct.get('total', 0)

        sm = seat_map.get(pt, {})
        bc_seats, bc_hrs = sm.get('Blue Collared',  (0, 0))
        wc_seats, wc_hrs = sm.get('White Collared', (0, 0))

        tm = tni_map.get(pt, {})
        bc_fixed, bc_cum = tm.get('Blue Collared',  (0, 0))
        wc_fixed, wc_cum = tm.get('White Collared', (0, 0))

        bc_cov  = round(bc_cum  / bc_fixed  * 100, 1) if bc_fixed  else 0
        wc_cov  = round(wc_cum  / wc_fixed  * 100, 1) if wc_fixed  else 0
        tot_cov = round((bc_cum + wc_cum) / (bc_fixed + wc_fixed) * 100, 1) if (bc_fixed + wc_fixed) else 0

        rows.append({
            'prog_type':    pt,
            'bc_progs':     bc_progs,    'wc_progs':  wc_progs,
            'common_progs': common_progs,'total_progs': total_progs,
            'int_prog':     int_prog,    'ext_prog':  ext_prog,
            'bc_seats':     bc_seats,    'wc_seats':  wc_seats,
            'total_seats':  bc_seats + wc_seats,
            'bc_hrs':       round(bc_hrs, 1), 'wc_hrs': round(wc_hrs, 1),
            'total_hrs':    round(bc_hrs + wc_hrs, 1),
            'bc_fixed':     bc_fixed,    'wc_fixed':  wc_fixed,
            'bc_cum':       bc_cum,      'wc_cum':    wc_cum,
            'bc_cov':       bc_cov,      'wc_cov':    wc_cov,
            'tot_cov':      tot_cov,
        })
    return rows


def _calc_totals(rows, db=None, plant_id=None):
    if not rows:
        return {}
    t = {k: 0 for k in rows[0]}
    t['prog_type'] = 'TOTAL'
    skip = {'prog_type', 'bc_cov', 'wc_cov', 'tot_cov', 'bc_fixed', 'wc_fixed', 'bc_cum', 'wc_cum'}
    for r in rows:
        for k, v in r.items():
            if k not in skip:
                t[k] = round(t.get(k, 0) + (v or 0), 1)
    if db is not None and plant_id is not None:
        fy = _fy_label()
        fy_start, fy_end = _current_fy()
        _noms, _trained = coverage_universe(db, plant_id, fy_start, fy_end, fy)
        _bc = [n for n in _noms if n['collar'] == 'Blue Collared']
        _wc = [n for n in _noms if n['collar'] == 'White Collared']
        t['bc_fixed'] = len(_bc)
        t['wc_fixed'] = len(_wc)
        t['bc_cum'] = sum(1 for n in _bc if (n['emp'], n['prog']) in _trained)
        t['wc_cum'] = sum(1 for n in _wc if (n['emp'], n['prog']) in _trained)
    else:
        t['bc_fixed'] = sum(r.get('bc_fixed', 0) or 0 for r in rows)
        t['wc_fixed'] = sum(r.get('wc_fixed', 0) or 0 for r in rows)
        t['bc_cum']   = sum(r.get('bc_cum',   0) or 0 for r in rows)
        t['wc_cum']   = sum(r.get('wc_cum',   0) or 0 for r in rows)
    t['bc_cov']  = round(t['bc_cum']  / t['bc_fixed']  * 100, 1) if t.get('bc_fixed')  else 0
    t['wc_cov']  = round(t['wc_cum']  / t['wc_fixed']  * 100, 1) if t.get('wc_fixed')  else 0
    t['tot_cov'] = round((t['bc_cum'] + t['wc_cum']) / (t['bc_fixed'] + t['wc_fixed']) * 100, 1) \
                   if (t.get('bc_fixed', 0) + t.get('wc_fixed', 0)) else 0
    return t


def _calc_compliance(plant_id, db):
    fy_start, fy_end = _current_fy()
    bc = db.execute(
        "SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='Blue Collared'",
        (plant_id,)).fetchone()[0]
    wc = db.execute(
        "SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1 AND collar='White Collared'",
        (plant_id,)).fetchone()[0]
    bc_act = db.execute('''SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t
        JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND e.collar='Blue Collared'
          AND t.start_date BETWEEN ? AND ?''',
        (plant_id, fy_start, fy_end)).fetchone()[0]
    wc_act = db.execute('''SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t
        JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
        WHERE t.plant_id=? AND e.collar='White Collared'
          AND t.start_date BETWEEN ? AND ?''',
        (plant_id, fy_start, fy_end)).fetchone()[0]
    from tms.config import get_config
    bc_target = get_config('mh_target_bc', 12, plant_id=plant_id)
    wc_target = get_config('mh_target_wc', 24, plant_id=plant_id)
    bc_mandate = bc * bc_target
    wc_mandate = wc * wc_target
    bc_pct = round(bc_act / bc_mandate * 100, 1) if bc_mandate else 0
    wc_pct = round(wc_act / wc_mandate * 100, 1) if wc_mandate else 0
    total_pct = round((bc_act + wc_act) / (bc_mandate + wc_mandate) * 100, 1) \
                 if (bc_mandate + wc_mandate) else 0
    # Headline = MIN(bc_pct, wc_pct), excluding any collar with 0 employees.
    candidates = []
    if bc > 0: candidates.append(bc_pct)
    if wc > 0: candidates.append(wc_pct)
    headline_pct = min(candidates) if candidates else 0
    if headline_pct >= 75:
        headline_rag = 'on-track'
    elif headline_pct >= 50:
        headline_rag = 'watch'
    else:
        headline_rag = 'critical'
    return {
        'bc_emp': bc, 'wc_emp': wc,
        'bc_mandate': bc_mandate, 'wc_mandate': wc_mandate,
        'bc_actual': round(bc_act, 1), 'wc_actual': round(wc_act, 1),
        'bc_pct': bc_pct,
        'wc_pct': wc_pct,
        'total_pct': total_pct,
        'headline_pct': headline_pct,
        'headline_rag': headline_rag,
        'worst_cells': _calc_worst_cells(plant_id, db),
    }


def _calc_worst_cells(plant_id, db, limit=3, min_nominated=3):
    """Return up to `limit` worst (prog_type, collar) cells by TNI coverage %.
    A cell's coverage = trained_cnt / nominated * 100, where a TNI row is 'trained'
    if any emp_training row exists for that emp on a session whose programme matches.
    Only cells with at least `min_nominated` TNI rows are considered.
    Each row also carries a `rag` label using the same 75/50 thresholds.
    """
    fy = _fy_label()
    fy_start, fy_end = _current_fy()
    try:
        noms, trained = coverage_universe(db, plant_id, fy_start, fy_end, fy)
    except Exception:
        return []
    from collections import defaultdict
    tot = defaultdict(int); cvd = defaultdict(int)   # key=(prog_type, collar)
    for n in noms:
        k = (n['prog_type'], n['collar'])
        tot[k] += 1
        if (n['emp'], n['prog']) in trained:
            cvd[k] += 1
    out = []
    for (pt, collar), nom in tot.items():
        if nom < min_nominated:
            continue
        c = cvd[(pt, collar)]
        pct = round(100.0 * c / nom, 1) if nom else 0
        rag = 'on-track' if pct >= 75 else ('watch' if pct >= 50 else 'critical')
        out.append({
            'prog_type': pt or '', 'collar': collar or '',
            'pct': pct, 'cov_pct': pct,
            'trained': c, 'trained_cnt': c, 'nominated': nom, 'rag': rag,
        })
    out.sort(key=lambda x: x['cov_pct'])
    return out[:limit]


# ── Excel error response ──────────────────────────────────────────────────────

def _error_excel_response(errors, inserted, download_name='Upload_Errors.xlsx'):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Failed Rows'
    ws.append([f'{inserted} rows imported successfully. {len(errors)} rows failed — details below.'])
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:C1')
    ws.append([])
    hdr = ['Row #', 'Error Reason', 'Tip']
    ws.append(hdr)
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=3, column=c)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='C0392B')
        cell.alignment = Alignment(horizontal='center')
    for err in errors:
        parts   = err.split(':', 1)
        row_ref = parts[0].strip() if len(parts) == 2 else ''
        reason  = parts[1].strip() if len(parts) == 2 else err
        tip = ''
        rl  = reason.lower()
        if 'not found in your plant'   in rl: tip = 'Check employee code is registered under this plant in Employee Master'
        elif 'required'                in rl: tip = 'This column must not be empty'
        elif 'month'                   in rl: tip = 'Use: April / May / June / July / August / September / October / November / December / January / February / March'
        elif 'type' in rl or 'prog'    in rl: tip = 'Use: Behavioural/Leadership | Cane | Commercial | EHS/HR | IT | Technical'
        elif 'mode'                    in rl: tip = 'Use: Classroom | OJT | SOP | Online'
        elif 'date'                    in rl: tip = 'Use format YYYY-MM-DD'
        elif 'hours' in rl or 'hrs'    in rl: tip = 'Must be a number e.g. 4 or 2.5'
        elif 'session'                 in rl: tip = 'Session Code must already exist in Training Calendar'
        elif 'employee'                in rl: tip = 'Employee Code must exist in Employee Master for this plant'
        ws.append([row_ref, reason, tip])
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 65
    for row in ws.iter_rows(min_row=4):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor='FFF5F5')
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out, download_name=download_name, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── DB cleanse helpers (called from init_db) ──────────────────────────────────

def _cleanse_master_spelling(db):
    rows = db.execute('SELECT id, name FROM programme_master').fetchall()
    for row in rows:
        cleaned = _smart_title(_apply_word_fixes(row['name']))
        if cleaned != row['name']:
            clash = db.execute(
                'SELECT id FROM programme_master WHERE plant_id=(SELECT plant_id FROM programme_master WHERE id=?) AND LOWER(name)=LOWER(?) AND id!=?',
                (row['id'], cleaned, row['id'])
            ).fetchone()
            if not clash:
                db.execute('UPDATE programme_master SET name=? WHERE id=?', (cleaned, row['id']))
    db.commit()


def _cleanse_programme_names(db, plant_id=None):
    from difflib import get_close_matches as gcm
    report = {}
    plants = [plant_id] if plant_id else [
        r[0] for r in db.execute('SELECT DISTINCT plant_id FROM programme_master').fetchall()]
    for pid in plants:
        master = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (pid,)).fetchall()]
        if not master:
            continue
        master_lower_map = {m.lower(): m for m in master}
        master_lower = list(master_lower_map.keys())
        fixed = 0; merged = 0
        for table in ('tni', 'emp_training', 'calendar'):
            rows = db.execute(f'SELECT id, programme_name FROM {table} WHERE plant_id=?', (pid,)).fetchall()
            for row in rows:
                raw = row['programme_name'] or ''
                if not raw:
                    continue
                raw_lower = raw.lower()
                if raw_lower in master_lower_map:
                    canonical = master_lower_map[raw_lower]
                    if canonical != raw:
                        db.execute(f'UPDATE {table} SET programme_name=? WHERE id=?', (canonical, row['id']))
                        fixed += 1
                else:
                    m = gcm(raw_lower, master_lower, n=1, cutoff=0.88)
                    if m:
                        canonical = master_lower_map[m[0]]
                        db.execute(f'UPDATE {table} SET programme_name=? WHERE id=?', (canonical, row['id']))
                        fixed += 1
        dupes = db.execute('''
            SELECT emp_code, programme_name, fy_year, MIN(id) as keep_id, COUNT(*) as cnt
            FROM tni WHERE plant_id=?
            GROUP BY emp_code, programme_name, fy_year HAVING cnt > 1
        ''', (pid,)).fetchall()
        for d in dupes:
            db.execute('DELETE FROM tni WHERE plant_id=? AND emp_code=? AND programme_name=? AND fy_year=? AND id != ?',
                       (pid, d['emp_code'], d['programme_name'], d['fy_year'], d['keep_id']))
            merged += 1
        db.commit()
        report[pid] = {'fixed': fixed, 'merged': merged}
    return report


_EMP_FIELD_NORM = {'designation': 'title', 'department': 'upper', 'section': 'upper'}


def _canonical_emp_field(raw, plant_id, db, column, cutoff=0.88, _existing=None):
    """Normalize + fuzzy-snap an employee field (designation/department/section)
    to an existing canonical value for this plant. Returns the canonical string,
    or the normalized raw if no fuzzy hit. Case-insensitive comparison."""
    if not raw or not str(raw).strip():
        return ''
    s = str(raw).strip()
    mode = _EMP_FIELD_NORM.get(column, 'title')
    if mode == 'upper':
        s = s.upper()
    else:
        s = _smart_title(s)
    if column not in _EMP_FIELD_NORM:
        return s
    if _existing is None:
        rows = db.execute(
            f"SELECT DISTINCT TRIM({column}) FROM employees "
            f"WHERE plant_id=? AND {column} IS NOT NULL AND {column}!=''",
            (plant_id,)).fetchall()
        _existing = [r[0] for r in rows if r[0]]
    if not _existing:
        return s
    from difflib import get_close_matches as gcm
    lower_map = {e.lower(): e for e in _existing}
    if s.lower() in lower_map:
        return lower_map[s.lower()]
    m = gcm(s.lower(), list(lower_map.keys()), n=1, cutoff=cutoff)
    if m:
        return lower_map[m[0]]
    return s


def _cleanse_emp_fields(db, plant_id=None):
    """One-time normalize designation (smart-title) + fuzzy-collapse drift across
    designation/department/section for each plant. Picks the most-common
    cluster representative as canonical. Idempotent — safe to run on startup."""
    from difflib import get_close_matches as gcm
    report = {}
    plants = [plant_id] if plant_id else [
        r[0] for r in db.execute('SELECT DISTINCT plant_id FROM employees').fetchall()]
    for pid in plants:
        plant_report = {}
        for column in ('designation', 'department', 'section'):
            mode = _EMP_FIELD_NORM[column]
            rows = db.execute(
                f"SELECT id, {column} FROM employees "
                f"WHERE plant_id=? AND {column} IS NOT NULL AND {column}!=''",
                (pid,)).fetchall()
            if not rows:
                continue
            normed = []
            for r in rows:
                v = (r[column] or '').strip()
                if not v:
                    normed.append((r['id'], '', ''))
                    continue
                if mode == 'upper':
                    v_norm = v.upper()
                else:
                    v_norm = _smart_title(v)
                normed.append((r['id'], v, v_norm))
            freq = {}
            for _, _, vn in normed:
                if vn:
                    freq[vn] = freq.get(vn, 0) + 1
            distinct = sorted(freq.keys(), key=lambda k: (-freq[k], -len(k)))
            canonical_map = {}
            for v in distinct:
                if v in canonical_map:
                    continue
                vl = v.lower()
                hit = None
                for c in canonical_map.values():
                    if c.lower() == vl:
                        hit = c
                        break
                if hit is None:
                    cand_lowers = list({c.lower(): c for c in canonical_map.values()}.keys())
                    if cand_lowers:
                        m = gcm(vl, cand_lowers, n=1, cutoff=0.88)
                        if m:
                            hit = {c.lower(): c for c in canonical_map.values()}[m[0]]
                canonical_map[v] = hit if hit else v
            updated = 0
            for rid, orig, vn in normed:
                if not vn:
                    continue
                canonical = canonical_map.get(vn, vn)
                if canonical != orig:
                    db.execute(
                        f'UPDATE employees SET {column}=? WHERE id=?',
                        (canonical, rid))
                    updated += 1
            if updated:
                plant_report[column] = updated
        db.commit()
        if plant_report:
            report[pid] = plant_report
    return report


def _cleanup_stale_analyze_files():
    pattern = os.path.join(BASE_DIR, 'data', 'tni_analyze_*.json')
    cutoff  = time.time() - 86400
    for path in glob.glob(pattern):
        try:
            if os.path.getmtime(path) < cutoff:
                os.remove(path)
        except Exception:
            pass


# ── Fresh TNI upload ──────────────────────────────────────────────────────────

def _poka_yoke_clean_prog(name):
    if not name:
        return ''
    s = re.sub(r'[\x00-\x1f\x7f]', '', str(name).strip())
    s = re.sub(r'\s+', ' ', s).strip()
    s = _apply_word_fixes(s)
    return _smart_title(s)


def _process_fresh_tni(df, plant_id, db):
    emp_rows = db.execute(
        'SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)
    ).fetchall()
    emp_map   = {r['emp_code']: r['name'] for r in emp_rows}
    emp_upper = {k.upper(): k for k in emp_map}

    cols     = df.columns.tolist()
    col_emp  = _detect_col(cols, ['emp code','employee code','empcode','staff code','emp id','employee id','code'])
    col_prog = _detect_col(cols, ['programme name','program name','training name','course name','training need','training'])
    col_type = _detect_col(cols, ['type of programme','type','programme type','prog type','training type','category'])
    col_mode = _detect_col(cols, ['mode','training mode','delivery mode'])
    col_hrs  = _detect_col(cols, ['planned hours','hours','hrs','duration'])

    def gv(row, col):
        if not col: return ''
        v = str(row.get(col, '') or '').strip()
        return '' if v.lower() in ('nan', 'none', '0', '') else v

    valid_rows = []; error_rows = []; name_corrections = {}
    seen = set(); duplicate_count = 0

    for i, row in df.iterrows():
        raw_emp  = gv(row, col_emp)
        raw_prog = gv(row, col_prog)
        prog_type = gv(row, col_type)
        mode      = gv(row, col_mode)
        hours     = _safe_float(gv(row, col_hrs)) or 0

        if not raw_emp and not raw_prog:
            continue
        if not raw_emp:
            error_rows.append({'row': i+2, 'emp_code': '', 'prog_name': raw_prog, 'reason': 'Employee Code missing'})
            continue
        if not raw_prog:
            error_rows.append({'row': i+2, 'emp_code': raw_emp, 'prog_name': '', 'reason': 'Programme Name missing'})
            continue

        clean_emp = raw_emp
        if raw_emp not in emp_map:
            if raw_emp.upper() in emp_upper:
                clean_emp = emp_upper[raw_emp.upper()]
            else:
                error_rows.append({'row': i+2, 'emp_code': raw_emp, 'prog_name': raw_prog,
                                   'reason': f'Employee "{raw_emp}" not found in this plant'})
                continue

        cleaned = _poka_yoke_clean_prog(raw_prog)
        if cleaned != raw_prog:
            name_corrections[raw_prog] = cleaned

        key = (clean_emp, cleaned.lower())
        if key in seen:
            duplicate_count += 1
            continue
        seen.add(key)
        valid_rows.append({
            'emp_code': clean_emp, 'programme_name': cleaned,
            'prog_type': prog_type, 'mode': mode, 'hours': hours,
        })

    return {
        'total_rows':       len(df),
        'valid_rows':       valid_rows,
        'error_rows':       error_rows,
        'name_corrections': name_corrections,
        'unique_progs':     sorted(set(r['programme_name'] for r in valid_rows)),
        'duplicate_count':  duplicate_count,
    }


# ── MS Forms import ───────────────────────────────────────────────────────────

_MSFORMS_SKIP_HEADERS = {'id','start time','completion time','email','name','responder'}


def _parse_msforms_excel(file_storage, plant_id, db):
    import pandas as pd
    raw = file_storage.read()
    try:
        df = pd.read_excel(io.BytesIO(raw), dtype=str).fillna('')
    except Exception as e:
        raise ValueError(f'Could not read file: {e}')

    emp_rows = db.execute(
        'SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)).fetchall()
    emp_map = {r['emp_code']: r['name'] for r in emp_rows}

    field_keywords = {
        'emp_code':       ['emp code','employee code','empcode','staff code','employee id'],
        'programme_name': ['programme name','program name','training name','course name','training need'],
        'prog_type':      ['type of programme','programme type','training type','type'],
        'mode':           ['mode','training mode','delivery mode'],
        'hours':          ['planned hours','hours','hrs','duration'],
    }
    col_map = {}
    for col in df.columns:
        cl = str(col).strip().lower()
        if cl in _MSFORMS_SKIP_HEADERS:
            continue
        for field, kws in field_keywords.items():
            if field not in col_map:
                for kw in kws:
                    if kw in cl or cl in kw:
                        col_map[field] = col
                        break

    inserted, errors = 0, []
    for i, row in df.iterrows():
        def gv(field):
            c = col_map.get(field)
            v = str(row.get(c, '') or '').strip() if c else ''
            return '' if v.lower() in ('nan','none') else v

        emp_code  = gv('emp_code')
        prog_name = gv('programme_name')
        prog_type = gv('prog_type')
        mode      = gv('mode')
        hours     = _safe_float(gv('hours')) or 0.0

        if not emp_code or not prog_name:
            errors.append(f'Row {i+2}: Employee Code and Programme Name are required.')
            continue
        if emp_code not in emp_map:
            errors.append(f'Row {i+2}: Employee code "{emp_code}" not found in your plant.')
            continue
        db.execute(
            'INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours,fy_year) VALUES(?,?,?,?,?,?,?)',
            (plant_id, emp_code, prog_name, prog_type, mode, hours, _fy_label()))
        inserted += 1

    db.commit()
    return inserted, errors


# ── Smart TNI Analyzer ────────────────────────────────────────────────────────

_ACRONYMS = {
    'PPE','SOP','EHS','OJT','DCS','UPS','VFD','DG','SLD','AC','DC','GST','ISO',
    'HR','IT','MBC','FFT','MIST','DM','ETP','CPU','CGCB','MSDS','OFSAM','ZFD',
    'STD2SD','FCS','RTD','TC','KNO3','MOP','PDM','5S','5-S','JCB','PM','R&M',
    'AI','ML','KPI','GMP','BOD','COD','TOC','TDS','ROI','MIS','SAP','ERR','ERB',
    'CCTV','GPS','QR','LED','LCD','CRM','ERP','LMS','HRM','WMS','PLC','SCADA',
}

_WORD_FIXES = {
    'techqnique':'Technique','tecnique':'Technique','techique':'Technique',
    'technqiue':'Technique','teqnique':'Technique','technque':'Technique',
    'grainding':'Grinding','graining':'Grinding','granding':'Grinding',
    'grindig':'Grinding','grindding':'Grinding',
    'hyigene':'Hygiene','hyegiene':'Hygiene','hygeine':'Hygiene','higiene':'Hygiene',
    'maitenance':'Maintenance','maintenace':'Maintenance','maintainance':'Maintenance',
    'mantenance':'Maintenance',
    'safty':'Safety','saftey':'Safety',
    'operartion':'Operation','operetion':'Operation','opertion':'Operation',
    'managment':'Management','managament':'Management','mangement':'Management',
    'trainning':'Training','traning':'Training',
    'awarness':'Awareness','awreness':'Awareness',
    'handeling':'Handling','handlng':'Handling',
    'electrcial':'Electrical','eletrical':'Electrical',
    'chemcial':'Chemical','chemicle':'Chemical',
    'equpiment':'Equipment','equipement':'Equipment','equipmnet':'Equipment',
    'proceudre':'Procedure','proceedure':'Procedure',
    'complience':'Compliance','compliace':'Compliance',
    'enviroment':'Environment','enviromental':'Environmental',
    'knowlege':'Knowledge','knwoledge':'Knowledge','knoweldge':'Knowledge',
    'monitorng':'Monitoring','monitering':'Monitoring',
    'buidling':'Building','buldling':'Building',
    'confind':'Confined','confinde':'Confined','condfined':'Confined',
    'chocking':'Choking','chocing':'Choking',
    'equipments':'Equipment',
    'lubricants':'Lubricants',
}

_STRIP_CHARS = '.,;:()/'


def _smart_title(s):
    _SMALL = frozenset({'a','an','the','and','or','but','nor','for','yet','so',
                        'at','by','in','of','on','to','as','is','it',
                        'with','from','into','onto','off','per','via'})

    def _tw(w, is_first):
        if not w:
            return w
        if '/' in w:
            parts = w.split('/')
            return '/'.join(_tw(p, is_first and j == 0) for j, p in enumerate(parts))
        lstripped = w.lstrip(_STRIP_CHARS)
        prefix    = w[:len(w) - len(lstripped)]
        core      = lstripped.rstrip(_STRIP_CHARS)
        suffix    = lstripped[len(core):]
        if not core:
            return w
        core_up  = core.upper()
        core_low = core.lower()
        if core_up in _ACRONYMS:
            return prefix + core_up + suffix
        if core == core_up and len(core) >= 2 and core.isalpha():
            return prefix + core_up + suffix
        if core_low == 'ph':
            return prefix + 'pH' + suffix
        if not is_first and core_low in _SMALL and not prefix and not suffix:
            return core_low
        return prefix + core.capitalize() + suffix

    result = ' '.join(_tw(w, i == 0) for i, w in enumerate(s.split()))
    return result.strip('.,;: ')


def _apply_word_fixes(s):
    if not s:
        return s
    from difflib import get_close_matches as _gcm
    words = s.split()
    out   = []
    for w in words:
        lstripped = w.lstrip(_STRIP_CHARS)
        prefix    = w[:len(w) - len(lstripped)]
        core      = lstripped.rstrip(_STRIP_CHARS)
        suffix    = lstripped[len(core):]
        core_low  = core.lower()

        if not core or len(core) < 4 or core.upper() in _ACRONYMS:
            out.append(w)
            continue

        fix = _WORD_FIXES.get(core_low)
        if fix:
            out.append(prefix + fix + suffix)
            continue

        if core_low in _MASTER_VOCAB:
            out.append(w)
            continue

        if _MASTER_VOCAB:
            m = _gcm(core_low, _MASTER_VOCAB.keys(), n=1, cutoff=0.82)
            if m:
                out.append(prefix + _MASTER_VOCAB[m[0]] + suffix)
                continue

        out.append(w)
    return ' '.join(out)


def _canonical_prog(raw_name, plant_id, db, strict=False, _master=None):
    """_master: pre-loaded list of programme names — pass to avoid per-row DB query in bulk loops."""
    if not raw_name or not raw_name.strip():
        return raw_name
    from difflib import get_close_matches as gcm
    if _master is None:
        master = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
        ).fetchall()] or []
    else:
        master = _master
    master_lower = [m.lower() for m in master]
    corrected = _apply_word_fixes(raw_name.strip())
    raw_lower = corrected.lower()
    if raw_lower in master_lower:
        return master[master_lower.index(raw_lower)]
    m = gcm(raw_lower, master_lower, n=1, cutoff=0.82)
    if m:
        return master[master_lower.index(m[0])]
    return None if strict else _smart_title(corrected)


def _tni_canon_candidates(plant_id, db, fy=None):
    """Canonical candidate names for TNI programme canonicalization:
    programme_master ∪ distinct names already in TNI.

    Folding in existing TNI names is what stops a *second* spelling variant of a
    NEW programme from entering once the first one is present — `programme_master`
    alone can't, because a brand-new programme isn't in master yet when its first
    spelling is saved. Pass the returned list as `_master=` to `_canonical_prog`,
    and append each accepted canonical back to it inside batch loops so variants
    later in the same file collapse onto the earlier spelling.
    """
    names = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()]
    seen = {n.lower() for n in names if n}
    q = 'SELECT DISTINCT programme_name FROM tni WHERE plant_id=?'
    params = [plant_id]
    if fy:
        q += ' AND fy_year=?'
        params.append(fy)
    for (n,) in db.execute(q, params).fetchall():
        if n and n.lower() not in seen:
            names.append(n)
            seen.add(n.lower())
    return names


def _fuzzy_fix(val, valid_list):
    if not val: return '', False
    vl = val.strip().lower()
    for v in valid_list:
        if v.lower() == vl: return v, False
    # Handle swapped slash-parts: "HR/EHS" → matches "EHS/HR"
    vl_sorted = '/'.join(sorted(vl.split('/')))
    for v in valid_list:
        if '/'.join(sorted(v.lower().split('/'))) == vl_sorted:
            return v, True
    for v in valid_list:
        if vl in v.lower() or v.lower() in vl: return v, True
    m = get_close_matches(vl, [v.lower() for v in valid_list], n=1, cutoff=0.55)
    if m:
        idx = [v.lower() for v in valid_list].index(m[0])
        return valid_list[idx], True
    return val, False


def _detect_col(columns, keywords):
    for col in columns:
        cl = str(col).strip().lower()
        for kw in keywords:
            if kw in cl or cl in kw:
                return col
    return None


MASTER_PROGRAMMES = [
    "5-S Management","5S Commercial","Advance Practice Pathology In Agriculture",
    "Advanced Excel","Alignment Of Pumps, Fans And Gear Boxes With Motors",
    "An Overview Of Bio-Pesticides And Its Classification",
    "Bagasse Feeding As Per Boiler Requirement","Bagasse Handling",
    "Basic Fire Safety Awareness","Basic Knowledge Of Hardware",
    "Basic Of DCS Maintenance","Basic Of Maintenance And Programming Of Electronic Governor",
    "Behaviour Based Safety","Boiler Operation & Maintenance","Breakdown Handling",
    "Budget Planning","CGCB Guideline","CPU Operation",
    "Cane Quality & Minimise The Cut To Crush Period",
    "Checking Of Tube Cleaning/Tube Choking","Chemical Safety","Communication Skill",
    "Concept Of Fitting Methodology","Condenser Maintenance And Testing",
    "Condition Monitoring & Log Book Maintenance",
    "Condition Monitoring System Of Equipment","Confined Space",
    "Control Of Insect Pest & Diseases","Control Of Maintenance Schedule",
    "DCS Maintenance & Programming","DM Plant Operation & Maintenance",
    "Dismantling And Fitting Of Pump, Gear Boxes And Fans",
    "ETP Operation","Economiser Maintenance","Electrical Safety",
    "Emergency Awareness During Running Plant","Emergency Management",
    "FFT Efficiency","Failure Analysis Of Process Material",
    "Fire Fighting Equipment Technique","Fire Safety","Flow Measurement",
    "GST","General Checking Of Equipment In Running Season",
    "General Safety Awareness","Good Knowledge Of Metals","HR SOP",
    "Handling Of All Testing Equipment","Handling Of Juice And Mud Removal System",
    "Health Monitoring/Condition Monitoring Of Running Equipment",
    "Hot Work (Gas Cutting, Building And Grinding)",
    "How To Collect Samples","How To Handle Emergency Situation During Running Plant",
    "How To Identify Cane Diseases","Hydraulic Testing And Vacuum Trial Of Pan",
    "ISO General Awareness","IT SOP","Implementation Of OFSAM",
    "Importance Of Machine Guarding","Improvement Of Fermentation Efficiency",
    "Improvement Steam Economy","Income Tax",
    "Industrial Hygiene And 5-S Management",
    "Inspections And Testing Of Lifting Tool And Tackles Lifting Operation",
    "Inventory Management","Irrigation Automation","JCB Operation","JCB Pumping",
    "Juice Analysis","Knowledge About Cleaning And Ability To Conduct Schedule Checking Of Engines",
    "Knowledge About Cleaning, Switch Gear Panels And Motors Checklist",
    "Knowledge About Lab Apparatus & Equipment",
    "Knowledge And Application Of Condensate Removal System",
    "Knowledge And Operations Of Turbine","Knowledge Of Boiler Water Treatment",
    "Knowledge Of DCS Hardware And Panel Wiring",
    "Knowledge Of Electrical Equipment (Motor, Transformers, DG), SLD And Electrical Logics",
    "Knowledge Of Field Instruments Like Pressure Transmitters, Temperature Transmitters, RTD/TC, I To P Converters, Control Valves, Loop Testing",
    "Knowledge Of Importance Of Aeration In Melt",
    "Knowledge Of Industrial Lubricants Properties",
    "Knowledge Of Lighting And AC Systems",
    "Knowledge Of MBC/Belt Conveyor Health Monitoring - Gearbox, Chain Condition, Rake Condition, Idlers And Belt",
    "Knowledge Of Maintenance Of Pumps","Knowledge Of Measuring Instruments",
    "Knowledge Of Molasses Brix & Purity",
    "Knowledge Of Operating Parameter And Quality Of Steam And Cooling Water",
    "Knowledge Of Operation Of MIST","Knowledge Of Pumps And Its Parts",
    "Knowledge Of Supersaturation Zones During Pan Boiling",
    "Knowledge Of Three Motion Hydraulic Cane Unloader Operations",
    "Knowledge Of Upgraded Technology Related IT",
    "Knowledge Of Wire Rope Sling Size By Weight And Job Wise",
    "Knowledge Of Working Tools, Tackles And Fasteners",
    "Labour Laws","Leadership Quality","Legal Compliance",
    "Maintain Brix Of Magma","Maintain Temperature From Juice Heater",
    "Maintaining Load Of Machine By Operating The Feed Valve",
    "Maintenance And Programming Of Electronic Governor",
    "Maintenance Effectiveness","Maintenance Of AC & DC Drives",
    "Maintenance Of Electrical Machine And Switch Gears",
    "Maintenance Of Power Turbine",
    "Maintenance Of Safety Valves And Checking For Its Perfection",
    "Maintenance Of UPS & Battery","Manufacturing Process",
    "Massecuite Curing Temperature And Its Impact","Material Handling",
    "Material Handling (Manual And Mechanical)","Measurement Maintenance",
    "Mill Efficiency","Molasses Conditioning Temperature And Brix",
    "Monitoring And Ensuring Proper Operations Of Bagasse Belt Conveyors",
    "New Methodology In Soil Testing","New Varietal Trial","New Wage Code",
    "Nil Safety","Operation & Maintenance Of Boiler",
    "Operation Of Boiler From Cold To Pressurization",
    "Operation Of Turbine Within Controlled Parameters",
    "Operational Behaviour Awareness",
    "Optimum Temperature And pH Adjustment During Defecation",
    "Organic Waste Management","Overhauling And Maintenance Of Centrifugal Machine",
    "Ownership","PPE Awareness Use Inspection And Handling","Personal Effectiveness",
    "Planning Of Different Massecuite Boiling To Control The Material Load",
    "PowerPoint","Premium Potash Fertilisers: SOP Vs KNO3 Vs MOP Vs PDM",
    "Preventive Maintenance",
    "Preventive Maintenance And Condition Monitoring Of Electrical Equipment",
    "Problem Solving","Process Parameter","Process Safety Management",
    "Purification Of Condensate Removal System","Qualitimetry",
    "Quality Communication & Industrial Security","Quality R&M And Operation",
    "Raw Material Analysis","Removal Of Juice And Filter Cake From The System",
    "Repair & Maintenance Of Workshop Machine","Reporting Of Non EHS Lapses",
    "Reporting Of Non Performance Of Chemicals","Road Safety & Defensive Driving",
    "SOP Boiler","SOP Distillation And Fermentation Operation","SOP ETP",
    "SOP Electrical","SOP Instrumentation","SOP Mill House","SOP PM Module",
    "SOP Sales","SOP Store","SOP Workshop","STD2SD","Safety Induction",
    "Sample Collection","Sample Collection Methods As Per SOP",
    "Sampling & Its Importance","Screen Checking And Molasses Purity Control",
    "Self Discipline","Start And Stop The Boiler","Switch Gear Maintenance & Testing",
    "Team Work","Theft Prevention","To Control Fermentation Process Parameter",
    "To Maintain Brix Of Massecuite Of Dropping Pan",
    "To Maintain Brix Of Syrup At Fix",
    "To Maintain Chemical Dosing As Per Requirement",
    "To Maintain Temperature & pH Of Juice",
    "To Maintain Temperature & pH Of Juice/Melt",
    "To Maintain Temperature And pH Of Juice","Treated Water Parameters",
    "Understanding Of DCS Logic And Graphics","Use Of MSDS",
    "Use Of Optimum Dose Of Flocculant","VFD Maintenance","VFD Operation & Maintenance",
    "Water Management","Withdrawal Of Scum From FCS Clarifier","ZFD",
]
_MASTER_LOWER = [p.lower() for p in MASTER_PROGRAMMES]


def _build_master_vocab():
    _skip = frozenset({'a','an','the','and','or','but','of','in','to','at','by','as','is','it',
                       'for','on','per','via','from','into','onto','off','with','how','its'})
    vocab = {}
    for prog in MASTER_PROGRAMMES:
        for w in re.split(r'[ /,;:()\-]+', prog):
            core = re.sub(r'[.,;:()/&\-]', '', w).strip()
            if core and len(core) >= 4 and core.upper() not in _ACRONYMS and core.lower() not in _skip:
                vocab[core.lower()] = core
    return vocab


_MASTER_VOCAB = _build_master_vocab()


def _stream_input_rows(file_storage, skip_rows=0):
    """Yield (columns_list, rows_iterator) from an uploaded .xlsx/.xls/.csv file.

    Streams via openpyxl read-only mode (for xlsx) or csv module — keeps memory
    flat regardless of file size. Replaces pandas DataFrame load that was OOM-
    killing the worker on 5000+ row uploads (Render Starter 512MB cap).

    Returns a tuple. The iterator yields dict-like row mappings keyed by header.
    """
    import io as _io
    fname = (file_storage.filename or '').lower()
    raw = file_storage.read()

    if fname.endswith('.csv'):
        import csv as _csv
        text = raw.decode('utf-8-sig', errors='replace')
        reader = _csv.reader(_io.StringIO(text))
        all_rows = list(reader)
        if skip_rows:
            all_rows = all_rows[skip_rows:]
        if not all_rows:
            return [], iter([])
        headers = [str(h or '').strip() for h in all_rows[0]]
        def _csv_iter():
            for row in all_rows[1:]:
                yield {headers[i]: (row[i] if i < len(row) else '') for i in range(len(headers))}
        return headers, _csv_iter()

    # XLSX path — openpyxl streaming
    import openpyxl
    wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    # Skip top N rows
    for _ in range(skip_rows):
        try:
            next(row_iter)
        except StopIteration:
            return [], iter([])
    try:
        header_row = next(row_iter)
    except StopIteration:
        return [], iter([])
    headers = [str(h or '').strip() for h in header_row]
    def _xlsx_iter():
        try:
            for row in row_iter:
                yield {headers[i]: ('' if row[i] is None else str(row[i]))
                       for i in range(min(len(headers), len(row)))}
        finally:
            wb.close()
    return headers, _xlsx_iter()


def _smart_analyze_rows(df, plant_id, db, columns=None):
    from difflib import get_close_matches as gcm
    from tms.data_hygiene import (
        analyze_programme_name as _hy_prog,
        analyze_prog_type as _hy_type,
        analyze_mode as _hy_mode,
        suggest_top_n as _hy_suggest,
        normalise as _hy_normalise,
        validate as _hy_validate,
        spellcheck_text as _hy_spellcheck,
    )
    # Build allowlist from existing programme_master vocab so spellchecker
    # never "corrects" canonical domain words back to wrong English neighbours.
    _spell_allowlist = set()
    for _name in [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=?', (plant_id,))]:
        for _w in re.findall(r'[A-Za-z]{2,}', _name or ''):
            _spell_allowlist.add(_w.lower())
    _prog_cache = {}
    _sugg_cache = {}

    _active_master       = [r[0] for r in db.execute(
        'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,))]
    _active_master_lower = [p.lower() for p in _active_master]
    # Fold in distinct names already in TNI so the analyzer collapses an incoming
    # spelling variant onto a programme that exists in TNI even before it reaches
    # programme_master — same duplicate-prevention rule as _tni_canon_candidates.
    _seen_lower = set(_active_master_lower)
    for (_n,) in db.execute('SELECT DISTINCT programme_name FROM tni WHERE plant_id=?', (plant_id,)).fetchall():
        if _n and _n.lower() not in _seen_lower:
            _active_master.append(_n)
            _active_master_lower.append(_n.lower())
            _seen_lower.add(_n.lower())
    _has_master          = len(_active_master) > 0

    def _match_master(raw_lower):
        if raw_lower in _prog_cache:
            return _prog_cache[raw_lower]
        m = gcm(raw_lower, _active_master_lower, n=1, cutoff=0.65)
        result = _active_master[_active_master_lower.index(m[0])] if m else None
        _prog_cache[raw_lower] = result
        return result

    def _suggestions_for(raw_text):
        """Top-5 master candidates for a raw programme string. Cached per row."""
        if not raw_text or not _has_master:
            return []
        key = raw_text.lower()
        if key in _sugg_cache:
            return _sugg_cache[key]
        sugg = _hy_suggest(raw_text, _active_master, n=5, min_score=0.50)
        _sugg_cache[key] = sugg
        return sugg

    emp_rows  = db.execute(
        'SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)).fetchall()
    emp_map   = {r['emp_code']: r['name'] for r in emp_rows}
    emp_upper = {k.upper(): k for k in emp_map}

    # Support both legacy pandas DataFrame and new streaming iterable-of-dicts.
    # New caller passes `columns` explicitly + `df` is an iterator of dict rows.
    _streaming = columns is not None
    if _streaming:
        cols = columns
        row_iter = enumerate(df)  # df is iterator of dicts
    else:
        cols = df.columns.tolist()
        row_iter = df.iterrows()  # pandas (index, Series) pairs
    col_emp   = _detect_col(cols, ['emp code','employee code','empcode','staff code','emp id','employee id','code'])
    col_prog  = _detect_col(cols, ['programme name','program name','training name','course name','training need','training'])
    col_type  = _detect_col(cols, ['type of programme','type','programme type','prog type','training type','category'])
    col_mode  = _detect_col(cols, ['mode','training mode','delivery mode'])
    col_hrs   = _detect_col(cols, ['planned hours','hours','hrs','duration'])

    if not col_emp and not col_prog:
        col_list = ', '.join(f'"{c}"' for c in cols[:15])
        raise ValueError(
            f'Could not detect Employee Code or Programme Name columns. '
            f'Columns found in file: {col_list}. '
            f'Try using "Skip top rows" if your file has a title row above the headers.'
        )

    def gv(row, col):
        if not col: return ''
        v = str(row.get(col, '') or '').strip()
        return '' if v.lower() in ('nan','none','0','') else v

    results = []
    for i, row in row_iter:
        # In streaming mode `row` is already a dict; in pandas mode it's a Series
        # — both support .get(col, default) so downstream code is unchanged.
        raw_emp  = gv(row, col_emp)
        raw_prog = gv(row, col_prog)
        raw_type = gv(row, col_type)
        raw_mode = gv(row, col_mode)
        raw_hrs  = gv(row, col_hrs)

        if not any([raw_emp, raw_prog, raw_type, raw_mode]):
            continue
        if raw_emp and raw_emp.startswith('⚠'):
            continue

        fixes  = []
        issues = []
        status = 'ok'

        clean_emp = raw_emp
        if not raw_emp:
            issues.append('Employee Code is missing')
            status = 'error'
        elif raw_emp in emp_map:
            pass
        elif raw_emp.upper() in emp_upper:
            clean_emp = emp_upper[raw_emp.upper()]
            fixes.append({'field':'Employee Code','original':raw_emp,'fixed':clean_emp,'how':'Capitalisation corrected'})
            if status == 'ok': status = 'fixed'
        else:
            issues.append(f'Employee code "{raw_emp}" not found in this plant')
            status = 'error'

        emp_name = emp_map.get(clean_emp, '')

        clean_prog = raw_prog
        prog_suggestions = []
        prog_garbage_class = None
        if not raw_prog:
            issues.append('Programme Name is missing')
            status = 'error'
        else:
            # Layer 1+2: normalise + validate via data_hygiene
            normalised = _hy_normalise(raw_prog)
            is_valid, reject_reason = _hy_validate(normalised)
            if not is_valid:
                issues.append(f'Programme Name invalid: {reject_reason} ("{raw_prog}")')
                status = 'error'
                prog_garbage_class = reject_reason
                clean_prog = ''
            else:
                word_fixed = _apply_word_fixes(normalised)
                # English-dictionary spell-check (domain-aware). Auto-fix any
                # misspelt word; abbreviations / acronyms / domain words skipped.
                spell_fixed, spell_corr = _hy_spellcheck(word_fixed, extra_allowlist=_spell_allowlist)
                if spell_corr:
                    for _orig, _new in spell_corr:
                        fixes.append({'field':'Programme Name','original':_orig,'fixed':_new,
                                      'how':'Spelling corrected (English dictionary)'})
                    word_fixed = spell_fixed
                    if status == 'ok': status = 'fixed'
                raw_lower  = word_fixed.lower()
                best = _match_master(raw_lower)
                if best is not None:
                    if best.lower() != raw_prog.strip().lower():
                        fixes.append({'field':'Programme Name','original':raw_prog,'fixed':best,'how':'Matched to master list'})
                        if status == 'ok': status = 'fixed'
                    clean_prog = best
                else:
                    # No 0.65-cutoff match → treat as genuinely new programme.
                    # Surface top-5 suggestions so SPOC can override if it's a typo
                    # below the cutoff (the silent-orphan gap).
                    titled = _smart_title(word_fixed)
                    if titled != raw_prog:
                        fixes.append({'field':'Programme Name','original':raw_prog,'fixed':titled,'how':'Spelling/case normalised — added as new programme'})
                        if status == 'ok': status = 'fixed'
                    else:
                        if _has_master:
                            fixes.append({'field':'Programme Name','original':raw_prog,'fixed':titled,'how':'New programme — will be added to Programme Master on import'})
                            if status == 'ok': status = 'fixed'
                    clean_prog = titled

                # Always compute suggestions for non-exact matches so the SPOC
                # has the option to pick a master entry instead of accepting "new".
                if best is None or best.lower() != raw_prog.strip().lower():
                    prog_suggestions = _suggestions_for(normalised)

        # Prog type — hygiene engine first (handles abbr + dot-stripping),
        # legacy _fuzzy_fix as fallback for substring matches.
        clean_type = ''
        if raw_type:
            hy_match, hy_conf, hy_method = _hy_type(raw_type, PROG_TYPES)
            if hy_match and hy_conf >= 0.85:
                clean_type = hy_match
                if hy_match.lower() != raw_type.strip().lower():
                    fixes.append({'field':'Type of Programme','original':raw_type,'fixed':hy_match,
                                  'how':f'Hygiene engine ({hy_method})'})
                    if status == 'ok': status = 'fixed'
            else:
                fuzzy_match, type_changed = _fuzzy_fix(raw_type, PROG_TYPES)
                if fuzzy_match in PROG_TYPES:
                    clean_type = fuzzy_match
                    if type_changed:
                        fixes.append({'field':'Type of Programme','original':raw_type,'fixed':fuzzy_match,
                                      'how':'Auto-matched to standard value'})
                        if status == 'ok': status = 'fixed'
                else:
                    issues.append(f'Unknown programme type: "{raw_type}" — could not auto-fix')
                    if status == 'ok': status = 'error'

        # Mode — same approach. Note: hygiene ABBR_MAP maps 'Offline' → 'Classroom'
        # which is the canonical fix for the existing data drift.
        clean_mode = ''
        if raw_mode:
            hy_match, hy_conf, hy_method = _hy_mode(raw_mode, MODES)
            if hy_match and hy_conf >= 0.85:
                clean_mode = hy_match
                if hy_match.lower() != raw_mode.strip().lower():
                    fixes.append({'field':'Mode','original':raw_mode,'fixed':hy_match,
                                  'how':f'Hygiene engine ({hy_method})'})
                    if status == 'ok': status = 'fixed'
            else:
                fuzzy_match, mode_changed = _fuzzy_fix(raw_mode, MODES)
                if fuzzy_match in MODES:
                    clean_mode = fuzzy_match
                    if mode_changed:
                        fixes.append({'field':'Mode','original':raw_mode,'fixed':fuzzy_match,
                                      'how':'Auto-matched to standard value'})
                        if status == 'ok': status = 'fixed'
                else:
                    issues.append(f'Unknown mode: "{raw_mode}" — could not auto-fix')
                    if status == 'ok': status = 'error'

        hours = _safe_float(raw_hrs) or 0.0

        results.append({
            'row_num':        i + 2,
            'status':         status,
            'fixes':          fixes,
            'issues':         issues,
            'emp_code':       clean_emp,
            'emp_name':       emp_name,
            'programme_name': clean_prog,
            'prog_type':      clean_type or raw_type,
            'mode':           clean_mode or raw_mode,
            'planned_hours':  hours,
            # Data Hygiene additions — non-breaking. UI may render these to
            # let SPOC override fuzzy-below-cutoff orphans.
            'prog_suggestions':  prog_suggestions,
            'prog_garbage_class': prog_garbage_class,
            'raw_prog':           raw_prog,
        })
    return results


def _ai_validate_programme_names(prog_summaries, master_progs):
    """
    Comprehensive AI check on TNI data using Google Gemini (free tier).
    Checks: name suffixes, semantic duplicates, type mismatch, mode mismatch,
    unrealistic hours — all human errors that rule-based checks miss.

    prog_summaries: list of {name, prog_type, mode, avg_hours}
    Returns {name_lower: [{'type': str, 'msg': str, 'fix': str|None}]}.
    Returns {} silently if GEMINI_API_KEY not set or call fails.
    """
    import os
    if not os.environ.get('GEMINI_API_KEY') or not prog_summaries:
        return {}
    try:
        import google.generativeai as genai, json as _j
        genai.configure(api_key=os.environ['GEMINI_API_KEY'])
        model = genai.GenerativeModel('gemini-1.5-flash')

        master_text  = '\n'.join(f'- {p}' for p in master_progs[:100]) or '(none yet)'
        summary_text = '\n'.join(
            f'{i+1}. Name: "{s["name"]}" | Type: {s["prog_type"] or "?"} | Mode: {s["mode"] or "?"} | Avg Hours: {s["avg_hours"]}'
            for i, s in enumerate(prog_summaries)
        )

        prompt = f"""You validate TNI (Training Needs Identification) data for BCML (Balrampur Chini Mills), an Indian sugar manufacturer (sugar mills, boilers, cane farming, EHS). SPOCs upload Excel files with employee training nominations.

Master list of canonical programme names:
{master_text}

Programmes in this upload (with their Type, Mode, Hours as entered by the SPOC):
{summary_text}

Check each entry for ALL of the following human errors:

1. SUFFIX — Programme name contains non-canonical suffix: FY codes (FY25-26, 25-26), year (2025), batch (Batch 1, B2), plant unit codes (BCM, GCM, RCM, TCM, MZP, ACM, KCM, BBN, HCM, MCM), quarters (Q1, Q2), months (Jan-Mar). Suggest clean name.

2. SEMANTIC_DUP — Two entries in this list clearly mean the same programme (>90% confident). Add dup_with: [other 1-based indices].
   Examples: "Fire Safety" vs "Fire Safety Training", "5S" vs "5S Housekeeping", "POSH" vs "POSH Awareness"

3. TYPE_MISMATCH — The Type of Programme entered is clearly wrong for this programme name.
   Rules:
   - Fire/Safety/EHS/Environment/POSH/First Aid → EHS/HR
   - Boiler/Turbine/Electrical/Mechanical/SOP/OJT/Technical operations → Technical
   - Excel/Computer/SAP/IT/Software → IT
   - Leadership/Communication/Behavioural/Soft Skills/Motivation → Behavioural/Leadership
   - Cane/Farming/Harvesting/Soil/Seed/Crop → Cane
   - Finance/Accounts/Commercial/Taxation → Commercial
   Only flag if you are very confident the entered type is wrong.

4. MODE_MISMATCH — The Mode is clearly wrong for this programme.
   Rules:
   - Field operations, equipment handling, practical skills → OJT (not Classroom)
   - Standard Operating Procedures → SOP (not Classroom)
   - E-learning, quiz, video-based → Online (not Classroom)
   Only flag if obviously wrong.

5. HOURS_FLAG — Planned hours are unrealistic for this type of training.
   Rules:
   - Any programme > 40 hours → flag as suspicious
   - SOP or OJT programme > 16 hours → flag
   - Online programme > 8 hours → flag
   - 0 hours → flag as missing

Be conservative — only flag when clearly wrong. Short clean names are fine.
Do NOT flag names just because they are not in master.

Respond with ONLY a compact JSON array, no markdown, no explanation:
[{{"idx":1,"issues":[]}},{{"idx":2,"issues":[{{"type":"type_mismatch","msg":"Fire Safety should be EHS/HR not Technical","fix":"EHS/HR"}},{{"type":"suffix","msg":"Name contains FY code","fix":"Fire Safety Training"}}]}}]

Issue types: suffix, semantic_dup, type_mismatch, mode_mismatch, hours_flag"""

        resp = model.generate_content(prompt)
        raw  = resp.text.strip()
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[1].rsplit('```', 1)[0].strip()

        data = _j.loads(raw)
        findings = {}
        for item in data:
            idx    = item.get('idx', 0) - 1
            issues = item.get('issues') or []
            for issue in issues:
                dw = issue.get('dup_with')
                if dw is not None and not isinstance(dw, list):
                    issue['dup_with'] = [dw]
            if issues and 0 <= idx < len(prog_summaries):
                findings[prog_summaries[idx]['name'].lower()] = issues
        return findings
    except Exception:
        return {}


def _error_excel_for_tni(error_rows, dup_rows=None, plant_id=None, db=None):
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rows To Fix'

    hdr_fill = PatternFill('solid', fgColor='7F1D1D')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    headers  = ['Row #','Employee Code','Programme Name','Type of Programme',
                'Mode','Planned Hours','Issue(s) Found']
    col_w    = [7, 16, 34, 22, 14, 14, 60]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    err_fill = PatternFill('solid', fgColor='FEF2F2')
    for ri, r in enumerate(error_rows, 2):
        vals = [r['row_num'], r['emp_code'], r['programme_name'],
                r['prog_type'], r['mode'], r['planned_hours'],
                '; '.join(r['issues'])]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = err_fill
            if ci == 7:
                cell.font = Font(color='DC2626', size=10, italic=True)
                cell.alignment = Alignment(wrap_text=True)

    last_row = len(error_rows) + 1

    if plant_id and db and last_row > 1:
        master_progs = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
        ).fetchall()] or []
        ws_pl = wb.create_sheet('_ProgList')
        ws_pl.sheet_state = 'hidden'
        for idx, v in enumerate(master_progs, 1):
            ws_pl.cell(row=idx, column=1, value=v)
        dv_prog = DataValidation(type='list',
                                 formula1=f'_ProgList!$A$1:$A${len(master_progs)}',
                                 allow_blank=True, showDropDown=False)
        dv_prog.sqref = f'C2:C{last_row}'
        ws.add_data_validation(dv_prog)

    if last_row > 1:
        dv_type = DataValidation(type='list', formula1=f'"{",".join(PROG_TYPES)}"', allow_blank=True)
        dv_mode = DataValidation(type='list', formula1=f'"{",".join(MODES)}"',      allow_blank=True)
        dv_type.sqref = f'D2:D{last_row}'
        dv_mode.sqref = f'E2:E{last_row}'
        ws.add_data_validation(dv_type)
        ws.add_data_validation(dv_mode)

    ws2 = wb.create_sheet('How To Fix')
    tips = [
        'These rows could NOT be auto-fixed. Please correct column by column:',
        '',
        'Employee Code — must exactly match the code in TMS employee master',
        'Programme Name — use the dropdown (click the cell) to pick from the master list',
        'Type of Programme — use the dropdown (click the cell)',
        'Mode — use the dropdown',
        '',
        'After fixing, save and re-upload on the TNI Analyzer page.',
        'Column G (Issues) shows exactly what was wrong — read it before fixing.',
    ]
    for ri, t in enumerate(tips, 1):
        c = ws2.cell(row=ri, column=1, value=t)
        if ri == 1: c.font = Font(bold=True, color='7F1D1D', size=12)
        ws2.column_dimensions['A'].width = 80

    if dup_rows:
        ws3 = wb.create_sheet('Duplicates')
        dup_hdr_fill = PatternFill('solid', fgColor='92400E')
        dup_headers  = ['Row #', 'Employee Code', 'Employee Name', 'Programme Name',
                        'Type', 'Mode', 'Duplicate Type']
        dup_col_w    = [7, 16, 28, 34, 22, 14, 60]
        for ci, (h, w) in enumerate(zip(dup_headers, dup_col_w), 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = dup_hdr_fill
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.row_dimensions[1].height = 24
        ws3.freeze_panes = 'A2'
        dup_fill = PatternFill('solid', fgColor='FFFBEB')
        for ri, r in enumerate(dup_rows, 2):
            vals = [r.get('row_num', ''), r['emp_code'], r.get('emp_name', ''),
                    r['programme_name'], r.get('prog_type', ''), r.get('mode', ''),
                    r['dup_type']]
            for ci, v in enumerate(vals, 1):
                cell = ws3.cell(row=ri, column=ci, value=v)
                cell.fill = dup_fill
                if ci == 7:
                    cell.font = Font(color='92400E', size=10, italic=True)
        ws3.cell(row=len(dup_rows)+3, column=1,
                 value='These rows were NOT imported. Fix or confirm before re-uploading.').font = \
            Font(bold=True, color='92400E')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── Cross-table calendar row validator ──────────────────────────────────────
# Centralises every cross-table check that calendar writes (add/edit/bulk) used
# to do (or skip) individually. Audit-driven: 8 of the Calendar audit's confirmed
# gaps reduce to "no cross-table validation at write time".
#
# Returns (errors, warnings).
#   errors   = hard blockers, write must NOT proceed
#   warnings = soft signals, write may proceed but UI should surface them
#
# Each item is a tuple (field_name_or_'_general_', message_string).
#
# row keys expected (all str unless noted): programme_name, prog_type, source,
# planned_month, plan_start, plan_end, time_from, time_to, duration_hrs (number),
# level, mode, target_audience, planned_pax (int), trainer_vendor, status (opt)
def validate_calendar_row(row, plant_id, db, is_edit=False, exclude_id=None,
                          is_central=False, prev_prog_type=None):
    from tms.constants import (PROG_TYPES, MODES, LEVELS, AUDIENCES,
                                MONTHS_FY, STATUSES)

    errors = []
    warnings = []

    def E(field, msg): errors.append((field, msg))
    def W(field, msg): warnings.append((field, msg))

    # ── Programme name + master gate (existing _canonical_prog logic, surfaced) ──
    prog_raw = (row.get('programme_name') or '').strip()
    if not prog_raw:
        E('programme_name', 'Programme Name is required.')
        return errors, warnings  # downstream checks all key off programme
    canonical = _canonical_prog(prog_raw, plant_id, db, strict=True)
    if canonical is None:
        E('programme_name',
          f'Programme "{prog_raw}" not in Programme Master. Add it first.')
        return errors, warnings
    prog_name = canonical

    # ── prog_type: hard enum check + cross-check vs programme_master ──
    prog_type = (row.get('prog_type') or '').strip()
    if not prog_type:
        E('prog_type', 'Type of Programme is required.')
    elif prog_type not in PROG_TYPES:
        E('prog_type',
          f'Invalid Type "{prog_type}". Allowed: {", ".join(PROG_TYPES)}')
    else:
        master_pt = db.execute(
            'SELECT prog_type FROM programme_master WHERE plant_id=? AND name=? LIMIT 1',
            (plant_id, prog_name)
        ).fetchone()
        if master_pt and master_pt[0] and master_pt[0] != prog_type:
            # Legacy edit grace: if user is EDITING and not actually changing
            # prog_type (it matches the previous saved value), don't block. The
            # mismatch is pre-existing data; SPOC may just be fixing date/pax.
            # Otherwise HARD BLOCK — prog_type must match Programme Master.
            if is_edit and prev_prog_type and prev_prog_type == prog_type:
                W('prog_type',
                  f'Type "{prog_type}" does not match Programme Master '
                  f'("{master_pt[0]}"). Pre-existing drift — fix via '
                  f'Programme Master sync.')
            else:
                E('prog_type',
                  f'Type "{prog_type}" does not match Programme Master '
                  f'("{master_pt[0]}") for "{prog_name}". '
                  f'Pick the master Type, or update Programme Master first.')

    # ── source enum (silently coerced to TNI Driven if invalid — keep behaviour
    # but warn so SPOC sees the silent fix) ──
    source = (row.get('source') or '').strip()
    if source and source not in ('TNI Driven', 'New Requirement'):
        W('source',
          f'Source "{source}" not allowed. Coerced to "TNI Driven".')

    # ── Date order + FY window + auto-derive planned_month ──
    ps = (row.get('plan_start') or '').strip()
    pe = (row.get('plan_end') or '').strip()
    fy_start, fy_end = _current_fy()
    if ps and not _in_current_fy(ps):
        E('plan_start',
          f'Plan Start date {ps} outside current FY ({fy_start} to {fy_end}).')
    if pe and not _in_current_fy(pe):
        E('plan_end',
          f'Plan End date {pe} outside current FY ({fy_start} to {fy_end}).')
    if ps and pe and ps > pe:
        E('plan_end',
          f'Plan End ({pe}) must be on or after Plan Start ({ps}).')

    # planned_month auto-derive from plan_start if blank or mismatched
    planned_month = (row.get('planned_month') or '').strip()
    if ps:
        try:
            mnum = int(ps[5:7])
            month_names = ['', 'January', 'February', 'March', 'April', 'May',
                           'June', 'July', 'August', 'September', 'October',
                           'November', 'December']
            derived = month_names[mnum]
            if not planned_month:
                row['planned_month'] = derived
            elif planned_month != derived:
                W('planned_month',
                  f'Planned Month "{planned_month}" does not match Plan Start month '
                  f'"{derived}". Will use derived value.')
                row['planned_month'] = derived
        except (ValueError, IndexError):
            pass

    # ── Times ordering ──
    tf = (row.get('time_from') or '').strip()
    tt = (row.get('time_to') or '').strip()
    if tf and tt and tt <= tf:
        E('time_to', f'End Time ({tt}) must be after Start Time ({tf}).')

    # ── duration_hrs bounds ──
    try:
        dur = float(row.get('duration_hrs') or 0)
    except (ValueError, TypeError):
        dur = 0
    if dur <= 0:
        E('duration_hrs', 'Duration must be greater than 0 hours.')
    elif dur > 80:
        E('duration_hrs', f'Duration {dur} hrs unrealistic (max 80). Check input.')

    # ── Time window vs duration cross-check ──
    # (End Time − Start Time) × days  must equal Duration (Hrs), ±15 min.
    # Catches: "9-5 = 8hrs but Duration=2", "1hr session but 9-12 window".
    ok, msg = _validate_time_vs_duration(tf, tt, dur, ps, pe)
    if not ok:
        E('duration_hrs', msg)

    # ── mode / level / audience enum (warn — keep behaviour permissive but flag) ──
    mode = (row.get('mode') or '').strip()
    if mode and mode not in MODES:
        E('mode', f'Invalid Mode "{mode}". Allowed: {", ".join(MODES)}')
    level = (row.get('level') or '').strip()
    if level and level not in LEVELS:
        W('level', f'Level "{level}" not in standard set. Allowed: {", ".join(LEVELS)}')

    # status (if provided — used by edit)
    status = (row.get('status') or '').strip()
    if status and status not in STATUSES:
        E('status', f'Invalid Status "{status}".')

    # ── planned_pax bounds ──
    try:
        ppx = int(row.get('planned_pax') or 0)
    except (ValueError, TypeError):
        ppx = 0
    if ppx < 0:
        E('planned_pax', 'Planned Pax cannot be negative.')
    elif ppx > 500:
        E('planned_pax', f'Planned Pax {ppx} unrealistic (max 500/session). Check input.')

    # ── Duplicate detection: same plant+programme+plan_start+time_from already scheduled ──
    if ps:
        dup_sql = (
            'SELECT id, session_code FROM calendar '
            'WHERE plant_id=? AND programme_name=? AND plan_start=? '
            'AND COALESCE(time_from,"")=COALESCE(?,"")'
        )
        params = [plant_id, prog_name, ps, tf]
        if is_edit and exclude_id is not None:
            dup_sql += ' AND id!=?'
            params.append(exclude_id)
        dup = db.execute(dup_sql, params).fetchone()
        if dup:
            W('plan_start',
              f'A session for "{prog_name}" already exists on {ps}'
              + (f' at {tf}' if tf else '')
              + f' (session {dup["session_code"]}). Confirm this is a second batch.')

    # ── Over-plan vs TNI demand ── (SKIP for central — no single-plant TNI)
    demand_row = None if is_central else db.execute(
        'SELECT COUNT(DISTINCT emp_code) AS d FROM tni '
        'WHERE plant_id=? AND LOWER(programme_name)=LOWER(?)',
        (plant_id, prog_name)
    ).fetchone()
    demand = (demand_row['d'] if demand_row else 0) or 0
    if demand > 0 and ppx > 0:
        existing_sql = (
            'SELECT COALESCE(SUM(planned_pax),0) AS s FROM calendar '
            'WHERE plant_id=? AND LOWER(programme_name)=LOWER(?) AND status!=?'
        )
        existing_params = [plant_id, prog_name, 'Cancelled']
        if is_edit and exclude_id is not None:
            existing_sql += ' AND id!=?'
            existing_params.append(exclude_id)
        existing_pax = db.execute(existing_sql, existing_params).fetchone()['s'] or 0
        new_total = existing_pax + ppx
        ratio = new_total / demand
        if ratio > 1.5:
            E('planned_pax',
              f'Over-plan: {prog_name} would be planned for {new_total} pax vs TNI demand '
              f'{demand} (ratio {ratio:.1f}x). Block — reduce pax or add TNI nominations.')
        elif ratio > 1.2:
            W('planned_pax',
              f'Soft over-plan: {prog_name} cumulative planned {new_total} pax vs TNI demand '
              f'{demand} (ratio {ratio:.1f}x). Confirm intentional.')

    return errors, warnings


def flash_validation(errors, warnings, flash_fn):
    """Helper: surface errors+warnings to the user via flash()."""
    for fld, msg in errors:
        flash_fn(f'❌ {msg}' if fld == '_general_' else f'❌ {fld}: {msg}', 'danger')
    for fld, msg in warnings:
        flash_fn(f'⚠ {msg}' if fld == '_general_' else f'⚠ {fld}: {msg}', 'warning')


def resync_calendar_audience(plant_id, prog_names, db):
    """When TNI changes for a programme, re-derive target_audience on every
    calendar row for that programme. Closes a Calendar audit gap where audience
    was stamped at calendar add time and never re-derived later.

    prog_names: iterable of programme names (case-insensitive). Pass None or
    empty to skip. Returns count of calendar rows updated.
    """
    if not prog_names:
        return 0
    seen = set()
    updated = 0
    for pn in prog_names:
        if not pn:
            continue
        key = pn.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        new_aud = _derive_audience(plant_id, pn, db)
        if not new_aud:
            continue  # nothing in TNI for this programme — leave as-is
        cur = db.execute(
            'UPDATE calendar SET target_audience=? '
            'WHERE plant_id=? AND LOWER(programme_name)=? '
            'AND COALESCE(target_audience,"") != ?',
            (new_aud, plant_id, key, new_aud)
        )
        updated += cur.rowcount
    if updated:
        db.commit()
    return updated
