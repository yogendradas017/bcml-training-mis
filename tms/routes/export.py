import io

from flask import redirect, url_for, session, flash, send_file, render_template, request

from tms.constants import PLANT_MAP, PLANTS, MONTHS_FY
from tms.db import get_db
from tms.decorators import login_required, central_required
from tms.helpers import _fy_label, _current_fy
from tms.config import get_config

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment


def _register(app):

    @app.route('/export/<int:plant_id>', methods=['GET', 'POST'])
    @login_required
    def export_excel(plant_id):
        if session.get('role') == 'spoc' and session.get('plant_id') != plant_id:
            flash('Access denied.', 'danger')
            return redirect(url_for('spoc_dashboard'))
        plant = PLANT_MAP.get(plant_id)
        if not plant:
            flash('Plant not found.', 'danger')
            return redirect(url_for('index'))

        if request.method == 'GET':
            db   = get_db()
            depts = [r[0] for r in db.execute(
                'SELECT DISTINCT department FROM employees WHERE plant_id=? AND department IS NOT NULL AND department!="" ORDER BY department',
                (plant_id,)).fetchall()]
            return render_template('export_config.html', plant=plant,
                                   months=MONTHS_FY, departments=depts)

        # ── POST: build filtered workbook ──────────────────────────────────────
        f        = request.form
        sheets   = f.getlist('sheets')
        month_f  = f.get('month', '')
        collar_f = f.get('collar', '')
        dept_f   = f.get('dept', '')
        src_f    = f.get('source', '')

        if not sheets:
            flash('Select at least one sheet to export.', 'warning')
            return redirect(url_for('export_excel', plant_id=plant_id))

        db    = get_db()
        fy    = _fy_label()
        wb    = openpyxl.Workbook(write_only=True)
        pname = plant['name'].upper()

        H_FONT  = Font(bold=True, color='FFFFFF', size=10)
        H_FILL  = PatternFill('solid', fgColor='1F4E79')
        H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        T_FONT  = Font(bold=True, size=11)

        def hc(ws, val):
            c = WriteOnlyCell(ws, value=val)
            c.font = H_FONT; c.fill = H_FILL; c.alignment = H_ALIGN
            return c

        def tc(ws, val):
            c = WriteOnlyCell(ws, value=val)
            c.font = T_FONT
            return c

        # ── Sheet 1: Employee Master ───────────────────────────────────────────
        if 'emp_master' in sheets:
            ws1 = wb.create_sheet('EMP_MASTER')
            ws1.append([tc(ws1, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws1.append([tc(ws1, f'{pname} — EMPLOYEE MASTER | FY {fy}')])
            ws1.append([])
            ws1.append([hc(ws1, h) for h in
                ['Sr.','Emp Code','Name','Designation','Grade','Collar',
                 'Department','Section','Category','Gender','PH',
                 'Exit Date','Exit Reason','Remarks']])
            where, params = ['plant_id=?'], [plant_id]
            if collar_f: where.append('collar=?'); params.append(collar_f)
            if dept_f:   where.append('department=?'); params.append(dept_f)
            for r, e in enumerate(db.execute(
                    f'SELECT * FROM employees WHERE {" AND ".join(where)} ORDER BY name', params), 1):
                ws1.append([r, e['emp_code'], e['name'], e['designation'] or '',
                            e['grade'] or '', e['collar'] or '', e['department'] or '',
                            e['section'] or '', e['category'] or '', e['gender'] or '',
                            e['physically_handicapped'] or '',
                            e['exit_date'] or '', e['exit_reason'] or '', e['remarks'] or ''])

        # ── Sheet 2: TNI Tracking ──────────────────────────────────────────────
        if 'tni' in sheets:
            ws2 = wb.create_sheet('TNI_Tracking')
            ws2.append([tc(ws2, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws2.append([tc(ws2, f'{pname} — TNI TRACKING | FY {fy}')])
            ws2.append([])
            ws2.append([hc(ws2, h) for h in
                ['Sr.','Emp Code','Name','Designation','Grade','Collar','Dept',
                 'Section','Programme Name','Type','Mode','Target Month','Planned Hrs','Completed?','Source']])
            done_set = set(
                (row[0], row[1]) for row in db.execute(
                    'SELECT emp_code, programme_name FROM emp_training WHERE plant_id=?', (plant_id,)))
            where  = ['t.plant_id=?']; params = [plant_id]
            if collar_f: where.append('e.collar=?');      params.append(collar_f)
            if dept_f:   where.append('e.department=?');  params.append(dept_f)
            if src_f:    where.append('t.source=?');      params.append(src_f)
            for r, t in enumerate(db.execute(f'''
                    SELECT t.*,e.name,e.designation,e.grade,e.collar,e.department,e.section
                    FROM tni t LEFT JOIN employees e
                      ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
                    WHERE {" AND ".join(where)}''', params), 1):
                ws2.append([r, t['emp_code'], t['name'] or '', t['designation'] or '',
                            t['grade'] or '', t['collar'] or '', t['department'] or '',
                            t['section'] or '', t['programme_name'], t['prog_type'] or '',
                            t['mode'] or '', t['target_month'] or '', t['planned_hours'],
                            'Yes' if (t['emp_code'], t['programme_name']) in done_set else 'No',
                            t['source'] or ''])

        # ── Sheet 3: Calendar ─────────────────────────────────────────────────
        if 'calendar' in sheets:
            ws3 = wb.create_sheet('Cal_Plan_vs_Actual')
            ws3.append([tc(ws3, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws3.append([tc(ws3, f'{pname} — TRAINING CALENDAR | FY {fy}')])
            ws3.append([])
            ws3.append([hc(ws3, h) for h in
                ['S/N','PROG CODE','SESSION CODE','Source','Programme Name','Type',
                 'Planned Month','Plan Start','Plan End','Duration(Hrs)','Level','Mode',
                 'Target Audience','Planned Pax','Trainer/Vendor','STATUS','Actual Date','Actual Pax']])
            act_pax_map = {row[0]: row[1] for row in db.execute(
                'SELECT session_code, COUNT(*) FROM emp_training WHERE plant_id=? GROUP BY session_code', (plant_id,))}
            pd_date_map = {row[0]: row[1] for row in db.execute(
                'SELECT session_code, start_date FROM programme_details WHERE plant_id=?', (plant_id,))}
            where  = ['plant_id=?']; params = [plant_id]
            if month_f: where.append('planned_month=?');  params.append(month_f)
            if src_f:   where.append('source=?');         params.append(src_f)
            for r, c in enumerate(db.execute(
                    f'SELECT * FROM calendar WHERE {" AND ".join(where)} ORDER BY id', params), 1):
                ws3.append([r, c['prog_code'], c['session_code'], c['source'] or '',
                            c['programme_name'], c['prog_type'] or '', c['planned_month'] or '',
                            c['plan_start'] or '', c['plan_end'] or '',
                            c['duration_hrs'], c['level'] or '', c['mode'] or '',
                            c['target_audience'] or '', c['planned_pax'],
                            c['trainer_vendor'] or '', c['status'] or '',
                            pd_date_map.get(c['session_code'], ''),
                            act_pax_map.get(c['session_code'], 0)])

        # ── Sheet 4: 2A Employee Training ─────────────────────────────────────
        if 'training_2a' in sheets:
            ws4 = wb.create_sheet('2A_Emp_Training')
            ws4.append([tc(ws4, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws4.append([tc(ws4, f'{pname} — 2A: EMPLOYEE TRAINING DETAILS | FY {fy}')])
            ws4.append([])
            ws4.append([hc(ws4, h) for h in
                ['Sr.','Emp Code','Name','Designation','Grade','Collar','Dept','Section',
                 'Session Code','Start Date','End Date','Hrs','Programme Name','Type','Level','Mode',
                 'Source','Pre Rating','Post Rating','Venue','Month']])
            where  = ['t.plant_id=?']; params = [plant_id]
            if month_f:  where.append('t.month=?');       params.append(month_f)
            if collar_f: where.append('e.collar=?');      params.append(collar_f)
            if dept_f:   where.append('e.department=?');  params.append(dept_f)
            for r, t in enumerate(db.execute(f'''
                    SELECT t.*,e.name as emp_name,e.designation,e.grade,e.collar,
                           e.department,e.section
                    FROM emp_training t LEFT JOIN employees e
                      ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
                    WHERE {" AND ".join(where)} ORDER BY t.id''', params), 1):
                src = 'New Requirement' if t['host_plant_id'] == 99 else (t['cal_new'] or '')
                ws4.append([r, t['emp_code'], t['emp_name'] or '', t['designation'] or '',
                            t['grade'] or '', t['collar'] or '', t['department'] or '',
                            t['section'] or '', t['session_code'] or '',
                            t['start_date'] or '', t['end_date'] or '',
                            t['hrs'], t['programme_name'], t['prog_type'] or '',
                            t['level'] or '', t['mode'] or '', src,
                            t['pre_rating'], t['post_rating'], t['venue'] or '', t['month'] or ''])

        # ── Sheet 5: 2C Programme Details ─────────────────────────────────────
        if 'prog_2c' in sheets:
            ws5 = wb.create_sheet('2C_Programme')
            ws5.append([tc(ws5, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws5.append([tc(ws5, f'{pname} — 2C: PROGRAMME-WISE DETAILS | FY {fy}')])
            ws5.append([])
            ws5.append([hc(ws5, h) for h in
                ['Sr.','Session Code','Programme Name','Type','Level','Cal/New','Mode',
                 'Start Date','End Date','Audience','Hours Actual','Faculty Name','Int/Ext',
                 'Cost (Rs.)','Venue','Course FB','Faculty FB','Trainer FB-Participants',
                 'Trainer FB-Facilities','Participants','Man-Hours']])
            pax_map = {row[0]: row[1] for row in db.execute(
                'SELECT session_code, COUNT(*) FROM emp_training WHERE plant_id=? GROUP BY session_code', (plant_id,))}
            hrs_map = {row[0]: row[1] for row in db.execute(
                'SELECT session_code, COALESCE(SUM(hrs),0) FROM emp_training WHERE plant_id=? GROUP BY session_code', (plant_id,))}
            where  = ['p.plant_id=?']; params = [plant_id]
            join   = 'LEFT JOIN calendar c ON c.session_code=p.session_code AND c.plant_id=p.plant_id'
            if month_f: where.append('c.planned_month=?'); params.append(month_f)
            row_num = 1
            for p in db.execute(
                    f'SELECT p.* FROM programme_details p {join} WHERE {" AND ".join(where)} ORDER BY p.id', params):
                ws5.append([row_num, p['session_code'], p['programme_name'], p['prog_type'] or '',
                            p['level'] or '', p['cal_new'] or '', p['mode'] or '',
                            p['start_date'] or '', p['end_date'] or '', p['audience'] or '',
                            p['hours_actual'], p['faculty_name'] or '', p['int_ext'] or '',
                            p['cost'], p['venue'] or '', p['course_feedback'],
                            p['faculty_feedback'], p['trainer_fb_participants'],
                            p['trainer_fb_facilities'],
                            pax_map.get(p['session_code'], 0),
                            round(hrs_map.get(p['session_code'], 0), 1)])
                row_num += 1
            # Central programme details where this plant's employees attended
            central_2c = db.execute('''
                SELECT DISTINCT p.* FROM programme_details p
                JOIN emp_training t ON t.session_code=p.session_code AND t.plant_id=?
                WHERE p.plant_id=99
                ORDER BY p.start_date
            ''', (plant_id,)).fetchall()
            for p in central_2c:
                ws5.append([row_num, p['session_code'], p['programme_name'], p['prog_type'] or '',
                            p['level'] or '', 'Central', p['mode'] or '',
                            p['start_date'] or '', p['end_date'] or '', p['audience'] or '',
                            p['hours_actual'], p['faculty_name'] or '', p['int_ext'] or '',
                            p['cost'], p['venue'] or '', p['course_feedback'],
                            p['faculty_feedback'], p['trainer_fb_participants'],
                            p['trainer_fb_facilities'],
                            pax_map.get(p['session_code'], 0),
                            round(hrs_map.get(p['session_code'], 0), 1)])
                row_num += 1

        # ── suffix for filename ────────────────────────────────────────────────
        suffix = f"_{month_f}" if month_f else ''
        suffix += f"_{collar_f.replace(' ','')}" if collar_f else ''
        output   = io.BytesIO()
        wb.save(output); output.seek(0)
        filename = f"BCML_{plant['unit_code']}_Training_MIS_{fy.replace('-','')}{suffix}.xlsx"
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Consolidated cross-plant export ──────────────────────────────────────

    @app.route('/central/export', methods=['GET', 'POST'])
    @central_required
    def central_export():
        if request.method == 'GET':
            return render_template('export_central.html', months=MONTHS_FY)

        sheets  = request.form.getlist('sheets')
        month_f = request.form.get('month', '')
        if not sheets:
            flash('Select at least one sheet to export.', 'warning')
            return redirect(url_for('central_export'))

        db  = get_db()
        fy  = _fy_label()
        fy_start, fy_end = _current_fy()
        wb  = openpyxl.Workbook(write_only=True)

        H_FONT  = Font(bold=True, color='FFFFFF', size=10)
        H_FILL  = PatternFill('solid', fgColor='1F4E79')
        H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        T_FONT  = Font(bold=True, size=11)

        def hc(ws, val):
            c = WriteOnlyCell(ws, value=val)
            c.font = H_FONT; c.fill = H_FILL; c.alignment = H_ALIGN
            return c

        def tc(ws, val):
            c = WriteOnlyCell(ws, value=val)
            c.font = T_FONT
            return c

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        if 'summary' in sheets:
            ws = wb.create_sheet('Summary')
            ws.append([tc(ws, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws.append([tc(ws, f'CONSOLIDATED TRAINING SUMMARY | FY {fy}')])
            ws.append([])
            ws.append([hc(ws, h) for h in [
                'Plant', 'Unit Code', 'BC Headcount', 'WC Headcount', 'Total Emp',
                'Sessions Planned', 'Sessions Conducted',
                'Total Man-hours', 'BC Man-hours', 'WC Man-hours',
                'BC Mandate (hrs)', 'WC Mandate (hrs)',
                'BC Compliance %', 'WC Compliance %']])
            # batch: 7 queries total instead of 70
            _bc_h   = {r[0]:r[1] for r in db.execute("SELECT plant_id,COUNT(*) FROM employees WHERE is_active=1 AND collar='Blue Collared' GROUP BY plant_id")}
            _wc_h   = {r[0]:r[1] for r in db.execute("SELECT plant_id,COUNT(*) FROM employees WHERE is_active=1 AND collar='White Collared' GROUP BY plant_id")}
            _pln    = {r[0]:r[1] for r in db.execute("SELECT plant_id,COUNT(*) FROM calendar WHERE plan_start BETWEEN ? AND ? GROUP BY plant_id", (fy_start, fy_end))}
            _cond   = {r[0]:r[1] for r in db.execute("SELECT plant_id,COUNT(*) FROM calendar WHERE status='Conducted' AND plan_start BETWEEN ? AND ? GROUP BY plant_id", (fy_start, fy_end))}
            _mh     = {r[0]:r[1] for r in db.execute("SELECT plant_id,COALESCE(SUM(hrs),0) FROM emp_training WHERE start_date BETWEEN ? AND ? GROUP BY plant_id", (fy_start, fy_end))}
            _bc_hrs = {r[0]:r[1] for r in db.execute("SELECT t.plant_id,COALESCE(SUM(t.hrs),0) FROM emp_training t JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id WHERE e.collar='Blue Collared' AND t.start_date BETWEEN ? AND ? GROUP BY t.plant_id", (fy_start, fy_end))}
            _wc_hrs = {r[0]:r[1] for r in db.execute("SELECT t.plant_id,COALESCE(SUM(t.hrs),0) FROM emp_training t JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id WHERE e.collar='White Collared' AND t.start_date BETWEEN ? AND ? GROUP BY t.plant_id", (fy_start, fy_end))}
            tot_bc = tot_wc = tot_mh = tot_planned = tot_conducted = 0
            for p in PLANTS:
                pid      = p['id']
                bc       = _bc_h.get(pid, 0)
                wc       = _wc_h.get(pid, 0)
                planned  = _pln.get(pid, 0)
                conducted= _cond.get(pid, 0)
                mh       = _mh.get(pid, 0)
                bc_hrs   = _bc_hrs.get(pid, 0)
                wc_hrs   = _wc_hrs.get(pid, 0)
                bc_mandate = bc * get_config('mh_target_bc', 12, plant_id=pid)
                wc_mandate = wc * get_config('mh_target_wc', 24, plant_id=pid)
                bc_pct = round(bc_hrs / bc_mandate * 100, 1) if bc_mandate else 0
                wc_pct = round(wc_hrs / wc_mandate * 100, 1) if wc_mandate else 0
                ws.append([p['name'], p['unit_code'], bc, wc, bc + wc,
                           planned, conducted,
                           round(mh, 1), round(bc_hrs, 1), round(wc_hrs, 1),
                           bc_mandate, wc_mandate, bc_pct, wc_pct])
                tot_bc += bc; tot_wc += wc; tot_mh += mh
                tot_planned += planned; tot_conducted += conducted
            ws.append([])
            ws.append([tc(ws, 'TOTAL'), '', tot_bc, tot_wc, tot_bc + tot_wc,
                       tot_planned, tot_conducted, round(tot_mh, 1)])

        # ── Sheet 2: Employee Master All Plants ──────────────────────────────
        if 'emp_master' in sheets:
            ws = wb.create_sheet('EMP_MASTER_All_Plants')
            ws.append([tc(ws, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws.append([tc(ws, f'EMPLOYEE MASTER — ALL PLANTS | FY {fy}')])
            ws.append([])
            ws.append([hc(ws, h) for h in [
                'Sr.', 'Plant', 'Unit', 'Emp Code', 'Name', 'Designation',
                'Grade', 'Collar', 'Department', 'Section',
                'Category', 'Gender', 'PH', 'Active',
                'Exit Date', 'Exit Reason', 'Remarks']])
            r = 1
            for e in db.execute(
                    'SELECT e.*, p.name AS plant_name, p.unit_code'
                    ' FROM employees e'
                    ' JOIN plants p ON p.id = e.plant_id'
                    ' ORDER BY e.plant_id, e.name'):
                ws.append([r, e['plant_name'], e['unit_code'],
                           e['emp_code'], e['name'], e['designation'] or '',
                           e['grade'] or '', e['collar'] or '',
                           e['department'] or '', e['section'] or '',
                           e['category'] or '', e['gender'] or '',
                           e['physically_handicapped'] or '',
                           'Yes' if e['is_active'] else 'No',
                           e['exit_date'] or '', e['exit_reason'] or '', e['remarks'] or ''])
                r += 1

        # ── Sheet 3: TNI All Plants ───────────────────────────────────────────
        if 'tni' in sheets:
            ws = wb.create_sheet('TNI_All_Plants')
            ws.append([tc(ws, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws.append([tc(ws, f'TNI TRACKING — ALL PLANTS | FY {fy}')])
            ws.append([])
            ws.append([hc(ws, h) for h in [
                'Sr.', 'Plant', 'Unit', 'Emp Code', 'Name', 'Designation',
                'Grade', 'Collar', 'Dept', 'Section',
                'Programme Name', 'Type', 'Mode', 'Target Month',
                'Planned Hrs', 'Completed?', 'Source']])
            done_set = set(
                (r[0], r[1], r[2]) for r in db.execute(
                    'SELECT emp_code, programme_name, plant_id FROM emp_training WHERE start_date BETWEEN ? AND ?',
                    (fy_start, fy_end)))
            where = []; params = []
            if month_f:
                where.append('t.target_month=?'); params.append(month_f)
            where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
            r = 1
            for t in db.execute(f'''
                    SELECT t.*, e.name, e.designation, e.grade, e.collar,
                           e.department, e.section
                    FROM tni t
                    LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
                    {where_sql}
                    ORDER BY t.plant_id, t.emp_code''', params):
                p = PLANT_MAP.get(t['plant_id'], {})
                ws.append([r, p.get('name',''), p.get('unit_code',''),
                           t['emp_code'], t['name'] or '', t['designation'] or '',
                           t['grade'] or '', t['collar'] or '',
                           t['department'] or '', t['section'] or '',
                           t['programme_name'], t['prog_type'] or '',
                           t['mode'] or '', t['target_month'] or '',
                           t['planned_hours'],
                           'Yes' if (t['emp_code'], t['programme_name'], t['plant_id']) in done_set else 'No',
                           t['source'] or ''])
                r += 1

        # ── Sheet 3: Calendar All Plants ──────────────────────────────────────
        if 'calendar' in sheets:
            ws = wb.create_sheet('Calendar_All_Plants')
            ws.append([tc(ws, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws.append([tc(ws, f'TRAINING CALENDAR — ALL PLANTS | FY {fy}')])
            ws.append([])
            ws.append([hc(ws, h) for h in [
                'S/N', 'Plant', 'Unit', 'Prog Code', 'Session Code', 'Source',
                'Programme Name', 'Type', 'Planned Month', 'Plan Start', 'Plan End',
                'Duration(Hrs)', 'Level', 'Mode', 'Target Audience',
                'Planned Pax', 'Trainer/Vendor', 'STATUS', 'Actual Date', 'Actual Pax']])
            act_pax_map = {(r[0], r[1]): r[2] for r in db.execute(
                'SELECT plant_id, session_code, COUNT(*) FROM emp_training WHERE start_date BETWEEN ? AND ? GROUP BY plant_id, session_code',
                (fy_start, fy_end))}
            pd_date_map = {(r[0], r[1]): r[2] for r in db.execute(
                'SELECT plant_id, session_code, start_date FROM programme_details WHERE start_date BETWEEN ? AND ?',
                (fy_start, fy_end))}
            where = ['plan_start BETWEEN ? AND ?']; params = [fy_start, fy_end]
            if month_f: where.append('planned_month=?'); params.append(month_f)
            r = 1
            for c in db.execute(
                    f'SELECT * FROM calendar WHERE {" AND ".join(where)} ORDER BY plant_id, id',
                    params):
                p = PLANT_MAP.get(c['plant_id'], {})
                ws.append([r, p.get('name',''), p.get('unit_code',''),
                           c['prog_code'], c['session_code'], c['source'] or '',
                           c['programme_name'], c['prog_type'] or '',
                           c['planned_month'] or '', c['plan_start'] or '', c['plan_end'] or '',
                           c['duration_hrs'], c['level'] or '', c['mode'] or '',
                           c['target_audience'] or '', c['planned_pax'],
                           c['trainer_vendor'] or '', c['status'] or '',
                           pd_date_map.get((c['plant_id'], c['session_code']), ''),
                           act_pax_map.get((c['plant_id'], c['session_code']), 0)])
                r += 1

        # ── Sheet 4: 2A Attendance All Plants ─────────────────────────────────
        if 'training_2a' in sheets:
            ws = wb.create_sheet('2A_All_Plants')
            ws.append([tc(ws, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws.append([tc(ws, f'2A EMPLOYEE TRAINING — ALL PLANTS | FY {fy}')])
            ws.append([])
            ws.append([hc(ws, h) for h in [
                'Sr.', 'Plant', 'Unit', 'Emp Code', 'Name', 'Designation',
                'Grade', 'Collar', 'Dept', 'Section',
                'Session Code', 'Start Date', 'End Date', 'Hrs',
                'Programme Name', 'Type', 'Level', 'Mode', 'Source',
                'Pre Rating', 'Post Rating', 'Venue', 'Month']])
            where = ['t.start_date BETWEEN ? AND ?']; params = [fy_start, fy_end]
            if month_f: where.append('t.month=?'); params.append(month_f)
            r = 1
            for t in db.execute(f'''
                    SELECT t.*, e.name AS emp_name, e.designation, e.grade, e.collar,
                           e.department, e.section
                    FROM emp_training t
                    LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
                    WHERE {" AND ".join(where)}
                    ORDER BY t.plant_id, t.id''', params):
                p   = PLANT_MAP.get(t['plant_id'], {})
                src = 'New Requirement' if t['host_plant_id'] == 99 else (t['cal_new'] or '')
                ws.append([r, p.get('name',''), p.get('unit_code',''),
                           t['emp_code'], t['emp_name'] or '', t['designation'] or '',
                           t['grade'] or '', t['collar'] or '',
                           t['department'] or '', t['section'] or '',
                           t['session_code'] or '', t['start_date'] or '', t['end_date'] or '',
                           t['hrs'], t['programme_name'], t['prog_type'] or '',
                           t['level'] or '', t['mode'] or '', src,
                           t['pre_rating'], t['post_rating'], t['venue'] or '', t['month'] or ''])
                r += 1

        # ── Sheet 5: 2C Programme Details All Plants ──────────────────────────
        if 'prog_2c' in sheets:
            ws = wb.create_sheet('2C_All_Plants')
            ws.append([tc(ws, 'BALRAMPUR CHINI MILLS LIMITED')])
            ws.append([tc(ws, f'2C PROGRAMME DETAILS — ALL PLANTS | FY {fy}')])
            ws.append([])
            ws.append([hc(ws, h) for h in [
                'Sr.', 'Plant', 'Unit', 'Session Code', 'Programme Name',
                'Type', 'Level', 'Cal/New', 'Mode', 'Start Date', 'End Date',
                'Audience', 'Hours Actual', 'Faculty Name', 'Int/Ext',
                'Cost (Rs.)', 'Venue', 'Course FB', 'Faculty FB',
                'Trainer FB-Pax', 'Trainer FB-Fac', 'Participants', 'Man-Hours']])
            pax_map = {(r[0], r[1]): r[2] for r in db.execute(
                'SELECT plant_id, session_code, COUNT(*) FROM emp_training WHERE start_date BETWEEN ? AND ? GROUP BY plant_id, session_code',
                (fy_start, fy_end))}
            hrs_map = {(r[0], r[1]): r[2] for r in db.execute(
                'SELECT plant_id, session_code, COALESCE(SUM(hrs),0) FROM emp_training WHERE start_date BETWEEN ? AND ? GROUP BY plant_id, session_code',
                (fy_start, fy_end))}
            where = []; params = []
            if month_f:
                where.append(
                    'p.session_code IN (SELECT session_code FROM calendar WHERE planned_month=?)')
                params.append(month_f)
            where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
            r = 1
            for pd in db.execute(
                    f'SELECT p.* FROM programme_details p {where_sql} ORDER BY p.plant_id, p.id',
                    params):
                p = PLANT_MAP.get(pd['plant_id'], {})
                ws.append([r, p.get('name',''), p.get('unit_code',''),
                           pd['session_code'], pd['programme_name'], pd['prog_type'] or '',
                           pd['level'] or '', pd['cal_new'] or '', pd['mode'] or '',
                           pd['start_date'] or '', pd['end_date'] or '', pd['audience'] or '',
                           pd['hours_actual'], pd['faculty_name'] or '', pd['int_ext'] or '',
                           pd['cost'], pd['venue'] or '', pd['course_feedback'],
                           pd['faculty_feedback'], pd['trainer_fb_participants'],
                           pd['trainer_fb_facilities'],
                           pax_map.get((pd['plant_id'], pd['session_code']), 0),
                           round(hrs_map.get((pd['plant_id'], pd['session_code']), 0), 1)])
                r += 1

        suffix  = f'_{month_f}' if month_f else ''
        output  = io.BytesIO()
        wb.save(output); output.seek(0)
        filename = f'BCML_Consolidated_Training_MIS_{fy.replace("-","")}{suffix}.xlsx'
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
