import io
from datetime import datetime as _dt, date as _d, timedelta as _td

from flask import render_template, request, redirect, url_for, session, flash, send_file

from tms.constants import MONTHS_FY, PROG_TYPES, MODES
from tms.db import get_db
from tms.decorators import spoc_required
from tms.helpers import (
    _is_ajax, _canonical_prog, _date_to_month, _safe_float,
    _read_upload_file, _clean, _error_excel_response,
    _current_fy, _in_current_fy, _parse_date_strict,
    _recompute_session_actuals, _today_ist,
)

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from tms.audit import log_action


def _register(app):

    @app.route('/training')
    @spoc_required
    def emp_training():
        plant_id = session['plant_id']
        db = get_db()
        records = db.execute('''
            SELECT t.*, e.name as emp_name, e.designation, e.grade, e.collar,
                   e.department, e.section,
                   COALESCE(c.source, cc.source) as cal_source,
                   COALESCE(c.category, cc.category, pm.category, 'General') as category,
                   CASE WHEN t.host_plant_id=99 THEN 1 ELSE 0 END as is_central
            FROM emp_training t
            LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            LEFT JOIN calendar c  ON c.session_code=t.session_code AND c.plant_id=t.plant_id
            LEFT JOIN calendar cc ON cc.session_code=t.session_code AND cc.plant_id=99
                                  AND t.host_plant_id=99
            LEFT JOIN programme_master pm ON pm.plant_id=t.plant_id AND LOWER(pm.name)=LOWER(t.programme_name)
            WHERE t.plant_id=?
            ORDER BY t.id DESC
        ''', (plant_id,)).fetchall()
        emps = db.execute(
            'SELECT emp_code, name, grade, collar, department FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name',
            (plant_id,)).fetchall()
        # Plant's own calendar + central calendar sessions (for cross-plant attendance entry)
        own_sessions = db.execute(
            "SELECT session_code, programme_name, prog_type, mode, duration_hrs, plan_start, plan_end,"
            " time_from, time_to,"
            " 0 as is_central FROM calendar WHERE plant_id=? ORDER BY session_code",
            (plant_id,)).fetchall()
        central_sessions = db.execute(
            "SELECT session_code, programme_name, prog_type, mode, duration_hrs, plan_start, plan_end,"
            " time_from, time_to,"
            " 1 as is_central FROM calendar WHERE plant_id=99 ORDER BY session_code"
        ).fetchall()
        sessions_list = list(own_sessions) + list(central_sessions)
        return render_template('training_2a.html', records=records, employees=emps,
                               sessions=sessions_list, months=MONTHS_FY,
                               prog_types=PROG_TYPES, modes=MODES)

    @app.route('/training/add', methods=['POST'])
    @spoc_required
    def add_emp_training():
        plant_id = session['plant_id']
        f = request.form
        db = get_db()
        emp_code       = f.get('emp_code', '').strip()
        if not emp_code:
            flash('Employee Code is required.', 'danger')
            return redirect(url_for('emp_training'))

        # GAP 4: validate employee exists + active + correct plant
        emp_row = db.execute(
            'SELECT name, collar FROM employees WHERE plant_id=? AND emp_code=? AND is_active=1',
            (plant_id, emp_code)).fetchone()
        if not emp_row:
            flash(f'Employee "{emp_code}" not found or inactive for this plant.', 'danger')
            return redirect(url_for('emp_training'))

        session_code   = f.get('session_code', '').strip()
        start_date     = f.get('start_date', '')
        end_date       = f.get('end_date', '')
        time_from      = f.get('time_from', '').strip()
        time_to        = f.get('time_to', '').strip()
        prog_name_raw  = f.get('programme_name', '').strip()

        # Validate time range if both provided
        if time_from and time_to and time_to <= time_from:
            flash('End Time must be after Start Time.', 'danger')
            return redirect(url_for('emp_training'))
        # Required-pair: both or neither
        if bool(time_from) != bool(time_to):
            flash('Start Time and End Time must both be provided (or both blank).', 'danger')
            return redirect(url_for('emp_training'))

        prog_name = None
        prog_type = level = mode = cal_new = ''
        host_plant_id = None
        cal = None
        if session_code:
            cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                             (session_code, plant_id)).fetchone()
            if not cal:
                cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=99',
                                 (session_code,)).fetchone()
                if cal:
                    host_plant_id = 99
            if cal:
                # GAP 1: block if calendar session is Cancelled/Re-Scheduled
                if cal['status'] in ('Cancelled', 'Re-Scheduled'):
                    flash(f'Session {session_code} is {cal["status"]}. Cannot record attendance.', 'danger')
                    return redirect(url_for('emp_training'))
                prog_name  = cal['programme_name']
                prog_type  = cal['prog_type']
                level      = cal['level']
                mode       = cal['mode']
                cal_new    = 'Calendar Program'
                if not start_date: start_date = cal['plan_start'] or ''
                if not end_date:   end_date   = cal['plan_end'] or ''
                if not time_from:  time_from  = cal['time_from'] or ''
                if not time_to:    time_to    = cal['time_to'] or ''

            else:
                flash(f'Session code "{session_code}" not found in calendar.', 'warning')

        # GAP 2 (time gate): block future-dated training. Runs AFTER cal back-fill
        # so a session_code with a future plan_start can't sneak through with
        # blank form date — back-filled value is checked too.
        today_iso = _today_ist().isoformat()
        if start_date and start_date > today_iso:
            flash(f'Cannot log future-dated training. Start date "{start_date}" is after today.', 'danger')
            return redirect(url_for('emp_training'))

        if not prog_name:
            prog_name = _canonical_prog(prog_name_raw, plant_id, db, strict=True)
            if prog_name is None:
                flash(f'Programme "{prog_name_raw}" not found in Programme Master. Add it to the master list first.', 'danger')
                return redirect(url_for('emp_training'))

        if not prog_type and prog_name:
            mr = db.execute('SELECT prog_type FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
                            (plant_id, prog_name)).fetchone()
            if mr and mr['prog_type']:
                prog_type = mr['prog_type']

        try:
            hrs = float(f.get('hrs') or 0)
        except (ValueError, TypeError):
            hrs = 0
        if hrs <= 0:
            flash('Training hours must be greater than 0.', 'danger')
            return redirect(url_for('emp_training'))

        # NOTE: time-vs-duration cross-check intentionally NOT applied to 2A.
        # Per-person hrs is apples-to-oranges vs session window (a half-day
        # attendee of an 8hr session is legitimate). Anomaly flagging below
        # already catches per-row hours > cap.

        # Anomaly flags (allow save, tag for Central review)
        anomaly_flags = []
        if cal and cal['duration_hrs'] and cal['duration_hrs'] > 0:
            cap = cal['duration_hrs'] * 1.25
            if hrs > cap:
                anomaly_flags.append(f'hours_over({hrs} vs cap {cap:.1f})')
        # 2A vs 2C bi-directional check: if 2C already exists, flag if 2A hrs differs >25%
        if session_code:
            pd_row = db.execute(
                'SELECT hours_actual FROM programme_details WHERE plant_id=? AND session_code=?',
                (host_plant_id or plant_id, session_code)).fetchone()
            if pd_row and pd_row['hours_actual'] and pd_row['hours_actual'] > 0:
                if abs(hrs - pd_row['hours_actual']) / pd_row['hours_actual'] > 0.25:
                    anomaly_flags.append(f'hours_mismatch(2A={hrs} vs 2C={pd_row["hours_actual"]})')
        if cal:
            tgt = (cal['target_audience'] or '').strip()
            if tgt in ('Blue Collared', 'White Collared') and emp_row['collar'] and emp_row['collar'] != tgt:
                anomaly_flags.append(f'collar_mismatch({emp_row["collar"]} vs {tgt})')
            if start_date and cal['plan_start'] and cal['plan_end']:
                try:
                    d_start = _d.fromisoformat(start_date)
                    win_lo  = _d.fromisoformat(cal['plan_start']) - _td(days=1)
                    win_hi  = _d.fromisoformat(cal['plan_end'])   + _td(days=1)
                    if not (win_lo <= d_start <= win_hi):
                        anomaly_flags.append(f'date_outside({start_date} vs {cal["plan_start"]}/{cal["plan_end"]})')
                except (ValueError, TypeError):
                    pass

        fy_start, fy_end = _current_fy()
        if start_date and not _in_current_fy(start_date):
            flash(f'Training date must be within the current financial year ({fy_start} to {fy_end}).', 'danger')
            return redirect(url_for('emp_training'))

        # Pre / Post score range: 0-100 percentage scale.
        pre_r  = _safe_float(f.get('pre_rating'))
        post_r = _safe_float(f.get('post_rating'))
        for lbl, v in [('Pre', pre_r), ('Post', post_r)]:
            if v is not None and v != 0 and not (0 <= v <= 100):
                flash(f'{lbl}-Session Score must be between 0 and 100 (got {v}).', 'danger')
                return redirect(url_for('emp_training'))

        anom_str = ','.join(anomaly_flags) if anomaly_flags else None
        month = _date_to_month(start_date)
        db.execute('''INSERT OR IGNORE INTO emp_training
            (plant_id,emp_code,session_code,programme_name,start_date,end_date,
             time_from,time_to,
             hrs,prog_type,level,mode,cal_new,pre_rating,post_rating,venue,month,host_plant_id,anomaly_flags)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (plant_id, emp_code, session_code, prog_name,
             start_date, end_date, time_from or None, time_to or None, hrs,
             prog_type, level, mode, cal_new,
             pre_r, post_r,
             f.get('venue',''), month, host_plant_id, anom_str))
        if db.execute('SELECT changes()').fetchone()[0] == 0:
            flash('Duplicate record — this employee already has a training entry for this programme on this date.', 'warning')
            return redirect(url_for('emp_training'))
        # Live refresh of calendar.actual_pax / actual_hrs
        if session_code:
            _recompute_session_actuals(host_plant_id or plant_id, session_code, db)
        db.commit()
        tag = '[central]' if host_plant_id == 99 else ''
        log_action('RECORD_ADD', f"2a{tag}:{emp_code}:{prog_name}" + (f":anom[{anom_str}]" if anom_str else ""))
        if anomaly_flags:
            flash(f'Training record added with {len(anomaly_flags)} anomaly flag(s) — Central L&D will review.', 'warning')
        else:
            flash('Training record added.', 'success')
        return redirect(url_for('emp_training'))

    @app.route('/training/<int:rec_id>/delete', methods=['POST'])
    @spoc_required
    def delete_emp_training(rec_id):
        db = get_db()
        rec = db.execute(
            'SELECT session_code, host_plant_id FROM emp_training WHERE id=? AND plant_id=?',
            (rec_id, session['plant_id'])).fetchone()
        db.execute('DELETE FROM emp_training WHERE id=? AND plant_id=?', (rec_id, session['plant_id']))
        if rec and rec['session_code']:
            _recompute_session_actuals(rec['host_plant_id'] or session['plant_id'], rec['session_code'], db)
        db.commit()
        log_action('RECORD_DELETE', f"2a:{rec_id}")
        if _is_ajax():
            return '', 204
        flash('Training record deleted.', 'warning')
        return redirect(url_for('emp_training'))

    @app.route('/training/bulk-delete', methods=['POST'])
    @spoc_required
    def training_bulk_delete():
        plant_id = session['plant_id']
        ids = request.form.getlist('ids[]')
        if ids:
            db = get_db()
            # Capture session_codes BEFORE delete so we can recompute calendar actuals
            ph_all = ','.join('?' * len(ids))
            affected = db.execute(
                f'SELECT DISTINCT session_code, host_plant_id FROM emp_training '
                f'WHERE id IN ({ph_all}) AND plant_id=? AND session_code IS NOT NULL AND session_code != ""',
                ids + [plant_id]).fetchall()
            deleted = 0
            for i in range(0, len(ids), 900):
                chunk = ids[i:i+900]
                ph = ','.join('?' * len(chunk))
                db.execute(f'DELETE FROM emp_training WHERE id IN ({ph}) AND plant_id=?', chunk + [plant_id])
                deleted += len(chunk)
            for r in affected:
                _recompute_session_actuals(r['host_plant_id'] or plant_id, r['session_code'], db)
            db.commit()
            log_action('BULK_DELETE', f"2a:{deleted}")
            flash(f'{deleted} training records deleted.', 'warning')
        return redirect(url_for('emp_training'))

    @app.route('/training/template')
    @spoc_required
    def training_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '2A_Bulk_Upload'
        headers = ['Employee Code', 'Session Code (optional)', 'Programme Name',
                   'Type of Programme', 'Start Date (DD-MM-YYYY)', 'End Date (DD-MM-YYYY)',
                   'Hours', 'Venue', 'Pre-Session Rating (1-4)', 'Post-Session Rating (1-4)']
        hdr_fill = PatternFill('solid', start_color='1F4E79')
        hdr_font = Font(bold=True, color='FFFFFF')
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font
            ws.column_dimensions[get_column_letter(i)].width = 26
        ws.append(['21700011', 'BCM/EHS/001/B01', 'Fire Safety Training', 'EHS/HR', '10-06-2026', '10-06-2026', 4, 'Training Hall', 2.5, 3.8])
        ws.append(['21101568', '', 'MS Office Basics', 'IT', '05-07-2026', '06-07-2026', 8, 'Computer Lab', '', 3.5])
        ws['A5'] = 'NOTE: Session Code is optional. If provided, Programme Name/Type/Mode auto-fill from Calendar. Dates must be DD-MM-YYYY.'
        out = io.BytesIO()
        wb.save(out); out.seek(0)
        return send_file(out, download_name='2A_Training_Bulk_Upload_Template.xlsx', as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/training/bulk', methods=['POST'])
    @spoc_required
    def training_bulk_upload():
        plant_id = session['plant_id']
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('emp_training'))
        try:
            df = _read_upload_file(f)
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('emp_training'))

        db = get_db(); inserted = 0; errors = []; total_rows = len(df)
        for i, row in df.iterrows():
            emp_code     = _clean(row, ['employee code', 'emp code', 'empcode'])
            session_code = _clean(row, ['session code', 'session code (optional)', 'sessioncode'])
            prog_name    = _clean(row, ['programme name', 'program name', 'training name'])
            file_type    = _clean(row, ['type of programme', 'type', 'prog type', 'programme type'])
            raw_start    = _clean(row, ['start date (dd-mm-yyyy)', 'start date', 'startdate', 'date'])
            raw_end      = _clean(row, ['end date (dd-mm-yyyy)', 'end date', 'enddate'])
            try:
                start_date = _parse_date_strict(raw_start)
                end_date   = _parse_date_strict(raw_end)
            except ValueError as e:
                errors.append(f'Row {i+2}: Date format error — {e}. Use DD-MM-YYYY (e.g. 15-06-2026).')
                continue
            hrs          = _safe_float(_clean(row, ['hours', 'hrs', 'duration'])) or 0
            venue        = _clean(row, ['venue'])
            pre_r        = _safe_float(_clean(row, ['pre-session rating', 'pre rating', 'pre_rating']))
            post_r       = _safe_float(_clean(row, ['post-session rating', 'post rating', 'post_rating']))
            # Rating bounds 1-4 — block bad rows so silent garbage doesn't poison stats.
            bad = next(((lbl, v) for lbl, v in [('Pre', pre_r), ('Post', post_r)]
                        if v is not None and v != 0 and not (1 <= v <= 4)), None)
            if bad:
                errors.append(f'Row {i+2}: {bad[0]}-Session Rating must be 1-4 (got {bad[1]}).')
                continue

            if not emp_code:
                errors.append(f'Row {i+2}: Employee Code is required.')
                continue
            emp = db.execute('SELECT collar FROM employees WHERE emp_code=? AND plant_id=? AND is_active=1',
                             (emp_code, plant_id)).fetchone()
            if not emp:
                errors.append(f'Row {i+2}: Employee {emp_code} not found.')
                continue

            # GAP 2 (time gate): block future-dated rows
            if start_date and start_date > _today_ist().isoformat():
                errors.append(f'Row {i+2}: Cannot log future-dated training (start_date {start_date} after today).')
                continue

            prog_type = level = mode = cal_new = ''
            host_plant_id_row = None
            cal = None
            if session_code:
                cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                                 (session_code, plant_id)).fetchone()
                if not cal:
                    cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=99',
                                     (session_code,)).fetchone()
                    if cal:
                        host_plant_id_row = 99
                if cal:
                    # GAP 1: block Cancelled/Re-Scheduled sessions (HARD block stays)
                    if cal['status'] in ('Cancelled', 'Re-Scheduled'):
                        errors.append(f'Row {i+2}: Session {session_code} is {cal["status"]} — cannot record attendance.')
                        continue
                    prog_name  = prog_name or cal['programme_name']
                    prog_type  = cal['prog_type']
                    level      = cal['level']
                    mode       = cal['mode']
                    cal_new    = 'Calendar Program'
                    start_date = start_date or (cal['plan_start'] or '')
                    end_date   = end_date or (cal['plan_end'] or '')
                    row_time_from = cal['time_from'] or ''
                    row_time_to   = cal['time_to'] or ''
                else:
                    row_time_from = row_time_to = ''
            else:
                row_time_from = row_time_to = ''

            # Anomaly flags (allow save, tag for Central review)
            row_anom = []
            if cal:
                tgt = (cal['target_audience'] or '').strip()
                if tgt in ('Blue Collared', 'White Collared') and emp['collar'] and emp['collar'] != tgt:
                    row_anom.append(f'collar_mismatch({emp["collar"]} vs {tgt})')
                if start_date and cal['plan_start'] and cal['plan_end']:
                    try:
                        d_start = _d.fromisoformat(start_date)
                        win_lo  = _d.fromisoformat(cal['plan_start']) - _td(days=1)
                        win_hi  = _d.fromisoformat(cal['plan_end'])   + _td(days=1)
                        if not (win_lo <= d_start <= win_hi):
                            row_anom.append(f'date_outside({start_date})')
                    except (ValueError, TypeError):
                        pass
                if cal['duration_hrs'] and cal['duration_hrs'] > 0 and hrs > 0:
                    cap = cal['duration_hrs'] * 1.25
                    if hrs > cap:
                        row_anom.append(f'hours_over({hrs} vs cap {cap:.1f})')
            # 2A vs 2C bi-directional check
            if session_code and hrs > 0:
                pd_row = db.execute(
                    'SELECT hours_actual FROM programme_details WHERE plant_id=? AND session_code=?',
                    (host_plant_id_row or plant_id, session_code)).fetchone()
                if pd_row and pd_row['hours_actual'] and pd_row['hours_actual'] > 0:
                    if abs(hrs - pd_row['hours_actual']) / pd_row['hours_actual'] > 0.25:
                        row_anom.append(f'hours_mismatch(2A={hrs} vs 2C={pd_row["hours_actual"]})')

            if not prog_name:
                errors.append(f'Row {i+2}: Programme Name required (no session code matched).')
                continue

            if not prog_type:
                prog_type = file_type  # use column from file if session/master didn't fill it
            if not prog_type and prog_name:
                mr = db.execute('SELECT prog_type FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
                                (plant_id, prog_name)).fetchone()
                if mr and mr['prog_type']:
                    prog_type = mr['prog_type']
            if not prog_type:
                errors.append(f'Row {i+2}: Type of Programme is mandatory — fill the "Type of Programme" column.')
                continue

            if not cal_new:
                canonical = _canonical_prog(prog_name, plant_id, db, strict=True)
                if canonical is None:
                    errors.append(f'Row {i+2}: Programme "{prog_name}" not in Programme Master — add it first or link a Session Code.')
                    continue
                prog_name = canonical
            # NOTE: time-vs-duration check intentionally NOT run for 2A bulk —
            # per-person hrs differs legitimately from session window. See add.
            month = _date_to_month(start_date)
            db.execute('''INSERT OR IGNORE INTO emp_training
                (plant_id,emp_code,session_code,programme_name,start_date,end_date,
                 time_from,time_to,
                 hrs,prog_type,level,mode,cal_new,pre_rating,post_rating,venue,month,host_plant_id,anomaly_flags)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (plant_id, emp_code, session_code, prog_name,
                 start_date, end_date,
                 row_time_from or None, row_time_to or None,
                 hrs, prog_type, level, mode, cal_new,
                 pre_r, post_r, venue, month, host_plant_id_row,
                 ','.join(row_anom) if row_anom else None))
            if db.execute('SELECT changes()').fetchone()[0]:
                inserted += 1
                if session_code:
                    _recompute_session_actuals(host_plant_id_row or plant_id, session_code, db)
        db.commit()
        skipped = total_rows - inserted - len(errors)
        skip_msg = f' {skipped} duplicate(s) skipped.' if skipped > 0 else ''
        if errors:
            if inserted:
                flash(f'Bulk upload complete: {inserted} records added.{skip_msg} {len(errors)} rows had errors — downloading error report.', 'warning')
            return _error_excel_response(errors, inserted, 'Training2A_Upload_Errors.xlsx')
        log_action('BULK_UPLOAD', f"2a:{inserted}")
        flash(f'Bulk upload complete: {inserted} training records added successfully.{skip_msg}', 'success')
        return redirect(url_for('emp_training'))
