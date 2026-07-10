import io

from flask import render_template, request, redirect, url_for, session, flash, send_file

from tms.constants import PROG_TYPES, MODES, LEVELS, AUDIENCES, MONTHS_FY, STATUSES
from tms.db import get_db
from tms.decorators import spoc_required
from tms.helpers import (
    _is_ajax, _canonical_prog, _get_or_create_prog_code, _new_session_code,
    _derive_audience, _sync_calendar_from_2c,
    _read_upload_file, _clean, _safe_float, _error_excel_response,
    _current_fy, _in_current_fy, _parse_date_strict, _fy_label,
    validate_calendar_row, flash_validation,
)

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from tms.audit import log_action, log_record_change


def _register(app):

    @app.route('/calendar')
    @spoc_required
    def training_calendar():
        plant_id = session['plant_id']
        db = get_db()
        _sync_calendar_from_2c(plant_id, db)

        # Tier 5: auto-archive (Lapsed) sessions from prior FY that are still
        # 'To Be Planned' — they will never be conducted.
        # Awaiting Verification rows are NOT lapsed at FY rollover — they
        # represent conducted sessions still pending SPOC confirmation and
        # must survive the cycle boundary until verified or rejected.
        # Idempotent — runs cheap UPDATE, no-op if nothing matches.
        fy_start, _fy_end = _current_fy()
        lapsed_rows = db.execute(
            "UPDATE calendar SET status='Lapsed' "
            "WHERE plant_id=? AND status IN ('To Be Planned') "
            "AND plan_start IS NOT NULL AND plan_start != '' AND plan_start < ?",
            (plant_id, fy_start)
        ).rowcount
        if lapsed_rows:
            db.commit()

        # Default view: hide Lapsed unless explicitly requested (?include_lapsed=1)
        show_lapsed = request.args.get('include_lapsed') == '1'
        if show_lapsed:
            sessions = db.execute(
                'SELECT * FROM calendar WHERE plant_id=? ORDER BY plan_start ASC, id ASC',
                (plant_id,)).fetchall()
        else:
            sessions = db.execute(
                "SELECT * FROM calendar WHERE plant_id=? AND status != 'Lapsed' "
                "ORDER BY plan_start ASC, id ASC",
                (plant_id,)).fetchall()
        lapsed_count = db.execute(
            "SELECT COUNT(*) FROM calendar WHERE plant_id=? AND status='Lapsed'",
            (plant_id,)).fetchone()[0]
        demand_map = {}
        # Current-FY 'TNI Driven' demand only, so it matches the current-FY Audience
        # column beside it (was all-FY, which inflated the gap).
        for row in db.execute("SELECT programme_name, COUNT(DISTINCT emp_code) as cnt FROM tni WHERE plant_id=? AND fy_year=? AND source='TNI Driven' GROUP BY programme_name", (plant_id, _fy_label())):
            demand_map[row['programme_name']] = row['cnt']

        master_programmes = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()] or []
        all_cal_programmes = master_programmes
        tni_programmes = [r[0] for r in db.execute(
            'SELECT DISTINCT programme_name FROM tni WHERE plant_id=? ORDER BY programme_name', (plant_id,))]

        cov_rows = []
        pax_map  = {}
        for s in sessions:
            p = s['programme_name']
            if p not in pax_map:
                pax_map[p] = {'sessions': 0, 'planned_pax': 0, 'conducted_pax': 0}
            pax_map[p]['sessions']     += 1
            pax_map[p]['planned_pax']  += (s['planned_pax'] or 0)
            if s['status'] == 'Conducted':
                pax_map[p]['conducted_pax'] += (s['planned_pax'] or 0)
        for prog, d in demand_map.items():
            pm = pax_map.get(prog, {'sessions': 0, 'planned_pax': 0, 'conducted_pax': 0})
            planned_pax = pm['planned_pax']
            gap         = max(0, d - planned_pax)
            pct         = min(100, round(planned_pax / d * 100)) if d > 0 else 0
            cov_rows.append({'name': prog, 'demand': d,
                             'sessions': pm['sessions'], 'planned_pax': planned_pax,
                             'conducted_pax': pm['conducted_pax'],
                             'gap': gap, 'pct': pct,
                             'over': max(0, planned_pax - d)})
        cov_rows.sort(key=lambda x: x['gap'], reverse=True)

        qr_rows = db.execute(
            'SELECT session_code, stage, token, is_active FROM session_qr WHERE plant_id=?',
            (plant_id,)
        ).fetchall()
        qr_map = {}
        for q in qr_rows:
            qr_map.setdefault(q['session_code'], {})[q['stage']] = dict(q)

        fy_y0, fy_y1 = _current_fy()
        return render_template('calendar.html', sessions=sessions, demand_map=demand_map,
                               tni_programmes=tni_programmes,
                               all_cal_programmes=all_cal_programmes, cov_rows=cov_rows,
                               prog_types=PROG_TYPES, modes=MODES, levels=LEVELS,
                               audiences=AUDIENCES, months=MONTHS_FY, statuses=STATUSES,
                               qr_map=qr_map,
                               lapsed_count=lapsed_count,
                               show_lapsed=show_lapsed,
                               fy_label=_fy_label(),
                               fy_start=f'{fy_y0}-04-01',
                               fy_end=f'{fy_y1}-03-31')

    @app.route('/calendar/add', methods=['POST'])
    @spoc_required
    def add_calendar():
        plant_id = session['plant_id']
        f = request.form
        db = get_db()

        # Centralised cross-table validation (Tier 1+4)
        row = {
            'programme_name': f.get('programme_name', ''),
            'prog_type':      f.get('prog_type', ''),
            'source':         f.get('source', ''),
            'planned_month':  f.get('planned_month', ''),
            'plan_start':     f.get('plan_start', ''),
            'plan_end':       f.get('plan_end', ''),
            'time_from':      f.get('time_from', ''),
            'time_to':        f.get('time_to', ''),
            'duration_hrs':   f.get('duration_hrs', 0),
            'level':          f.get('level', ''),
            'mode':           f.get('mode', ''),
            'target_audience': f.get('target_audience', ''),
            'planned_pax':    f.get('planned_pax', 0),
            'trainer_vendor': f.get('trainer_vendor', ''),
        }
        errors, warnings = validate_calendar_row(row, plant_id, db, is_edit=False)
        if errors:
            flash_validation(errors, warnings, flash)
            return redirect(url_for('training_calendar'))

        # All validation passed — extract canonicalised values
        prog_name = _canonical_prog(row['programme_name'], plant_id, db, strict=True)
        prog_type = row['prog_type']
        dur       = float(row['duration_hrs'] or 0)
        source    = row['source'] if row['source'] in ('TNI Driven', 'New Requirement') else 'TNI Driven'

        prog_code    = _get_or_create_prog_code(plant_id, prog_name, prog_type, db)
        session_code = _new_session_code(plant_id, prog_code, db)

        tni_audience  = _derive_audience(plant_id, prog_name, db)
        form_audience = row['target_audience']
        audience      = tni_audience if tni_audience else form_audience

        # Category is inherited from Programme Master (single source of truth).
        # SPOC cannot override per-session — prevents drift.
        cat_row = db.execute(
            'SELECT category FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
            (plant_id, prog_name)).fetchone()
        category = (cat_row['category'] if cat_row and cat_row['category'] else 'General')

        db.execute('''INSERT INTO calendar
            (plant_id,prog_code,session_code,source,programme_name,prog_type,category,
             planned_month,plan_start,plan_end,time_from,time_to,duration_hrs,
             level,mode,target_audience,planned_pax,trainer_vendor,status)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (plant_id, prog_code, session_code, source,
             prog_name, prog_type, category,
             row['planned_month'], row['plan_start'], row['plan_end'],
             row['time_from'], row['time_to'], dur,
             row['level'], row['mode'], audience,
             int(row['planned_pax'] or 0),
             row['trainer_vendor'], 'To Be Planned'))
        db.commit()
        # Audit Tier 3: snapshot the new row in payload
        new_snap = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                              (session_code, plant_id)).fetchone()
        log_record_change('RECORD_ADD', session_code, 'calendar',
                          before=None, after=dict(new_snap) if new_snap else None)
        msg = f'Session {session_code} added.'
        if tni_audience and form_audience and form_audience != tni_audience:
            msg += f' Audience set to "{tni_audience}" (locked from TNI).'
        flash(msg, 'success')
        # Surface non-blocking warnings as well
        if warnings:
            flash_validation([], warnings, flash)
        return redirect(url_for('training_calendar'))

    @app.route('/calendar/<int:cal_id>/delete', methods=['POST'])
    @spoc_required
    def delete_calendar(cal_id):
        db = get_db()
        cal = db.execute('SELECT * FROM calendar WHERE id=? AND plant_id=?',
                         (cal_id, session['plant_id'])).fetchone()
        if cal and cal['status'] == 'Conducted':
            if _is_ajax():
                return 'Conducted sessions cannot be deleted.', 403
            flash('Conducted sessions cannot be deleted.', 'danger')
            return redirect(url_for('training_calendar'))
        before_snap_dict = dict(cal) if cal else None
        if cal:
            sc = cal['session_code']
            pid = session['plant_id']
            # Cascade cleanup — atomic with calendar delete to avoid orphans.
            # Safe because delete is blocked for Conducted sessions above; these
            # downstream rows should normally be empty for non-Conducted ones.
            db.execute('DELETE FROM session_qr WHERE plant_id=? AND session_code=?', (pid, sc))
            db.execute('DELETE FROM emp_training WHERE plant_id=? AND session_code=?', (pid, sc))
            db.execute('DELETE FROM effectiveness_review WHERE plant_id=? AND session_code=?', (pid, sc))
        db.execute('DELETE FROM calendar WHERE id=? AND plant_id=?', (cal_id, session['plant_id']))
        db.commit()
        log_record_change('RECORD_DELETE', cal_id, 'calendar',
                          before=before_snap_dict, after=None)
        if _is_ajax():
            return '', 204
        flash('Calendar entry deleted.', 'warning')
        return redirect(url_for('training_calendar'))

    @app.route('/calendar/<int:cal_id>/edit', methods=['POST'])
    @spoc_required
    def edit_calendar(cal_id):
        plant_id = session['plant_id']
        f = request.form
        db = get_db()
        existing = db.execute('SELECT status, session_code FROM calendar WHERE id=? AND plant_id=?',
                              (cal_id, plant_id)).fetchone()
        if existing and existing['status'] == 'Conducted':
            flash('Conducted sessions cannot be edited.', 'danger')
            return redirect(url_for('training_calendar'))
        # Defence-in-depth: block downgrade away from a post-conducted state
        # (Conducted/Awaiting Verification) while pending effectiveness_review
        # rows exist. Prevents orphan eff_review rows pointing at a session
        # whose status no longer warrants them. SPOC must clear/handle via the
        # verify_reject path (which cascades) before downgrading here.
        new_status = f.get('status', 'To Be Planned')
        if (existing and existing['status'] in ('Conducted', 'Awaiting Verification')
                and new_status not in ('Conducted', 'Awaiting Verification')):
            sc_chk = existing['session_code']
            pending_eff = db.execute(
                'SELECT COUNT(*) FROM effectiveness_review '
                'WHERE plant_id=? AND session_code=? AND completed_date IS NULL',
                (plant_id, sc_chk)).fetchone()[0]
            if pending_eff > 0:
                flash(
                    f"Can't downgrade — {pending_eff} pending effectiveness review(s) "
                    f"exist for {sc_chk}. Use Verify Sessions → Reject to cascade-clean, "
                    f"or complete the reviews first.",
                    'danger')
                return redirect(url_for('training_calendar'))
        # Privilege boundary: a SPOC must NOT be able to set 'Conducted' (or
        # 'Awaiting Verification') directly via the calendar edit form. That would
        # bypass the Central verification chokepoint — skipping the verification
        # log, the anomaly review, and effectiveness seeding that verify_approve /
        # 2C-save perform. The only legitimate path is: record Programme Details
        # (2C) -> status becomes 'Awaiting Verification' -> Central approves ->
        # 'Conducted'. Admin (trusted) may still set status directly.
        if session.get('role') != 'admin' and new_status in ('Conducted', 'Awaiting Verification'):
            flash(
                "SPOCs can't set this status directly. Record Programme Details (2C) — "
                "it is submitted to Central L&D for verification, and Central approval "
                "marks the session Conducted.",
                'danger')
            return redirect(url_for('training_calendar'))
        row = {
            'programme_name': f.get('programme_name', ''),
            'prog_type':      f.get('prog_type', ''),
            'source':         f.get('source', ''),
            'planned_month':  f.get('planned_month', ''),
            'plan_start':     f.get('plan_start', ''),
            'plan_end':       f.get('plan_end', ''),
            'time_from':      f.get('time_from', ''),
            'time_to':        f.get('time_to', ''),
            'duration_hrs':   f.get('duration_hrs', 0),
            'level':          f.get('level', ''),
            'mode':           f.get('mode', ''),
            'target_audience': f.get('target_audience', ''),
            'planned_pax':    f.get('planned_pax', 0),
            'trainer_vendor': f.get('trainer_vendor', ''),
            'status':         f.get('status', 'To Be Planned'),
        }
        prev = db.execute('SELECT prog_type FROM calendar WHERE id=?', (cal_id,)).fetchone()
        prev_pt = prev['prog_type'] if prev else None
        errors, warnings = validate_calendar_row(row, plant_id, db, is_edit=True,
                                                  exclude_id=cal_id, prev_prog_type=prev_pt)
        if errors:
            flash_validation(errors, warnings, flash)
            return redirect(url_for('training_calendar'))

        # Audit Tier 3: capture before-snapshot for field-level diff
        before_snap = db.execute('SELECT * FROM calendar WHERE id=? AND plant_id=?',
                                  (cal_id, plant_id)).fetchone()
        before_snap_dict = dict(before_snap) if before_snap else None

        edit_prog = _canonical_prog(row['programme_name'], plant_id, db, strict=True)
        dur       = float(row['duration_hrs'] or 0)
        source    = row['source'] if row['source'] in ('TNI Driven', 'New Requirement') else 'TNI Driven'

        tni_audience_edit  = _derive_audience(plant_id, edit_prog, db)
        form_audience_edit = row['target_audience']
        edit_audience      = tni_audience_edit if tni_audience_edit else form_audience_edit

        # Re-derive Category from Programme Master (in case programme changed)
        cat_row = db.execute(
            'SELECT category FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
            (plant_id, edit_prog)).fetchone()
        edit_category = (cat_row['category'] if cat_row and cat_row['category'] else 'General')

        db.execute('''UPDATE calendar SET
            programme_name=?, prog_type=?, source=?, category=?, planned_month=?,
            plan_start=?, plan_end=?, time_from=?, time_to=?,
            duration_hrs=?, level=?, mode=?, target_audience=?,
            planned_pax=?, trainer_vendor=?, status=?
            WHERE id=? AND plant_id=?''',
            (edit_prog, row['prog_type'], source, edit_category,
             row['planned_month'], row['plan_start'], row['plan_end'],
             row['time_from'], row['time_to'], dur,
             row['level'], row['mode'], edit_audience,
             int(row['planned_pax'] or 0), row['trainer_vendor'],
             row['status'],
             cal_id, plant_id))
        db.commit()
        after_snap = db.execute('SELECT * FROM calendar WHERE id=? AND plant_id=?',
                                 (cal_id, plant_id)).fetchone()
        log_record_change('RECORD_EDIT', cal_id, 'calendar',
                          before=before_snap_dict,
                          after=dict(after_snap) if after_snap else None)

        # Tier 5: persist reschedule history when dates or Re-Scheduled status changes
        if before_snap_dict and after_snap:
            date_changed = (
                before_snap_dict.get('plan_start') != after_snap['plan_start'] or
                before_snap_dict.get('plan_end')   != after_snap['plan_end']
            )
            became_rescheduled = (
                before_snap_dict.get('status') != 'Re-Scheduled' and
                after_snap['status'] == 'Re-Scheduled'
            )
            if date_changed or became_rescheduled:
                db.execute(
                    'INSERT INTO calendar_reschedule_history '
                    '(plant_id, cal_id, session_code, old_plan_start, old_plan_end, '
                    ' new_plan_start, new_plan_end, old_status, new_status, actor, reason) '
                    'VALUES(?,?,?,?,?,?,?,?,?,?,?)',
                    (plant_id, cal_id,
                     after_snap['session_code'],
                     before_snap_dict.get('plan_start'),
                     before_snap_dict.get('plan_end'),
                     after_snap['plan_start'],
                     after_snap['plan_end'],
                     before_snap_dict.get('status'),
                     after_snap['status'],
                     session.get('username', 'unknown'),
                     (f.get('reschedule_reason') or '').strip()[:500]))
                db.commit()
        msg = 'Session updated.'
        if tni_audience_edit and form_audience_edit and form_audience_edit != tni_audience_edit:
            msg += f' Audience locked to "{tni_audience_edit}" from TNI.'
        flash(msg, 'success')
        if warnings:
            flash_validation([], warnings, flash)
        return redirect(url_for('training_calendar'))

    @app.route('/calendar/bulk-delete', methods=['POST'])
    @spoc_required
    def calendar_bulk_delete():
        plant_id = session['plant_id']
        ids = request.form.getlist('ids[]')
        if ids:
            db = get_db()
            deleted = 0
            for i in range(0, len(ids), 900):
                chunk = ids[i:i+900]
                ph = ','.join('?' * len(chunk))
                # Fetch rows eligible for delete (non-Conducted only) with full
                # snapshot for per-record audit + session_code for cascade.
                rows = db.execute(
                    f'SELECT * FROM calendar WHERE id IN ({ph}) AND plant_id=? AND status != "Conducted"',
                    chunk + [plant_id]
                ).fetchall()
                for r in rows:
                    sc = r['session_code']
                    # Cascade cleanup — atomic with calendar delete to avoid orphans.
                    db.execute('DELETE FROM session_qr WHERE plant_id=? AND session_code=?',
                               (plant_id, sc))
                    db.execute('DELETE FROM emp_training WHERE plant_id=? AND session_code=?',
                               (plant_id, sc))
                    db.execute('DELETE FROM effectiveness_review WHERE plant_id=? AND session_code=?',
                               (plant_id, sc))
                    db.execute('DELETE FROM calendar WHERE id=? AND plant_id=?',
                               (r['id'], plant_id))
                    log_record_change('RECORD_DELETE', r['id'], 'calendar',
                                      before=dict(r), after=None)
                    deleted += 1
            db.commit()
            log_action('BULK_DELETE', f"cal:{deleted}")
            flash(f'{deleted} calendar sessions deleted.', 'warning')
        return redirect(url_for('training_calendar'))

    @app.route('/calendar/template')
    @spoc_required
    def calendar_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Calendar_Bulk_Upload'
        headers = ['Programme Name', 'Type of Programme', 'Source', 'Planned Month',
                   'Plan Start (DD-MM-YYYY)', 'Plan End (DD-MM-YYYY)', 'Duration (Hrs)',
                   'Level', 'Mode', 'Target Audience', 'Planned Pax', 'Trainer/Vendor']
        hdr_fill = PatternFill('solid', fgColor='1F4E79')
        hdr_font = Font(bold=True, color='FFFFFF')
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font
            ws.column_dimensions[get_column_letter(i)].width = 24
        ws.append(['Fire Safety Training', 'EHS/HR', 'TNI Driven', 'June', '10-06-2026', '10-06-2026', 4, 'General', 'Classroom', 'Blue Collared', 30, 'Internal Faculty'])
        ws.append(['Leadership Skills', 'Behavioural/Leadership', 'New Requirement', 'July', '05-07-2026', '06-07-2026', 8, 'Specialized', 'Classroom', 'White Collared', 20, 'External Vendor'])
        ws['A5'] = 'NOTE: Dates MUST be DD-MM-YYYY (e.g. 15-06-2026).'
        ws['A6'] = 'VALID Types: Behavioural/Leadership | Cane | Commercial | EHS/HR | IT | Technical'
        ws['A7'] = 'VALID Modes: Classroom | OJT | SOP | Online'
        ws['A8'] = 'VALID Audience: Blue Collared | White Collared | Common'
        ws['A9'] = 'VALID Months: April | May | June | July | August | September | October | November | December | January | February | March'
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, download_name='Calendar_Bulk_Upload_Template.xlsx', as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/calendar/bulk', methods=['POST'])
    @spoc_required
    def calendar_bulk_upload():
        plant_id = session['plant_id']
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('training_calendar'))
        try:
            df = _read_upload_file(f)
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('training_calendar'))
        db = get_db(); inserted = 0; errors = []; warnings_all = []
        for i, row_in in df.iterrows():
            prog_name = _clean(row_in, ['programme name', 'programme_name', 'program name'])
            prog_type = _clean(row_in, ['type of programme', 'type', 'prog type'])
            raw_src   = _clean(row_in, ['source']) or ''
            month     = _clean(row_in, ['planned month', 'month'])
            raw_start = _clean(row_in, ['plan start (dd-mm-yyyy)', 'plan start (yyyy-mm-dd)', 'plan start', 'start date'])
            raw_end   = _clean(row_in, ['plan end (dd-mm-yyyy)', 'plan end (yyyy-mm-dd)', 'plan end', 'end date'])
            try:
                plan_start = _parse_date_strict(raw_start)
                plan_end   = _parse_date_strict(raw_end)
            except ValueError as e:
                errors.append(f'Row {i+2}: Date format error — {e}. Use DD-MM-YYYY.')
                continue
            duration  = _safe_float(_clean(row_in, ['duration (hrs)', 'duration', 'hrs'])) or 0
            level     = _clean(row_in, ['level'])
            mode      = _clean(row_in, ['mode'])
            audience  = _clean(row_in, ['target audience', 'audience'])
            pax       = int(_safe_float(_clean(row_in, ['planned pax', 'pax'])) or 0)
            trainer   = _clean(row_in, ['trainer/vendor', 'trainer', 'vendor'])
            time_from = _clean(row_in, ['time from', 'start time'])
            time_to   = _clean(row_in, ['time to', 'end time'])

            # Run same centralised validator as single-row add/edit
            validate_input = {
                'programme_name': prog_name, 'prog_type': prog_type,
                'source': raw_src, 'planned_month': month,
                'plan_start': plan_start, 'plan_end': plan_end,
                'time_from': time_from, 'time_to': time_to,
                'duration_hrs': duration, 'level': level, 'mode': mode,
                'target_audience': audience, 'planned_pax': pax,
                'trainer_vendor': trainer,
            }
            row_errors, row_warnings = validate_calendar_row(validate_input, plant_id, db, is_edit=False)
            if row_errors:
                for fld, msg in row_errors:
                    errors.append(f'Row {i+2} [{fld}]: {msg}')
                continue
            for fld, msg in row_warnings:
                warnings_all.append(f'Row {i+2} [{fld}]: {msg}')

            # Use validator-corrected values (e.g. auto-derived planned_month)
            prog_name    = _canonical_prog(prog_name, plant_id, db, strict=True)
            month        = validate_input['planned_month']
            source       = raw_src if raw_src in ('TNI Driven', 'New Requirement') else 'TNI Driven'
            tni_aud      = _derive_audience(plant_id, prog_name, db)
            audience     = tni_aud if tni_aud else audience
            prog_code    = _get_or_create_prog_code(plant_id, prog_name, prog_type, db)
            session_code = _new_session_code(plant_id, prog_code, db)
            db.execute('''INSERT INTO calendar
                (plant_id,prog_code,session_code,source,programme_name,prog_type,
                 planned_month,plan_start,plan_end,time_from,time_to,duration_hrs,
                 level,mode,target_audience,planned_pax,trainer_vendor,status)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'To Be Planned')''',
                (plant_id, prog_code, session_code, source, prog_name, prog_type,
                 month, plan_start, plan_end, time_from, time_to,
                 duration, level, mode, audience, pax, trainer))
            inserted += 1
        db.commit()
        # Surface warnings (non-blocking) as well
        for w in warnings_all[:20]:  # cap to avoid flash flood
            flash(f'⚠ {w}', 'warning')
        if len(warnings_all) > 20:
            flash(f'⚠ +{len(warnings_all) - 20} more warnings suppressed.', 'warning')
        if errors:
            if inserted:
                flash(f'Bulk upload complete: {inserted} sessions added. {len(errors)} rows had errors — downloading error report.', 'warning')
            return _error_excel_response(errors, inserted, 'Calendar_Upload_Errors.xlsx')
        log_action('BULK_UPLOAD', f"cal:{inserted}")
        flash(f'Bulk upload complete: {inserted} sessions added to calendar.', 'success')
        return redirect(url_for('training_calendar'))
