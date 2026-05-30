import io as _io
import os as _os
from flask import render_template, request, redirect, url_for, session, flash, jsonify

from datetime import date as _date

from tms.constants import PLANTS, PLANT_MAP, MONTHS_FY
from tms.db import get_db
from tms.decorators import central_required
from tms.helpers import _calc_summary, _calc_totals, _calc_compliance, _current_fy


# --- SPOC error categorisation ---------------------------------------------
# Group raw issue strings by root cause so the UI shows actionable categories,
# not one row per emp_code-not-found incident.
ERROR_CATEGORIES = [
    ('emp_not_found', 'Employee Not Found in Plant',
     'SPOC pulling from stale emp master or wrong plant'),
    ('emp_missing',   'Employee Code Missing',
     'Required field blank in upload'),
    ('prog_missing',  'Programme Name Missing',
     'Required field blank in upload'),
    ('prog_garbage',  'Invalid / Junk Programme Name',
     'Placeholder text (NA, TBD, ?) or special-char garbage'),
    ('type_invalid',  'Invalid Programme Type',
     'Value not in dropdown (Behavioural/Cane/EHS/HR/IT/Technical/Commercial)'),
    ('mode_invalid',  'Invalid Mode',
     'Value not in dropdown (Classroom/OJT/SOP/Online)'),
    ('duplicate',     'Duplicate Within File',
     'Same employee+programme appears twice in same upload'),
    ('other',         'Other / Uncategorised',
     'Did not match any known pattern — review manually'),
]
ERROR_CAT_LABELS = {c[0]: c[1] for c in ERROR_CATEGORIES}
ERROR_CAT_DESC   = {c[0]: c[2] for c in ERROR_CATEGORIES}

def _categorize_error(issues_text, row_status):
    """Map an error row to one of ERROR_CATEGORIES based on issues text + status."""
    if row_status == 'duplicate':
        return 'duplicate'
    t = (issues_text or '').lower()
    if 'not found in this plant' in t: return 'emp_not_found'
    if 'employee code is missing' in t: return 'emp_missing'
    if 'programme name is missing' in t: return 'prog_missing'
    if 'programme name invalid' in t:   return 'prog_garbage'
    if 'unknown programme type' in t:   return 'type_invalid'
    if 'unknown mode' in t:             return 'mode_invalid'
    return 'other'


def _by_plant(rows, key='plant_id', val='cnt'):
    return {r[key]: r[val] for r in rows}


def _register(app):

    @app.route('/central')
    @central_required
    def central_dashboard():
        db = get_db()
        fy_start, fy_end = _current_fy()

        # ── 7 batched queries replace ~80 per-plant queries ─────────────────
        hc = {r['plant_id']: {'bc': r['bc'], 'wc': r['wc']} for r in db.execute(
            "SELECT plant_id, "
            "  SUM(CASE WHEN collar='Blue Collared'  THEN 1 ELSE 0 END) AS bc, "
            "  SUM(CASE WHEN collar='White Collared' THEN 1 ELSE 0 END) AS wc "
            "FROM employees WHERE is_active=1 GROUP BY plant_id"
        ).fetchall()}
        own_planned = _by_plant(db.execute(
            "SELECT plant_id, COUNT(*) AS cnt FROM calendar "
            "WHERE plan_start BETWEEN ? AND ? GROUP BY plant_id",
            (fy_start, fy_end)).fetchall())
        own_cond = _by_plant(db.execute(
            "SELECT plant_id, COUNT(*) AS cnt FROM calendar "
            "WHERE status='Conducted' AND plan_start BETWEEN ? AND ? GROUP BY plant_id",
            (fy_start, fy_end)).fetchall())
        cen_att = _by_plant(db.execute(
            "SELECT plant_id, COUNT(DISTINCT session_code) AS cnt FROM emp_training "
            "WHERE host_plant_id=99 AND session_code IS NOT NULL AND session_code!='' "
            "AND start_date>=? AND start_date<=? GROUP BY plant_id",
            (fy_start, fy_end)).fetchall())
        mh_all = _by_plant(db.execute(
            "SELECT plant_id, COALESCE(SUM(hrs),0) AS cnt FROM emp_training "
            "WHERE start_date>=? AND start_date<=? GROUP BY plant_id",
            (fy_start, fy_end)).fetchall())
        mh_bc = _by_plant(db.execute(
            "SELECT t.plant_id, COALESCE(SUM(t.hrs),0) AS cnt "
            "FROM emp_training t JOIN employees e "
            "  ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id "
            "WHERE e.collar='Blue Collared' AND t.start_date>=? AND t.start_date<=? "
            "GROUP BY t.plant_id", (fy_start, fy_end)).fetchall())
        mh_wc = _by_plant(db.execute(
            "SELECT t.plant_id, COALESCE(SUM(t.hrs),0) AS cnt "
            "FROM emp_training t JOIN employees e "
            "  ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id "
            "WHERE e.collar='White Collared' AND t.start_date>=? AND t.start_date<=? "
            "GROUP BY t.plant_id", (fy_start, fy_end)).fetchall())

        plant_summaries = []
        for p in PLANTS:
            pid = p['id']
            bc  = hc.get(pid, {}).get('bc', 0)
            wc  = hc.get(pid, {}).get('wc', 0)
            own_s = own_planned.get(pid, 0)
            own_c = own_cond.get(pid, 0)
            cen_c = cen_att.get(pid, 0)
            manhours = mh_all.get(pid, 0)
            bc_hrs   = mh_bc.get(pid, 0)
            wc_hrs   = mh_wc.get(pid, 0)
            bc_mandate = bc * 12
            wc_mandate = wc * 24
            bc_pct = round(bc_hrs / bc_mandate * 100, 1) if bc_mandate else 0
            wc_pct = round(wc_hrs / wc_mandate * 100, 1) if wc_mandate else 0
            plant_summaries.append({**p,
                'blue_collar': bc, 'white_collar': wc, 'total_emp': bc + wc,
                'sessions':  own_s + cen_c, 'conducted': own_c + cen_c,
                'own_sessions': own_s, 'own_conducted': own_c,
                'manhours': round(manhours, 1),
                'bc_pct': bc_pct, 'wc_pct': wc_pct,
            })

        plant_summaries.sort(key=lambda p: (p['bc_pct'] + p['wc_pct']) / 2, reverse=True)

        grand_central = db.execute(
            "SELECT COUNT(DISTINCT session_code) FROM emp_training "
            "WHERE host_plant_id=99 AND session_code IS NOT NULL AND session_code!='' "
            "AND start_date>=? AND start_date<=?",
            (fy_start, fy_end)).fetchone()[0]
        grand = {
            'total_emp': sum(p['total_emp'] for p in plant_summaries),
            'manhours':  round(sum(p['manhours'] for p in plant_summaries), 1),
            'sessions':  sum(p['own_sessions'] for p in plant_summaries) + grand_central,
            'conducted': sum(p['own_conducted'] for p in plant_summaries) + grand_central,
        }

        return render_template('central.html', plants=plant_summaries, grand=grand)

    @app.route('/central/duplicates', methods=['GET', 'POST'])
    @central_required
    def central_duplicates():
        """Programme master duplicate scanner + bulk merger.

        GET: scan all plants (or selected) and render clusters for review.
        POST: apply selected merges with cascade rename across tni/calendar/
              programme_details/emp_training.
        """
        from tms.master_dedup import find_duplicates, merge_cluster
        from tms.audit import log_action
        db = get_db()

        if request.method == 'POST':
            # Form fields:
            #   cluster_idx: 0,1,2,...
            #   winner_<idx>: master id of canonical row
            #   canonical_<idx>: free-text canonical name (defaults to winner's name)
            #   losers_<idx>: comma-separated master ids to merge in
            #   plant_<idx>: plant_id for this cluster
            plant_id = int(request.form.get('plant_id') or 0)
            merged_clusters = 0
            total_counts = {'tni': 0, 'calendar': 0, 'programme_details': 0, 'emp_training': 0,
                            'winner_renamed': 0, 'losers_deleted': 0}
            for key in request.form:
                if not key.startswith('winner_'):
                    continue
                idx = key[len('winner_'):]
                winner_id = int(request.form.get(f'winner_{idx}') or 0)
                losers_raw = request.form.get(f'losers_{idx}', '').strip()
                canonical  = request.form.get(f'canonical_{idx}', '').strip()
                cluster_pid = int(request.form.get(f'plant_{idx}') or plant_id or 0)
                if not winner_id or not losers_raw or not canonical or not cluster_pid:
                    continue
                loser_ids = [int(x) for x in losers_raw.split(',') if x.strip().isdigit()]
                if not loser_ids:
                    continue
                counts = merge_cluster(cluster_pid, winner_id, loser_ids, canonical, db,
                                       audit_log_fn=log_action)
                merged_clusters += 1
                for k, v in counts.items():
                    total_counts[k] = total_counts.get(k, 0) + v
            db.commit()
            if merged_clusters:
                flash(
                    f'Merged {merged_clusters} cluster(s). '
                    f'Cascaded: TNI {total_counts["tni"]}, Calendar {total_counts["calendar"]}, '
                    f'2C {total_counts["programme_details"]}, 2A {total_counts["emp_training"]}. '
                    f'Deleted {total_counts["losers_deleted"]} duplicate master rows.',
                    'success')
            else:
                flash('No merges selected.', 'warning')
            return redirect(url_for('central_duplicates'))

        # GET — scan
        plant_filter = request.args.get('plant_id', '').strip()
        try:
            plant_filter_id = int(plant_filter) if plant_filter else None
        except ValueError:
            plant_filter_id = None
        try:
            threshold = float(request.args.get('threshold', '0.85'))
            threshold = max(0.70, min(0.99, threshold))
        except (ValueError, TypeError):
            threshold = 0.85

        plant_clusters = []
        plants_to_scan = [plant_filter_id] if plant_filter_id else [
            p['id'] for p in PLANTS if p['id'] != 99]
        for pid in plants_to_scan:
            if pid not in PLANT_MAP:
                continue
            dupes = find_duplicates(pid, db, threshold=threshold)
            if dupes:
                plant_clusters.append({
                    'plant_id':   pid,
                    'plant_name': PLANT_MAP[pid]['name'],
                    'clusters':   dupes,
                })

        return render_template('central_duplicates.html',
                               plant_clusters=plant_clusters,
                               threshold=threshold,
                               plants=[p for p in PLANTS if p['id'] != 99],
                               plant_filter=plant_filter_id)

    @app.route('/central/tni-errors')
    @central_required
    def central_tni_errors():
        db = get_db()

        # ─── Filters: FY-driven by default, with optional unit / status / category ───
        fy_start, fy_end = _current_fy()
        fy_year_str = fy_start[:4]              # e.g. "2026"
        fy_label    = f'FY {fy_year_str[2:]}-{int(fy_year_str[2:])+1}'  # "FY 26-27"

        # Allow FY override via ?fy=2025 (means FY 25-26)
        fy_param = request.args.get('fy', '').strip()
        if fy_param.isdigit() and len(fy_param) == 4:
            fy_start = f'{fy_param}-04-01'
            fy_end   = f'{int(fy_param)+1}-03-31'
            fy_label = f'FY {fy_param[2:]}-{int(fy_param[2:])+1}'

        unit_filter   = request.args.get('plant', '').strip()
        status_filter = request.args.get('status', '').strip()
        cat_filter    = request.args.get('cat', '').strip()

        where  = ['date(ts) BETWEEN ? AND ?']
        params = [fy_start, fy_end]
        if unit_filter.isdigit():
            where.append('plant_id=?'); params.append(int(unit_filter))
        if status_filter in ('error', 'duplicate'):
            where.append('row_status=?'); params.append(status_filter)
        wsql = ' AND '.join(where)

        # ─── Pull all rows for the FY in one query, do aggregation in Python ───
        # SQLite has no PIVOT — Python pivot is cleaner than nested CASE WHENs.
        all_rows = db.execute(
            f'SELECT ts, plant_id, username, row_status, issues FROM tni_upload_errors WHERE {wsql}',
            params
        ).fetchall()

        # Build active plant list (10 units; skip Central pseudo-plant 99)
        active_plants = [p for p in PLANTS if p['id'] != 99]
        plant_ids     = [p['id'] for p in active_plants]
        plant_name_by = {p['id']: p['name'] for p in active_plants}

        # Pivot containers
        month_unit_matrix = {m: {pid: 0 for pid in plant_ids} for m in MONTHS_FY}
        month_cat_matrix  = {m: {c[0]: 0 for c in ERROR_CATEGORIES} for m in MONTHS_FY}
        per_plant_total   = {pid: 0 for pid in plant_ids}
        per_plant_cat     = {pid: {c[0]: 0 for c in ERROR_CATEGORIES} for pid in plant_ids}
        per_plant_user    = {pid: {} for pid in plant_ids}  # pid -> {username: count}
        cat_totals        = {c[0]: 0 for c in ERROR_CATEGORIES}
        per_user          = {}  # (username, pid) -> count

        # FY month order — index 0 = April through index 11 = March
        month_num_to_label = {
            4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August', 9: 'September',
            10: 'October', 11: 'November', 12: 'December',
            1: 'January', 2: 'February', 3: 'March',
        }

        for r in all_rows:
            ts = r['ts'] or ''
            try:
                mnum = int(ts[5:7])
            except (ValueError, IndexError):
                continue
            mlabel = month_num_to_label.get(mnum)
            if not mlabel or mlabel not in month_unit_matrix:
                continue
            pid = r['plant_id']
            if pid not in per_plant_total:
                continue
            cat = _categorize_error(r['issues'], r['row_status'])
            if cat_filter and cat != cat_filter:
                continue
            month_unit_matrix[mlabel][pid] += 1
            month_cat_matrix[mlabel][cat]  += 1
            per_plant_total[pid]           += 1
            per_plant_cat[pid][cat]        += 1
            cat_totals[cat]                += 1
            uname = r['username'] or 'unknown'
            per_plant_user[pid][uname] = per_plant_user[pid].get(uname, 0) + 1
            key = (uname, pid)
            per_user[key] = per_user.get(key, 0) + 1

        total_errors = sum(per_plant_total.values())

        # KPI: best / worst unit (only units with at least 1 error count; if none, blank)
        units_with_errors = {pid: cnt for pid, cnt in per_plant_total.items() if cnt > 0}
        if units_with_errors:
            best_pid = min(units_with_errors, key=units_with_errors.get)
            worst_pid = max(units_with_errors, key=units_with_errors.get)
            best_unit  = {'name': plant_name_by[best_pid],  'cnt': units_with_errors[best_pid]}
            worst_unit = {'name': plant_name_by[worst_pid], 'cnt': units_with_errors[worst_pid]}
        else:
            best_unit = worst_unit = {'name': '—', 'cnt': 0}

        # KPI trend: this month vs avg of prior months in FY
        today = _date.today()
        this_month_label = month_num_to_label.get(today.month, MONTHS_FY[0])
        this_month_cnt = sum(month_unit_matrix.get(this_month_label, {}).values())
        prior_months = [m for m in MONTHS_FY if m != this_month_label]
        prior_total  = sum(sum(month_unit_matrix[m].values()) for m in prior_months)
        prior_with_data = sum(1 for m in prior_months if sum(month_unit_matrix[m].values()) > 0)
        fy_avg = (prior_total / prior_with_data) if prior_with_data else 0
        if fy_avg == 0:
            trend_dir = 'flat'; trend_pct = 0
        elif this_month_cnt > fy_avg * 1.1:
            trend_dir = 'up'; trend_pct = int((this_month_cnt - fy_avg) / fy_avg * 100)
        elif this_month_cnt < fy_avg * 0.9:
            trend_dir = 'down'; trend_pct = int((fy_avg - this_month_cnt) / fy_avg * 100)
        else:
            trend_dir = 'flat'; trend_pct = 0

        # Unit comparison table — sort worst-first so training priority obvious
        # Fleet avg used for cell colouring on heatmap
        nonzero_cell_totals = []
        for m in MONTHS_FY:
            for pid in plant_ids:
                v = month_unit_matrix[m][pid]
                if v > 0:
                    nonzero_cell_totals.append(v)
        fleet_cell_avg = (sum(nonzero_cell_totals) / len(nonzero_cell_totals)) if nonzero_cell_totals else 0

        unit_comparison = []
        for pid in plant_ids:
            total = per_plant_total[pid]
            if total == 0:
                continue
            cat_dist = per_plant_cat[pid]
            top_cat_id, top_cat_cnt = max(cat_dist.items(), key=lambda x: x[1])
            top_cat_pct = round((top_cat_cnt / total) * 100) if total else 0
            users = per_plant_user[pid]
            top_user = max(users.items(), key=lambda x: x[1])[0] if users else '—'
            # Per-unit trend: last 3 months vs prior 3
            month_idx_now = MONTHS_FY.index(this_month_label) if this_month_label in MONTHS_FY else 0
            recent_months = MONTHS_FY[max(0, month_idx_now-2):month_idx_now+1]
            prior3_months = MONTHS_FY[max(0, month_idx_now-5):max(0, month_idx_now-2)]
            recent_sum = sum(month_unit_matrix[m][pid] for m in recent_months)
            prior_sum  = sum(month_unit_matrix[m][pid] for m in prior3_months)
            if prior_sum == 0:
                trend = 'new' if recent_sum > 0 else 'flat'
            elif recent_sum > prior_sum * 1.1:
                trend = 'up'
            elif recent_sum < prior_sum * 0.9:
                trend = 'down'
            else:
                trend = 'flat'
            unit_comparison.append({
                'plant_id':   pid,
                'plant_name': plant_name_by[pid],
                'total':      total,
                'top_cat':    ERROR_CAT_LABELS[top_cat_id],
                'top_cat_id': top_cat_id,
                'top_cat_pct': top_cat_pct,
                'top_user':   top_user,
                'trend':      trend,
            })
        unit_comparison.sort(key=lambda x: x['total'], reverse=True)

        # Category totals as ordered list with labels + descriptions
        category_summary = []
        for cat_id, label, desc in ERROR_CATEGORIES:
            cnt = cat_totals[cat_id]
            if cnt == 0:
                continue
            category_summary.append({
                'id': cat_id, 'label': label, 'desc': desc, 'cnt': cnt,
                'pct': round((cnt / total_errors) * 100) if total_errors else 0,
            })

        # Auto-insight — rule-based, no AI
        insights = []
        if worst_unit['cnt'] > 0 and best_unit['cnt'] >= 0 and len(units_with_errors) > 1:
            ratio = (worst_unit['cnt'] / max(1, best_unit['cnt']))
            if ratio >= 2:
                worst_row = next((u for u in unit_comparison if u['plant_name'] == worst_unit['name']), None)
                if worst_row:
                    insights.append(
                        f"<strong>{worst_unit['name']}</strong> has {ratio:.1f}× more errors than "
                        f"<strong>{best_unit['name']}</strong> this FY. "
                        f"Top issue: <strong>{worst_row['top_cat']}</strong> ({worst_row['top_cat_pct']}% of errors). "
                        f"Recommend: targeted training session for SPOC <strong>{worst_row['top_user']}</strong>."
                    )
        if category_summary and category_summary[0]['pct'] >= 50:
            top_cat = category_summary[0]
            insights.append(
                f"<strong>{top_cat['label']}</strong> accounts for {top_cat['pct']}% of all errors fleet-wide. "
                f"Root cause: {top_cat['desc'].lower()}. Add this to next SPOC refresher deck."
            )
        if trend_dir == 'up' and this_month_cnt > 0:
            insights.append(
                f"This month's error count ({this_month_cnt}) is <strong>{trend_pct}% above</strong> "
                f"the FY monthly average ({fy_avg:.0f}). Investigate which unit drove the spike."
            )
        if not insights:
            insights.append("No critical patterns detected. Error volume within normal range across units.")

        # FY options for dropdown (current + last 2)
        cur_yr = int(fy_year_str)
        fy_options = []
        for yr in range(cur_yr, cur_yr - 3, -1):
            fy_options.append({
                'value': str(yr),
                'label': f'FY {str(yr)[2:]}-{int(str(yr)[2:])+1}',
            })

        return render_template(
            'central_tni_errors.html',
            fy_label=fy_label,
            fy_options=fy_options,
            total_errors=total_errors,
            best_unit=best_unit,
            worst_unit=worst_unit,
            this_month_cnt=this_month_cnt,
            fy_avg=round(fy_avg, 1),
            trend_dir=trend_dir,
            trend_pct=trend_pct,
            this_month_label=this_month_label,
            months_fy=MONTHS_FY,
            active_plants=active_plants,
            month_unit_matrix=month_unit_matrix,
            month_cat_matrix=month_cat_matrix,
            fleet_cell_avg=fleet_cell_avg,
            per_plant_total=per_plant_total,
            unit_comparison=unit_comparison,
            category_summary=category_summary,
            error_categories=ERROR_CATEGORIES,
            insights=insights,
            filters={'fy': fy_param or fy_year_str, 'plant': unit_filter,
                     'status': status_filter, 'cat': cat_filter},
        )

    def _build_monthly_error_xlsx(month_start, month_end, db):
        """Build a multi-sheet Excel of SPOC errors for the given month.

        Sheets:
          - Summary: KPIs + per-unit + per-category counts
          - One sheet per plant that has errors that month
        Returns BytesIO buffer + dict of summary stats.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        rows = db.execute(
            '''SELECT ts, plant_id, username, row_status, row_num, emp_code, emp_name,
                      programme_name, prog_type, mode, planned_hours, issues
               FROM tni_upload_errors
               WHERE date(ts) BETWEEN ? AND ?
               ORDER BY plant_id, ts''',
            (month_start, month_end)
        ).fetchall()

        per_plant = {}
        per_cat   = {c[0]: 0 for c in ERROR_CATEGORIES}
        total     = len(rows)
        for r in rows:
            pid = r['plant_id']
            per_plant.setdefault(pid, []).append(r)
            per_cat[_categorize_error(r['issues'], r['row_status'])] += 1

        wb = openpyxl.Workbook()
        hdr_fill = PatternFill('solid', fgColor='1A3A5C')
        hdr_font = Font(color='FFFFFF', bold=True)

        # ─── Summary sheet ───
        s = wb.active
        s.title = 'Summary'
        s.append([f'BCML TMS — SPOC Upload Errors  ({month_start} to {month_end})'])
        s['A1'].font = Font(bold=True, size=14, color='1A3A5C')
        s.append([])
        s.append(['Total errors', total])
        s.append(['Units affected', len(per_plant)])
        s.append(['SPOCs involved', len({r['username'] for r in rows if r['username']})])
        s.append([])
        s.append(['Per-Unit Breakdown'])
        s['A7'].font = Font(bold=True, color='1A3A5C')
        s.append(['Unit', 'Errors'])
        for c in ('A', 'B'):
            s[f'{c}8'].fill = hdr_fill; s[f'{c}8'].font = hdr_font
        ranked = sorted(per_plant.items(), key=lambda x: len(x[1]), reverse=True)
        for pid, prows in ranked:
            s.append([PLANT_MAP.get(pid, {}).get('name', f'Plant {pid}'), len(prows)])
        s.append([])
        s.append(['Error Categories'])
        s[f'A{s.max_row}'].font = Font(bold=True, color='1A3A5C')
        s.append(['Category', 'Count', '%'])
        for c in ('A', 'B', 'C'):
            s[f'{c}{s.max_row}'].fill = hdr_fill; s[f'{c}{s.max_row}'].font = hdr_font
        for cid, lbl, _d in ERROR_CATEGORIES:
            if per_cat[cid] > 0:
                pct = round((per_cat[cid] / total) * 100) if total else 0
                s.append([lbl, per_cat[cid], f'{pct}%'])
        s.column_dimensions['A'].width = 36
        s.column_dimensions['B'].width = 12
        s.column_dimensions['C'].width = 10

        # ─── Per-plant sheets ───
        headers = ['Timestamp', 'SPOC', 'Status', 'Row#', 'Emp Code', 'Emp Name',
                   'Programme Name', 'Type', 'Mode', 'Planned Hrs', 'Category', 'Issue']
        for pid, prows in ranked:
            name = PLANT_MAP.get(pid, {}).get('name', f'Plant {pid}')
            # Sheet names limited to 31 chars in xlsx
            sheet_name = name[:31]
            ws = wb.create_sheet(sheet_name)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal='center')
            for r in prows:
                cat = _categorize_error(r['issues'], r['row_status'])
                ws.append([
                    r['ts'], r['username'] or '—', r['row_status'], r['row_num'],
                    r['emp_code'] or '—', r['emp_name'] or '—',
                    r['programme_name'] or '—', r['prog_type'] or '—',
                    r['mode'] or '—', r['planned_hours'] or 0,
                    ERROR_CAT_LABELS.get(cat, cat),
                    r['issues'] or '—',
                ])
            for ci, w in enumerate([20, 14, 10, 8, 14, 22, 32, 18, 14, 12, 28, 60], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, {
            'total': total,
            'unit_count': len(per_plant),
            'top_units': [(PLANT_MAP.get(p, {}).get('name', f'Plant {p}'), len(rows))
                          for p, rows in ranked[:3]],
            'top_cats': sorted(
                [(ERROR_CAT_LABELS[c], n) for c, n in per_cat.items() if n > 0],
                key=lambda x: x[1], reverse=True
            )[:3],
        }

    @app.route('/cron/monthly-error-report')
    def cron_monthly_error_report():
        """Triggered by Render Cron Job on the 1st of each month at 03:30 UTC (09:00 IST).
        Generates last month's SPOC error report and emails it. Auth: ?token=<CRON_SECRET>.
        Optional ?month=YYYY-MM to backfill any prior month. Optional ?to=email override.
        """
        from tms.email_util import send_email
        import hmac as _hmac
        secret = _os.environ.get('CRON_SECRET', '').strip()
        if not secret:
            return jsonify({'ok': False, 'error': 'CRON_SECRET not configured'}), 500
        # Constant-time compare prevents network-timing byte recovery.
        if not _hmac.compare_digest(request.args.get('token', ''), secret):
            return jsonify({'ok': False, 'error': 'invalid token'}), 403

        # Determine reporting month — default = previous calendar month.
        # IST is UTC+5:30 — for the 1st-of-month cron at 03:30 UTC the previous
        # month is unambiguous regardless of timezone.
        from datetime import date as _d, timedelta as _td
        manual = request.args.get('month', '').strip()
        if manual and len(manual) == 7 and manual[4] == '-':
            try:
                yr, mo = int(manual[:4]), int(manual[5:7])
                if 1 <= mo <= 12:
                    first = _d(yr, mo, 1)
                else:
                    raise ValueError
            except ValueError:
                return jsonify({'ok': False, 'error': 'month must be YYYY-MM'}), 400
        else:
            today = _d.today()
            # First day of CURRENT month, then subtract one day to get last day of previous month
            first_this = _d(today.year, today.month, 1)
            last_prev  = first_this - _td(days=1)
            first      = _d(last_prev.year, last_prev.month, 1)

        # Compute month_end
        if first.month == 12:
            next_first = _d(first.year + 1, 1, 1)
        else:
            next_first = _d(first.year, first.month + 1, 1)
        month_end = next_first - _td(days=1)
        month_label = first.strftime('%B %Y')

        db = get_db()
        buf, summary = _build_monthly_error_xlsx(
            first.isoformat(), month_end.isoformat(), db
        )

        recipient = (request.args.get('to') or
                     _os.environ.get('REPORT_TO_EMAIL') or
                     'yogendradas017@gmail.com').strip()

        # Body
        if summary['total'] == 0:
            body_html = f'''
            <p>Hi,</p>
            <p>No SPOC upload errors recorded across any unit in <strong>{month_label}</strong>. Clean month.</p>
            <p>— BCML TMS</p>
            '''
        else:
            top_units = ''.join(f'<li><strong>{n}</strong>: {c} errors</li>' for n, c in summary['top_units'])
            top_cats  = ''.join(f'<li><strong>{n}</strong>: {c}</li>' for n, c in summary['top_cats'])
            body_html = f'''
            <p>Hi,</p>
            <p>Monthly SPOC upload error report for <strong>{month_label}</strong>:</p>
            <ul style="font-size:14px;line-height:1.7;">
              <li>Total errors: <strong>{summary['total']}</strong></li>
              <li>Units affected: <strong>{summary['unit_count']}</strong></li>
            </ul>
            <p><strong>Top units needing attention:</strong></p>
            <ol>{top_units}</ol>
            <p><strong>Top error categories:</strong></p>
            <ol>{top_cats}</ol>
            <p>Full breakdown attached. Use it to plan next month's SPOC refresher training.</p>
            <p>— BCML TMS</p>
            '''

        ok, detail = send_email(
            to_addrs=recipient,
            subject=f'BCML TMS — SPOC Upload Errors Report — {month_label}',
            body_html=body_html,
            attachments=[(
                f'SPOC_Errors_{first.strftime("%Y-%m")}.xlsx',
                buf.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )],
        )
        status_code = 200 if ok else 500
        return jsonify({
            'ok': ok, 'detail': detail,
            'month': month_label, 'total_errors': summary['total'],
            'recipient': recipient,
        }), status_code

    @app.route('/cron/backup')
    def cron_backup():
        """Nightly DB backup → emailed as gzipped attachment.
        Auth: ?token=<CRON_SECRET>. Recipient: BACKUP_RECIPIENT env (defaults to
        SMTP_USER). Uses sqlite3 online .backup API — safe with live writers.
        """
        import hmac as _hmac, sqlite3 as _sq3, gzip as _gz, tempfile as _tmp
        import os as _os2
        from datetime import datetime as _dt
        from tms.email_util import send_email
        from tms.constants import DB_PATH

        secret = _os.environ.get('CRON_SECRET', '').strip()
        if not secret:
            return jsonify({'ok': False, 'error': 'CRON_SECRET not configured'}), 500
        if not _hmac.compare_digest(request.args.get('token', ''), secret):
            return jsonify({'ok': False, 'error': 'invalid token'}), 403

        recipient = (request.args.get('to', '').strip()
                     or _os.environ.get('BACKUP_RECIPIENT', '').strip()
                     or _os.environ.get('SMTP_USER', '').strip())
        if not recipient:
            return jsonify({'ok': False, 'error': 'no recipient'}), 500

        stamp = _dt.now().strftime('%Y-%m-%d')
        tmpd = _tmp.mkdtemp()
        try:
            snap_path = _os2.path.join(tmpd, 'snap.db')
            # Online-safe snapshot — works while gunicorn holds connection.
            src = _sq3.connect(DB_PATH)
            dst = _sq3.connect(snap_path)
            with dst:
                src.backup(dst)
            dst.close(); src.close()

            with open(snap_path, 'rb') as f:
                raw = f.read()
            gz_bytes = _gz.compress(raw, compresslevel=9)
            size_mb = len(gz_bytes) / (1024 * 1024)
        finally:
            try:
                import shutil as _sh; _sh.rmtree(tmpd, ignore_errors=True)
            except Exception:
                pass

        body = f"""
        <p>TMS nightly backup — {stamp}</p>
        <ul>
            <li>DB source: <code>{DB_PATH}</code></li>
            <li>Compressed size: <strong>{size_mb:.2f} MB</strong></li>
            <li>Retention: kept in your Gmail; search "BCML TMS Backup" to find</li>
            <li>Restore: gunzip the attachment, replace DATABASE_PATH file, restart</li>
        </ul>
        <p>— BCML TMS</p>
        """
        ok, detail = send_email(
            to_addrs=recipient,
            subject=f'BCML TMS Backup — {stamp}',
            body_html=body,
            attachments=[(f'tms_backup_{stamp}.db.gz', gz_bytes, 'application/gzip')],
        )
        return jsonify({
            'ok': ok, 'detail': detail, 'date': stamp,
            'size_mb': round(size_mb, 2), 'recipient': recipient,
        }), (200 if ok else 500)

    @app.route('/central/plant/<int:plant_id>')
    @central_required
    def central_plant_view(plant_id):
        if plant_id not in PLANT_MAP:
            flash('Plant not found.', 'danger')
            return redirect(url_for('central_dashboard'))
        plant        = PLANT_MAP[plant_id]
        db           = get_db()
        sel_month    = request.args.get('month', '')
        summary_rows = _calc_summary(plant_id, sel_month, db)
        totals       = _calc_totals(summary_rows, db=db, plant_id=plant_id)
        compliance   = _calc_compliance(plant_id, db)
        return render_template('central_plant.html', plant=plant,
                               summary_rows=summary_rows, totals=totals,
                               compliance=compliance, months=MONTHS_FY,
                               selected_month=sel_month)
