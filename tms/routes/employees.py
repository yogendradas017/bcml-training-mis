import io
import os
from datetime import date
from flask import render_template, request, redirect, url_for, session, flash, send_file
import sqlite3

from tms.constants import GENDERS, GRADES, CATEGORIES, COLLARS, PH_OPTIONS, TEMP_UPLOAD_DIR, EMPLOYEES_PAGE_CAP
from tms.db import get_db
from tms.decorators import spoc_required
from tms.helpers import normalise_collar, _today_ist, _canonical_emp_field, _smart_title
from tms.audit import log_action, log_record_change

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    _XLSX = True
except ImportError:
    _XLSX = False


def _plant_depts(db, plant_id):
    rows = db.execute(
        "SELECT DISTINCT UPPER(TRIM(department)) FROM employees "
        "WHERE plant_id=? AND department IS NOT NULL AND department!='' ORDER BY 1",
        (plant_id,)).fetchall()
    return [r[0] for r in rows]


def _plant_sections(db, plant_id):
    rows = db.execute(
        "SELECT DISTINCT UPPER(TRIM(section)) FROM employees "
        "WHERE plant_id=? AND section IS NOT NULL AND section!='' ORDER BY 1",
        (plant_id,)).fetchall()
    return [r[0] for r in rows]


def _plant_designations(db, plant_id):
    rows = db.execute(
        "SELECT DISTINCT TRIM(designation) FROM employees "
        "WHERE plant_id=? AND designation IS NOT NULL AND designation!='' ORDER BY 1",
        (plant_id,)).fetchall()
    return [r[0] for r in rows]


def _validate_emp_fields(f, departments, sections):
    """Return list of error strings. Empty list = all good.
    All fields mandatory except Remarks."""
    errs = []
    emp_code = (f.get('emp_code') or '').strip()
    name     = (f.get('name') or '').strip()
    desig    = (f.get('designation') or '').strip()
    grade    = (f.get('grade') or '').strip().upper()
    collar   = normalise_collar(f.get('collar', ''))
    dept     = (f.get('department') or '').strip().upper()
    sect     = (f.get('section') or '').strip().upper()
    cat      = (f.get('category') or '').strip().upper()
    gender   = (f.get('gender') or '').strip()
    ph       = (f.get('physically_handicapped') or '').strip()

    if not emp_code: errs.append('Employee Code is required.')
    if not name:     errs.append('Full Name is required.')
    if not desig:    errs.append('Designation is required.')
    if not grade:    errs.append('Grade is required.')
    if not collar:   errs.append('Collar Type is required.')
    if not dept:     errs.append('Department is required.')
    if not sect:     errs.append('Section is required.')
    if not cat:      errs.append('Category is required.')
    if not gender:   errs.append('Gender is required.')
    if not ph:       errs.append('Physically Handicapped is required.')

    if grade and grade not in GRADES:
        errs.append(f"Invalid grade '{f.get('grade')}'. Must be one of the predefined grades.")
    if collar and collar not in COLLARS:
        errs.append(f"Invalid collar '{f.get('collar')}'. Must be Blue Collared or White Collared.")
    if cat and cat not in [c.upper() for c in CATEGORIES]:
        errs.append(f"Invalid category '{f.get('category')}'.")
    if gender and gender not in GENDERS:
        errs.append(f"Invalid gender '{gender}'. Must be Male, Female, or Others.")
    if ph and ph not in PH_OPTIONS:
        errs.append(f"Physically Handicapped must be Yes or No.")
    return errs


def _register(app):

    @app.route('/employees')
    @spoc_required
    def employees():
        plant_id = session['plant_id']
        db = get_db()
        show_exited = request.args.get('show_exited', '0') == '1'
        show_all = request.args.get('show_all', '0') == '1'

        base_sql = 'SELECT * FROM employees WHERE plant_id=?' + ('' if show_exited else ' AND is_active=1')
        total_count = db.execute(
            f'SELECT COUNT(*) FROM ({base_sql})', (plant_id,)).fetchone()[0]

        # Default render cap from constants (override per-org via org_config later).
        LIMIT = EMPLOYEES_PAGE_CAP
        if show_all or total_count <= LIMIT:
            emps = db.execute(f'{base_sql} ORDER BY name', (plant_id,)).fetchall()
            truncated = False
        else:
            emps = db.execute(f'{base_sql} ORDER BY name LIMIT {LIMIT}',
                              (plant_id,)).fetchall()
            truncated = True

        recent_exited = db.execute(
            "SELECT * FROM employees WHERE plant_id=? AND is_active=0 AND exit_date >= date('now','-7 days') ORDER BY exit_date DESC",
            (plant_id,)).fetchall()
        departments = _plant_depts(db, plant_id)
        sections = _plant_sections(db, plant_id)
        designations = _plant_designations(db, plant_id)
        # Pending effectiveness reviews per emp_code → shown in the Exit modal up
        # front, so the SPOC sees the block reason before submitting (not after).
        pending_reviews = {r['emp_code']: r['c'] for r in db.execute(
            "SELECT emp_code, COUNT(*) c FROM effectiveness_review "
            "WHERE plant_id=? AND (completed_date IS NULL OR completed_date='') "
            "GROUP BY emp_code", (plant_id,)).fetchall()}
        return render_template('employees.html',
                               employees=emps, show_exited=show_exited,
                               recent_exited=recent_exited,
                               genders=GENDERS, grades=GRADES, categories=CATEGORIES,
                               collars=COLLARS, ph_options=PH_OPTIONS,
                               departments=departments, sections=sections,
                               designations=designations,
                               pending_reviews=pending_reviews,
                               total_count=total_count, truncated=truncated,
                               show_all=show_all,
                               today=str(_today_ist()))

    @app.route('/employees/check-code')
    @spoc_required
    def emp_check_code():
        """Live duplicate-code check for the Add Employee form (debounced client-side)."""
        from flask import jsonify
        plant_id = session['plant_id']
        code = (request.args.get('code') or '').strip()
        if not code:
            return jsonify({'exists': False})
        row = get_db().execute(
            'SELECT name, is_active FROM employees WHERE plant_id=? AND emp_code=?',
            (plant_id, code)).fetchone()
        if row:
            return jsonify({'exists': True, 'name': row['name'], 'active': bool(row['is_active'])})
        return jsonify({'exists': False})

    @app.route('/employees/suggest-similar')
    @spoc_required
    def emp_suggest_similar():
        """UI fuzzy-suggest before user creates new designation/dept/section value.
        Returns {similar: <canonical>} if a close existing match exists, else {}."""
        from flask import jsonify
        plant_id = session['plant_id']
        field = (request.args.get('field') or '').strip()
        value = (request.args.get('value') or '').strip()
        if field not in ('designation', 'department', 'section') or not value:
            return jsonify({})
        db = get_db()
        canonical = _canonical_emp_field(value, plant_id, db, field)
        # Only flag if canonical differs (case-insensitive) from raw input
        if canonical and canonical.lower() != value.lower():
            return jsonify({'similar': canonical, 'raw': value})
        return jsonify({})

    @app.route('/employees/add', methods=['POST'])
    @spoc_required
    def add_employee():
        plant_id = session['plant_id']
        f = request.form
        db = get_db()
        departments = _plant_depts(db, plant_id)
        sections = _plant_sections(db, plant_id)
        errs = _validate_emp_fields(f, departments, sections)
        if errs:
            for e in errs:
                flash(e, 'danger')
            return redirect(url_for('employees'))
        collar = normalise_collar(f.get('collar', ''))
        grade = (f.get('grade') or '').strip().upper() or ''
        desig = _canonical_emp_field(f.get('designation', ''), plant_id, db, 'designation')
        dept = _canonical_emp_field(f.get('department', ''), plant_id, db, 'department')
        sect = _canonical_emp_field(f.get('section', ''), plant_id, db, 'section')
        cat = (f.get('category') or '').strip().upper() or ''
        try:
            db.execute('''INSERT INTO employees
                (plant_id,emp_code,name,designation,grade,collar,department,section,
                 category,gender,physically_handicapped,remarks)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
                (plant_id, f['emp_code'].strip(), f['name'].strip(),
                 desig, grade, collar,
                 dept, sect, cat,
                 f.get('gender', ''), f.get('physically_handicapped', 'No'),
                 f.get('remarks', '')))
            db.commit()
            log_action('RECORD_ADD', f"emp:{f['emp_code'].strip()}")
            flash(f"Employee {f['name'].strip()} added successfully.", 'success')
        except sqlite3.IntegrityError:
            flash(f"Employee code {f['emp_code'].strip()} already exists.", 'danger')
        return redirect(url_for('employees'))

    @app.route('/employees/<int:emp_id>/edit', methods=['POST'])
    @spoc_required
    def edit_employee(emp_id):
        plant_id = session['plant_id']
        db = get_db()
        before = db.execute('SELECT * FROM employees WHERE id=? AND plant_id=?',
                            (emp_id, plant_id)).fetchone()
        if not before:
            flash('Employee not found.', 'danger')
            return redirect(url_for('employees'))
        f = request.form
        departments = _plant_depts(db, plant_id)
        sections    = _plant_sections(db, plant_id)
        # emp_code is immutable — replay current value through validator so it passes.
        f_with_code = dict(f); f_with_code['emp_code'] = before['emp_code']
        errs = _validate_emp_fields(f_with_code, departments, sections)
        if errs:
            for e in errs:
                flash(e, 'danger')
            return redirect(url_for('employees'))
        collar = normalise_collar(f.get('collar', ''))
        grade  = (f.get('grade') or '').strip().upper()
        desig  = _canonical_emp_field(f.get('designation', ''), plant_id, db, 'designation')
        dept   = _canonical_emp_field(f.get('department', ''),  plant_id, db, 'department')
        sect   = _canonical_emp_field(f.get('section', ''),     plant_id, db, 'section')
        cat    = (f.get('category') or '').strip().upper()
        db.execute('''UPDATE employees SET
            name=?, designation=?, grade=?, collar=?, department=?, section=?,
            category=?, gender=?, physically_handicapped=?, remarks=?
            WHERE id=? AND plant_id=?''',
            (f['name'].strip(), desig, grade, collar,
             dept, sect, cat, f.get('gender',''),
             f.get('physically_handicapped','No'), f.get('remarks',''),
             emp_id, plant_id))
        db.commit()
        after = db.execute('SELECT * FROM employees WHERE id=? AND plant_id=?',
                           (emp_id, plant_id)).fetchone()
        log_record_change('RECORD_EDIT', emp_id, 'employees',
                          before=dict(before), after=dict(after),
                          extra_detail=f"emp_edit:{before['emp_code']}")
        flash(f"Employee {after['name']} updated.", 'success')
        return redirect(url_for('employees'))

    @app.route('/employees/<int:emp_id>/exit', methods=['POST'])
    @spoc_required
    def exit_employee(emp_id):
        db = get_db()
        plant_id = session['plant_id']
        exit_date   = request.form.get('exit_date', str(_today_ist()))
        exit_reason = request.form.get('exit_reason', '')
        confirm_pending = request.form.get('confirm_pending', '0') == '1'
        if exit_date > str(_today_ist()):
            flash('Exit date cannot be a future date.', 'danger')
            return redirect(url_for('employees'))
        if not exit_reason.strip():
            flash('Exit reason is mandatory for attrition analysis.', 'danger')
            return redirect(url_for('employees'))

        # Load the employee row first — needed for emp_code lookup + audit before-snapshot
        before = db.execute(
            'SELECT * FROM employees WHERE id=? AND plant_id=?',
            (emp_id, plant_id)).fetchone()
        if not before:
            flash('Employee not found.', 'danger')
            return redirect(url_for('employees'))
        emp_code = before['emp_code']

        # Downstream pending-rows guard: block exit if attendance / effectiveness
        # rows still reference this employee, unless SPOC explicitly confirms.
        # emp_training rows are immutable history (always present for trained staff)
        # so we only block on effectiveness_review rows that are NOT yet completed.
        train_n = db.execute(
            'SELECT COUNT(*) FROM emp_training WHERE emp_code=? AND plant_id=?',
            (emp_code, plant_id)).fetchone()[0]
        eff_pending = db.execute(
            "SELECT COUNT(*) FROM effectiveness_review "
            "WHERE emp_code=? AND plant_id=? "
            "AND (completed_date IS NULL OR completed_date='')",
            (emp_code, plant_id)).fetchone()[0]

        if eff_pending > 0 and not confirm_pending:
            flash(
                f"Cannot exit {emp_code}: {eff_pending} pending effectiveness "
                f"review(s) open. Close them first, or resubmit with the "
                f"'Confirm exit despite pending reviews' checkbox ticked.",
                'danger')
            return redirect(url_for('employees'))

        db.execute('UPDATE employees SET is_active=0, exit_date=?, exit_reason=? WHERE id=? AND plant_id=?',
                   (exit_date, exit_reason, emp_id, plant_id))
        db.commit()
        after = db.execute(
            'SELECT * FROM employees WHERE id=? AND plant_id=?',
            (emp_id, plant_id)).fetchone()
        extra = (f"emp_exit:{emp_code} train_rows={train_n} "
                 f"eff_pending={eff_pending} confirmed={confirm_pending}")
        log_record_change('RECORD_EDIT', emp_id, 'employees',
                          before=dict(before) if before else None,
                          after=dict(after) if after else None,
                          extra_detail=extra)
        if eff_pending > 0:
            flash(
                f'Employee exited. WARNING: {eff_pending} pending effectiveness '
                f'review(s) remain open against {emp_code}.',
                'warning')
        else:
            flash('Employee marked as exited.', 'warning')
        return redirect(url_for('employees'))

    @app.route('/employees/<int:emp_id>/reactivate', methods=['POST'])
    @spoc_required
    def reactivate_employee(emp_id):
        db = get_db()
        db.execute('UPDATE employees SET is_active=1, exit_date=NULL, exit_reason=NULL WHERE id=? AND plant_id=?',
                   (emp_id, session['plant_id']))
        db.commit()
        log_action('RECORD_EDIT', f"emp_reactivate:{emp_id}")
        flash('Employee reactivated.', 'success')
        return redirect(url_for('employees') + '?show_exited=1')

    @app.route('/employees/bulk-template')
    @spoc_required
    def emp_bulk_template():
        if not _XLSX:
            flash('openpyxl not installed.', 'danger')
            return redirect(url_for('employees'))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employees'

        # All fields mandatory. Remarks kept optional (free-text note).
        headers = ['Emp Code *', 'Full Name *', 'Designation *', 'Grade *', 'Collar *',
                   'Department *', 'Section *', 'Category *', 'Gender *',
                   'Physically Handicapped *', 'Remarks']
        hdr_fill = PatternFill('solid', fgColor='1A3A5C')
        hdr_font = Font(color='FFFFFF', bold=True)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center')

        # Reference sheet with allowed values (used as dropdown source)
        ref = wb.create_sheet('Reference')
        ref_data = {
            'Grade': GRADES,
            'Collar': COLLARS,
            'Gender': GENDERS,
            'Physically Handicapped': PH_OPTIONS,
            'Category': CATEGORIES,
        }
        ref_col_letter = {}  # remember which column each list lives in
        col = 1
        for header, vals in ref_data.items():
            letter = get_column_letter(col)
            ref_col_letter[header] = letter
            ref.cell(row=1, column=col, value=header).font = Font(bold=True)
            for ri, v in enumerate(vals, 2):
                ref.cell(row=ri, column=col, value=v)
            col += 1

        # Excel dropdown validations — apply to first 5000 rows of each column
        # Column index per header (1-based): Grade=4, Collar=5, Category=8,
        # Gender=9, Physically Handicapped=10
        dv_specs = [
            ('Grade', 4, len(GRADES)),
            ('Collar', 5, len(COLLARS)),
            ('Category', 8, len(CATEGORIES)),
            ('Gender', 9, len(GENDERS)),
            ('Physically Handicapped', 10, len(PH_OPTIONS)),
        ]
        for label, col_idx, n_items in dv_specs:
            letter = ref_col_letter[label]
            # Reference!<letter>2:<letter><n_items+1>  (skip header row in Reference sheet)
            formula = f"=Reference!${letter}$2:${letter}${n_items + 1}"
            dv = DataValidation(type='list', formula1=formula, allow_blank=False,
                                showErrorMessage=True,
                                errorTitle='Invalid value',
                                error=f'Pick a value from the {label} dropdown.')
            target_letter = get_column_letter(col_idx)
            dv.add(f'{target_letter}2:{target_letter}5001')
            ws.add_data_validation(dv)

        for col_idx, width in zip(range(1, len(headers) + 1),
                                  [14, 30, 22, 26, 18, 22, 22, 26, 12, 26, 24]):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row + autofilter
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

        # Brief instructions row beneath header — italic, light grey
        instr = ('Fill ONE row per employee. All starred (*) fields are mandatory. '
                 'Use the dropdowns for Grade, Collar, Category, Gender, '
                 'Physically Handicapped — typed-in values that do not match the '
                 'allowed list will be rejected at upload.')
        c = ws.cell(row=5003, column=1, value=instr)
        c.font = Font(italic=True, color='6B7280', size=10)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='employee_bulk_upload_template.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/employees/bulk-upload', methods=['POST'])
    @spoc_required
    def emp_bulk_upload():
        if not _XLSX:
            flash('openpyxl not installed.', 'danger')
            return redirect(url_for('employees'))

        plant_id = session['plant_id']
        f = request.files.get('bulk_file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload a valid .xlsx file.', 'danger')
            return redirect(url_for('employees'))

        db = get_db()
        departments = _plant_depts(db, plant_id)
        sections = _plant_sections(db, plant_id)
        grades_upper = [g.upper() for g in GRADES]

        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('employees'))

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        inserted = 0
        errors = []

        # Pre-load existing emp codes for this plant to detect DB duplicates upfront
        existing_codes = {
            r[0] for r in db.execute(
                'SELECT emp_code FROM employees WHERE plant_id=?', (plant_id,)).fetchall()
        }
        seen_in_file = {}  # emp_code -> first row number (detect within-file duplicates)

        # Pre-load distinct values per fuzzy-cleansed field. Grows as new rows
        # are accepted in this batch so within-file drift also snaps.
        existing_desig = list({r[0] for r in db.execute(
            "SELECT DISTINCT TRIM(designation) FROM employees WHERE plant_id=? "
            "AND designation IS NOT NULL AND designation!=''", (plant_id,)).fetchall() if r[0]})
        existing_dept = list({r[0] for r in db.execute(
            "SELECT DISTINCT TRIM(department) FROM employees WHERE plant_id=? "
            "AND department IS NOT NULL AND department!=''", (plant_id,)).fetchall() if r[0]})
        existing_sect = list({r[0] for r in db.execute(
            "SELECT DISTINCT TRIM(section) FROM employees WHERE plant_id=? "
            "AND section IS NOT NULL AND section!=''", (plant_id,)).fetchall() if r[0]})

        for i, row in enumerate(rows, start=2):
            if not any(row):
                continue
            emp_code   = str(row[0]).strip() if row[0] else ''
            name       = str(row[1]).strip() if row[1] else ''
            desig      = str(row[2]).strip() if row[2] else ''
            grade_raw  = str(row[3]).strip() if row[3] else ''
            collar_raw = str(row[4]).strip() if row[4] else ''
            dept_raw   = str(row[5]).strip() if row[5] else ''
            sect_raw   = str(row[6]).strip() if row[6] else ''
            cat_raw    = str(row[7]).strip() if row[7] else ''
            gender_raw = str(row[8]).strip() if row[8] else ''
            ph_raw     = str(row[9]).strip() if row[9] else 'No'
            remarks    = str(row[10]).strip() if row[10] else ''

            row_errors = []

            # All fields mandatory except Remarks
            if not emp_code:    row_errors.append('Emp Code is required')
            if not name:        row_errors.append('Full Name is required')
            if not desig:       row_errors.append('Designation is required')
            if not grade_raw:   row_errors.append('Grade is required')
            if not collar_raw:  row_errors.append('Collar is required')
            if not dept_raw:    row_errors.append('Department is required')
            if not sect_raw:    row_errors.append('Section is required')
            if not cat_raw:     row_errors.append('Category is required')
            if not gender_raw:  row_errors.append('Gender is required')
            if not ph_raw:      row_errors.append('Physically Handicapped is required')

            grade  = grade_raw.upper()
            collar = normalise_collar(collar_raw)
            desig  = _canonical_emp_field(desig, plant_id, db, 'designation', _existing=existing_desig)
            dept   = _canonical_emp_field(dept_raw, plant_id, db, 'department', _existing=existing_dept)
            sect   = _canonical_emp_field(sect_raw, plant_id, db, 'section', _existing=existing_sect)
            cat    = cat_raw.upper()
            gender = gender_raw
            ph     = ph_raw if ph_raw in PH_OPTIONS else ''

            if collar_raw and collar not in COLLARS:
                row_errors.append(f"Invalid collar '{collar_raw}' (must be Blue Collared / White Collared)")

            if grade_raw and grade not in grades_upper:
                row_errors.append(f"Invalid grade '{grade_raw}'")

            if cat_raw and cat not in [c.upper() for c in CATEGORIES]:
                row_errors.append(f"Invalid category '{cat_raw}'")

            if gender_raw and gender not in GENDERS:
                row_errors.append(f"Invalid gender '{gender_raw}' (must be Male / Female / Others)")

            if ph_raw and ph_raw not in PH_OPTIONS:
                row_errors.append(f"Invalid PH value '{ph_raw}' (must be Yes / No)")

            # Duplicate checks — report before attempting insert
            if emp_code:
                if emp_code in existing_codes:
                    row_errors.append(f"Emp code {emp_code} already exists in the system — skipped")
                elif emp_code in seen_in_file:
                    row_errors.append(f"Emp code {emp_code} appears again (first at row {seen_in_file[emp_code]}) — skipped")
                else:
                    seen_in_file[emp_code] = i

            if row_errors:
                errors.append(f"Row {i} ({emp_code or '?'}): {'; '.join(row_errors)}")
                continue

            db.execute('''INSERT INTO employees
                (plant_id,emp_code,name,designation,grade,collar,department,section,
                 category,gender,physically_handicapped,remarks)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
                (plant_id, emp_code, name, desig, grade, collar,
                 dept, sect, cat, gender, ph or 'No', remarks))
            existing_codes.add(emp_code)
            if desig and desig not in existing_desig: existing_desig.append(desig)
            if dept  and dept  not in existing_dept:  existing_dept.append(dept)
            if sect  and sect  not in existing_sect:  existing_sect.append(sect)
            inserted += 1

        db.commit()

        if inserted:
            log_action('BULK_UPLOAD', f"employees:{inserted}")
            flash(f'{inserted} employee(s) uploaded successfully.', 'success')
        if errors:
            for e in errors:
                flash(e, 'upload_error')
        if not inserted and not errors:
            flash('No data rows found in the file.', 'warning')

        return redirect(url_for('employees'))

    # ---- Partial Bulk Update (any subset of columns) ----------------------
    # User downloads template with Emp Code + all updatable columns. They fill
    # only the columns they want to change (e.g. just Designation for promotions)
    # and leave the rest blank. Blank cell = keep existing value.

    UPDATABLE_COLS = {
        # excel header (lower) -> (db_col, validator_key or None)
        'emp code':                ('emp_code', None),       # key column (required)
        'full name':               ('name', None),
        'designation':             ('designation', None),
        'grade':                   ('grade', 'grade'),
        'collar':                  ('collar', 'collar'),
        'department':              ('department', None),
        'section':                 ('section', None),
        'category':                ('category', 'category'),
        'gender':                  ('gender', 'gender'),
        'physically handicapped':  ('physically_handicapped', 'ph'),
        'remarks':                 ('remarks', None),
    }

    @app.route('/employees/bulk-update-template')
    @spoc_required
    def emp_bulk_update_template():
        if not _XLSX:
            flash('openpyxl not installed.', 'danger')
            return redirect(url_for('employees'))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Updates'

        headers = ['Emp Code *', 'Full Name', 'Designation', 'Grade', 'Collar',
                   'Department', 'Section', 'Category', 'Gender',
                   'Physically Handicapped', 'Remarks']
        hdr_fill = PatternFill('solid', fgColor='1A3A5C')
        hdr_font = Font(color='FFFFFF', bold=True)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center')

        ref = wb.create_sheet('Reference')
        ref_data = {
            'Grade': GRADES,
            'Collar': COLLARS,
            'Gender': GENDERS,
            'Physically Handicapped': PH_OPTIONS,
            'Category': CATEGORIES,
        }
        ref_col_letter = {}
        col = 1
        for header, vals in ref_data.items():
            letter = get_column_letter(col)
            ref_col_letter[header] = letter
            ref.cell(row=1, column=col, value=header).font = Font(bold=True)
            for ri, v in enumerate(vals, 2):
                ref.cell(row=ri, column=col, value=v)
            col += 1

        dv_specs = [
            ('Grade', 4, len(GRADES)),
            ('Collar', 5, len(COLLARS)),
            ('Category', 8, len(CATEGORIES)),
            ('Gender', 9, len(GENDERS)),
            ('Physically Handicapped', 10, len(PH_OPTIONS)),
        ]
        for label, col_idx, n_items in dv_specs:
            letter = ref_col_letter[label]
            formula = f"=Reference!${letter}$2:${letter}${n_items + 1}"
            # allow_blank=True — blank cell means "don't update this column"
            dv = DataValidation(type='list', formula1=formula, allow_blank=True,
                                showErrorMessage=True,
                                errorTitle='Invalid value',
                                error=f'Pick a value from the {label} dropdown, or leave blank to keep existing.')
            target_letter = get_column_letter(col_idx)
            dv.add(f'{target_letter}2:{target_letter}5001')
            ws.add_data_validation(dv)

        for col_idx, width in zip(range(1, len(headers) + 1),
                                  [14, 30, 22, 26, 18, 22, 22, 26, 12, 26, 24]):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

        instr = ('Emp Code is REQUIRED. For every other column: fill the new value '
                 'to update it, or leave BLANK to keep the existing value. '
                 'Example — to update designation for 10 promoted employees, '
                 'fill their Emp Code + new Designation only; leave all other cells blank.')
        c = ws.cell(row=5003, column=1, value=instr)
        c.font = Font(italic=True, color='6B7280', size=10)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='employee_bulk_update_template.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/employees/bulk-update', methods=['POST'])
    @spoc_required
    def emp_bulk_update():
        if not _XLSX:
            flash('openpyxl not installed.', 'danger')
            return redirect(url_for('employees'))

        plant_id = session['plant_id']
        f = request.files.get('bulk_update_file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload a valid .xlsx file.', 'danger')
            return redirect(url_for('employees'))

        db = get_db()
        grades_upper = [g.upper() for g in GRADES]
        cats_upper = [c.upper() for c in CATEGORIES]

        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('employees'))

        # Read header row -> map to (db_col, validator_key)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            flash('File is empty.', 'danger')
            return redirect(url_for('employees'))

        col_idx_map = {}  # excel col idx (0-based) -> (db_col, validator)
        emp_code_idx = None
        unknown_headers = []
        for idx, h in enumerate(header_row):
            if h is None:
                continue
            key = str(h).strip().lower().rstrip('*').strip()
            if key in UPDATABLE_COLS:
                db_col, vk = UPDATABLE_COLS[key]
                col_idx_map[idx] = (db_col, vk)
                if db_col == 'emp_code':
                    emp_code_idx = idx
            else:
                unknown_headers.append(str(h))

        if emp_code_idx is None:
            flash('Header "Emp Code" not found — required to match rows.', 'danger')
            return redirect(url_for('employees'))

        if unknown_headers:
            flash(f'Ignored unknown column(s): {", ".join(unknown_headers)}', 'warning')

        # Pre-load emp_code -> id for this plant
        emp_lookup = {
            r['emp_code']: r['id']
            for r in db.execute(
                'SELECT id, emp_code FROM employees WHERE plant_id=?', (plant_id,)).fetchall()
        }

        updated = 0
        skipped = 0
        errors = []
        seen_codes = set()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            emp_code = str(row[emp_code_idx]).strip() if row[emp_code_idx] else ''
            if not emp_code:
                errors.append(f'Row {i}: Emp Code blank — skipped')
                continue
            if emp_code in seen_codes:
                errors.append(f'Row {i}: Emp Code {emp_code} duplicated in file — skipped')
                continue
            seen_codes.add(emp_code)

            if emp_code not in emp_lookup:
                skipped += 1
                errors.append(f'Row {i}: Emp Code {emp_code} not found in this plant — skipped')
                continue

            updates = {}  # db_col -> value
            row_errors = []

            for idx, (db_col, vk) in col_idx_map.items():
                if db_col == 'emp_code':
                    continue
                raw = row[idx] if idx < len(row) else None
                if raw is None or str(raw).strip() == '':
                    continue  # blank = skip
                val = str(raw).strip()

                if vk == 'grade':
                    v = val.upper()
                    if v not in grades_upper:
                        row_errors.append(f"invalid grade '{val}'")
                        continue
                    updates[db_col] = v
                elif vk == 'collar':
                    v = normalise_collar(val)
                    if v not in COLLARS:
                        row_errors.append(f"invalid collar '{val}'")
                        continue
                    updates[db_col] = v
                elif vk == 'category':
                    v = val.upper()
                    if v not in cats_upper:
                        row_errors.append(f"invalid category '{val}'")
                        continue
                    updates[db_col] = v
                elif vk == 'gender':
                    if val not in GENDERS:
                        row_errors.append(f"invalid gender '{val}'")
                        continue
                    updates[db_col] = val
                elif vk == 'ph':
                    if val not in PH_OPTIONS:
                        row_errors.append(f"invalid PH '{val}'")
                        continue
                    updates[db_col] = val
                else:
                    # Free-text — fuzzy-canonicalize designation/department/section
                    if db_col in ('department', 'section', 'designation'):
                        updates[db_col] = _canonical_emp_field(val, plant_id, db, db_col)
                    else:
                        updates[db_col] = val

            if row_errors:
                errors.append(f'Row {i} ({emp_code}): {"; ".join(row_errors)}')
                continue

            if not updates:
                skipped += 1
                continue  # nothing to change

            set_clause = ', '.join(f'{c}=?' for c in updates.keys())
            params = list(updates.values()) + [emp_lookup[emp_code], plant_id]
            db.execute(f'UPDATE employees SET {set_clause} WHERE id=? AND plant_id=?', params)
            updated += 1
            log_action('RECORD_EDIT', f"emp_bulk_update:{emp_code}:{','.join(updates.keys())}")

        db.commit()

        if updated:
            flash(f'{updated} employee(s) updated successfully.', 'success')
        if skipped:
            flash(f'{skipped} row(s) skipped (no changes or emp not found).', 'warning')
        if errors:
            for e in errors:
                flash(e, 'upload_error')
        if not updated and not errors and not skipped:
            flash('No data rows found in the file.', 'warning')

        return redirect(url_for('employees'))
