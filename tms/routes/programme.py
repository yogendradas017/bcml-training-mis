import io
import logging

from flask import render_template, request, redirect, url_for, session, flash, send_file, jsonify

from tms.constants import PROG_TYPES, MODES, AUDIENCES, MONTHS_FY, INT_EXT
from tms.db import get_db
from tms.decorators import spoc_required
from tms.helpers import (
    _is_ajax, _smart_title, _prog_in_use, _canonical_prog,
    _read_upload_file, _clean, _safe_float, _error_excel_response,
    _sync_master_from_tni, _current_fy, _in_current_fy, _parse_date_strict,
    _validate_time_vs_duration,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tms.audit import log_action, log_record_change


def _register(app):

    @app.route('/programme-master')
    @spoc_required
    def programme_master():
        plant_id = session['plant_id']
        db = get_db()
        db.execute('''CREATE TABLE IF NOT EXISTS programme_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plant_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            prog_type TEXT, mode TEXT,
            created_at TEXT DEFAULT (date('now')),
            UNIQUE(plant_id, name))''')
        progs = db.execute(
            'SELECT * FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()
        return render_template('programme_master.html', progs=progs,
                               prog_types=PROG_TYPES, modes=MODES)

    @app.route('/programme-master/add', methods=['POST'])
    @spoc_required
    def programme_master_add():
        plant_id  = session['plant_id']
        name      = _smart_title(request.form.get('name', '').strip())
        prog_type = request.form.get('prog_type', '').strip()
        raw_src   = request.form.get('source', '').strip()
        source    = raw_src if raw_src in ('TNI Requirement', 'New Requirement') else 'TNI Requirement'
        raw_cat   = request.form.get('category', '').strip()
        category  = raw_cat if raw_cat in ('Specialized', 'General') else 'General'
        if not name:
            flash('Programme name is required.', 'danger')
            return redirect(url_for('programme_master'))
        db = get_db()
        try:
            existing = db.execute('SELECT id FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
                                  (plant_id, name)).fetchone()
            if existing:
                db.execute('UPDATE programme_master SET prog_type=?, source=?, category=? WHERE id=?',
                           (prog_type or None, source, category, existing['id']))
                db.commit()
                log_action('RECORD_EDIT', f"pm:{name}:cat={category}")
                flash(f'"{name}" updated.', 'success')
            else:
                db.execute('INSERT INTO programme_master(plant_id,name,prog_type,source,category) VALUES(?,?,?,?,?)',
                           (plant_id, name, prog_type or None, source, category))
                db.commit()
                log_action('RECORD_ADD', f"pm:{name}:{source}:cat={category}")
                flash(f'"{name}" added to master as {source} · {category}.', 'success')
        except Exception as e:
            logging.error(f'programme_master_add error: {e}')
            flash(f'Error: {e}', 'danger')
        return redirect(url_for('programme_master'))

    @app.route('/programme-master/<int:prog_id>/delete', methods=['POST'])
    @spoc_required
    def programme_master_delete(prog_id):
        plant_id = session['plant_id']
        db = get_db()
        prog = db.execute('SELECT name FROM programme_master WHERE id=? AND plant_id=?',
                          (prog_id, plant_id)).fetchone()
        if not prog:
            flash('Programme not found.', 'danger')
            return redirect(url_for('programme_master'))
        if _prog_in_use(prog['name'], plant_id, db):
            flash(f'Cannot delete "{prog["name"]}" — it is referenced in TNI, Calendar, or Training Records.', 'danger')
            return redirect(url_for('programme_master'))
        db.execute('DELETE FROM programme_master WHERE id=? AND plant_id=?', (prog_id, plant_id))
        db.commit()
        log_action('RECORD_DELETE', f"pm:{prog['name']}")
        flash(f'"{prog["name"]}" removed from master list.', 'warning')
        return redirect(url_for('programme_master'))

    @app.route('/programme-master/<int:prog_id>/set-type', methods=['POST'])
    @spoc_required
    def programme_master_set_type(prog_id):
        plant_id = session['plant_id']
        data     = request.get_json(silent=True) or {}
        prog_type = data.get('prog_type', '').strip()
        from tms.constants import PROG_TYPES
        if prog_type and prog_type not in PROG_TYPES:
            return jsonify({'ok': False, 'error': 'Invalid type'}), 400
        db = get_db()
        row = db.execute('SELECT id FROM programme_master WHERE id=? AND plant_id=?',
                         (prog_id, plant_id)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': 'Not found'}), 404
        db.execute('UPDATE programme_master SET prog_type=? WHERE id=? AND plant_id=?',
                   (prog_type or None, prog_id, plant_id))
        db.commit()
        log_action('RECORD_EDIT', f"pm_type:{prog_id}:{prog_type}")
        return jsonify({'ok': True, 'prog_type': prog_type})

    @app.route('/programme-master/<int:prog_id>/set-source', methods=['POST'])
    @spoc_required
    def programme_master_set_source(prog_id):
        plant_id = session['plant_id']
        data     = request.get_json(silent=True) or {}
        source   = data.get('source', '').strip()
        if source not in ('TNI Requirement', 'New Requirement'):
            return jsonify({'ok': False, 'error': 'Invalid source'}), 400
        db = get_db()
        row = db.execute('SELECT id FROM programme_master WHERE id=? AND plant_id=?',
                         (prog_id, plant_id)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': 'Not found'}), 404
        db.execute('UPDATE programme_master SET source=? WHERE id=? AND plant_id=?',
                   (source, prog_id, plant_id))
        db.commit()
        log_action('RECORD_EDIT', f"pm_source:{prog_id}:{source}")
        return jsonify({'ok': True, 'source': source})

    @app.route('/programme-master/<int:prog_id>/set-category', methods=['POST'])
    @spoc_required
    def programme_master_set_category(prog_id):
        plant_id = session['plant_id']
        data     = request.get_json(silent=True) or {}
        category = data.get('category', '').strip()
        if category not in ('Specialized', 'General'):
            return jsonify({'ok': False, 'error': 'Invalid category'}), 400
        db = get_db()
        row = db.execute('SELECT id FROM programme_master WHERE id=? AND plant_id=?',
                         (prog_id, plant_id)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': 'Not found'}), 404
        db.execute('UPDATE programme_master SET category=? WHERE id=? AND plant_id=?',
                   (category, prog_id, plant_id))
        db.commit()
        log_action('RECORD_EDIT', f"pm_category:{prog_id}:{category}")
        return jsonify({'ok': True, 'category': category})

    @app.route('/programme-master/bulk-delete', methods=['POST'])
    @spoc_required
    def programme_master_bulk_delete():
        plant_id = session['plant_id']
        ids = request.form.getlist('ids[]')
        if not ids:
            flash('No programmes selected.', 'warning')
            return redirect(url_for('programme_master'))
        try:
            ids_int = [int(i) for i in ids]
        except ValueError:
            flash('Invalid selection.', 'danger')
            return redirect(url_for('programme_master'))
        db = get_db()
        blocked = []; deleted = 0
        for i in ids_int:
            prog = db.execute('SELECT name FROM programme_master WHERE id=? AND plant_id=?',
                              (i, plant_id)).fetchone()
            if not prog:
                continue
            if _prog_in_use(prog['name'], plant_id, db):
                blocked.append(prog['name'])
            else:
                db.execute('DELETE FROM programme_master WHERE id=? AND plant_id=?', (i, plant_id))
                deleted += 1
        db.commit()
        if deleted:
            log_action('BULK_DELETE', f"pm:{deleted}")
            flash(f'{deleted} programme(s) deleted.', 'warning')
        if blocked:
            flash(f'{len(blocked)} programme(s) could not be deleted (in use in TNI/Calendar/Training): {", ".join(blocked[:5])}', 'danger')
        return redirect(url_for('programme_master'))

    @app.route('/programme-master/sync-from-tni', methods=['POST'])
    @spoc_required
    def programme_master_sync_from_tni():
        plant_id = session['plant_id']
        db = get_db()
        tni_progs = [r[0] for r in db.execute(
            'SELECT DISTINCT programme_name FROM tni WHERE plant_id=? AND programme_name IS NOT NULL AND programme_name != "" ORDER BY programme_name',
            (plant_id,)).fetchall()]
        if not tni_progs:
            flash('No TNI data found — master list unchanged.', 'warning')
            return redirect(url_for('programme_master'))
        existing = {r['name']: r['prog_type'] for r in db.execute(
            'SELECT name, prog_type FROM programme_master WHERE plant_id=?', (plant_id,)).fetchall()}
        tni_types = {r['programme_name']: r['top_type'] for r in db.execute('''
            SELECT programme_name,
                   (SELECT prog_type FROM tni t2
                    WHERE t2.plant_id=t.plant_id AND t2.programme_name=t.programme_name
                      AND t2.prog_type IS NOT NULL AND t2.prog_type != ""
                    GROUP BY prog_type ORDER BY COUNT(*) DESC LIMIT 1) AS top_type
            FROM tni t WHERE plant_id=? AND programme_name IS NOT NULL AND programme_name != ""
            GROUP BY programme_name
        ''', (plant_id,)).fetchall()}
        # Only wipe TNI-sourced rows — New Requirement entries added manually survive
        db.execute("DELETE FROM programme_master WHERE plant_id=? AND source='TNI Requirement'", (plant_id,))
        for name in tni_progs:
            prog_type = tni_types.get(name) or existing.get(name)
            db.execute('INSERT OR IGNORE INTO programme_master(plant_id, name, prog_type, source) VALUES(?,?,?,?)',
                       (plant_id, name, prog_type, 'TNI Requirement'))
        db.commit()
        log_action('BULK_UPDATE', f"pm_sync_tni:{len(tni_progs)}")
        flash(f'Programme Master synced from TNI — {len(tni_progs)} TNI programme(s) updated. New Requirement entries preserved.', 'success')
        return redirect(url_for('programme_master'))

    @app.route('/programme-master/bulk', methods=['POST'])
    @spoc_required
    def programme_master_bulk():
        plant_id = session['plant_id']
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('programme_master'))
        try:
            import pandas as _pd
            fname = f.filename.lower()
            if fname.endswith('.csv'):
                df = _pd.read_csv(f, dtype=str).fillna('')
            else:
                df = _pd.read_excel(f, dtype=str).fillna('')
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('programme_master'))

        cols_lower = {c.strip().lower(): c for c in df.columns}
        name_col = next((cols_lower[k] for k in ['programme name','program name','name','training name','course name'] if k in cols_lower), None)
        type_col = next((cols_lower[k] for k in ['type of programme','type','prog type','programme type'] if k in cols_lower), None)

        if not name_col:
            flash(f'Could not find a "Programme Name" column. Columns found: {", ".join(df.columns.tolist()[:10])}', 'danger')
            return redirect(url_for('programme_master'))

        src_col = next((cols_lower[k] for k in ['source','requirement','req type'] if k in cols_lower), None)
        db = get_db()
        inserted = skipped = 0
        for _, row in df.iterrows():
            name = str(row.get(name_col, '')).strip()
            if not name or name.lower() in ('nan', 'none', ''):
                continue
            prog_type = str(row.get(type_col, '')).strip() if type_col else ''
            raw_src   = str(row.get(src_col, '')).strip() if src_col else ''
            source    = raw_src if raw_src in ('TNI Requirement', 'New Requirement') else 'TNI Requirement'
            try:
                db.execute('INSERT INTO programme_master(plant_id,name,prog_type,source) VALUES(?,?,?,?)',
                           (plant_id, name, prog_type or None, source))
                inserted += 1
            except Exception:
                skipped += 1
        db.commit()
        log_action('BULK_UPLOAD', f"pm:{inserted}")
        flash(f'{inserted} programmes added. {skipped} already existed (skipped).', 'success' if inserted else 'warning')
        return redirect(url_for('programme_master'))

    @app.route('/programme-master/template')
    @spoc_required
    def programme_master_template():
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Programme Master'
        hdr_fill = PatternFill('solid', fgColor='1A1F35')
        headers = ['Programme Name', 'Type of Programme', 'Source']
        widths  = [45, 30, 20]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        for row in [('Fire Safety', 'EHS/HR', 'TNI Requirement'),
                    ('5-S Management', 'EHS/HR', 'TNI Requirement'),
                    ('English Communication', 'Behavioural/Leadership', 'New Requirement')]:
            ws.append(row)
        dv_type = DataValidation(type='list', formula1=f'"{",".join(PROG_TYPES)}"', allow_blank=True)
        dv_type.sqref = 'B2:B500'
        dv_src = DataValidation(type='list', formula1='"TNI Requirement,New Requirement"', allow_blank=True)
        dv_src.sqref = 'C2:C500'
        ws.add_data_validation(dv_type)
        ws.add_data_validation(dv_src)
        ws.freeze_panes = 'A2'
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='Programme_Master_Template.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/programme-master/export')
    @spoc_required
    def programme_master_export():
        plant_id   = session['plant_id']
        plant_name = session.get('plant_name', 'Plant')
        db = get_db()
        progs = db.execute('SELECT name, prog_type, source, created_at FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Programme Master'
        hdr_fill = PatternFill('solid', fgColor='1A1F35')
        headers = ['#', 'Programme Name', 'Type of Programme', 'Source', 'Added On']
        widths  = [5, 45, 30, 20, 14]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'
        for i, r in enumerate(progs, 1):
            ws.append([i, r['name'], r['prog_type'] or '', r['source'] or 'TNI Requirement', r['created_at'] or ''])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'Programme_Master_{plant_name}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Programme Details (2C) ────────────────────────────────────────────────

    @app.route('/programme')
    @spoc_required
    def programme_details():
        plant_id = session['plant_id']
        db = get_db()
        records = db.execute('''
            SELECT p.*,
                   c.source as cal_source,
                   (SELECT COUNT(*) FROM emp_training t WHERE t.session_code=p.session_code AND t.plant_id=p.plant_id) as participants,
                   (SELECT COALESCE(SUM(t.hrs),0) FROM emp_training t WHERE t.session_code=p.session_code AND t.plant_id=p.plant_id) as man_hours
            FROM programme_details p
            LEFT JOIN calendar c ON c.session_code=p.session_code AND c.plant_id=p.plant_id
            WHERE p.plant_id=?
            ORDER BY p.id DESC
        ''', (plant_id,)).fetchall()

        # Central 2C rows where this plant's employees attended — read-only
        central_records = db.execute('''
            SELECT p.*,
                   c.source as cal_source,
                   COUNT(t.id) as participants,
                   COALESCE(SUM(t.hrs),0) as man_hours
            FROM programme_details p
            JOIN emp_training t ON t.session_code=p.session_code AND t.plant_id=?
            LEFT JOIN calendar c ON c.session_code=p.session_code AND c.plant_id=99
            WHERE p.plant_id=99
            GROUP BY p.id
            ORDER BY p.start_date DESC
        ''', (plant_id,)).fetchall()

        cal_sessions = db.execute(
            "SELECT session_code, programme_name, prog_type, duration_hrs, plan_start, plan_end,"
            " time_from, time_to"
            " FROM calendar WHERE plant_id=? ORDER BY session_code",
            (plant_id,)).fetchall()
        return render_template('programme_2c.html', records=records,
                               central_records=central_records,
                               cal_sessions=cal_sessions,
                               int_ext=INT_EXT, audiences=AUDIENCES, months=MONTHS_FY)

    @app.route('/programme/add', methods=['POST'])
    @spoc_required
    def add_programme_details():
        plant_id = session['plant_id']
        f = request.form
        db = get_db()
        session_code = f.get('session_code', '').strip()

        if db.execute('SELECT 1 FROM programme_details WHERE session_code=? AND plant_id=?',
                      (session_code, plant_id)).fetchone():
            flash(f'Session {session_code} already recorded. Edit the existing entry.', 'warning')
            return redirect(url_for('programme_details'))

        cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?',
                         (session_code, plant_id)).fetchone()
        if not cal:
            is_central = db.execute('SELECT 1 FROM calendar WHERE session_code=? AND plant_id=99',
                                    (session_code,)).fetchone()
            if is_central:
                flash(f'Session {session_code} is a centrally managed programme. Programme details (2C) are entered by the Central L&D team. Record your employees\' attendance in Training Records (2A) using this session code.', 'warning')
            else:
                flash(f'Session code "{session_code}" not found in calendar. Only planned sessions can be recorded in 2C.', 'danger')
            return redirect(url_for('programme_details'))

        # GAP 1: block 2C entry if session is Cancelled/Re-Scheduled
        if cal['status'] in ('Cancelled', 'Re-Scheduled'):
            flash(f'Session {session_code} is {cal["status"]}. Cannot record 2C details.', 'danger')
            return redirect(url_for('programme_details'))

        # GAP 2 (time gate): 2C cannot be entered before plan_end
        from datetime import date as _d
        today_iso = _d.today().isoformat()
        if cal['plan_end'] and today_iso < cal['plan_end']:
            flash(f'Session not yet ended. 2C can be entered from {cal["plan_end"]} onwards.', 'danger')
            return redirect(url_for('programme_details'))

        # GAP 2 (Phase 4): minimum 1 attendance must exist before 2C save
        attended = db.execute(
            'SELECT COUNT(*) FROM emp_training WHERE plant_id=? AND session_code=?',
            (plant_id, session_code)).fetchone()[0]
        if attended < 1:
            flash(f'No attendance records found for session {session_code}. Add at least 1 employee in 2A before saving 2C.', 'danger')
            return redirect(url_for('programme_details'))

        prog_name = cal['programme_name']
        prog_type = cal['prog_type']
        level     = cal['level']
        cal_new   = 'Calendar Program'
        mode      = cal['mode']
        audience  = cal['target_audience'] or ''

        try:
            hours = float(f.get('hours_actual') or 0)
        except (ValueError, TypeError):
            hours = 0
        if hours <= 0:
            flash('Actual hours must be greater than 0.', 'danger')
            return redirect(url_for('programme_details'))

        # Validate hard rules FIRST (FY window, time window). Anomaly flagging
        # comes after, so we don't compute+flash anomalies for a row that's
        # about to be rejected.
        fy_start, fy_end = _current_fy()
        start_date = f.get('start_date', '')
        if start_date and not _in_current_fy(start_date):
            flash(f'Start date must be within the current financial year ({fy_start} to {fy_end}).', 'danger')
            return redirect(url_for('programme_details'))

        time_from = (f.get('time_from', '') or '').strip() or (cal['time_from'] or '')
        time_to   = (f.get('time_to', '')   or '').strip() or (cal['time_to']   or '')
        ok, tmsg = _validate_time_vs_duration(
            time_from, time_to, hours,
            f.get('start_date', ''), f.get('end_date', ''))
        if not ok:
            flash(tmsg, 'danger')
            return redirect(url_for('programme_details'))

        # Hours mismatch — 25% threshold standard. For centrally-hosted
        # sessions, attendees live under their own plant_id; widen the AVG
        # query to include host_plant_id matches.
        avg_2a = db.execute(
            'SELECT AVG(hrs) FROM emp_training '
            'WHERE session_code=? AND hrs > 0 '
            'AND (plant_id=? OR host_plant_id=?)',
            (session_code, plant_id, plant_id)).fetchone()[0]
        pd_anomalies = []
        if avg_2a and avg_2a > 0 and abs(hours - avg_2a) / avg_2a > 0.25:
            pd_anomalies.append(f'hours_mismatch(2C={hours} vs avg2A={avg_2a:.1f})')
            flash(f'Note: 2C hours ({hours}) differ from average 2A hours ({avg_2a:.1f}) by >25%. Saved — please verify.', 'warning')

        # Phase 5: compute extended anomalies for verification path
        from datetime import datetime as _dt
        sum_hrs = db.execute(
            'SELECT COALESCE(SUM(hrs),0) FROM emp_training WHERE plant_id=? AND session_code=?',
            (plant_id, session_code)).fetchone()[0]
        anomalies = list(pd_anomalies)  # carry over hours_mismatch if any
        planned_pax = cal['planned_pax'] or 0
        if planned_pax > 0 and attended < planned_pax * 0.5:
            anomalies.append(f'low_attendance({attended}/{planned_pax})')
        if audience in ('Blue Collared', 'White Collared'):
            mc = db.execute(
                'SELECT COUNT(*) FROM emp_training t '
                'JOIN employees e ON e.plant_id=t.plant_id AND e.emp_code=t.emp_code '
                'WHERE t.plant_id=? AND t.session_code=? AND e.collar IS NOT NULL AND e.collar != ?',
                (plant_id, session_code, audience)).fetchone()[0]
            if mc > 0:
                anomalies.append(f'collar_mismatch({mc}_attendees)')

        anom_pd = ','.join(anomalies) if anomalies else None

        # Audit Tier 2 fix: a stub programme_details may already exist (created
        # by feedback aggregator when QR responses arrived before 2C save).
        # In that case UPDATE preserving feedback fields if SPOC left them blank.
        existing_pd = db.execute(
            'SELECT id, course_feedback, faculty_feedback, '
            '       trainer_fb_participants, trainer_fb_facilities '
            'FROM programme_details WHERE plant_id=? AND session_code=? LIMIT 1',
            (plant_id, session_code)
        ).fetchone()

        form_course   = _safe_float(f.get('course_feedback'))
        form_faculty  = _safe_float(f.get('faculty_feedback'))
        form_partic   = _safe_float(f.get('trainer_fb_participants'))
        form_facil    = _safe_float(f.get('trainer_fb_facilities'))

        if existing_pd:
            # Prefer non-blank form value; otherwise keep stub's pre-computed value.
            cf = form_course  if form_course  not in (None, 0) else existing_pd['course_feedback']
            ff = form_faculty if form_faculty not in (None, 0) else existing_pd['faculty_feedback']
            tp = form_partic  if form_partic  not in (None, 0) else existing_pd['trainer_fb_participants']
            tf = form_facil   if form_facil   not in (None, 0) else existing_pd['trainer_fb_facilities']
            db.execute('''UPDATE programme_details SET
                programme_name=?, prog_type=?, level=?, cal_new=?, mode=?,
                start_date=?, end_date=?, time_from=?, time_to=?, audience=?,
                hours_actual=?, faculty_name=?, int_ext=?, cost=?,
                venue=?, course_feedback=?, faculty_feedback=?,
                trainer_fb_participants=?, trainer_fb_facilities=?,
                anomaly_flags=?
                WHERE id=?''',
                (prog_name, prog_type, level, cal_new, mode,
                 f.get('start_date',''), f.get('end_date',''),
                 time_from or None, time_to or None, audience,
                 float(f.get('hours_actual') or 0), f.get('faculty_name',''),
                 f.get('int_ext',''), float(f.get('cost') or 0),
                 f.get('venue',''), cf, ff, tp, tf,
                 anom_pd, existing_pd['id']))
        else:
            db.execute('''INSERT INTO programme_details
                (plant_id,session_code,programme_name,prog_type,level,cal_new,mode,
                 start_date,end_date,time_from,time_to,audience,hours_actual,faculty_name,int_ext,cost,
                 venue,course_feedback,faculty_feedback,trainer_fb_participants,trainer_fb_facilities,
                 anomaly_flags)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (plant_id, session_code, prog_name, prog_type, level, cal_new, mode,
                 f.get('start_date',''), f.get('end_date',''),
                 time_from or None, time_to or None, audience,
                 float(f.get('hours_actual') or 0), f.get('faculty_name',''),
                 f.get('int_ext',''), float(f.get('cost') or 0),
                 f.get('venue',''),
                 form_course, form_faculty, form_partic, form_facil,
                 anom_pd))
        db.commit()

        now_iso  = _dt.now().isoformat(timespec='seconds')
        user_id  = session.get('user_id')
        username = session.get('username', '')
        user_role = session.get('role')

        # Central/Admin save → instant Conducted. SPOC → Awaiting Verification.
        if user_role in ('central', 'admin'):
            new_status = 'Conducted'
            v_at, v_by, stage = now_iso, user_id, 'verified_on_2c'
        else:
            new_status = 'Awaiting Verification'
            v_at, v_by, stage = None, None, '2c_added'

        db.execute(
            'UPDATE calendar SET status=?, conducted_at=?, conducted_by=?, '
            'verified_at=?, verified_by=?, actual_pax=?, actual_hrs=? '
            'WHERE session_code=? AND plant_id=?',
            (new_status, now_iso, user_id, v_at, v_by, attended, sum_hrs,
             session_code, plant_id))
        db.execute(
            'INSERT INTO verification_log (session_code, plant_id, stage, actor, actor_id, detail) '
            'VALUES (?,?,?,?,?,?)',
            (session_code, plant_id, stage, username, user_id,
             '; '.join(anomalies) if anomalies else 'clean'))
        db.commit()

        # Audit Tier 3: snapshot the 2C row in payload (INSERT or UPDATE-of-stub)
        new_pd = db.execute(
            'SELECT * FROM programme_details WHERE plant_id=? AND session_code=? LIMIT 1',
            (plant_id, session_code)).fetchone()
        log_record_change(
            'RECORD_EDIT' if existing_pd else 'RECORD_ADD',
            session_code, 'programme_details',
            before=dict(existing_pd) if existing_pd else None,
            after=dict(new_pd) if new_pd else None,
            extra_detail=f'status:{new_status}')
        if new_status == 'Conducted':
            flash(f'Programme {session_code} saved and verified.', 'success')
        else:
            n_anom = len(anomalies)
            anom_str = f' ({n_anom} anomal{"y" if n_anom == 1 else "ies"} flagged)' if anomalies else ''
            flash(f'Programme {session_code} saved → Awaiting Verification by Central L&D{anom_str}.', 'warning')
        return redirect(url_for('programme_details'))

    @app.route('/programme/<int:rec_id>/delete', methods=['POST'])
    @spoc_required
    def delete_programme(rec_id):
        db = get_db()
        rec = db.execute('SELECT * FROM programme_details WHERE id=? AND plant_id=?',
                         (rec_id, session['plant_id'])).fetchone()
        if rec:
            before_snap_dict = dict(rec)
            db.execute('DELETE FROM programme_details WHERE id=? AND plant_id=?', (rec_id, session['plant_id']))
            db.execute(
                "UPDATE calendar SET status='To Be Planned', conducted_at=NULL, conducted_by=NULL, "
                "verified_at=NULL, verified_by=NULL, actual_pax=0, actual_hrs=0 "
                "WHERE session_code=? AND plant_id=?",
                (rec['session_code'], session['plant_id']))
            db.execute(
                'INSERT INTO verification_log (session_code, plant_id, stage, actor, actor_id, detail) '
                'VALUES (?,?,?,?,?,?)',
                (rec['session_code'], session['plant_id'], '2c_deleted',
                 session.get('username', ''), session.get('user_id'),
                 'Reverted to To Be Planned'))
            db.commit()
            log_record_change('RECORD_DELETE', rec_id, 'programme_details',
                              before=before_snap_dict, after=None)
        if _is_ajax():
            return '', 204
        flash('Programme record deleted.', 'warning')
        return redirect(url_for('programme_details'))

    @app.route('/programme/bulk-delete', methods=['POST'])
    @spoc_required
    def programme_bulk_delete():
        plant_id = session['plant_id']
        ids = request.form.getlist('ids[]')
        if ids:
            ph = ','.join('?' * len(ids))
            db = get_db()
            recs = db.execute(f'SELECT session_code FROM programme_details WHERE id IN ({ph}) AND plant_id=?',
                              ids + [plant_id]).fetchall()
            for r in recs:
                db.execute(
                    "UPDATE calendar SET status='To Be Planned', conducted_at=NULL, conducted_by=NULL, "
                    "verified_at=NULL, verified_by=NULL, actual_pax=0, actual_hrs=0 "
                    "WHERE session_code=? AND plant_id=?",
                    (r['session_code'], plant_id))
            db.execute(f'DELETE FROM programme_details WHERE id IN ({ph}) AND plant_id=?', ids + [plant_id])
            db.commit()
            log_action('BULK_DELETE', f"2c:{len(ids)}")
            flash(f'{len(ids)} programme records deleted.', 'warning')
        return redirect(url_for('programme_details'))

    @app.route('/programme/template')
    @spoc_required
    def programme_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '2C_Bulk_Upload'
        headers = ['Session Code', 'Actual Start Date (DD-MM-YYYY)', 'Actual End Date (DD-MM-YYYY)',
                   'Actual Hours', 'Faculty Name', 'Internal/External', 'Cost (Rs.)', 'Venue',
                   'Course Feedback (1-5)', 'Faculty Feedback (1-5)',
                   'Trainer FB Participants (1-5)', 'Trainer FB Facilities (1-5)']
        hdr_fill = PatternFill('solid', fgColor='6B3FA0')
        hdr_font = Font(bold=True, color='FFFFFF')
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font
            ws.column_dimensions[get_column_letter(i)].width = 26
        ws.append(['BCM/EHS/001/B01', '10-06-2026', '10-06-2026', 4, 'Mr. Ramesh Kumar', 'Internal', 0, 'Training Hall', 4.2, 4.0, 3.8, 4.1])
        ws['A4'] = 'NOTE: Dates MUST be DD-MM-YYYY (e.g. 15-06-2026). Session Code must exist in Training Calendar. Internal/External: Internal | External | Online'
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, download_name='2C_Programme_Bulk_Upload_Template.xlsx', as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/programme/bulk', methods=['POST'])
    @spoc_required
    def programme_bulk_upload():
        plant_id = session['plant_id']
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('programme_details'))
        try:
            df = _read_upload_file(f)
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('programme_details'))
        db = get_db(); inserted = 0; errors = []
        for i, row in df.iterrows():
            sc        = _clean(row, ['session code', 'session_code'])
            raw_start = _clean(row, ['actual start date (dd-mm-yyyy)', 'actual start date (yyyy-mm-dd)', 'start date', 'actual start date'])
            raw_end   = _clean(row, ['actual end date (dd-mm-yyyy)', 'actual end date (yyyy-mm-dd)', 'end date', 'actual end date'])
            try:
                start_date = _parse_date_strict(raw_start)
                end_date   = _parse_date_strict(raw_end)
            except ValueError as e:
                errors.append(f'Row {i+2}: Date format error — {e}. Use DD-MM-YYYY (e.g. 15-06-2026).')
                continue
            hrs        = _safe_float(_clean(row, ['actual hours', 'hours', 'hrs'])) or 0
            faculty    = _clean(row, ['faculty name', 'faculty'])
            int_ext    = _clean(row, ['internal/external', 'int/ext', 'internal external'])
            cost       = _safe_float(_clean(row, ['cost (rs.)', 'cost'])) or 0
            venue      = _clean(row, ['venue'])
            cfb        = _safe_float(_clean(row, ['course feedback (1-5)', 'course feedback', 'course fb']))
            ffb        = _safe_float(_clean(row, ['faculty feedback (1-5)', 'faculty feedback', 'faculty fb']))
            tfbp       = _safe_float(_clean(row, ['trainer fb participants (1-5)', 'trainer fb participants']))
            tfbf       = _safe_float(_clean(row, ['trainer fb facilities (1-5)', 'trainer fb facilities']))
            if not sc:
                errors.append(f'Row {i+2}: Session Code is required.')
                continue
            cal = db.execute('SELECT * FROM calendar WHERE session_code=? AND plant_id=?', (sc, plant_id)).fetchone()
            if not cal:
                is_central = db.execute('SELECT 1 FROM calendar WHERE session_code=? AND plant_id=99', (sc,)).fetchone()
                if is_central:
                    errors.append(f'Row {i+2}: Session {sc} is centrally managed — record attendance in 2A, not 2C.')
                else:
                    errors.append(f'Row {i+2}: Session Code {sc} not found in Calendar.')
                continue

            # GAP 1: block Cancelled/Re-Scheduled
            if cal['status'] in ('Cancelled', 'Re-Scheduled'):
                errors.append(f'Row {i+2}: Session {sc} is {cal["status"]} — cannot record 2C.')
                continue

            if db.execute('SELECT 1 FROM programme_details WHERE session_code=? AND plant_id=?', (sc, plant_id)).fetchone():
                errors.append(f'Row {i+2}: Session {sc} already has programme details recorded.')
                continue
            if not start_date:
                errors.append(f'Row {i+2}: Actual Start Date is required.')
                continue

            # GAP 9: FY date check (was missing in bulk path)
            if not _in_current_fy(start_date):
                fy_s, fy_e = _current_fy()
                errors.append(f'Row {i+2}: Start date {start_date} is outside current FY ({fy_s} to {fy_e}).')
                continue

            # GAP 2 (time gate): 2C cannot be entered before plan_end
            from datetime import date as _d
            today_iso = _d.today().isoformat()
            if cal['plan_end'] and today_iso < cal['plan_end']:
                errors.append(f'Row {i+2}: Session {sc} not yet ended (plan_end {cal["plan_end"]}). 2C cannot be saved before session ends.')
                continue

            # GAP 2 (Phase 4): minimum 1 attendance must exist
            attended = db.execute(
                'SELECT COUNT(*) FROM emp_training WHERE plant_id=? AND session_code=?',
                (plant_id, sc)).fetchone()[0]
            if attended < 1:
                errors.append(f'Row {i+2}: Session {sc} has no attendance — add 2A records first.')
                continue

            # Phase 5: anomaly check + status routing (25% standard)
            from datetime import datetime as _dt2
            sum_hrs = db.execute(
                'SELECT COALESCE(SUM(hrs),0) FROM emp_training WHERE plant_id=? AND session_code=?',
                (plant_id, sc)).fetchone()[0]
            avg_hrs_2a = db.execute(
                'SELECT AVG(hrs) FROM emp_training WHERE plant_id=? AND session_code=? AND hrs > 0',
                (plant_id, sc)).fetchone()[0] or 0
            row_anomalies = []
            planned_pax = cal['planned_pax'] or 0
            if planned_pax > 0 and attended < planned_pax * 0.5:
                row_anomalies.append(f'low_attendance({attended}/{planned_pax})')
            if avg_hrs_2a > 0 and abs(hrs - avg_hrs_2a) / avg_hrs_2a > 0.25:
                row_anomalies.append(f'hours_mismatch(2C={hrs} vs avg2A={avg_hrs_2a:.1f})')
            tgt_aud = cal['target_audience'] or ''
            if tgt_aud in ('Blue Collared', 'White Collared'):
                mc = db.execute(
                    'SELECT COUNT(*) FROM emp_training t '
                    'JOIN employees e ON e.plant_id=t.plant_id AND e.emp_code=t.emp_code '
                    'WHERE t.plant_id=? AND t.session_code=? AND e.collar IS NOT NULL AND e.collar != ?',
                    (plant_id, sc, tgt_aud)).fetchone()[0]
                if mc > 0:
                    row_anomalies.append(f'collar_mismatch({mc}_attendees)')

            # Cross-check time window vs hrs × days
            ok_t, t_msg = _validate_time_vs_duration(
                cal['time_from'] or '', cal['time_to'] or '', hrs,
                start_date, end_date)
            if not ok_t:
                errors.append(f'Row {i+2}: {t_msg}')
                continue

            anom_pd_bulk = ','.join(row_anomalies) if row_anomalies else None
            db.execute('''INSERT INTO programme_details
                (plant_id,session_code,programme_name,prog_type,level,cal_new,mode,
                 start_date,end_date,time_from,time_to,audience,hours_actual,faculty_name,int_ext,cost,
                 venue,course_feedback,faculty_feedback,trainer_fb_participants,trainer_fb_facilities,
                 anomaly_flags)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (plant_id, sc, cal['programme_name'], cal['prog_type'], cal['level'],
                 'Calendar Program', cal['mode'], start_date, end_date,
                 cal['time_from'] or None, cal['time_to'] or None,
                 cal['target_audience'], hrs, faculty, int_ext, cost, venue, cfb, ffb, tfbp, tfbf,
                 anom_pd_bulk))

            now_iso  = _dt2.now().isoformat(timespec='seconds')
            user_id  = session.get('user_id')
            username = session.get('username', '')
            user_role = session.get('role')

            if user_role in ('central', 'admin'):
                new_status = 'Conducted'
                v_at, v_by, stage = now_iso, user_id, 'verified_on_2c'
            else:
                new_status = 'Awaiting Verification'
                v_at, v_by, stage = None, None, '2c_added'

            db.execute(
                'UPDATE calendar SET status=?, conducted_at=?, conducted_by=?, '
                'verified_at=?, verified_by=?, actual_pax=?, actual_hrs=? '
                'WHERE session_code=? AND plant_id=?',
                (new_status, now_iso, user_id, v_at, v_by, attended, sum_hrs,
                 sc, plant_id))
            db.execute(
                'INSERT INTO verification_log (session_code, plant_id, stage, actor, actor_id, detail) '
                'VALUES (?,?,?,?,?,?)',
                (sc, plant_id, stage, username, user_id,
                 '; '.join(row_anomalies) if row_anomalies else 'clean'))
            inserted += 1
        db.commit()
        if errors:
            if inserted:
                flash(f'Bulk upload complete: {inserted} programme records saved. {len(errors)} rows had errors — downloading error report.', 'warning')
            return _error_excel_response(errors, inserted, 'Programme2C_Upload_Errors.xlsx')
        log_action('BULK_UPLOAD', f"2c:{inserted}")
        flash(f'Bulk upload complete: {inserted} programme records saved.', 'success')
        return redirect(url_for('programme_details'))
