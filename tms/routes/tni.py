import os
import io
import secrets
import json as _json
import uuid as _uuid

from flask import render_template, request, redirect, url_for, session, flash, send_file, jsonify

from tms.constants import (
    PROG_TYPES, MODES, MONTHS_FY, BASE_DIR, TEMP_UPLOAD_DIR, NON_TNI_SOURCES
)
from tms.db import get_db
from tms.decorators import spoc_required
from tms.helpers import (
    _is_ajax, _get_programme_names, _canonical_prog, _tni_canon_candidates, _sync_master_from_tni,
    _read_upload_file, _read_upload_file_path, _clean, _safe_float,
    _error_excel_response, _process_fresh_tni, _parse_msforms_excel,
    _smart_analyze_rows, _fy_label, _tni_is_locked,
    _error_excel_for_tni, _cleanse_programme_names,
    resync_calendar_audience,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tms.audit import log_action


def _register(app):

    @app.route('/tni')
    @spoc_required
    def tni():
        plant_id = session['plant_id']
        db = get_db()
        fy    = _fy_label()
        total = db.execute(
            'SELECT COUNT(DISTINCT emp_code || "|" || programme_name) FROM tni WHERE plant_id=? AND fy_year=?',
            (plant_id, fy)).fetchone()[0]
        emps = db.execute('SELECT emp_code, name, grade, collar, department FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name', (plant_id,)).fetchall()
        programmes = _get_programme_names(plant_id, db)
        depts = [r[0] for r in db.execute(
            'SELECT DISTINCT department FROM employees WHERE plant_id=? AND department IS NOT NULL AND department != "" ORDER BY department',
            (plant_id,)).fetchall()]

        master_lower = set(r[0].lower() for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=?', (plant_id,)).fetchall())
        dirty_names = []
        if master_lower:
            tni_names = [r[0] for r in db.execute(
                'SELECT DISTINCT programme_name FROM tni WHERE plant_id=? AND fy_year=?', (plant_id, fy)).fetchall()]
            dirty_names = [n for n in tni_names if n.lower() not in master_lower]

        dup_count = db.execute('''
            SELECT COALESCE(SUM(cnt - 1), 0)
            FROM (SELECT COUNT(*) as cnt FROM tni WHERE plant_id=? AND fy_year=?
                  GROUP BY emp_code, programme_name HAVING cnt > 1)
        ''', (plant_id, fy)).fetchone()[0]

        # Programme Master rows for the quick-add-for-employee bulk form
        master_progs = db.execute(
            'SELECT id, name, prog_type, COALESCE(category, "General") as category '
            'FROM programme_master WHERE plant_id=? ORDER BY prog_type, name',
            (plant_id,)).fetchall()

        # TNI breakdown by programme type + per-type drill (relocated from dashboard)
        tni_by_type = db.execute(
            'SELECT prog_type, COUNT(DISTINCT emp_code || "|" || programme_name) as cnt'
            ' FROM tni WHERE plant_id=? AND fy_year=? AND prog_type IS NOT NULL AND prog_type!=""'
            ' GROUP BY prog_type ORDER BY cnt DESC',
            (plant_id, fy)).fetchall()
        _drill = db.execute(
            'SELECT prog_type, programme_name, COUNT(DISTINCT emp_code) as emp_cnt'
            ' FROM tni WHERE plant_id=? AND fy_year=? AND prog_type IS NOT NULL AND prog_type!=""'
            ' GROUP BY prog_type, programme_name ORDER BY prog_type, emp_cnt DESC',
            (plant_id, fy)).fetchall()
        tni_drill = {}
        for r in _drill:
            tni_drill.setdefault(r['prog_type'], []).append((r['programme_name'], r['emp_cnt']))

        return render_template('tni.html', total=total,
                               employees=emps, programmes=programmes,
                               prog_types=PROG_TYPES, modes=MODES, months=MONTHS_FY,
                               departments=depts, dirty_names=dirty_names,
                               dup_count=dup_count,
                               master_progs=master_progs,
                               tni_by_type=tni_by_type, tni_drill=tni_drill)

    @app.route('/tni/data')
    @spoc_required
    def tni_data():
        plant_id = session['plant_id']
        db       = get_db()
        page     = max(1, int(request.args.get('page', 1)))
        per_page = 30

        q         = request.args.get('q', '').strip()
        collar    = request.args.get('collar', '')
        dept      = request.args.get('dept', '')
        ptype     = request.args.get('type', '')
        mode      = request.args.get('mode', '')
        completed = request.args.get('completed', '')

        where  = ['t.plant_id=?', 't.fy_year=?']
        params = [plant_id, _fy_label()]
        if collar: where.append('e.collar=?');       params.append(collar)
        if dept:   where.append('e.department=?');   params.append(dept)
        if ptype:  where.append('t.prog_type=?');    params.append(ptype)
        if mode:   where.append('t.mode=?');         params.append(mode)
        if q:
            where.append('(COALESCE(e.name,"") LIKE ? OR t.emp_code LIKE ? OR t.programme_name LIKE ?)')
            like = f'%{q}%'; params += [like, like, like]
        _et_exists = 'EXISTS(SELECT 1 FROM emp_training et WHERE et.plant_id=t.plant_id AND et.emp_code=t.emp_code AND et.programme_name=t.programme_name)'
        if completed == 'Yes':     where.append(_et_exists)
        elif completed == 'Pending': where.append(f'NOT {_et_exists}')

        where_clause = ' AND '.join(where)
        join_sql = f'''
            FROM tni t
            LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE {where_clause}
        '''

        total  = db.execute(f'SELECT COUNT(*) {join_sql}', params).fetchone()[0]
        offset = (page - 1) * per_page
        rows_raw = db.execute(
            f'''SELECT t.id, t.emp_code, t.programme_name, t.prog_type, t.mode,
                       t.planned_hours, t.source,
                       e.name, e.collar, e.department,
                       CASE WHEN {_et_exists} THEN 'Yes' ELSE 'Pending' END AS completed
                {join_sql}
                ORDER BY t.id DESC LIMIT ? OFFSET ?''',
            params + [per_page, offset]
        ).fetchall()

        rows = [{
            'id':             r['id'],
            'emp_code':       r['emp_code'],
            'name':           r['name'] or r['emp_code'],
            'collar':         r['collar'] or '',
            'department':     r['department'] or '',
            'programme_name': r['programme_name'],
            'prog_type':      r['prog_type'] or '',
            'mode':           r['mode'] or '',
            'planned_hours':  r['planned_hours'],
            'source':         r['source'] or 'TNI Driven',
            'completed':      r['completed'],
            'delete_url':     url_for('delete_tni', tni_id=r['id']),
        } for r in rows_raw]

        return jsonify({'total': total, 'page': page, 'per_page': per_page, 'rows': rows})

    @app.route('/tni/export')
    @spoc_required
    def tni_export():
        """Export current filtered TNI view as Excel. Honors q/collar/dept/type/mode/completed."""
        plant_id = session['plant_id']
        db       = get_db()
        q         = request.args.get('q', '').strip()
        collar    = request.args.get('collar', '')
        dept      = request.args.get('dept', '')
        ptype     = request.args.get('type', '')
        mode      = request.args.get('mode', '')
        completed = request.args.get('completed', '')

        where  = ['t.plant_id=?', 't.fy_year=?']
        params = [plant_id, _fy_label()]
        if collar: where.append('e.collar=?');       params.append(collar)
        if dept:   where.append('e.department=?');   params.append(dept)
        if ptype:  where.append('t.prog_type=?');    params.append(ptype)
        if mode:   where.append('t.mode=?');         params.append(mode)
        if q:
            where.append('(COALESCE(e.name,"") LIKE ? OR t.emp_code LIKE ? OR t.programme_name LIKE ?)')
            like = f'%{q}%'; params += [like, like, like]
        _et_exists = 'EXISTS(SELECT 1 FROM emp_training et WHERE et.plant_id=t.plant_id AND et.emp_code=t.emp_code AND et.programme_name=t.programme_name)'
        if completed == 'Yes':     where.append(_et_exists)
        elif completed == 'Pending': where.append(f'NOT {_et_exists}')

        rows = db.execute(
            f'''SELECT t.emp_code, e.name, e.collar, e.department,
                       t.programme_name, t.prog_type, t.mode, t.planned_hours, t.source,
                       CASE WHEN {_et_exists} THEN 'Yes' ELSE 'Pending' END AS completed
                FROM tni t
                LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
                WHERE {' AND '.join(where)}
                ORDER BY e.name, t.programme_name''', params).fetchall()

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'TNI Export'
        headers = ['Emp Code','Name','Collar','Department','Programme','Type','Mode','Planned Hours','Source','Completed']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1E3A5F')
            c.alignment = Alignment(horizontal='center')
        for ri, r in enumerate(rows, 2):
            ws.cell(row=ri, column=1, value=r['emp_code'])
            ws.cell(row=ri, column=2, value=r['name'])
            ws.cell(row=ri, column=3, value=r['collar'])
            ws.cell(row=ri, column=4, value=r['department'])
            ws.cell(row=ri, column=5, value=r['programme_name'])
            ws.cell(row=ri, column=6, value=r['prog_type'])
            ws.cell(row=ri, column=7, value=r['mode'])
            ws.cell(row=ri, column=8, value=r['planned_hours'])
            ws.cell(row=ri, column=9, value=r['source'])
            ws.cell(row=ri, column=10, value=r['completed'])
        for col_idx in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fy = _fy_label()
        return send_file(buf, as_attachment=True,
                         download_name=f'TNI_{session.get("unit_code","plant")}_{fy}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/tni/add', methods=['POST'])
    @spoc_required
    def add_tni():
        plant_id = session['plant_id']
        if _tni_is_locked() and session.get('role') != 'admin':
            flash('TNI window is closed (FY ended March 31). Contact admin to submit an override request.', 'danger')
            return redirect(url_for('tni'))
        f = request.form
        db = get_db()

        emp_code      = f.get('emp_code', '').strip()
        prog_name_raw = f.get('programme_name', '').strip()
        prog_type     = f.get('prog_type', '').strip()
        mode          = f.get('mode', '').strip()
        raw_source = f.get('source', 'TNI Driven').strip()
        source     = raw_source if raw_source in ('TNI Driven', 'New Requirement') else 'TNI Driven'

        if not emp_code or not prog_name_raw:
            flash('Employee Code and Programme Name are required.', 'danger')
            return redirect(url_for('tni'))

        emp = db.execute('SELECT 1 FROM employees WHERE emp_code=? AND plant_id=? AND is_active=1',
                         (emp_code, plant_id)).fetchone()
        if not emp:
            flash(f'Employee code "{emp_code}" not found in active employees for this plant.', 'danger')
            return redirect(url_for('tni'))

        if prog_type and prog_type not in PROG_TYPES:
            flash(f'Invalid programme type "{prog_type}". Choose from: {", ".join(PROG_TYPES)}.', 'danger')
            return redirect(url_for('tni'))
        if mode and mode not in MODES:
            flash(f'Invalid mode "{mode}". Choose from: {", ".join(MODES)}.', 'danger')
            return redirect(url_for('tni'))

        try:
            planned_hours = float(f.get('planned_hours') or 0)
            if planned_hours < 0:
                raise ValueError
        except ValueError:
            flash('Planned hours must be a non-negative number (e.g. 4 or 2.5).', 'danger')
            return redirect(url_for('tni'))

        auto_add = f.get('auto_add_to_master') == '1'
        # Canonicalize against master ∪ existing TNI names so a spelling variant
        # collapses onto a programme already present in TNI (not just master).
        prog_name = _canonical_prog(prog_name_raw, plant_id, db,
                                    _master=_tni_canon_candidates(plant_id, db))
        in_master = db.execute(
            'SELECT 1 FROM programme_master WHERE plant_id=? AND LOWER(name)=LOWER(?)',
            (plant_id, prog_name)).fetchone()
        if not in_master:
            if not auto_add:
                flash(f'"{prog_name}" is not in Programme Master. Add it to Programme Master first, or use the "Add to Programme Master & Continue" option.', 'danger')
                return redirect(url_for('tni'))
            source = 'New Requirement'
            db.execute(
                'INSERT OR IGNORE INTO programme_master(plant_id, name, prog_type, source) VALUES(?,?,?,?)',
                (plant_id, prog_name, prog_type or None, 'New Requirement'))
            db.commit()

        cur = db.execute(
            '''INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours,source,fy_year)
               VALUES(?,?,?,?,?,?,?,?)''',
            (plant_id, emp_code, prog_name, prog_type, mode, planned_hours, source, _fy_label())
        )
        if cur.rowcount == 0:
            flash(f'TNI entry for "{prog_name}" already exists for employee {emp_code}.', 'warning')
            return redirect(url_for('tni'))
        _sync_master_from_tni(plant_id, db)
        db.commit()
        # Re-derive audience on calendar rows for this programme (TNI changed → audience may shift)
        resync_calendar_audience(plant_id, [prog_name], db)
        log_action('RECORD_ADD', f"tni:{emp_code}:{prog_name}")
        flash('TNI entry added.', 'success')
        return redirect(url_for('tni'))

    @app.route('/tni/quick-add-for-employee', methods=['POST'])
    @spoc_required
    def tni_quick_add_for_employee():
        """Bulk-add multiple TNI rows for ONE employee in a single submit.
        Designed for late-joiner onboarding: pick employee, tick programmes from
        Programme Master, set planned hours (default 4 if blank), submit.
        All programmes inserted as 'TNI Driven'. Skips dupes via UNIQUE constraint."""
        plant_id = session['plant_id']
        if _tni_is_locked() and session.get('role') != 'admin':
            flash('TNI window is closed (FY ended March 31). Contact admin.', 'danger')
            return redirect(url_for('tni'))
        db = get_db()
        emp_code = (request.form.get('emp_code') or '').strip()
        # int-cast prog_ids — reject non-numeric tampering, surface count loss
        raw_prog_ids = request.form.getlist('prog_ids')
        prog_ids = []
        for rid in raw_prog_ids:
            try:
                prog_ids.append(int(rid))
            except (ValueError, TypeError):
                pass
        if raw_prog_ids and len(prog_ids) != len(raw_prog_ids):
            flash(f'{len(raw_prog_ids) - len(prog_ids)} invalid programme id(s) ignored.', 'warning')
        default_hours = request.form.get('default_hours') or '4'
        try:
            default_hours = float(default_hours)
            if default_hours <= 0:
                flash('Default hours was 0 or negative — coerced to 4.', 'warning')
                default_hours = 4.0
        except (ValueError, TypeError):
            default_hours = 4.0

        if not emp_code:
            flash('Pick an employee first.', 'danger')
            return redirect(url_for('tni'))
        if not prog_ids:
            flash('Select at least one programme.', 'danger')
            return redirect(url_for('tni'))

        emp = db.execute(
            'SELECT 1 FROM employees WHERE emp_code=? AND plant_id=? AND is_active=1',
            (emp_code, plant_id)).fetchone()
        if not emp:
            flash(f'Employee "{emp_code}" not found in active employees.', 'danger')
            return redirect(url_for('tni'))

        # Lookup chosen programmes by id (from programme_master)
        placeholders = ','.join(['?'] * len(prog_ids))
        progs = db.execute(
            f'SELECT id, name, prog_type FROM programme_master '
            f'WHERE plant_id=? AND id IN ({placeholders})',
            [plant_id] + prog_ids).fetchall()

        added, dupes = 0, 0
        fy = _fy_label()
        for p in progs:
            cur = db.execute(
                '''INSERT OR IGNORE INTO tni
                   (plant_id, emp_code, programme_name, prog_type, mode, planned_hours, source, fy_year)
                   VALUES(?,?,?,?,?,?,?,?)''',
                (plant_id, emp_code, p['name'], p['prog_type'], None,
                 default_hours, 'TNI Driven', fy))
            if cur.rowcount:
                added += 1
            else:
                dupes += 1
        db.commit()
        # Re-derive audience for any calendar rows that referenced these programmes
        if added:
            resync_calendar_audience(plant_id, [p['name'] for p in progs], db)
            _prog_list = ','.join(str(p['id']) for p in progs)[:300]
            log_action('BULK_ADD', f"tni-quick-for-emp:{emp_code}:added={added}:progs=[{_prog_list}]")
        msg = f'{added} TNI entries added for {emp_code}.'
        if dupes:
            msg += f' ({dupes} already existed — skipped)'
        flash(msg, 'success' if added else 'warning')
        return redirect(url_for('tni'))

    @app.route('/tni/<int:tni_id>/set-source', methods=['POST'])
    @spoc_required
    def tni_set_source(tni_id):
        plant_id = session['plant_id']
        data     = request.get_json(silent=True) or {}
        source   = (data.get('source') or '').strip()
        if source not in NON_TNI_SOURCES:
            return jsonify({'ok': False, 'error': 'Invalid source'}), 400
        db  = get_db()
        row = db.execute('SELECT programme_name FROM tni WHERE id=? AND plant_id=?',
                         (tni_id, plant_id)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': 'Not found'}), 404
        db.execute('UPDATE tni SET source=? WHERE id=? AND plant_id=?',
                   (source, tni_id, plant_id))
        db.execute('INSERT OR IGNORE INTO programme_master(plant_id,name) VALUES(?,?)',
                   (plant_id, row['programme_name']))
        db.commit()
        log_action('RECORD_EDIT', f"tni_source:{tni_id}:{source}")
        return jsonify({'ok': True})

    @app.route('/tni/<int:tni_id>/delete', methods=['POST'])
    @spoc_required
    def delete_tni(tni_id):
        if _tni_is_locked() and session.get('role') != 'admin':
            flash('TNI window is closed. Contact admin to submit an override request.', 'danger')
            return redirect(url_for('tni'))
        db = get_db()
        affected = db.execute('SELECT programme_name FROM tni WHERE id=? AND plant_id=?',
                              (tni_id, session['plant_id'])).fetchone()
        db.execute('DELETE FROM tni WHERE id=? AND plant_id=?', (tni_id, session['plant_id']))
        db.commit()
        if affected:
            resync_calendar_audience(session['plant_id'], [affected['programme_name']], db)
        log_action('RECORD_DELETE', f"tni:{tni_id}")
        if _is_ajax():
            return '', 204
        flash('TNI entry deleted.', 'warning')
        return redirect(url_for('tni'))

    @app.route('/tni/bulk-delete', methods=['POST'])
    @spoc_required
    def tni_bulk_delete():
        plant_id = session['plant_id']
        if _tni_is_locked() and session.get('role') != 'admin':
            flash('TNI window is closed. Contact admin to submit an override request.', 'danger')
            return redirect(url_for('tni'))
        db = get_db()
        if request.form.get('select_all') == '1':
            q         = request.form.get('q', '').strip()
            collar    = request.form.get('collar', '')
            dept      = request.form.get('dept', '')
            ptype     = request.form.get('type', '')
            mode      = request.form.get('mode', '')
            completed = request.form.get('completed', '')
            where  = ['t.plant_id=?']
            params = [plant_id]
            if collar: where.append('e.collar=?');     params.append(collar)
            if dept:   where.append('e.department=?'); params.append(dept)
            if ptype:  where.append('t.prog_type=?');  params.append(ptype)
            if mode:   where.append('t.mode=?');       params.append(mode)
            if q:
                where.append('(COALESCE(e.name,"") LIKE ? OR t.emp_code LIKE ? OR t.programme_name LIKE ?)')
                like = f'%{q}%'; params += [like, like, like]
            _et_ex = 'EXISTS(SELECT 1 FROM emp_training et WHERE et.plant_id=t.plant_id AND et.emp_code=t.emp_code AND et.programme_name=t.programme_name)'
            if completed == 'Yes':       where.append(_et_ex)
            elif completed == 'Pending': where.append(f'NOT {_et_ex}')
            join_sql = f'FROM tni t LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id WHERE {" AND ".join(where)}'
            count = db.execute(f'SELECT COUNT(*) {join_sql}', params).fetchone()[0]
            if count:
                affected_progs = [r[0] for r in db.execute(
                    f'SELECT DISTINCT t.programme_name {join_sql}', params).fetchall()]
                db.execute(f'DELETE FROM tni WHERE id IN (SELECT t.id {join_sql})', params)
                db.commit()
                resync_calendar_audience(plant_id, affected_progs, db)
                log_action('BULK_DELETE', f"tni:{count}")
                flash(f'{count} TNI entries deleted.', 'warning')
        else:
            ids = request.form.getlist('ids[]')
            if ids:
                deleted = 0
                affected_progs = set()
                for i in range(0, len(ids), 900):
                    chunk = ids[i:i+900]
                    ph = ','.join('?' * len(chunk))
                    for r in db.execute(
                        f'SELECT DISTINCT programme_name FROM tni WHERE id IN ({ph}) AND plant_id=?',
                        chunk + [plant_id]).fetchall():
                        affected_progs.add(r[0])
                    db.execute(f'DELETE FROM tni WHERE id IN ({ph}) AND plant_id=?', chunk + [plant_id])
                    deleted += len(chunk)
                db.commit()
                resync_calendar_audience(plant_id, affected_progs, db)
                log_action('BULK_DELETE', f"tni:{deleted}")
                flash(f'{deleted} TNI entries deleted.', 'warning')
        return redirect(url_for('tni'))

    @app.route('/tni/cleanse', methods=['GET', 'POST'])
    @spoc_required
    def tni_cleanse():
        plant_id = session['plant_id']
        db = get_db()

        if request.args.get('quick') == '1':
            result = _cleanse_programme_names(db, plant_id=plant_id)
            r = result.get(plant_id, {'fixed': 0, 'merged': 0})
            log_action('BULK_UPDATE', f"tni_cleanse:fixed={r['fixed']}_merged={r['merged']}")
            flash(f'Data cleanse complete: {r["fixed"]} programme name(s) corrected, '
                  f'{r["merged"]} duplicate(s) merged.', 'success')
            return redirect(url_for('tni'))

        if request.method == 'POST':
            result = _cleanse_programme_names(db, plant_id=plant_id)
            r = result.get(plant_id, {'fixed': 0, 'merged': 0})
            log_action('BULK_UPDATE', f"tni_cleanse:fixed={r['fixed']}_merged={r['merged']}")
            flash(f'Data cleanse complete: {r["fixed"]} programme name(s) corrected, '
                  f'{r["merged"]} duplicate(s) merged.', 'success')
            return redirect(url_for('tni'))

        from difflib import get_close_matches as gcm
        master = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)).fetchall()]
        preview = []
        if master:
            master_lower_map = {m.lower(): m for m in master}
            master_lower = list(master_lower_map.keys())
            seen = set()
            rows = db.execute(
                'SELECT DISTINCT programme_name FROM tni WHERE plant_id=? AND fy_year=?', (plant_id, _fy_label())).fetchall()
            for row in rows:
                raw = row['programme_name'] or ''
                if not raw or raw in seen:
                    continue
                seen.add(raw)
                raw_lower = raw.lower()
                if raw_lower in master_lower_map:
                    canonical = master_lower_map[raw_lower]
                    if canonical != raw:
                        preview.append({'original': raw, 'fixed': canonical, 'how': 'Case correction'})
                else:
                    m = gcm(raw_lower, master_lower, n=1, cutoff=0.88)
                    if m:
                        canonical = master_lower_map[m[0]]
                        preview.append({'original': raw, 'fixed': canonical, 'how': 'Spelling correction'})
        return render_template('tni_cleanse.html', preview=preview)

    @app.route('/tni/duplicates')
    @spoc_required
    def tni_duplicates():
        plant_id = session['plant_id']
        db = get_db()
        fy   = _fy_label()
        rows = db.execute('''
            SELECT t.emp_code,
                   MAX(e.name) as emp_name,
                   t.programme_name,
                   COUNT(*) as cnt,
                   GROUP_CONCAT(t.mode, ' / ') as modes,
                   GROUP_CONCAT(t.id) as ids
            FROM tni t
            LEFT JOIN employees e ON e.emp_code=t.emp_code AND e.plant_id=t.plant_id
            WHERE t.plant_id=? AND t.fy_year=?
            GROUP BY t.emp_code, t.programme_name
            HAVING cnt > 1
            ORDER BY cnt DESC, emp_name
        ''', (plant_id, fy)).fetchall()
        total_extra = sum(r['cnt'] - 1 for r in rows)
        return render_template('tni_duplicates.html', rows=rows, total_extra=total_extra)

    @app.route('/tni/duplicates/delete', methods=['POST'])
    @spoc_required
    def tni_duplicates_delete():
        plant_id = session['plant_id']
        db = get_db()
        deleted = 0
        fy   = _fy_label()
        rows = db.execute('''
            SELECT GROUP_CONCAT(id ORDER BY id) as ids
            FROM tni
            WHERE plant_id=? AND fy_year=?
            GROUP BY emp_code, programme_name
            HAVING COUNT(*) > 1
        ''', (plant_id, fy)).fetchall()
        for r in rows:
            id_list = [int(x) for x in r['ids'].split(',')]
            keep = id_list[0]
            remove = id_list[1:]
            ph = ','.join('?' * len(remove))
            db.execute(f'DELETE FROM tni WHERE id IN ({ph}) AND plant_id=?', remove + [plant_id])
            deleted += len(remove)
        db.commit()
        log_action('BULK_DELETE', f"tni_dups:{deleted}")
        flash(f'{deleted} duplicate TNI entries removed.', 'success')
        return redirect(url_for('tni'))

    @app.route('/tni/template')
    @spoc_required
    def tni_template():
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.workbook.defined_name import DefinedName

        plant_id = session['plant_id']
        db       = get_db()
        emps = db.execute(
            'SELECT emp_code, name FROM employees WHERE plant_id=? AND is_active=1 ORDER BY name',
            (plant_id,)).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'TNI Data'

        headers  = ['Employee Code', 'Employee Name (auto)', 'Programme Name',
                    'Type of Programme', 'Mode', 'Target Month', 'Planned Hours']
        hdr_fill = PatternFill('solid', fgColor='1F4E79')
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        widths = [18, 28, 36, 24, 14, 16, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        ws_emp = wb.create_sheet('_EmpList')
        ws_emp.sheet_state = 'hidden'
        for r, emp in enumerate(emps, 1):
            c = ws_emp.cell(row=r, column=1, value=str(emp['emp_code']))
            c.number_format = '@'
            ws_emp.cell(row=r, column=2, value=emp['name'])
        emp_count = len(emps)

        ws_vals = wb.create_sheet('_ValidValues')
        ws_vals.sheet_state = 'hidden'
        for r, v in enumerate(PROG_TYPES, 1): ws_vals.cell(row=r, column=1, value=v)
        for r, v in enumerate(MODES, 1):      ws_vals.cell(row=r, column=2, value=v)
        for r, v in enumerate(MONTHS_FY, 1):  ws_vals.cell(row=r, column=3, value=v)

        master_progs = [r[0] for r in db.execute(
            'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
        ).fetchall()] or []
        ws_prog = wb.create_sheet('_ProgList')
        ws_prog.sheet_state = 'hidden'
        for r, v in enumerate(master_progs, 1):
            ws_prog.cell(row=r, column=1, value=v)

        if emp_count:
            wb.defined_names['EmpCodes'] = DefinedName(
                'EmpCodes', attr_text=f'_EmpList!$A$1:$A${emp_count}')
        if master_progs:
            wb.defined_names['ProgList'] = DefinedName(
                'ProgList', attr_text=f'_ProgList!$A$1:$A${len(master_progs)}')

        max_rows = 2000
        if emp_count:
            dv_emp = DataValidation(type='list', formula1='EmpCodes', allow_blank=False,
                                    showErrorMessage=True, errorTitle='Invalid Employee',
                                    error='Select a valid Employee Code from the dropdown.',
                                    showDropDown=False)
            dv_emp.sqref = f'A2:A{max_rows}'
            ws.add_data_validation(dv_emp)
        if master_progs:
            dv_prog = DataValidation(type='list', formula1='ProgList', allow_blank=False,
                                     showErrorMessage=True, errorTitle='Programme Not in Master',
                                     error='Select a programme from the dropdown. If missing, add it to Programme Master first.',
                                     showDropDown=False)
            dv_prog.sqref = f'C2:C{max_rows}'
            ws.add_data_validation(dv_prog)

        dv_type = DataValidation(type='list', formula1=f'"{",".join(PROG_TYPES)}"',
                                 allow_blank=False, showErrorMessage=True,
                                 errorTitle='Invalid Type', error='Select from: ' + ', '.join(PROG_TYPES),
                                 showDropDown=False)
        dv_type.sqref = f'D2:D{max_rows}'
        ws.add_data_validation(dv_type)

        dv_mode = DataValidation(type='list', formula1=f'"{",".join(MODES)}"',
                                 allow_blank=True, showErrorMessage=True,
                                 errorTitle='Invalid Mode', error='Select from: ' + ', '.join(MODES),
                                 showDropDown=False)
        dv_mode.sqref = f'E2:E{max_rows}'
        ws.add_data_validation(dv_mode)

        dv_month = DataValidation(type='list', formula1=f'"{",".join(MONTHS_FY)}"',
                                  allow_blank=True, showErrorMessage=True,
                                  errorTitle='Invalid Month', error='Select a valid month.',
                                  showDropDown=False)
        dv_month.sqref = f'F2:F{max_rows}'
        ws.add_data_validation(dv_month)

        dv_hrs = DataValidation(type='decimal', operator='greaterThan', formula1='0',
                                allow_blank=True, showErrorMessage=True,
                                errorTitle='Invalid Hours', error='Enter a number > 0, e.g. 4 or 2.5')
        dv_hrs.sqref = f'G2:G{max_rows}'
        ws.add_data_validation(dv_hrs)

        for r in range(2, max_rows + 1):
            ws.cell(row=r, column=2).value = (
                f'=IF(A{r}="","",IFERROR(VLOOKUP(TEXT(A{r},"0"),_EmpList!$A:$B,2,0),"Not Found"))')
            ws.cell(row=r, column=2).font  = Font(color='1F4E79', italic=True)
            ws.cell(row=r, column=2).fill  = PatternFill('solid', fgColor='EFF6FF')
            ws.cell(row=r, column=7).value = 0

        ws_info = wb.create_sheet('Instructions')
        ws_info.cell(row=1, column=1, value='Instructions').font = Font(bold=True, size=12)
        ws_info.cell(row=2, column=1, value='• Fill data in the "TNI Data" sheet only — rows 2 onwards.')
        ws_info.cell(row=3, column=1, value='• Do NOT add or delete columns.')
        ws_info.cell(row=4, column=1, value='• Column B (Employee Name) auto-fills from the Employee Code you enter in Column A.')
        ws_info.cell(row=5, column=1, value='• Do NOT delete the hidden sheets (_EmpList, _ValidValues, _ProgList).')
        ws_info.column_dimensions['A'].width = 80

        out = io.BytesIO()
        wb.save(out); out.seek(0)
        return send_file(out, download_name='TNI_Bulk_Upload_Template.xlsx', as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/tni/bulk', methods=['POST'])
    @spoc_required
    def tni_bulk_upload():
        plant_id = session['plant_id']
        if _tni_is_locked() and session.get('role') != 'admin':
            flash('TNI window is closed (FY ended March 31). Contact admin to submit an override request.', 'danger')
            return redirect(url_for('tni'))
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('tni'))
        try:
            df = _read_upload_file(f)
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('tni'))

        db = get_db(); inserted = 0; errors = []
        affected_progs = set()
        active_emps = {r[0] for r in db.execute(
            'SELECT emp_code FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)
        ).fetchall()}
        fy = _fy_label()
        # Canonical candidates = master ∪ existing TNI names. GROWS as new programmes
        # are accepted below, so a 2nd spelling variant later in THIS file collapses
        # onto the first one (closes the stale-snapshot duplicate leak).
        master_cache = _tni_canon_candidates(plant_id, db, fy)
        _cand_lower = {n.lower() for n in master_cache}
        for i, row in enumerate(df.to_dict('records')):
            emp_code  = _clean(row, ['employee code', 'emp code', 'empcode', 'employee_code'])
            prog_name = _clean(row, ['programme name', 'program name', 'programme_name', 'training name'])
            prog_type = _clean(row, ['type of programme', 'type', 'prog type', 'programme type'])
            mode      = _clean(row, ['mode'])
            hours     = _safe_float(_clean(row, ['planned hours', 'hours', 'hrs'])) or 0

            if not emp_code or not prog_name:
                errors.append(f'Row {i+2}: Employee Code and Programme Name are required.')
                continue
            if emp_code not in active_emps:
                errors.append(f'Row {i+2}: Employee {emp_code} not found in your plant.')
                continue
            prog_name = _canonical_prog(prog_name, plant_id, db, _master=master_cache)
            if prog_name and prog_name.lower() not in _cand_lower:
                master_cache.append(prog_name)        # let later rows canonicalize onto it
                _cand_lower.add(prog_name.lower())
            db.execute('INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours,fy_year) VALUES(?,?,?,?,?,?,?)',
                       (plant_id, emp_code, prog_name, prog_type, mode, hours, fy))
            affected_progs.add(prog_name)
            inserted += 1
        _sync_master_from_tni(plant_id, db)
        db.commit()
        resync_calendar_audience(plant_id, affected_progs, db)
        if errors:
            if inserted:
                flash(f'Bulk upload complete: {inserted} TNI entries added. {len(errors)} rows had errors — downloading error report.', 'warning')
            return _error_excel_response(errors, inserted, 'TNI_Upload_Errors.xlsx')
        log_action('BULK_UPLOAD', f"tni:{inserted}")
        flash(f'Bulk upload complete: {inserted} TNI entries added successfully.', 'success')
        return redirect(url_for('tni'))

    @app.route('/tni/fresh-upload', methods=['GET', 'POST'])
    @spoc_required
    def tni_fresh_upload():
        plant_id = session['plant_id']
        if _tni_is_locked() and session.get('role') != 'admin':
            flash('TNI window is closed (FY ended March 31). Only admin can upload TNI after the financial year end.', 'danger')
            return redirect(url_for('tni'))
        db = get_db()

        if request.method == 'GET':
            return render_template('tni_fresh_upload.html')

        confirm_token = request.form.get('confirm')
        if confirm_token:
            ext      = session.get('fresh_upload_ext', '.xlsx')
            tmp_path = os.path.join(TEMP_UPLOAD_DIR, f'tni_fresh_{confirm_token}{ext}')
            if not os.path.exists(tmp_path):
                flash('Session expired — please re-upload the file.', 'danger')
                return redirect(url_for('tni_fresh_upload'))
            try:
                df = _read_upload_file_path(tmp_path)
            except Exception as e:
                flash(f'Could not read file: {e}', 'danger')
                return redirect(url_for('tni_fresh_upload'))

            result = _process_fresh_tni(df, plant_id, db)
            rows   = result['valid_rows']

            fy = _fy_label()
            import datetime as _dt
            archived_at = _dt.datetime.now().isoformat(timespec='seconds')
            db.execute('''
                INSERT INTO tni_archive
                    (archive_token, archived_at, plant_id, emp_code, programme_name,
                     prog_type, mode, target_month, planned_hours, source, fy_year)
                SELECT ?, ?, plant_id, emp_code, programme_name,
                       prog_type, mode, target_month, planned_hours, source, fy_year
                FROM tni WHERE plant_id=? AND fy_year=?
            ''', (confirm_token, archived_at, plant_id, fy))
            archived_count = db.execute(
                'SELECT COUNT(*) FROM tni_archive WHERE archive_token=?', (confirm_token,)
            ).fetchone()[0]
            # Save New Requirement programmes before wiping TNI-sourced master entries
            new_req_progs = db.execute(
                "SELECT name, prog_type, mode FROM programme_master WHERE plant_id=? AND source='New Requirement'",
                (plant_id,)).fetchall()
            db.execute('DELETE FROM tni WHERE plant_id=? AND fy_year=?', (plant_id, fy))
            db.execute("DELETE FROM programme_master WHERE plant_id=? AND source='TNI Requirement'", (plant_id,))
            for r in rows:
                db.execute(
                    'INSERT INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours,fy_year) VALUES(?,?,?,?,?,?,?)',
                    (plant_id, r['emp_code'], r['programme_name'], r['prog_type'], r['mode'], r['hours'], fy))
            _sync_master_from_tni(plant_id, db)
            # Restore New Requirement programmes that survived
            for nr in new_req_progs:
                db.execute('INSERT OR IGNORE INTO programme_master(plant_id, name, prog_type, mode, source) VALUES(?,?,?,?,?)',
                           (plant_id, nr['name'], nr['prog_type'], nr['mode'], 'New Requirement'))
            db.commit()
            # Full TNI replacement → audience for every calendar row may have shifted.
            # Resync across all programmes touched by the new upload.
            resync_calendar_audience(plant_id, {r['programme_name'] for r in rows}, db)
            try: os.remove(tmp_path)
            except Exception: pass
            log_action('BULK_UPLOAD', f"tni_fresh:{len(rows)}")
            flash(f'Fresh upload complete: {len(rows)} TNI entries saved. '
                  f'{len(result["unique_progs"])} unique programmes are now your master list. '
                  f'Previous {archived_count} rows archived (ref: {confirm_token[:8]})', 'success')
            return redirect(url_for('tni'))

        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('tni_fresh_upload'))

        os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)
        token    = secrets.token_hex(16)
        ext      = os.path.splitext(f.filename)[1].lower() or '.xlsx'
        tmp_path = os.path.join(TEMP_UPLOAD_DIR, f'tni_fresh_{token}{ext}')
        f.save(tmp_path)
        session['fresh_upload_token'] = token
        session['fresh_upload_ext']   = ext

        try:
            df = _read_upload_file_path(tmp_path)
        except Exception as e:
            try: os.remove(tmp_path)
            except Exception: pass
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('tni_fresh_upload'))

        result = _process_fresh_tni(df, plant_id, db)
        return render_template('tni_fresh_upload.html', preview=True, token=token,
                               total_rows=result['total_rows'], valid_rows=result['valid_rows'],
                               error_rows=result['error_rows'],
                               name_corrections=result['name_corrections'],
                               unique_progs=result['unique_progs'],
                               duplicate_count=result['duplicate_count'])

    @app.route('/tni/msforms', methods=['GET'])
    @spoc_required
    def tni_msforms():
        plant_id   = session['plant_id']
        plant_name = session.get('plant_name', '')
        db = get_db()
        emp_count = db.execute(
            'SELECT COUNT(*) FROM employees WHERE plant_id=? AND is_active=1', (plant_id,)).fetchone()[0]
        return render_template('tni_msforms.html', plant_name=plant_name,
                               emp_count=emp_count, prog_types=PROG_TYPES,
                               modes=MODES, months=MONTHS_FY)

    @app.route('/tni/msforms/import', methods=['POST'])
    @spoc_required
    def tni_msforms_import():
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('tni_msforms'))
        plant_id = session['plant_id']
        db = get_db()
        try:
            inserted, errors = _parse_msforms_excel(f, plant_id, db)
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('tni_msforms'))
        if errors:
            return _error_excel_response(errors, inserted, 'MSForms_Import_Errors.xlsx')
        log_action('BULK_UPLOAD', f"tni_msforms:{inserted}")
        flash(f'Microsoft Forms import complete: {inserted} TNI entries added.', 'success')
        return redirect(url_for('tni'))

    @app.route('/tni/analyze', methods=['GET', 'POST'])
    @spoc_required
    def tni_analyze():
        if request.method == 'GET':
            return render_template('tni_analyze.html', step='upload')

        import traceback as _tb, logging as _log
        try:
            f = request.files.get('file')
            if not f or f.filename == '':
                flash('No file selected.', 'danger')
                return render_template('tni_analyze.html', step='upload')
            try:
                skip = int(request.form.get('skip_rows') or 0)
            except (ValueError, TypeError):
                skip = 0
            # Stream rows via openpyxl/csv instead of loading entire file via
            # pandas — pandas DataFrame load was OOM-killing the gunicorn worker
            # on 5000+ row uploads (Render Starter 512MB).
            try:
                from tms.helpers import _stream_input_rows
                columns, row_iter = _stream_input_rows(f, skip_rows=skip)
            except Exception as e:
                flash(f'Could not read file: {e}', 'danger')
                return render_template('tni_analyze.html', step='upload')

            plant_id = session['plant_id']
            db       = get_db()
            try:
                rows = _smart_analyze_rows(row_iter, plant_id, db, columns=columns)
            except Exception as e:
                _log.error('_smart_analyze_rows error:\n' + _tb.format_exc())
                flash(f'Analysis error: {e}', 'danger')
                return render_template('tni_analyze.html', step='upload')

            if not rows:
                col_list = ', '.join(f'"{c}"' for c in columns[:15])
                flash(f'No data rows found. Columns detected: {col_list}. '
                      f'Try increasing "Skip top rows" if headers are not on row 1.', 'warning')
                return render_template('tni_analyze.html', step='upload')

            master_progs = [r[0] for r in db.execute(
                'SELECT name FROM programme_master WHERE plant_id=? ORDER BY name', (plant_id,)
            ).fetchall()] or []

            for r in rows:
                r['row_num']   = int(r['row_num']) if r.get('row_num') is not None else 0
                if r.get('planned_hours') is None:
                    r['planned_hours'] = 0.0

            aid = str(_uuid.uuid4())
            os.makedirs(os.path.join(BASE_DIR, 'data'), exist_ok=True)
            with open(os.path.join(BASE_DIR, 'data', f'tni_analyze_{aid}.json'), 'w') as fp:
                _json.dump(rows, fp, default=str)

            ok_count    = sum(1 for r in rows if r['status'] == 'ok')
            fixed_count = sum(1 for r in rows if r['status'] == 'fixed')
            err_count   = sum(1 for r in rows if r['status'] == 'error')

            upload_progs_lower = set(
                str(r['programme_name']).lower() for r in rows
                if r['status'] in ('ok', 'fixed') and r.get('programme_name'))

            # Compact review data for client-side virtual rendering.
            # Shipping inline-styled HTML for 5000+ rows freezes the browser;
            # JSON is ~10× smaller and JS only renders visible page.
            review_data = []
            for r in rows:
                if r['status'] == 'ok':
                    continue
                review_data.append({
                    'n':  r['row_num'],
                    's':  r['status'],
                    'ec': r.get('emp_code') or '',
                    'en': r.get('emp_name') or '',
                    'pn': r.get('programme_name') or '',
                    'pt': r.get('prog_type') or '',
                    'm':  r.get('mode') or '',
                    'h':  r.get('planned_hours') or 0,
                    'fx': r.get('fixes') or [],
                    'is': r.get('issues') or [],
                    'gc': r.get('prog_garbage_class') or '',
                })

            return render_template('tni_analyze.html', step='review',
                                   review_data=review_data, aid=aid,
                                   ok_count=ok_count, fixed_count=fixed_count,
                                   err_count=err_count,
                                   master_progs=master_progs,
                                   upload_progs_lower=upload_progs_lower,
                                   prog_types=PROG_TYPES, modes=MODES, months=MONTHS_FY)

        except Exception as e:
            _log.error('tni_analyze POST unhandled:\n' + _tb.format_exc())
            flash(f'Unexpected error: {e}', 'danger')
            return render_template('tni_analyze.html', step='upload')

    @app.route('/tni/analyze/confirm', methods=['POST'])
    @spoc_required
    def tni_analyze_confirm():
        plant_id = session['plant_id']
        if _tni_is_locked() and session.get('role') != 'admin':
            flash('TNI window is closed (FY ended March 31). Contact admin to submit an override request.', 'danger')
            return redirect(url_for('tni'))
        aid  = request.form.get('aid', '')
        path = os.path.join(BASE_DIR, 'data', f'tni_analyze_{aid}.json')
        if not aid or not os.path.exists(path):
            flash('Session expired. Please re-upload.', 'danger')
            return redirect(url_for('tni_analyze'))

        with open(path) as fp:
            rows = _json.load(fp)

        plant_id = session['plant_id']
        db       = get_db()
        inserted = 0

        # Legacy support: tolerate fix_prog_* fields if they ever arrive (e.g. an
        # old browser tab with cached HTML), but the analyzer no longer emits
        # 'warning' rows so corrections is normally empty.
        corrections = {}
        for k, v in request.form.items():
            if k.startswith('fix_prog_') and v.strip():
                try:
                    corrections[int(k[9:])] = v.strip()
                except ValueError:
                    pass

        err_rows = [r for r in rows if r['status'] == 'error']

        fy = _fy_label()
        existing = set()
        for er in db.execute('SELECT emp_code, programme_name FROM tni WHERE plant_id=? AND fy_year=?', (plant_id, fy)):
            existing.add((er['emp_code'].strip().upper(), er['programme_name'].strip().lower()))

        dup_rows = []; updated = 0; seen_batch = {}
        affected_progs = set()

        for row in rows:
            if row['status'] == 'error':
                continue
            # Honour any manual override that arrived from an older review page,
            # otherwise just use the analyzer's cleaned name.
            fix = corrections.get(row['row_num'])
            if fix and fix != '__new__':
                prog_name = fix
            else:
                prog_name = row['programme_name']

            key = (row['emp_code'].strip().upper(), prog_name.strip().lower())
            if key in seen_batch:
                dup_rows.append({**row, 'programme_name': prog_name,
                    'dup_type': f'Same employee+programme already at Row {seen_batch[key]} in this file — first entry was imported, this row skipped'})
                continue
            seen_batch[key] = row['row_num']
            if key in existing:
                db.execute('''UPDATE tni SET prog_type=?, mode=?, planned_hours=?
                              WHERE plant_id=? AND fy_year=? AND UPPER(emp_code)=? AND LOWER(programme_name)=?''',
                           (row['prog_type'], row['mode'], row['planned_hours'],
                            plant_id, fy, row['emp_code'].upper(), prog_name.lower()))
                updated += 1
            else:
                db.execute('INSERT OR IGNORE INTO tni(plant_id,emp_code,programme_name,prog_type,mode,planned_hours,fy_year) VALUES(?,?,?,?,?,?,?)',
                           (plant_id, row['emp_code'], prog_name,
                            row['prog_type'], row['mode'], row['planned_hours'], fy))
                inserted += 1
            affected_progs.add(prog_name)

        _sync_master_from_tni(plant_id, db)
        db.commit()
        resync_calendar_audience(plant_id, affected_progs, db)
        try: os.remove(path)
        except OSError: pass
        # Also clean up any other stale analyze files to prevent disk accumulation
        try:
            from tms.helpers import _cleanup_stale_analyze_files
            _cleanup_stale_analyze_files()
        except Exception:
            pass

        log_action('BULK_UPLOAD', f"tni_analyze:new={inserted}_updated={updated}")

        # Persist error + duplicate rows for Central R&D / training analysis
        try:
            uname = session.get('username', 'unknown')
            for er in err_rows:
                db.execute('''INSERT INTO tni_upload_errors(
                    plant_id, username, aid, row_status, row_num,
                    emp_code, emp_name, programme_name, prog_type, mode,
                    planned_hours, issues, garbage_class)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (plant_id, uname, aid, 'error', er.get('row_num'),
                     er.get('emp_code') or '', er.get('emp_name') or '',
                     er.get('programme_name') or '', er.get('prog_type') or '',
                     er.get('mode') or '', er.get('planned_hours') or 0,
                     ' | '.join(er.get('issues') or []),
                     er.get('prog_garbage_class') or ''))
            for dr in dup_rows:
                db.execute('''INSERT INTO tni_upload_errors(
                    plant_id, username, aid, row_status, row_num,
                    emp_code, emp_name, programme_name, prog_type, mode,
                    planned_hours, issues, garbage_class)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (plant_id, uname, aid, 'duplicate', dr.get('row_num'),
                     dr.get('emp_code') or '', dr.get('emp_name') or '',
                     dr.get('programme_name') or '', dr.get('prog_type') or '',
                     dr.get('mode') or '', dr.get('planned_hours') or 0,
                     dr.get('dup_type') or 'duplicate',
                     dr.get('prog_garbage_class') or ''))
            db.commit()
        except Exception as _e:
            import logging as _l
            _l.warning(f'tni_upload_errors persist failed: {_e}')

        if err_rows or dup_rows:
            buf = _error_excel_for_tni(err_rows, dup_rows=dup_rows, plant_id=plant_id, db=db)
            # Persist to disk so user can land on /tni and grab the report from there.
            err_path = os.path.join(BASE_DIR, 'data', f'tni_errors_{aid}.xlsx')
            with open(err_path, 'wb') as _fp:
                _fp.write(buf.getvalue())
            parts = []
            if err_rows:  parts.append(f'{len(err_rows)} errors')
            if dup_rows:  parts.append(f'{len(dup_rows)} duplicates')
            from markupsafe import Markup as _Markup
            dl_url = url_for('tni_errors_download', aid=aid)
            flash(
                _Markup(
                    f'Import complete — {inserted} new + {updated} updated. '
                    f'{" & ".join(parts)} could not be imported. '
                    f'<a href="{dl_url}" class="alert-link"><i class="bi bi-download"></i> Download issues report</a>'
                ),
                'warning'
            )
            return redirect(url_for('tni'))

        flash(f'Import complete — {inserted} new entries added, {updated} existing entries updated. All clean!', 'success')
        return redirect(url_for('tni'))

    @app.route('/tni/analyze/errors/<aid>')
    @spoc_required
    def tni_errors_download(aid):
        # Plant-scoped + simple filename validation to prevent path traversal
        if not aid or not aid.replace('-', '').isalnum():
            flash('Invalid report token.', 'danger')
            return redirect(url_for('tni'))
        path = os.path.join(BASE_DIR, 'data', f'tni_errors_{aid}.xlsx')
        if not os.path.exists(path):
            flash('Report no longer available — re-upload to regenerate.', 'warning')
            return redirect(url_for('tni'))
        return send_file(path, as_attachment=True,
                         download_name='TNI_Import_Issues.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
